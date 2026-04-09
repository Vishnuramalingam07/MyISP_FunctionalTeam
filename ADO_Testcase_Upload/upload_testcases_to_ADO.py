"""
Azure DevOps Test Case Uploader
================================
Reads test cases from an Excel file and uploads them to a specified
ADO Test Plan / Suite, assigns them to a Lead, and links each test case
to a User Story.

Target:
  Organization : accenturecio08
  Project      : AutomationProcess_29697
  Test Plan ID : 4319862
  Test Suite ID: 4330501
  User Story ID: 4345110

Excel columns expected (sheet: "Test Cases"):
  ID | Work Item Type | Title | Test Step | Step Action | Step Expected |
  Area Path | Initiative | TextVerification | Assigned To | State |
  TextVerification1 | Automation status | Tags

Usage:
  1. Set PAT_TOKEN below (or export ADO_PAT environment variable).
  2. Set LEAD_NAME to the full display name / email of the lead.
  3. Set EXCEL_FILE to the path of your Excel workbook.
  4. Run:  python upload_testcases_to_ADO.py
"""

import os
import sys
import json
import base64
import html
import concurrent.futures
import threading
import requests
import openpyxl
from typing import List, Dict, Optional

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION  – update these values before running
# ──────────────────────────────────────────────────────────────────────────────
ORGANIZATION  = "accenturecio08"
PROJECT       = "AutomationProcess_29697"
TEST_PLAN_ID  = 4319862
TEST_SUITE_ID = 4319870
# USER_STORY_ID is now read per-row from the Excel 'User_StoryID' column.
# Set a fallback here only if your Excel does not have that column.
USER_STORY_ID_FALLBACK = 4345110  # used only when the column is absent

# Personal Access Token – use environment variable from ADO_SECRETS.env
from dotenv import load_dotenv
load_dotenv('ADO_SECRETS.env')
PAT_TOKEN = os.getenv("ADO_PAT_MAIN", "")
if not PAT_TOKEN:
    raise ValueError("ADO_PAT_MAIN not found in ADO_SECRETS.env file")

# Full display name OR email of the Lead (leave blank to skip AssignedTo)
LEAD_NAME = ""

# Path to the Excel workbook containing test cases
EXCEL_FILE = r"C:\Users\vishnu.ramalingam\TFS_Upload\US_4328593.xlsx"

# Sheet name inside the workbook
SHEET_NAME = "Test Cases"

# ──────────────────────────────────────────────────────────────────────────────


# ── Derived constants ─────────────────────────────────────────────────────────
BASE_URL        = f"https://dev.azure.com/{ORGANIZATION}/{PROJECT}"
ORG_URL         = f"https://dev.azure.com/{ORGANIZATION}"
WORK_ITEMS_URL  = f"{BASE_URL}/_apis/wit/workitems"
TEST_SUITES_URL = (
    f"{BASE_URL}/_apis/test/plans/{TEST_PLAN_ID}"
    f"/suites/{TEST_SUITE_ID}/testcases"
)

_encoded_pat = base64.b64encode(f":{PAT_TOKEN}".encode()).decode()
HEADERS = {
    "Authorization": f"Basic {_encoded_pat}",
    "Content-Type":  "application/json",
}
PATCH_HEADERS = {
    "Authorization": f"Basic {_encoded_pat}",
    "Content-Type":  "application/json-patch+json",
}
# ─────────────────────────────────────────────────────────────────────────────


def resolve_identity(search_term: str) -> str:
    """
    Search ADO for a user identity matching the given name/email fragment.
    Returns the full 'DisplayName <email>' string ADO accepts for AssignedTo.
    Falls back to the original search_term if no match is found.
    """
    import urllib.parse
    url = (
        f"https://vssps.dev.azure.com/{ORGANIZATION}/_apis/identities"
        f"?searchFilter=General"
        f"&filterValue={urllib.parse.quote(search_term)}"
        f"&queryMembership=None&api-version=7.0"
    )
    resp = requests.get(url, headers=HEADERS, timeout=30)
    if not resp.ok:
        print(f"[WARN] Identity lookup failed (HTTP {resp.status_code}), using name as-is.")
        return search_term

    identities = resp.json().get("value", [])
    if not identities:
        print(f"[WARN] No identity found for '{search_term}', using as-is.")
        return search_term

    for identity in identities:
        if not identity.get("isActive"):
            continue
        display_name = identity.get("providerDisplayName", "")
        props = identity.get("properties", {})
        email = (
            props.get("Mail", {}).get("$value", "")
            or props.get("Account", {}).get("$value", "")
        )
        if display_name and email:
            full_id = f"{display_name} <{email}>"
            print(f"[INFO] Resolved identity: {full_id}")
            return full_id
        if display_name:
            print(f"[INFO] Resolved identity (no email): {display_name}")
            return display_name

    # fallback – use first result's display name
    fallback = identities[0].get("providerDisplayName", search_term)
    print(f"[INFO] Using identity (fallback): {fallback}")
    return fallback


def fetch_user_story_fields(user_story_id: int) -> Dict:
    """
    Fetch AreaPath and Initiative from the target User Story so test cases
    inherit the same values (required by project rules).
    Returns a dict with keys 'area_path' and 'initiative'.
    """
    url = (
        f"{WORK_ITEMS_URL}/{user_story_id}"
        f"?fields=System.AreaPath,Custom.Initiative&api-version=7.0"
    )
    resp = requests.get(url, headers=HEADERS, timeout=30)
    if not resp.ok:
        print(f"[WARN] Could not fetch User Story {user_story_id} fields "
              f"(HTTP {resp.status_code}). Using project defaults.")
        return {"area_path": PROJECT, "initiative": ""}

    fields = resp.json().get("fields", {})
    area   = fields.get("System.AreaPath", PROJECT)
    init   = fields.get("Custom.Initiative", "")
    print(f"[INFO] Inherited from US {user_story_id} -> "
          f"AreaPath='{area}', Initiative='{init}'")
    return {"area_path": area, "initiative": init}


def validate_pat() -> None:
    """Quick connectivity check – verifies PAT is set and the org is reachable."""
    if PAT_TOKEN in ("YOUR_PAT_TOKEN_HERE", "", None):
        print("[ERROR] PAT_TOKEN is not set. "
              "Export the ADO_PAT environment variable or update the script.")
        sys.exit(1)

    url = f"{ORG_URL}/_apis/projects?api-version=7.0"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    if resp.status_code == 401:
        print("[ERROR] Authentication failed – check your PAT token.")
        sys.exit(1)
    resp.raise_for_status()
    print("[OK] Connected to Azure DevOps organisation:", ORGANIZATION)


# ── Excel parsing ─────────────────────────────────────────────────────────────

def _split_steps(text: str) -> List[str]:
    """Split a multi-line numbered step block into individual step strings."""
    if not text:
        return []
    # Remove Excel carriage-return encoding
    text = text.replace("_x000D_", "").replace("\r", "")
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    return lines


def parse_excel(file_path: str, sheet_name: str) -> List[Dict]:
    """
    Parse the Excel workbook and return a list of test-case dicts.

    Supports two formats detected automatically from headers:

    Format A (legacy – multi-row):
      ID | Work Item Type | Title | Test Step | Step Action | Step Expected |
      Area Path | Assigned To | State | Automation status | Tags | ...

    Format B (new – single-row per test case):
      Test Scenarios | Test Case | Action | Expected Result
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] Sheet '{sheet_name}' not found. "
              f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]

    # Resolve column indices from header row
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    def get(row_values, column_name: str, default=""):
        idx = col.get(column_name)
        if idx is None:
            return default
        val = row_values[idx]
        return str(val).strip() if val is not None else default

    test_cases: List[Dict] = []

    # ── Format A: has "Work Item Type" column ─────────────────────────────────
    if "Work Item Type" in col:
        current_tc: Optional[Dict] = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            work_item_type = get(row, "Work Item Type")
            if work_item_type.lower() == "test case":
                if current_tc is not None:
                    test_cases.append(current_tc)
                current_tc = {
                    "title":             get(row, "Title"),
                    "area_path":         get(row, "Area Path"),
                    "assigned_to":       get(row, "Assigned To"),
                    "state":             get(row, "State") or "Design",
                    "automation_status": get(row, "Automation status") or "Not Automated",
                    "tags":              get(row, "Tags"),
                    "steps":             [],
                }
            elif current_tc is not None and get(row, "Test Step"):
                action   = get(row, "Step Action").strip('"').replace('""', '"')
                expected = get(row, "Step Expected").strip('"').replace('""', '"')
                current_tc["steps"].append({"action": action, "expected": expected})
        if current_tc is not None:
            test_cases.append(current_tc)

    # ── Format B: User_StoryID | Test Scenarios | Test Case | Action | Expected Result ──
    elif "Test Case" in col:
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = get(row, "Test Case")
            if not title:
                continue  # skip blank rows

            # Read User Story ID from the column; fall back to config constant
            raw_us = col.get("User_StoryID")
            us_val = row[raw_us] if raw_us is not None else None
            try:
                user_story_id = int(us_val) if us_val is not None else USER_STORY_ID_FALLBACK
            except (TypeError, ValueError):
                user_story_id = USER_STORY_ID_FALLBACK

            action_text   = get(row, "Action")
            expected_text = get(row, "Expected Result")

            action_lines   = _split_steps(action_text)
            expected_lines = _split_steps(expected_text)

            max_len = max(len(action_lines), len(expected_lines), 1)
            steps = []
            for i in range(max_len):
                steps.append({
                    "action":   action_lines[i]   if i < len(action_lines)   else "",
                    "expected": expected_lines[i] if i < len(expected_lines) else "",
                })

            test_cases.append({
                "title":             title,
                "area_path":         PROJECT,
                "assigned_to":       "",
                "state":             "Design",
                "automation_status": "Not Automated",
                "tags":              "",
                "steps":             steps,
                "user_story_id":     user_story_id,
            })

    else:
        print(f"[ERROR] Unrecognised Excel format. Headers found: {headers}")
        sys.exit(1)

    # Format A rows have no user_story_id key — add the fallback
    for tc in test_cases:
        tc.setdefault("user_story_id", USER_STORY_ID_FALLBACK)

    print(f"[INFO] Parsed {len(test_cases)} test case(s) from '{file_path}'.")
    return test_cases


# ── Step XML builder ──────────────────────────────────────────────────────────

def build_steps_xml(steps: List[Dict]) -> str:
    """
    Build the proprietary ADO XML string for test-case steps.
    Both the action and expected strings are HTML-escaped to be safe
    inside XML attributes.
    """
    if not steps:
        return '<steps id="0" last="0"></steps>'

    parts = [f'<steps id="0" last="{len(steps)}">']
    for idx, step in enumerate(steps, start=1):
        action   = html.escape(step.get("action", ""), quote=False)
        expected = html.escape(step.get("expected", ""), quote=False)
        parts.append(
            f'<step id="{idx}" type="ActionStep">'
            f'<parameterizedString isformatted="true">{action}</parameterizedString>'
            f'<parameterizedString isformatted="true">{expected}</parameterizedString>'
            f'<description/>'
            f'</step>'
        )
    parts.append('</steps>')
    return "".join(parts)


# ── ADO REST helpers ──────────────────────────────────────────────────────────

def create_test_case(tc: Dict, lead_name: str, us_fields: Dict,
                     user_story_id: int = 0) -> int:
    """
    Create a new Test Case work item in ADO and return its ID.
    Includes the US relation in the same POST body — eliminates a separate
    PATCH round-trip per test case (saves ~1-2 s per TC).
    """
    steps_xml = build_steps_xml(tc["steps"])
    assigned   = lead_name if lead_name.strip() else tc.get("assigned_to", "")
    area_path  = us_fields.get("area_path") or tc.get("area_path") or PROJECT
    initiative = us_fields.get("initiative") or tc.get("initiative", "")

    body = [
        {"op": "add", "path": "/fields/System.Title",
         "value": tc["title"]},
        {"op": "add", "path": "/fields/System.AreaPath",
         "value": area_path},
        {"op": "add", "path": "/fields/System.State",
         "value": tc["state"]},
        {"op": "add", "path": "/fields/Microsoft.VSTS.TCM.Steps",
         "value": steps_xml},
        {"op": "add", "path": "/fields/Microsoft.VSTS.TCM.AutomationStatus",
         "value": tc["automation_status"]},
    ]

    if initiative:
        body.append({"op": "add",
                     "path": "/fields/Custom.Initiative",
                     "value": initiative})

    if assigned:
        body.append({"op": "add",
                     "path": "/fields/System.AssignedTo",
                     "value": assigned})

    if tc.get("tags"):
        body.append({"op": "add",
                     "path": "/fields/System.Tags",
                     "value": tc["tags"]})

    # Embed the User Story relation in the create call (saves 1 PATCH per TC)
    if user_story_id:
        body.append({
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "Microsoft.VSTS.Common.TestedBy-Reverse",
                "url": f"{ORG_URL}/_apis/wit/workitems/{user_story_id}",
                "attributes": {"comment": f"Tested by User Story {user_story_id}"},
            },
        })

    url  = f"{WORK_ITEMS_URL}/$Test Case?api-version=7.0"
    resp = requests.post(url, headers=PATCH_HEADERS,
                         data=json.dumps(body), timeout=30)
    _raise_for_status(resp, f"create test case '{tc['title']}'")
    tc_id = resp.json()["id"]
    return tc_id


def add_to_suite(test_case_id: int) -> None:
    """Add a single test case to the configured test suite (fallback)."""
    resp = requests.post(
        f"{TEST_SUITES_URL}/{test_case_id}?api-version=7.0",
        headers=HEADERS,
        timeout=30,
    )
    _raise_for_status(resp, f"add TC {test_case_id} to suite {TEST_SUITE_ID}")


def batch_add_to_suite(test_case_ids: list) -> None:
    """Add all test cases to the suite in a single API call — much faster than one-by-one."""
    if not test_case_ids:
        return
    ids_str = ",".join(str(tc_id) for tc_id in test_case_ids)
    resp = requests.post(
        f"{TEST_SUITES_URL}/{ids_str}?api-version=7.0",
        headers=HEADERS,
        timeout=90,
    )
    _raise_for_status(resp, f"batch add {len(test_case_ids)} TCs to suite {TEST_SUITE_ID}")


def link_to_user_story(test_case_id: int, user_story_id: int) -> None:
    """
    Add a 'Tested By' relation on the Test Case pointing to the User Story.
    (rel: Microsoft.VSTS.Common.TestedBy-Reverse — TC -> US direction)
    """
    body = [
        {
            "op":    "add",
            "path":  "/relations/-",
            "value": {
                "rel": "Microsoft.VSTS.Common.TestedBy-Reverse",
                "url": f"{ORG_URL}/_apis/wit/workitems/{user_story_id}",
                "attributes": {"comment": f"Tested by User Story {user_story_id}"},
            },
        }
    ]
    url  = f"{WORK_ITEMS_URL}/{test_case_id}?api-version=7.0"
    resp = requests.patch(url, headers=PATCH_HEADERS,
                          data=json.dumps(body), timeout=30)
    _raise_for_status(resp, f"link TC {test_case_id} -> US {user_story_id}")


def _raise_for_status(resp: requests.Response, context: str) -> None:
    """Raise with a helpful message on non-2xx responses."""
    if not resp.ok:
        try:
            detail = resp.json().get("message", resp.text[:300])
        except Exception:
            detail = resp.text[:300]
        raise RuntimeError(
            f"[FAILED] {context} – HTTP {resp.status_code}: {detail}"
        )

def fetch_suite_user_stories(plan_id: int, suite_id: int) -> set:
    """
    Return the set of User Story IDs linked to suite_id or any of its descendants.

    Uses the LEGACY  _apis/test/plans/{planId}/suites  endpoint (api-version 5.0)
    which reliably includes 'requirementId' for each requirement-based suite in the
    list response – unlike the newer testplan API which omits it from list items.

    A BFS walk from suite_id collects requirementId from all suites in the
    subtree (the target suite + all its nested children at any depth).

    Falls back to the newer testplan API if the legacy call fails.

    Returns an empty set only when NO requirement linkage is found at all,
    in which case the caller skips validation (plain static suite scenario).
    """
    valid_us_ids: set = set()

    # ── Primary: legacy test API (list includes requirementId) ───────────
    legacy_url = (
        f"https://dev.azure.com/{ORGANIZATION}/{PROJECT}"
        f"/_apis/test/plans/{plan_id}/suites?api-version=5.0"
    )
    legacy_resp = requests.get(legacy_url, headers=HEADERS, timeout=30)

    if legacy_resp.ok:
        suites = legacy_resp.json().get("value", [])
        print(f"[INFO] Legacy suite list: {len(suites)} total suite(s) in plan {plan_id}.")

        # Build parent-id → [child-id, ...] map.
        # Old API uses 'parent', new API uses 'parentSuite' – handle both.
        children_map: Dict[int, list] = {}
        for s in suites:
            par = s.get("parent") or s.get("parentSuite") or {}
            par_id = int(par.get("id") or 0)
            s_id   = int(s.get("id") or 0)
            children_map.setdefault(par_id, []).append(s_id)

        # BFS: collect all suite IDs in the subtree rooted at suite_id
        relevant: set = set()
        queue = [int(suite_id)]
        while queue:
            cur = queue.pop(0)
            relevant.add(cur)
            queue.extend(children_map.get(cur, []))

        print(f"[INFO] Suites in scope (suite {suite_id} + all descendants): {sorted(relevant)}")

        # Collect requirementId from every suite in the subtree
        for s in suites:
            s_id = int(s.get("id") or 0)
            if s_id in relevant:
                req_id = s.get("requirementId")
                if req_id is not None:
                    valid_us_ids.add(int(req_id))
                    print(f"[INFO]   Suite {s_id} ('{s.get('name', '')}') "
                          f"→ requirementId={req_id}")
    else:
        print(
            f"[WARN] Legacy suite list failed (HTTP {legacy_resp.status_code}). "
            f"Trying newer testplan API..."
        )

    # ── Fallback: newer testplan API – individual GET for suite itself ────
    if not valid_us_ids:
        new_url = (
            f"https://dev.azure.com/{ORGANIZATION}/{PROJECT}"
            f"/_apis/testplan/plans/{plan_id}/suites/{suite_id}?api-version=7.0"
        )
        r = requests.get(new_url, headers=HEADERS, timeout=30)
        if r.ok:
            req_id = r.json().get("requirementId")
            if req_id is not None:
                valid_us_ids.add(int(req_id))
                print(f"[INFO] Suite {suite_id} (new API): requirementId={req_id}")
        else:
            print(f"[WARN] New API GET suite {suite_id} failed (HTTP {r.status_code}).")

    if valid_us_ids:
        print(f"[INFO] Valid User Stories for suite {suite_id}: {sorted(valid_us_ids)}")
    else:
        print(
            f"[WARN] Suite {suite_id}: no requirement linkage discoverable – "
            f"User Story validation will be skipped (plain static suite)."
        )

    return valid_us_ids


def validate_user_story_in_suite(
    user_story_id:  int,
    plan_id:        int,
    suite_id:       int,
    suite_us_cache: Dict,
) -> None:
    """
    Raise RuntimeError if user_story_id is NOT among the User Stories
    linked to suite_id (checked at suite level and child-suite level).
    Uses suite_us_cache to avoid redundant API calls.
    """
    if suite_id not in suite_us_cache:
        suite_us_cache[suite_id] = fetch_suite_user_stories(plan_id, suite_id)

    valid_ids = suite_us_cache[suite_id]

    if not valid_ids:
        # No requirement linkage found at all – cannot enforce, warn and allow
        print(
            f"[WARN] Suite {suite_id}: no User Story linkage discoverable "
            f"– skipping validation for US {user_story_id}."
        )
        return

    if user_story_id not in valid_ids:
        raise RuntimeError(
            f"User Story {user_story_id} is NOT available in Suite {suite_id}. "
            f"Suite {suite_id} is linked to User Story ID(s): {sorted(valid_ids)}. "
            f"Please use the correct suite for this User Story."
        )

# ── Parallel worker ──────────────────────────────────────────────────────────

_print_lock = threading.Lock()

def _upload_one(args):
    """Worker executed in a thread pool – creates one TC and returns result."""
    idx, total, tc, resolved_lead, us_fields_cache = args
    title = tc["title"]
    us_id = tc["user_story_id"]
    us_fields = us_fields_cache[us_id]
    try:
        tc_id = create_test_case(tc, resolved_lead, us_fields, us_id)
        with _print_lock:
            print(f"[{idx}/{total}] ✓ Created TC {tc_id}: {title}")
        return tc_id, title, us_id, None
    except RuntimeError as err:
        with _print_lock:
            print(f"[{idx}/{total}] ✗ FAILED: {title} | {err}")
        return None, title, us_id, str(err)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 65)
    print("  ADO Test Case Uploader")
    print("=" * 65)
    print(f"  Organisation : {ORGANIZATION}")
    print(f"  Project      : {PROJECT}")
    print(f"  Test Plan    : {TEST_PLAN_ID}")
    print(f"  Test Suite   : {TEST_SUITE_ID}")
    print(f"  User Story   : (per row from Excel)")
    print(f"  Lead         : {LEAD_NAME or '(from Excel)'}")
    print(f"  Excel file   : {EXCEL_FILE}")
    print("=" * 65)

    validate_pat()

    # Resolve the lead's full ADO identity before uploading
    resolved_lead = resolve_identity(LEAD_NAME) if LEAD_NAME.strip() else ""
    print(f"  Resolved Lead: {resolved_lead or '(none)'}")
    print("=" * 65)

    test_cases = parse_excel(EXCEL_FILE, SHEET_NAME)
    if not test_cases:
        print("[WARN] No test cases found – nothing to upload.")
        return

    # ── Step 1: Pre-fetch AreaPath/Initiative for all unique User Stories ─────
    us_fields_cache: Dict[int, Dict] = {}
    unique_us_ids = list({tc["user_story_id"] for tc in test_cases})
    print(f"[INFO] Pre-fetching fields for {len(unique_us_ids)} unique User Story ID(s)...")
    for us_id in unique_us_ids:
        us_fields_cache[us_id] = fetch_user_story_fields(us_id)

    # ── Step 2: Create all test cases in parallel (8 concurrent threads) ──────
    total = len(test_cases)
    print(f"[INFO] Uploading {total} test case(s) with up to 8 parallel threads...")

    upload_args = [
        (idx, total, tc, resolved_lead, us_fields_cache)
        for idx, tc in enumerate(test_cases, start=1)
    ]

    success_count = 0
    failed_titles = []
    created_tc_ids = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        for tc_id, title, us_id, err in executor.map(_upload_one, upload_args):
            if tc_id is not None:
                created_tc_ids.append(tc_id)
                success_count += 1
            else:
                failed_titles.append(f"{title} | Error: {err}")

    # ── Step 3: Batch-add all created TCs to the suite in ONE API call ────────
    if created_tc_ids:
        print(f"\n[INFO] Adding {len(created_tc_ids)} test case(s) to suite {TEST_SUITE_ID} (batch)...")
        try:
            batch_add_to_suite(created_tc_ids)
            print(f"[INFO] All {len(created_tc_ids)} test case(s) added to suite successfully.")
        except RuntimeError as batch_err:
            print(f"[WARN] Batch add failed ({batch_err}). Retrying individually...")
            for tc_id in created_tc_ids:
                try:
                    add_to_suite(tc_id)
                    print(f"       Added TC {tc_id} individually.")
                except RuntimeError as e2:
                    print(f"       [WARN] Could not add TC {tc_id} to suite: {e2}")

    print("\n" + "=" * 65)
    print(f"  Upload complete: {success_count}/{len(test_cases)} test case(s) succeeded.")
    if failed_titles:
        print(f"  Failed ({len(failed_titles)}):")
        for t in failed_titles:
            print(f"    - {t}")
    print("=" * 65)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Upload test cases from Excel to Azure DevOps.")
    parser.add_argument("--excel-file", dest="excel_file", default=None,
                        help="Path to the Excel workbook. Overrides the EXCEL_FILE constant.")
    parser.add_argument("--suite-id", dest="suite_id", default=None,
                        help="Test Suite ID to upload to. Overrides TEST_SUITE_ID constant.")
    parser.add_argument("--plan-id", dest="plan_id", default=None,
                        help="Test Plan ID. Overrides TEST_PLAN_ID constant.")
    parser.add_argument("--pat-token", dest="pat_token", default=None,
                        help="PAT token for ADO authentication. Overrides PAT_TOKEN constant.")
    args = parser.parse_args()

    # Apply CLI overrides to module-level constants so all functions see them
    if args.excel_file:
        EXCEL_FILE = args.excel_file
    if args.suite_id:
        TEST_SUITE_ID = int(args.suite_id)
        TEST_SUITES_URL = (
            f"{BASE_URL}/_apis/test/plans/{TEST_PLAN_ID}"
            f"/suites/{TEST_SUITE_ID}/testcases"
        )
    if args.plan_id:
        TEST_PLAN_ID = int(args.plan_id)
    if args.pat_token:
        PAT_TOKEN = args.pat_token
        _encoded_pat = base64.b64encode(f":{PAT_TOKEN}".encode()).decode()
        HEADERS["Authorization"] = f"Basic {_encoded_pat}"
        PATCH_HEADERS["Authorization"] = f"Basic {_encoded_pat}"

    main()

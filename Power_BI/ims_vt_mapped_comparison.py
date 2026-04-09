"""
IMS VT Mapped Column Comparison Tool
=====================================
This tool compares two Excel files where column names differ between files,
using the mapping defined in "IMS VT Automation Mappings.xlsx"
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from pathlib import Path
from typing import Dict, List, Tuple

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class IMSVTComparer:
    """
    Compares two Excel files using column name mappings
    """
    
    def __init__(self, mapping_file: str = "IMS VT Automation Mappings.xlsx"):
        self.mapping_file = mapping_file
        self.mappings = self.load_mappings()
        self.fileA_data = None
        self.fileB_data = None
        self.comparison_results = None
    
    def load_mappings(self) -> List[Dict]:
        """Load column mappings from the mapping file"""
        logger.info(f"Loading mappings from {self.mapping_file}...")
        
        try:
            df = pd.read_excel(self.mapping_file, header=None)
        except Exception as e:
            logger.error(f"Cannot read mapping file: {e}")
            logger.info("\nTIP: Close the Excel file if it's open and try again")
            return []
        
        mappings = []
        
        for i in range(len(df)):
            imsvt_col = df.iloc[i, 3]  # Column 3: IMSVT column names
            imsvt_sub = df.iloc[i, 4]  # Column 4: IMSVT sub-columns
            ims_col = df.iloc[i, 6]    # Column 6: IMS KDA Report column names
            ims_sub = df.iloc[i, 7]    # Column 7: IMS KDA Report sub-columns
            
            # Skip headers
            if pd.notna(imsvt_col):
                imsvt_str = str(imsvt_col).strip()
                if any(x in imsvt_str for x in ['MainHeader', 'IMSVT', 'Columns']):
                    continue
                
            # Main column mapping
            if pd.notna(imsvt_col) and pd.notna(ims_col):
                imsvt_str = str(imsvt_col).strip()
                ims_str = str(ims_col).strip()
                
                if imsvt_str and ims_str:
                    mapping_entry = {
                        'imsvt_main': imsvt_str,
                        'ims_main': ims_str,
                        'sub_columns': []
                    }
                    
                    # Check for sub-columns
                    if pd.notna(imsvt_sub) and pd.notna(ims_sub):
                        mapping_entry['imsvt_sub'] = str(imsvt_sub).strip()
                        mapping_entry['ims_sub'] = str(ims_sub).strip()
                    
                    mappings.append(mapping_entry)
        
        logger.info(f"Loaded {len(mappings)} column mappings")
        return mappings
    
    def load_file(self, file_path: str, sheet_name: str = None, is_fileA: bool = True):
        """
        Load an Excel file
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (if None, use first sheet)
            is_fileA: True if this is File A (IMSVT), False if File B (IMS KDA)
        """
        label = "File A (IMSVT)" if is_fileA else "File B (IMS KDA Report)"
        logger.info(f"\nLoading {label}: {file_path}...")
        
        try:
            # Try pandas first
            # IMSVT file has headers in row 2 (index 2), IMS KDA has headers in row 0
            header_row = 2 if is_fileA else 0
            
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=header_row)
            else:
                xl = pd.ExcelFile(file_path, engine='openpyxl')
                sheet_name = xl.sheet_names[0]
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=header_row)
            
            logger.info(f"  Sheet: {sheet_name}")
            logger.info(f"  Shape: {df.shape}")
            logger.info(f"  Columns: {len(df.columns)}")
            
            if is_fileA:
                self.fileA_data = df
                self.fileA_sheet = sheet_name
            else:
                self.fileB_data = df
                self.fileB_sheet = sheet_name
            
            return df
        
        except Exception as e:
            logger.error(f"Failed to load {label}: {e}")
            logger.info("\nTroubleshooting tips:")
            logger.info("  1. Close the file if it's open in Excel")
            logger.info("  2. Check if the file is corrupted")
            logger.info("  3. Try saving as a new .xlsx file")
            logger.info("  4. Verify the file path is correct")
            return None
    
    def compare_files(self, key_column: str = None):
        """
        Compare two files using the mapped columns
        
        Args:
            key_column: Column name to use as key for matching rows
                       If None, will compare only column presence and structure
        """
        logger.info("\n" + "="*80)
        logger.info("COMPARING FILES WITH COLUMN MAPPING")
        logger.info("="*80)
        
        if self.fileA_data is None or self.fileB_data is None:
            logger.error("Both files must be loaded before comparison")
            return None
        
        results = []
        
        for mapping in self.mappings:
            imsvt_col = mapping['imsvt_main']
            ims_col = mapping['ims_main']
            
            # Check if columns exist in respective files
            imsvt_exists = imsvt_col in self.fileA_data.columns
            ims_exists = ims_col in self.fileB_data.columns
            
            # Try case-insensitive match if not found
            if not imsvt_exists:
                for col in self.fileA_data.columns:
                    if col and str(col).strip().lower() == imsvt_col.lower():
                        imsvt_exists = True
                        imsvt_col = col
                        break
            
            if not ims_exists:
                for col in self.fileB_data.columns:
                    if col and str(col).strip().lower() == ims_col.lower():
                        ims_exists = True
                        ims_col = col
                        break
            
            result = {
                'IMSVT_Column': mapping['imsvt_main'],
                'IMS_Column': mapping['ims_main'],
                'IMSVT_Found': imsvt_exists,
                'IMS_Found': ims_exists
            }
            
            # Compare values if both columns exist
            if imsvt_exists and ims_exists:
                imsvt_vals = self.fileA_data[imsvt_col]
                ims_vals = self.fileB_data[ims_col]
                
                # Get first non-null values as samples
                imsvt_sample = imsvt_vals.dropna().iloc[0] if len(imsvt_vals.dropna()) > 0 else None
                ims_sample = ims_vals.dropna().iloc[0] if len(ims_vals.dropna()) > 0 else None
                
                result['IMSVT_Sample'] = imsvt_sample
                result['IMS_Sample'] = ims_sample
                result['Match_Status'] = 'Both columns found'
                
                # Basic comparison statistics
                result['IMSVT_NonNull_Count'] = imsvt_vals.notna().sum()
                result['IMS_NonNull_Count'] = ims_vals.notna().sum()
                
            elif imsvt_exists:
                result['Match_Status'] = 'Only in IMSVT'
                result['IMSVT_Sample'] = self.fileA_data[imsvt_col].dropna().iloc[0] if len(self.fileA_data[imsvt_col].dropna()) > 0 else None
            elif ims_exists:
                result['Match_Status'] = 'Only in IMS KDA'
                result['IMS_Sample'] = self.fileB_data[ims_col].dropna().iloc[0] if len(self.fileB_data[ims_col].dropna()) > 0 else None
            else:
                result['Match_Status'] = 'Neither file has column'
            
            results.append(result)
        
        self.comparison_results = pd.DataFrame(results)
        
        # Print summary
        logger.info("\n" + "="*80)
        logger.info("COMPARISON SUMMARY")
        logger.info("="*80)
        logger.info(f"Total mapped columns: {len(self.mappings)}")
        logger.info(f"Both found: {len(self.comparison_results[self.comparison_results['Match_Status'] == 'Both columns found'])}")
        logger.info(f"Only in IMSVT: {len(self.comparison_results[self.comparison_results['Match_Status'] == 'Only in IMSVT'])}")
        logger.info(f"Only in IMS KDA: {len(self.comparison_results[self.comparison_results['Match_Status'] == 'Only in IMS KDA'])}")
        logger.info(f"Neither found: {len(self.comparison_results[self.comparison_results['Match_Status'] == 'Neither file has column'])}")
        
        return self.comparison_results
    
    def save_report(self, output_path: str = None):
        """Save comparison report to Excel"""
        if self.comparison_results is None:
            logger.error("No comparison results to save")
            return None
        
        if output_path is None:
            output_path = f"IMS_VT_Column_Mapping_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"\nSaving report to {output_path}...")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main comparison results
            self.comparison_results.to_excel(writer, sheet_name='Column Mapping Analysis', index=False)
            
            # Column lists
            if self.fileA_data is not None:
                df_fileA_cols = pd.DataFrame({
                    'File A (IMSVT) Columns': list(self.fileA_data.columns)
                })
                df_fileA_cols.to_excel(writer, sheet_name='File A Columns', index=False)
            
            if self.fileB_data is not None:
                df_fileB_cols = pd.DataFrame({
                    'File B (IMS KDA) Columns': list(self.fileB_data.columns)
                })
                df_fileB_cols.to_excel(writer, sheet_name='File B Columns', index=False)
            
            # Mapping reference
            df_mappings = pd.DataFrame([
                {'IMSVT_Column': m['imsvt_main'], 'IMS_KDA_Column': m['ims_main']}
                for m in self.mappings
            ])
            df_mappings.to_excel(writer, sheet_name='Mapping Reference', index=False)
        
        # Apply formatting
        wb = openpyxl.load_workbook(output_path)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            ws.freeze_panes = ws['A2']
        
        wb.save(output_path)
        logger.info(f"✓ Report saved successfully")
        
        return output_path


def main():
    """Main execution"""
    print("\n")
    print("="*80)
    print("IMS VT MAPPED COLUMN COMPARISON TOOL")
    print("="*80)
    print("\n")
    
    print("This tool compares two Excel files where column names differ,")
    print("using mappings defined in 'IMS VT Automation Mappings.xlsx'")
    print("\n")
    
    comparer = IMSVTComparer()
    
    if not comparer.mappings:
        print("\n❌ Failed to load mappings. Please check the mapping file.")
        return
    
    print(f"✓ Loaded {len(comparer.mappings)} column mappings\n")
    
    # Instructions
    print("="*80)
    print("INSTRUCTIONS")
    print("="*80)
    print("\n1. Ensure both Excel files are CLOSED (not open in Excel)")
    print("2. Place files in the same folder as this script")
    print("3. File A should be the IMSVT file")
    print("4. File B should be the IMS Managed Security KDA Report file")
    print("\n")
    
    # Try to load IMSVT file
    print("="*80)
    print("LOADING FILES")
    print("="*80)
    
    # Check if IMSVT.xlsx exists
    if Path("IMSVT.xlsx").exists():
        print("\nFound IMSVT.xlsx - attempting to load...")
        fileA = comparer.load_file("IMSVT.xlsx", is_fileA=True)
        
        if fileA is None:
            print("\n❌ Cannot open IMSVT.xlsx")
            print("\nPlease:")
            print("  1. Close the file if it's open in Excel")
            print("  2. Try opening it in Excel and saving as a new .xlsx file")
            print("  3. Then run this script again")
            return
    else:
        print("\n⚠ IMSVT.xlsx not found in current directory")
        print(f"Current directory: {Path.cwd()}")
        print("\nPlease place IMSVT.xlsx in this directory and run again")
        return
    
    # Look for IMS KDA file
    print("\nLooking for IMS Managed Security KDA Report file...")
    possible_files = list(Path(".").glob("*IMS*.xlsx"))
    kda_files = [f for f in possible_files if 'Mapping' not in f.name and 'IMSVT' not in f.name]
    
    if kda_files:
        print(f"Found potential IMS KDA file: {kda_files[0]}")
        fileB = comparer.load_file(str(kda_files[0]), is_fileA=False)
    else:
        print("\n⚠ IMS Managed Security KDA Report file not found")
        print("\nSince we only have IMSVT.xlsx, we'll analyze its structure")
        print("against the mapping to show which columns exist.")
        
        # Analyze single file
        results = []
        for mapping in comparer.mappings:
            imsvt_col = mapping['imsvt_main']
            imsvt_exists = imsvt_col in fileA.columns
            
            result = {
                'IMSVT_Column': imsvt_col,
                'IMS_Column': mapping['ims_main'],
                'Found_in_IMSVT': imsvt_exists
            }
            
            if imsvt_exists:
                sample = fileA[imsvt_col].dropna().iloc[0] if len(fileA[imsvt_col].dropna()) > 0 else None
                result['Sample_Value'] = sample
            
            results.append(result)
        
        df_results = pd.DataFrame(results)
        output_file = f"IMSVT_Column_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        df_results.to_excel(output_file, index=False)
        
        print(f"\n✓ Analysis saved to: {output_file}")
        print(f"\nFound {len(df_results[df_results['Found_in_IMSVT']])} out of {len(comparer.mappings)} mapped columns in IMSVT.xlsx")
        return
    
    # Compare files
    print("\n" + "="*80)
    print("PERFORMING COMPARISON")
    print("="*80)
    
    comparer.compare_files()
    output_file = comparer.save_report()
    
    print("\n" + "="*80)
    print("COMPLETE!")
    print("="*80)
    print(f"\n✓ Comparison report saved to: {output_file}")
    print("\nOpen the Excel file to see:")
    print("  • Column Mapping Analysis - Which columns were found")
    print("  • File A Columns - All columns in IMSVT file")
    print("  • File B Columns - All columns in IMS KDA file")
    print("  • Mapping Reference - Complete mapping list")


if __name__ == "__main__":
    main()

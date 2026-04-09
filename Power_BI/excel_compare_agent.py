"""
Excel Compare Agent
===================
A comprehensive tool to compare values between two or more Excel workbooks based on
user-provided requirements.

Author: Excel Compare Agent System
Date: March 5, 2026
Version: 1.0.0
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
import logging
from dataclasses import dataclass, asdict
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# DATA CLASSES FOR CONFIGURATION
# ============================================================================

@dataclass
class ComparisonRule:
    """Rules for comparing values"""
    text_mode: str = "case_insensitive_trimmed"  # exact, case_insensitive, case_insensitive_trimmed
    numeric_tolerance: float = 0.0  # absolute tolerance
    numeric_tolerance_percent: Optional[float] = None  # percentage tolerance
    date_format: Optional[str] = None  # date format if needed
    date_only: bool = False  # compare date only, ignore time
    treat_blank_as_zero: bool = False  # treat blank cells as 0
    treat_blank_as_na: bool = False  # treat blank cells as "NA"


@dataclass
class ComparisonConfig:
    """Configuration for Excel comparison"""
    fileA: str
    fileB: str
    sheetA: Optional[str] = None
    sheetB: Optional[str] = None
    keyColumns: List[str] = None
    compareColumns: List[str] = None
    rules: ComparisonRule = None
    outputPath: Optional[str] = None
    includeAllRows: bool = True  # include matched rows or only mismatches
    highlightDifferences: bool = True
    timestamp: str = None
    
    def __post_init__(self):
        if self.rules is None:
            self.rules = ComparisonRule()
        if self.timestamp is None:
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if self.outputPath is None:
            self.outputPath = f"Comparison_Report_{self.timestamp}.xlsx"
        if self.keyColumns is None:
            self.keyColumns = []
        if self.compareColumns is None:
            self.compareColumns = []


# ============================================================================
# REQUIREMENT PARSER
# ============================================================================

class RequirementParser:
    """Parse natural language requirements into ComparisonConfig"""
    
    @staticmethod
    def parse(requirement_text: str) -> ComparisonConfig:
        """
        Parse user requirement text into a structured ComparisonConfig
        
        Args:
            requirement_text: Natural language requirement
            
        Returns:
            ComparisonConfig object
        """
        logger.info("Parsing requirement text...")
        
        # Initialize with default empty values
        config = ComparisonConfig(fileA="", fileB="")
        
        # Extract file names (including files with spaces like "Mar 28th_Release.xlsx")
        # Remove common keywords first to isolate file names
        temp_text = requirement_text
        for keyword in ['Compare', 'compare', 'vs', 'vs.', 'versus', 'with', 'and']:
            temp_text = temp_text.replace(keyword, ' ')
        
        # Pattern matches: paths with alphanumeric, spaces, hyphens, underscores, slashes before .xlsx
        file_pattern = r'([\w\s\-/\\]+\.xlsx)'
        files = re.findall(file_pattern, temp_text, re.IGNORECASE)
        
        if len(files) >= 2:
            # Clean up file names (remove extra whitespace)
            config.fileA = ' '.join(files[0].split()).strip()  # baseline/source
            config.fileB = ' '.join(files[1].split()).strip()  # target/updated
            logger.info(f"Detected File A (baseline): {config.fileA}")
            logger.info(f"Detected File B (target): {config.fileB}")
        
        # Extract sheet names
        sheet_pattern = r'[Ss]heet[:\s]+"([^"]+)"|[Ss]heet[:\s]+(\w+)'
        sheets = re.findall(sheet_pattern, requirement_text)
        if sheets:
            config.sheetA = sheets[0][0] or sheets[0][1]
            config.sheetB = config.sheetA  # assume same sheet name
            logger.info(f"Detected Sheet: {config.sheetA}")
        
        # Extract key columns (match using)
        key_pattern = r'[Mm]atch.*?using[:\s]+([^\.\n]+)'
        key_match = re.search(key_pattern, requirement_text)
        if key_match:
            key_text = key_match.group(1)
            # Split by common separators
            keys = re.split(r'[,;|]|\s+and\s+|\s+&\s+', key_text)
            config.keyColumns = [k.strip() for k in keys if k.strip()]
            logger.info(f"Detected Key Columns: {config.keyColumns}")
        
        # Extract compare fields
        compare_pattern = r'[Cc]ompare\s+fields?[:\s]+([^\.\n]+)'
        compare_match = re.search(compare_pattern, requirement_text)
        if compare_match:
            compare_text = compare_match.group(1)
            # Split by comma, semicolon, or "and"
            fields = re.split(r'[,;]|\s+and\s+(?!-)', compare_text)
            config.compareColumns = [f.strip() for f in fields if f.strip()]
            logger.info(f"Detected Compare Columns: {config.compareColumns}")
        
        # Parse comparison rules
        rules = ComparisonRule()
        
        # Text comparison mode
        if re.search(r'case[- ]insensitive', requirement_text, re.IGNORECASE):
            rules.text_mode = "case_insensitive_trimmed"
        elif re.search(r'exact.*match', requirement_text, re.IGNORECASE):
            rules.text_mode = "exact"
        
        # Numeric tolerance
        tolerance_pattern = r'tolerance[:\s]+([\d.]+)'
        tolerance_match = re.search(tolerance_pattern, requirement_text)
        if tolerance_match:
            rules.numeric_tolerance = float(tolerance_match.group(1))
            logger.info(f"Detected Numeric Tolerance: {rules.numeric_tolerance}")
        
        config.rules = rules
        
        return config
    
    @staticmethod
    def save_config(config: ComparisonConfig, output_path: str = "comparison_config.json"):
        """Save configuration to JSON file"""
        config_dict = asdict(config)
        with open(output_path, 'w') as f:
            json.dump(config_dict, f, indent=2)
        logger.info(f"Configuration saved to {output_path}")


# ============================================================================
# DATA LOADER AND NORMALIZER
# ============================================================================

class ExcelDataLoader:
    """Load and normalize Excel data"""
    
    @staticmethod
    def find_header_row(df: pd.DataFrame, expected_columns: List[str]) -> int:
        """
        Find the row that contains the best match for expected column names
        
        Args:
            df: DataFrame to search
            expected_columns: List of expected column names
            
        Returns:
            Row index (0-based) of header row
        """
        max_matches = 0
        header_row = 0
        
        for idx in range(min(10, len(df))):  # Check first 10 rows
            row_values = df.iloc[idx].astype(str).str.lower().tolist()
            matches = sum(1 for col in expected_columns 
                         if any(col.lower() in val for val in row_values))
            if matches > max_matches:
                max_matches = matches
                header_row = idx
        
        return header_row
    
    @staticmethod
    def load_excel(file_path: str, sheet_name: Optional[str] = None, 
                   expected_columns: Optional[List[str]] = None) -> Tuple[pd.DataFrame, str]:
        """
        Load Excel file and return DataFrame with normalized column names
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (if None, use first sheet)
            expected_columns: Expected column names for header detection
            
        Returns:
            Tuple of (DataFrame, actual_sheet_name)
        """
        logger.info(f"Loading {file_path}...")
        
        # Load workbook to get sheet names
        xl_file = pd.ExcelFile(file_path)
        
        # Determine sheet to use
        if sheet_name:
            # Try exact match first
            if sheet_name in xl_file.sheet_names:
                actual_sheet = sheet_name
            else:
                # Try case-insensitive match
                matching_sheets = [s for s in xl_file.sheet_names 
                                 if s.lower() == sheet_name.lower()]
                if matching_sheets:
                    actual_sheet = matching_sheets[0]
                else:
                    logger.warning(f"Sheet '{sheet_name}' not found. Using first sheet.")
                    actual_sheet = xl_file.sheet_names[0]
        else:
            actual_sheet = xl_file.sheet_names[0]
        
        logger.info(f"Using sheet: {actual_sheet}")
        
        # Load data without header first
        df = pd.read_excel(file_path, sheet_name=actual_sheet, header=None)
        
        # Find header row if expected columns provided
        header_row = 0
        if expected_columns:
            header_row = ExcelDataLoader.find_header_row(df, expected_columns)
            logger.info(f"Detected header row: {header_row + 1}")
        
        # Reload with proper header
        df = pd.read_excel(file_path, sheet_name=actual_sheet, header=header_row)
        
        # Normalize column names: strip whitespace, handle duplicates
        df.columns = [str(col).strip() for col in df.columns]
        
        # Handle duplicate column names
        cols = pd.Series(df.columns)
        for dup in cols[cols.duplicated()].unique():
            cols[cols == dup] = [f"{dup}_{i}" if i != 0 else dup 
                                for i in range(sum(cols == dup))]
        df.columns = cols
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns")
        logger.info(f"Columns: {list(df.columns)}")
        
        return df, actual_sheet
    
    @staticmethod
    def normalize_value(value: Any, rules: ComparisonRule) -> Any:
        """
        Normalize a value based on comparison rules
        
        Args:
            value: Value to normalize
            rules: Comparison rules
            
        Returns:
            Normalized value
        """
        # Handle NaN/None
        if pd.isna(value):
            if rules.treat_blank_as_zero:
                return 0
            elif rules.treat_blank_as_na:
                return "NA"
            else:
                return None
        
        # String normalization
        if isinstance(value, str):
            if rules.text_mode in ["case_insensitive", "case_insensitive_trimmed"]:
                value = value.lower()
            if rules.text_mode == "case_insensitive_trimmed":
                value = value.strip()
            return value
        
        # Numeric normalization
        if isinstance(value, (int, float)):
            return float(value)
        
        return value


# ============================================================================
# COMPARISON ENGINE
# ============================================================================

class ComparisonEngine:
    """Core comparison logic"""
    
    def __init__(self, config: ComparisonConfig):
        self.config = config
        self.dfA = None
        self.dfB = None
        self.sheetA_name = None
        self.sheetB_name = None
        self.results = None
        self.summary = None
    
    def load_data(self):
        """Load data from both Excel files"""
        logger.info("=" * 80)
        logger.info("LOADING DATA")
        logger.info("=" * 80)
        
        # Combine key and compare columns for header detection
        expected_cols = self.config.keyColumns + self.config.compareColumns
        
        # Load File A
        self.dfA, self.sheetA_name = ExcelDataLoader.load_excel(
            self.config.fileA, 
            self.config.sheetA,
            expected_cols
        )
        
        # Load File B
        self.dfB, self.sheetB_name = ExcelDataLoader.load_excel(
            self.config.fileB,
            self.config.sheetB,
            expected_cols
        )
        
        # Validate key columns exist
        self._validate_columns()
    
    def _validate_columns(self):
        """Validate that required columns exist in both dataframes"""
        missing_in_A = []
        missing_in_B = []
        
        for col in self.config.keyColumns:
            if col not in self.dfA.columns:
                missing_in_A.append(col)
            if col not in self.dfB.columns:
                missing_in_B.append(col)
        
        if missing_in_A or missing_in_B:
            error_msg = "Column validation failed:\n"
            if missing_in_A:
                error_msg += f"  Missing in File A: {missing_in_A}\n"
                error_msg += f"  Available columns in A: {list(self.dfA.columns)}\n"
            if missing_in_B:
                error_msg += f"  Missing in File B: {missing_in_B}\n"
                error_msg += f"  Available columns in B: {list(self.dfB.columns)}\n"
            raise ValueError(error_msg)
        
        # Check compare columns (warn but don't fail)
        for col in self.config.compareColumns:
            if col not in self.dfA.columns:
                logger.warning(f"Compare column '{col}' not found in File A")
            if col not in self.dfB.columns:
                logger.warning(f"Compare column '{col}' not found in File B")
    
    def _create_composite_key(self, df: pd.DataFrame) -> pd.Series:
        """Create composite key from key columns"""
        if len(self.config.keyColumns) == 1:
            return df[self.config.keyColumns[0]].astype(str)
        else:
            return df[self.config.keyColumns].astype(str).agg('||'.join, axis=1)
    
    def _compare_values(self, val_a: Any, val_b: Any, column: str) -> Dict[str, Any]:
        """
        Compare two values and return comparison result
        
        Returns:
            Dict with keys: result, match, notes
        """
        rules = self.config.rules
        
        # Normalize values
        norm_a = ExcelDataLoader.normalize_value(val_a, rules)
        norm_b = ExcelDataLoader.normalize_value(val_b, rules)
        
        # Both blank
        if norm_a is None and norm_b is None:
            return {
                'result': 'Both_Blank',
                'match': True,
                'notes': ''
            }
        
        # One blank
        if norm_a is None or norm_b is None:
            return {
                'result': 'Mismatch',
                'match': False,
                'notes': 'One value is blank'
            }
        
        # Type mismatch check
        if type(norm_a) != type(norm_b):
            # Try numeric conversion
            try:
                norm_a = float(norm_a)
                norm_b = float(norm_b)
            except (ValueError, TypeError):
                return {
                    'result': 'Type_Mismatch',
                    'match': False,
                    'notes': f'Type mismatch: {type(val_a).__name__} vs {type(val_b).__name__}'
                }
        
        # Numeric comparison
        if isinstance(norm_a, (int, float)) and isinstance(norm_b, (int, float)):
            diff = abs(norm_a - norm_b)
            
            # Check absolute tolerance
            if diff <= rules.numeric_tolerance:
                result = 'Match' if diff == 0 else 'Tolerance_Match'
                return {
                    'result': result,
                    'match': True,
                    'notes': f'Diff: {diff:.4f} (tolerance: {rules.numeric_tolerance})' if result == 'Tolerance_Match' else ''
                }
            
            # Check percentage tolerance
            if rules.numeric_tolerance_percent is not None:
                if norm_a != 0:
                    pct_diff = abs(diff / norm_a) * 100
                    if pct_diff <= rules.numeric_tolerance_percent:
                        return {
                            'result': 'Tolerance_Match',
                            'match': True,
                            'notes': f'Pct diff: {pct_diff:.2f}% (tolerance: {rules.numeric_tolerance_percent}%)'
                        }
            
            return {
                'result': 'Mismatch',
                'match': False,
                'notes': f'Diff: {diff:.4f}'
            }
        
        # String comparison
        if isinstance(norm_a, str) and isinstance(norm_b, str):
            match = norm_a == norm_b
            return {
                'result': 'Match' if match else 'Mismatch',
                'match': match,
                'notes': ''
            }
        
        # Default exact match
        match = norm_a == norm_b
        return {
            'result': 'Match' if match else 'Mismatch',
            'match': match,
            'notes': ''
        }
    
    def compare(self):
        """Execute the comparison"""
        logger.info("=" * 80)
        logger.info("EXECUTING COMPARISON")
        logger.info("=" * 80)
        
        # Create composite keys
        self.dfA['__CompKey__'] = self._create_composite_key(self.dfA)
        self.dfB['__CompKey__'] = self._create_composite_key(self.dfB)
        
        # Check for duplicate keys
        dup_a = self.dfA['__CompKey__'].duplicated()
        dup_b = self.dfB['__CompKey__'].duplicated()
        
        if dup_a.any():
            logger.warning(f"Found {dup_a.sum()} duplicate keys in File A")
        if dup_b.any():
            logger.warning(f"Found {dup_b.sum()} duplicate keys in File B")
        
        # Create comparison results
        results_list = []
        
        # Get all unique keys
        all_keys = set(self.dfA['__CompKey__'].unique()) | set(self.dfB['__CompKey__'].unique())
        logger.info(f"Total unique keys to compare: {len(all_keys)}")
        
        for idx, key in enumerate(all_keys, 1):
            if idx % 100 == 0:
                logger.info(f"Processed {idx}/{len(all_keys)} keys...")
            
            # Get rows for this key
            rows_a = self.dfA[self.dfA['__CompKey__'] == key]
            rows_b = self.dfB[self.dfB['__CompKey__'] == key]
            
            # Determine match status
            if len(rows_a) == 0:
                match_status = 'Missing_in_A'
            elif len(rows_b) == 0:
                match_status = 'Missing_in_B'
            elif len(rows_a) > 1:
                match_status = 'Duplicate_Key_A'
            elif len(rows_b) > 1:
                match_status = 'Duplicate_Key_B'
            else:
                match_status = 'Matched'
            
            # Build result row
            result_row = {
                'Key': key,
                'Match_Status': match_status
            }
            
            # Add key column values
            for key_col in self.config.keyColumns:
                if len(rows_a) > 0:
                    result_row[f'{key_col}_A'] = rows_a.iloc[0].get(key_col, None)
                else:
                    result_row[f'{key_col}_A'] = None
                
                if len(rows_b) > 0:
                    result_row[f'{key_col}_B'] = rows_b.iloc[0].get(key_col, None)
                else:
                    result_row[f'{key_col}_B'] = None
            
            # Compare each field
            mismatch_count = 0
            
            for col in self.config.compareColumns:
                # Get values
                val_a = rows_a.iloc[0].get(col, None) if len(rows_a) > 0 else None
                val_b = rows_b.iloc[0].get(col, None) if len(rows_b) > 0 else None
                
                # Check if column exists
                col_exists_a = col in self.dfA.columns
                col_exists_b = col in self.dfB.columns
                
                if not col_exists_a or not col_exists_b:
                    result_row[f'{col}_A'] = val_a if col_exists_a else 'COLUMN_MISSING'
                    result_row[f'{col}_B'] = val_b if col_exists_b else 'COLUMN_MISSING'
                    result_row[f'{col}_Result'] = 'Missing_Column'
                    result_row[f'{col}_Notes'] = f"Column missing in {'A' if not col_exists_a else 'B'}"
                    mismatch_count += 1
                    continue
                
                # Skip comparison if either row is missing
                if match_status in ['Missing_in_A', 'Missing_in_B']:
                    result_row[f'{col}_A'] = val_a
                    result_row[f'{col}_B'] = val_b
                    result_row[f'{col}_Result'] = 'N/A'
                    result_row[f'{col}_Notes'] = 'Row not found in one file'
                    continue
                
                # Perform comparison
                comp_result = self._compare_values(val_a, val_b, col)
                
                result_row[f'{col}_A'] = val_a
                result_row[f'{col}_B'] = val_b
                result_row[f'{col}_Result'] = comp_result['result']
                result_row[f'{col}_Notes'] = comp_result['notes']
                
                if not comp_result['match']:
                    mismatch_count += 1
            
            # Overall row result
            if match_status != 'Matched':
                result_row['Row_Result'] = 'Unmatched'
            elif mismatch_count > 0:
                result_row['Row_Result'] = 'Has_Mismatch'
            else:
                result_row['Row_Result'] = 'All_Match'
            
            result_row['Mismatch_Count'] = mismatch_count
            
            results_list.append(result_row)
        
        self.results = pd.DataFrame(results_list)
        logger.info(f"Comparison complete. {len(self.results)} rows processed.")
        
        # Generate summary
        self._generate_summary()
    
    def _generate_summary(self):
        """Generate summary statistics"""
        logger.info("Generating summary statistics...")
        
        summary = {
            'Total_Rows_A': len(self.dfA),
            'Total_Rows_B': len(self.dfB),
            'Matched_Keys': len(self.results[self.results['Match_Status'] == 'Matched']),
            'Missing_in_A': len(self.results[self.results['Match_Status'] == 'Missing_in_A']),
            'Missing_in_B': len(self.results[self.results['Match_Status'] == 'Missing_in_B']),
            'Duplicate_Keys_A': len(self.results[self.results['Match_Status'] == 'Duplicate_Key_A']),
            'Duplicate_Keys_B': len(self.results[self.results['Match_Status'] == 'Duplicate_Key_B']),
            'Rows_All_Match': len(self.results[self.results['Row_Result'] == 'All_Match']),
            'Rows_With_Mismatch': len(self.results[self.results['Row_Result'] == 'Has_Mismatch']),
        }
        
        # Field-wise statistics
        field_stats = []
        for col in self.config.compareColumns:
            result_col = f'{col}_Result'
            if result_col in self.results.columns:
                stats = self.results[result_col].value_counts().to_dict()
                field_stats.append({
                    'Field': col,
                    'Match': stats.get('Match', 0) + stats.get('Both_Blank', 0),
                    'Mismatch': stats.get('Mismatch', 0),
                    'Tolerance_Match': stats.get('Tolerance_Match', 0),
                    'Type_Mismatch': stats.get('Type_Mismatch', 0),
                    'Missing_Column': stats.get('Missing_Column', 0),
                })
        
        summary['Field_Statistics'] = field_stats
        
        # Match rate
        total_comparisons = summary['Matched_Keys'] * len(self.config.compareColumns)
        if total_comparisons > 0:
            total_matches = sum(fs['Match'] + fs['Tolerance_Match'] for fs in field_stats)
            summary['Overall_Match_Rate_%'] = round((total_matches / total_comparisons) * 100, 2)
        else:
            summary['Overall_Match_Rate_%'] = 0
        
        self.summary = summary
        
        logger.info("Summary generated successfully")


# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ReportGenerator:
    """Generate Excel reports with formatting"""
    
    def __init__(self, engine: ComparisonEngine):
        self.engine = engine
        self.config = engine.config
    
    def generate_excel_report(self):
        """Generate comprehensive Excel report"""
        logger.info("=" * 80)
        logger.info("GENERATING EXCEL REPORT")
        logger.info("=" * 80)
        
        output_path = self.config.outputPath
        logger.info(f"Output file: {output_path}")
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1. Summary sheet
            self._write_summary_sheet(writer)
            
            # 2. Compared Data sheet
            self._write_compared_data_sheet(writer)
            
            # 3. Unmatched_In_A sheet
            unmatched_a = self.engine.results[
                self.engine.results['Match_Status'] == 'Missing_in_A'
            ]
            if len(unmatched_a) > 0:
                unmatched_a.to_excel(writer, sheet_name='Unmatched_In_A', index=False)
                logger.info(f"Unmatched_In_A sheet: {len(unmatched_a)} rows")
            
            # 4. Unmatched_In_B sheet
            unmatched_b = self.engine.results[
                self.engine.results['Match_Status'] == 'Missing_in_B'
            ]
            if len(unmatched_b) > 0:
                unmatched_b.to_excel(writer, sheet_name='Unmatched_In_B', index=False)
                logger.info(f"Unmatched_In_B sheet: {len(unmatched_b)} rows")
            
            # 5. Config sheet
            self._write_config_sheet(writer)
        
        # Apply formatting
        self._apply_formatting(output_path)
        
        logger.info(f"✓ Excel report generated: {output_path}")
        return output_path
    
    def _write_summary_sheet(self, writer):
        """Write summary statistics sheet"""
        summary = self.engine.summary
        
        # Overall statistics
        overall_data = [
            ['Metric', 'Value'],
            ['Total Rows in File A', summary['Total_Rows_A']],
            ['Total Rows in File B', summary['Total_Rows_B']],
            ['Matched Keys', summary['Matched_Keys']],
            ['Missing in A', summary['Missing_in_A']],
            ['Missing in B', summary['Missing_in_B']],
            ['Duplicate Keys in A', summary['Duplicate_Keys_A']],
            ['Duplicate Keys in B', summary['Duplicate_Keys_B']],
            ['Rows - All Match', summary['Rows_All_Match']],
            ['Rows - Has Mismatch', summary['Rows_With_Mismatch']],
            ['Overall Match Rate (%)', summary['Overall_Match_Rate_%']],
        ]
        
        df_overall = pd.DataFrame(overall_data[1:], columns=overall_data[0])
        df_overall.to_excel(writer, sheet_name='Summary', startrow=0, index=False)
        
        # Field statistics
        if summary['Field_Statistics']:
            df_fields = pd.DataFrame(summary['Field_Statistics'])
            df_fields.to_excel(writer, sheet_name='Summary', startrow=len(overall_data) + 2, index=False)
        
        logger.info("Summary sheet written")
    
    def _write_compared_data_sheet(self, writer):
        """Write compared data with all details"""
        df = self.engine.results
        
        # Filter rows if needed
        if not self.config.includeAllRows:
            df = df[df['Row_Result'] != 'All_Match']
        
        df.to_excel(writer, sheet_name='Compared_Data', index=False)
        logger.info(f"Compared_Data sheet: {len(df)} rows")
    
    def _write_config_sheet(self, writer):
        """Write configuration details"""
        config_data = [
            ['Configuration', 'Value'],
            ['Timestamp', self.config.timestamp],
            ['File A', self.config.fileA],
            ['File B', self.config.fileB],
            ['Sheet A', self.engine.sheetA_name],
            ['Sheet B', self.engine.sheetB_name],
            ['Key Columns', ', '.join(self.config.keyColumns)],
            ['Compare Columns', ', '.join(self.config.compareColumns)],
            ['Text Comparison Mode', self.config.rules.text_mode],
            ['Numeric Tolerance', str(self.config.rules.numeric_tolerance)],
            ['Numeric Tolerance %', str(self.config.rules.numeric_tolerance_percent) if self.config.rules.numeric_tolerance_percent else 'N/A'],
            ['Treat Blank as Zero', str(self.config.rules.treat_blank_as_zero)],
        ]
        
        df_config = pd.DataFrame(config_data[1:], columns=config_data[0])
        df_config.to_excel(writer, sheet_name='Config', index=False)
        logger.info("Config sheet written")
    
    def _apply_formatting(self, output_path: str):
        """Apply Excel formatting to the report"""
        logger.info("Applying formatting...")
        
        wb = openpyxl.load_workbook(output_path)
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        missing_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Conditional formatting for Compared_Data sheet
            if sheet_name == 'Compared_Data' and self.config.highlightDifferences:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.column > 2:  # Skip Key and Match_Status columns
                            value = str(cell.value) if cell.value else ''
                            if 'Mismatch' in value or 'Type_Mismatch' in value:
                                cell.fill = mismatch_fill
                            elif value in ['Match', 'Both_Blank', 'All_Match']:
                                cell.fill = match_fill
                            elif 'Missing' in value or 'Unmatched' in value:
                                cell.fill = missing_fill
            
            # Freeze top row
            ws.freeze_panes = ws['A2']
        
        wb.save(output_path)
        logger.info("Formatting applied successfully")
    
    def generate_text_report(self) -> str:
        """Generate human-readable text summary"""
        summary = self.engine.summary
        
        report = []
        report.append("=" * 80)
        report.append("EXCEL COMPARISON REPORT")
        report.append("=" * 80)
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        report.append(f"File A (Baseline): {self.config.fileA}")
        report.append(f"File B (Target):   {self.config.fileB}")
        report.append(f"Sheet A: {self.engine.sheetA_name}")
        report.append(f"Sheet B: {self.engine.sheetB_name}")
        report.append("")
        report.append(f"Key Columns:     {', '.join(self.config.keyColumns)}")
        report.append(f"Compare Columns: {', '.join(self.config.compareColumns)}")
        report.append("")
        report.append("-" * 80)
        report.append("OVERALL STATISTICS")
        report.append("-" * 80)
        report.append(f"Total Rows in File A:     {summary['Total_Rows_A']:,}")
        report.append(f"Total Rows in File B:     {summary['Total_Rows_B']:,}")
        report.append(f"Matched Keys:             {summary['Matched_Keys']:,}")
        report.append(f"Missing in A:             {summary['Missing_in_A']:,}")
        report.append(f"Missing in B:             {summary['Missing_in_B']:,}")
        report.append(f"Duplicate Keys in A:      {summary['Duplicate_Keys_A']:,}")
        report.append(f"Duplicate Keys in B:      {summary['Duplicate_Keys_B']:,}")
        report.append("")
        report.append(f"Rows - All Match:         {summary['Rows_All_Match']:,}")
        report.append(f"Rows - Has Mismatch:      {summary['Rows_With_Mismatch']:,}")
        report.append(f"Overall Match Rate:       {summary['Overall_Match_Rate_%']:.2f}%")
        report.append("")
        report.append("-" * 80)
        report.append("FIELD-WISE COMPARISON")
        report.append("-" * 80)
        
        if summary['Field_Statistics']:
            # Header
            report.append(f"{'Field':<40} {'Match':<10} {'Mismatch':<10} {'Tolerance':<10}")
            report.append("-" * 80)
            
            for fs in summary['Field_Statistics']:
                field_name = fs['Field'][:38]
                report.append(
                    f"{field_name:<40} {fs['Match']:<10} {fs['Mismatch']:<10} {fs['Tolerance_Match']:<10}"
                )
        
        report.append("")
        report.append("=" * 80)
        report.append(f"Detailed results saved to: {self.config.outputPath}")
        report.append("=" * 80)
        
        return '\n'.join(report)


# ============================================================================
# MAIN EXCEL COMPARE AGENT
# ============================================================================

class ExcelCompareAgent:
    """Main agent for Excel comparison"""
    
    def __init__(self, requirement_text: str = None, config: ComparisonConfig = None):
        """
        Initialize agent with either requirement text or config object
        
        Args:
            requirement_text: Natural language requirement
            config: Pre-built ComparisonConfig object
        """
        if config:
            self.config = config
        elif requirement_text:
            self.config = RequirementParser.parse(requirement_text)
        else:
            raise ValueError("Either requirement_text or config must be provided")
        
        self.engine = ComparisonEngine(self.config)
        self.report_gen = None
    
    def validate_config(self) -> List[str]:
        """
        Validate configuration and return list of issues/warnings
        
        Returns:
            List of validation messages (empty if all good)
        """
        issues = []
        
        if not self.config.fileA or not Path(self.config.fileA).exists():
            issues.append(f"File A not found: {self.config.fileA}")
        
        if not self.config.fileB or not Path(self.config.fileB).exists():
            issues.append(f"File B not found: {self.config.fileB}")
        
        if not self.config.keyColumns:
            issues.append("No key columns specified - will attempt to auto-detect")
        
        if not self.config.compareColumns:
            issues.append("No compare columns specified - will attempt to auto-detect")
        
        return issues
    
    def auto_detect_columns(self):
        """Auto-detect key and compare columns if not specified"""
        logger.info("Auto-detecting columns...")
        
        # Load minimal data
        df_a, _ = ExcelDataLoader.load_excel(self.config.fileA, self.config.sheetA)
        
        # Auto-detect key column
        if not self.config.keyColumns:
            id_columns = [col for col in df_a.columns 
                         if any(keyword in col.lower() 
                               for keyword in ['id', '_id', 'key', 'us_id', 'feature', 'external'])]
            if id_columns:
                self.config.keyColumns = [id_columns[0]]
                logger.info(f"Auto-detected key column: {id_columns[0]}")
            else:
                self.config.keyColumns = [df_a.columns[0]]
                logger.warning(f"No ID column found, using first column: {df_a.columns[0]}")
        
        # Auto-detect compare columns (all except key)
        if not self.config.compareColumns:
            self.config.compareColumns = [col for col in df_a.columns 
                                         if col not in self.config.keyColumns]
            logger.info(f"Auto-detected {len(self.config.compareColumns)} compare columns")
    
    def run(self) -> str:
        """
        Execute the complete comparison process
        
        Returns:
            Path to output Excel file
        """
        logger.info("=" * 80)
        logger.info("EXCEL COMPARE AGENT - STARTING")
        logger.info("=" * 80)
        
        # Validate configuration
        issues = self.validate_config()
        if issues:
            for issue in issues:
                logger.warning(issue)
        
        # Auto-detect if needed
        if not self.config.keyColumns or not self.config.compareColumns:
            self.auto_detect_columns()
        
        # Load data
        self.engine.load_data()
        
        # Perform comparison
        self.engine.compare()
        
        # Generate reports
        self.report_gen = ReportGenerator(self.engine)
        
        # Generate text report
        text_report = self.report_gen.generate_text_report()
        print("\n")
        print(text_report)
        
        # Generate Excel report
        output_file = self.report_gen.generate_excel_report()
        
        logger.info("=" * 80)
        logger.info("EXCEL COMPARE AGENT - COMPLETED")
        logger.info("=" * 80)
        
        return output_file


# ============================================================================
# CLI INTERFACE
# ============================================================================

def main():
    """Command-line interface"""
    import sys
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Excel Compare Agent - Compare Excel workbooks based on requirements',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example usage:
    
    1. Using requirement text:
       python excel_compare_agent.py --requirement "Compare file1.xlsx vs file2.xlsx..."
    
    2. Interactive mode:
       python excel_compare_agent.py --interactive
    
    3. Using config file:
       python excel_compare_agent.py --config comparison_config.json
        """
    )
    
    parser.add_argument('--requirement', '-r', type=str, help='Natural language requirement text')
    parser.add_argument('--config', '-c', type=str, help='Path to JSON config file')
    parser.add_argument('--interactive', '-i', action='store_true', help='Interactive mode')
    parser.add_argument('--output', '-o', type=str, help='Output file path')
    parser.add_argument('--save-config', type=str, help='Save parsed config to JSON file')
    
    args = parser.parse_args()
    
    # Interactive mode
    if args.interactive or (not args.requirement and not args.config):
        print("=" * 80)
        print("EXCEL COMPARE AGENT - INTERACTIVE MODE")
        print("=" * 80)
        print("\nPlease provide your comparison requirement:")
        print("(Paste multi-line text, then press Ctrl+Z (Windows) or Ctrl+D (Unix) on a new line)\n")
        
        try:
            requirement_lines = []
            while True:
                line = input()
                requirement_lines.append(line)
        except EOFError:
            requirement_text = '\n'.join(requirement_lines)
        
        if not requirement_text.strip():
            print("Error: No requirement provided")
            sys.exit(1)
        
        agent = ExcelCompareAgent(requirement_text=requirement_text)
    
    # Config file mode
    elif args.config:
        logger.info(f"Loading config from {args.config}")
        with open(args.config, 'r') as f:
            config_dict = json.load(f)
        
        # Reconstruct config object
        config = ComparisonConfig(**config_dict)
        agent = ExcelCompareAgent(config=config)
    
    # Requirement text mode
    else:
        agent = ExcelCompareAgent(requirement_text=args.requirement)
    
    # Override output path if specified
    if args.output:
        agent.config.outputPath = args.output
    
    # Save config if requested
    if args.save_config:
        RequirementParser.save_config(agent.config, args.save_config)
    
    # Run comparison
    try:
        output_file = agent.run()
        print(f"\n✓ Success! Comparison report saved to: {output_file}")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Comparison failed: {str(e)}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()

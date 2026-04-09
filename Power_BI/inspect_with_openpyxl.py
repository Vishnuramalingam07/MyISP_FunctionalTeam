"""
Inspect files using openpyxl directly to avoid conversion errors
"""
import pandas as pd
from openpyxl import load_workbook

print("="*120)
print("INSPECTING IMSVT.xlsx STRUCTURE")
print("="*120)

try:
    wb_imsvt = load_workbook('IMSVT.xlsx', data_only=True)
    ws_imsvt = wb_imsvt.active
    print(f"✓ Loaded IMSVT.xlsx")
    print(f"  Active sheet: {ws_imsvt.title}")
    print(f"  Dimensions: {ws_imsvt.dimensions}")
    
    print(f"\nFirst 2 header rows:")
    for row_idx in range(1, 3):
        row_data = []
        for col_idx in range(1, min(15, ws_imsvt.max_column + 1)):
            cell_value = ws_imsvt.cell(row_idx, col_idx).value
            if cell_value:
                row_data.append(f"Col{col_idx}='{cell_value}'")
        print(f"  Row {row_idx}: {', '.join(row_data)}")
    
    print(f"\nSearching for 'Offshore Ratio' in IMSVT:")
    found_positions = []
    for row_idx in range(1, min(5, ws_imsvt.max_row + 1)):
        for col_idx in range(1, ws_imsvt.max_column + 1):
            cell_value = ws_imsvt.cell(row_idx, col_idx).value
            if cell_value and 'Offshore Ratio' in str(cell_value):
                found_positions.append((row_idx, col_idx, cell_value))
                print(f"  Found at Row {row_idx}, Col {col_idx}: '{cell_value}'")
    
    # If found, show the structure
    if found_positions:
        row, col, val = found_positions[0]
        print(f"\n  Structure for '{val}' column:")
        print(f"    Row {1}: {ws_imsvt.cell(1, col).value}")  # Main header
        print(f"    Row {2}: {ws_imsvt.cell(2, col).value}")  # Sub header
        print(f"    Row {3}: {ws_imsvt.cell(3, col).value}")  # First data value
        
except Exception as e:
    print(f"✗ Error loading IMSVT: {e}")

print("\n" + "="*120)
print("INSPECTING IMS Mgd Security_OCPKDA_0012026420.xlsm STRUCTURE")
print("="*120)

try:
    wb_ims = load_workbook('IMS Mgd Security_OCPKDA_0012026420.xlsm', data_only=True)
    print(f"✓ Loaded IMS file")
    print(f"  Sheet names: {wb_ims.sheetnames}")
    
    ws_ims = wb_ims['KeyDealAttributes']
    print(f"  Active sheet: {ws_ims.title}")
    print(f"  Dimensions: {ws_ims.dimensions}")
    
    print(f"\nFirst 3 rows (to understand header structure):")
    for row_idx in range(1, 4):
        row_data = []
        for col_idx in range(1, min(15, ws_ims.max_column + 1)):
            cell_value = ws_ims.cell(row_idx, col_idx).value
            if cell_value:
                row_data.append(f"Col{col_idx}='{cell_value}'")
        if row_data:
            print(f"  Row {row_idx}: {', '.join(row_data)}")
    
    print(f"\nSearching for 'Offshore Ratio' in IMS:")
    found_ims = []
    for row_idx in range(1, min(10, ws_ims.max_row + 1)):
        for col_idx in range(1, ws_ims.max_column + 1):
            cell_value = ws_ims.cell(row_idx, col_idx).value
            if cell_value and 'Offshore Ratio' in str(cell_value):
                found_ims.append((row_idx, col_idx, cell_value))
                print(f"  Found at Row {row_idx}, Col {col_idx}: '{cell_value}'")
    
    # If found, show the structure
    if found_ims:
        row, col, val = found_ims[0]
        print(f"\n  Structure for '{val}' column:")
        print(f"    Row {1}: {ws_ims.cell(1, col).value}")  # Main header
        print(f"    Row {2}: {ws_ims.cell(2, col).value}")  # Sub header (if exists)
        print(f"    Row {3}: {ws_ims.cell(3, col).value}")  # First data value
        
        # Look for sub-columns (Solution Standards, Actual Value, Variation From Standard)
        print(f"\n  Checking adjacent columns for sub-headers:")
        for offset in range(-2, 3):
            if col + offset >= 1:
                sub_val = ws_ims.cell(2, col + offset).value
                data_val = ws_ims.cell(3, col + offset).value
                if sub_val:
                    print(f"    Col {col+offset}: Sub='{sub_val}', Data='{data_val}'")
        
except Exception as e:
    print(f"✗ Error loading IMS file: {e}")

print("\n" + "="*120)
print("MAPPING VERIFICATION SUMMARY")
print("="*120)
print("The mapping file shows:")
print("  IMSVT: 'Managed Security - Offshore Ratio (%)' -> 'Guidance'")
print("  IMS:   'Offshore Ratio (%)' -> 'Solution Standards'")
print("\nWe need to verify:")
print("  1. Both files have these exact column structures")
print("  2. The values in corresponding cells match")

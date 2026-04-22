"""
Azure DevOps Test Plan Execution Report Generator - Custom Format (OPTIMIZED)
"""

import requests
import json
from datetime import datetime
import base64
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import shutil
import os
import importlib.util
from requests.adapters import HTTPAdapter
from openpyxl import load_workbook

# ============================================================================
# CONFIGURATION
# ============================================================================

from dotenv import load_dotenv
load_dotenv('ADO_SECRETS.env')

ADO_CONFIG = {
    'organization': 'accenturecio08',
    'project': 'AutomationProcess_29697',
    'plan_id': '4443950',
    'suite_id': '4443958',  # PT Execution suite ID
    'target_suite_name': 'PT Execution', 
    'pat_token': os.getenv('ADO_PAT_MAIN', ''),
    'max_workers': 20,  # Parallel API calls
}

if not ADO_CONFIG['pat_token']:
    raise ValueError("ADO_PAT_MAIN not found in ADO_SECRETS.env file")

# Get the directory where this script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Full path to the SharePoint Excel file (downloaded by download_sharepoint_file.py)
XLSX_FILE = os.path.join(SCRIPT_DIR, "PT Status excel.xlsx")

# Directory where all output reports and files are saved
OUTPUT_DIR = SCRIPT_DIR

# Skip SharePoint download if the Excel file is newer than this many hours
XLSX_MAX_AGE_HOURS = 6


# ============================================================================
# AZURE DEVOPS API CLIENT (OPTIMIZED)
# ============================================================================

class AzureDevOpsClient:
    def __init__(self, config):
        self.org = config['organization']
        self.project = config['project']
        self.plan_id = config['plan_id']
        self.suite_id = config['suite_id']
        self.target_suite_name = config.get('target_suite_name', 'PT Execution')
        self.pat = config['pat_token']
        self.max_workers = config.get('max_workers', 10)
        
        # Encode PAT for Basic Auth
        auth_bytes = f":{self.pat}".encode('ascii')
        encoded_pat = base64.b64encode(auth_bytes).decode('ascii')
        
        self.headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Basic {encoded_pat}',
            'Accept': 'application/json'
        }
        
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        # Increase connection pool for faster parallel requests
        _adapter = HTTPAdapter(pool_connections=30, pool_maxsize=30)
        self.session.mount('https://', _adapter)
        
        # Cache for work items to avoid duplicate fetches
        self.work_item_cache = {}
        self.cache_lock = threading.Lock()
        # Cache for all_suites to avoid duplicate API calls
        self._cached_suites = None
    
    def test_connection(self):
        """Test basic connection"""
        try:
            url = f"https://dev.azure.com/{self.org}/_apis/projects/{self.project}?api-version=7.0"
            print(f"\n🔍 Testing Connection...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                print(f"   ✅ Connected to: {data.get('name', 'Unknown')}")
                return True
            else:
                print(f"   ❌ Failed: {response.status_code}")
                return False
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return False
    
    def get_test_plan(self):
        """Fetch test plan details"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/testplan/plans/{self.plan_id}?api-version=7.0"
            
            print(f"\n📋 Fetching Test Plan {self.plan_id}...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                print(f"   ✅ Plan: {data.get('name', 'N/A')}")
                return data
            else:
                url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}?api-version=7.0"
                response = self.session.get(url, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    print(f"   ✅ Plan: {data.get('name', 'N/A')}")
                    return data
                    
                print(f"   ❌ Failed: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return None
    
    def get_all_suites_in_plan(self):
        """Get all suites in the plan with pagination support (cached)"""
        if self._cached_suites is not None:
            print(f"\n📦 Using cached suites ({len(self._cached_suites)} suites)")
            return self._cached_suites
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/testplan/plans/{self.plan_id}/suites?api-version=7.0"
            
            print(f"\n📦 Fetching All Suites...")
            
            all_suites = []
            continuation_token = None
            
            while True:
                request_url = url
                if continuation_token:
                    request_url = f"{url}&continuationToken={continuation_token}"
                
                response = self.session.get(request_url, timeout=30)
                
                if response.status_code != 200:
                    url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}/suites?api-version=7.0"
                    response = self.session.get(url, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    suites = data.get('value', [])
                    all_suites.extend(suites)
                    
                    continuation_token = response.headers.get('x-ms-continuationtoken')
                    if not continuation_token:
                        break
                else:
                    break
            
            print(f"   ✅ Found {len(all_suites)} suites")
            self._cached_suites = all_suites
            return all_suites
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return []
    
    def verify_suite_exists(self, suite_id):
        """Verify if a specific suite exists"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}/suites/{suite_id}?api-version=7.0"
            
            print(f"\n🔍 Verifying Suite {suite_id}...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                suite_name = data.get('name', 'Unknown')
                print(f"   ✅ Suite exists: {suite_name}")
                return data
            else:
                print(f"   ❌ Suite not accessible (Status: {response.status_code})")
                return None
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return None
    
    def get_test_points_from_suite(self, suite_id):
        """Fetch ALL test points from a suite with $top=1000 and pagination"""
        try:
            base_url = (f"https://dev.azure.com/{self.org}/{self.project}"
                        f"/_apis/test/plans/{self.plan_id}/suites/{suite_id}"
                        f"/points?$top=1000&api-version=7.0")
            all_points = []
            continuation = None

            while True:
                current_url = (f"{base_url}&continuationToken={continuation}"
                               if continuation else base_url)
                response = self.session.get(current_url, timeout=30)
                if response.status_code == 200:
                    data = response.json()
                    all_points.extend(data.get('value', []))
                    continuation = response.headers.get('x-ms-continuationtoken')
                    if not continuation:
                        break
                else:
                    break

            return all_points
        except Exception:
            return []
    
    def get_work_items_batch(self, work_item_ids):
        """Fetch multiple work items in a single API call (OPTIMIZED)"""
        if not work_item_ids:
            return {}
        
        try:
            # Remove duplicates and None values
            unique_ids = list(set([id for id in work_item_ids if id]))
            
            if not unique_ids:
                return {}
            
            # Azure DevOps supports batch requests with up to 200 IDs
            batch_size = 200
            all_work_items = {}
            batches = [unique_ids[i:i+batch_size] for i in range(0, len(unique_ids), batch_size)]

            def _fetch_wi_batch(batch_ids):
                ids_param = ','.join(map(str, batch_ids))
                url = (f"https://dev.azure.com/{self.org}/{self.project}"
                       f"/_apis/wit/workitems?ids={ids_param}"
                       f"&fields=System.Title,System.Tags,"
                       f"Microsoft.VSTS.TCM.AutomationStatus,System.State"
                       f"&api-version=7.0")
                try:
                    resp = self.session.get(url, timeout=30)
                    if resp.status_code == 200:
                        return {wi.get('id'): wi for wi in resp.json().get('value', [])}
                except Exception:
                    pass
                return {}

            with ThreadPoolExecutor(max_workers=min(5, len(batches))) as _ex:
                for result in _ex.map(_fetch_wi_batch, batches):
                    all_work_items.update(result)

            return all_work_items
                
        except Exception as e:
            print(f"      ⚠️  Batch fetch error: {e}")
            return {}
    
    def get_child_suites_from_cache(self, parent_suite_id, all_suites):
        """Get child suites from cached all_suites list (OPTIMIZED)"""
        try:
            child_suites = []
            parent_suite_id_int = int(parent_suite_id)
            
            for suite in all_suites:
                parent = suite.get('parent', {}) or suite.get('parentSuite', {})
                parent_id = None
                
                if isinstance(parent, dict):
                    parent_id = parent.get('id')
                elif isinstance(parent, (int, str)):
                    try:
                        parent_id = int(parent)
                    except:
                        pass
                
                try:
                    parent_id_int = int(parent_id) if parent_id else None
                except:
                    parent_id_int = parent_id
                
                if parent_id_int == parent_suite_id_int:
                    child_suites.append(suite)
            
            return child_suites
                
        except Exception as e:
            return []
    
    def _build_suite_tree(self, root_suite_id, root_suite_name, all_suites):
        """Build complete suite tree with metadata"""
        suite_tree = []
        
        def traverse_suite(suite_id, suite_name, parent_lead=None, parent_module=None, test_type=None, depth=0):
            """
            Traverse suite hierarchy:
            - Depth 0: Root (PT Execution)
            - Depth 1: Lead folders (Kavi, Pirtheebaa, etc.)
            - Depth 2: Deeper Nesting (test cases with Automation Status)
            
            Note: Test type (Planned/Not Automated) will be determined from 
                  work item's 'Microsoft.VSTS.TCM.AutomationStatus' field
            """
            
            # Determine current level type
            current_lead = parent_lead
            current_module = parent_module
            current_test_type = test_type  # Not used, determined from work item field
            
            if depth == 1:
                # This is a Lead folder
                current_lead = suite_name
                current_module = None
            elif depth >= 2:
                # Deeper nesting - use suite name as module
                current_module = suite_name
            
            suite_tree.append({
                'id': suite_id,
                'name': suite_name,
                'parent_lead': current_lead,
                'parent_module': current_module,
                'test_type': current_test_type,
                'depth': depth
            })
            
            # Only recurse into children up to depth 1 (stop at depth 2)
            if depth < 2:
                children = self.get_child_suites_from_cache(suite_id, all_suites)
                
                for child in children:
                    child_id = child.get('id')
                    child_name = child.get('name', 'Unknown')
                    
                    traverse_suite(child_id, child_name, 
                                 parent_lead=current_lead, 
                                 parent_module=current_module,
                                 test_type=current_test_type,
                                 depth=depth+1)
        
        traverse_suite(root_suite_id, root_suite_name)
        return suite_tree
    
    def _collect_test_points_parallel(self, suite_tree):
        """Collect test points from all suites in parallel"""
        all_test_items = []
        
        def fetch_suite_tests(suite_info):
            suite_id = suite_info['id']
            suite_name = suite_info['name']
            parent_lead = suite_info['parent_lead']
            parent_module = suite_info['parent_module']
            test_type = suite_info['test_type']
            
            test_items = self.get_test_points_from_suite(suite_id)
            
            result = []
            for item in test_items:
                # Pass lead, module, and test type information
                result.append((item, suite_name, parent_lead, parent_module, test_type))
            
            return result
        
        # Use thread pool for parallel fetching
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(fetch_suite_tests, suite): suite for suite in suite_tree}
            
            completed = 0
            total = len(suite_tree)
            
            for future in as_completed(futures):
                completed += 1
                if completed % 10 == 0 or completed == total:
                    print(f"      Progress: {completed}/{total} suites processed", end='\r')
                
                try:
                    items = future.result()
                    all_test_items.extend(items)
                except Exception as e:
                    pass
        
        print()  # New line after progress
        return all_test_items

    def get_all_test_data_from_suite(self, suite_id, suite_name):
        """Get all test data from suite and children (OPTIMIZED)"""
        print(f"\n📊 Collecting Test Data from: {suite_name}...")
        print(f"   ⚡ Using optimized batch processing...")
        
        # Fetch all suites once
        print(f"\n   Step 1: Fetching suite hierarchy...")
        all_suites = self.get_all_suites_in_plan()
        
        # Build suite hierarchy
        print(f"   Step 2: Building suite tree...")
        suite_tree = self._build_suite_tree(suite_id, suite_name, all_suites)
        
        # Collect all test points in parallel
        print(f"   Step 3: Collecting test points from {len(suite_tree)} suites...")
        all_test_items = self._collect_test_points_parallel(suite_tree)
        
        if not all_test_items:
            print(f"   ⚠️  No test points found")
            return []
        
        print(f"   ✓ Found {len(all_test_items)} test items")
        
        # Extract unique work item IDs
        print(f"   Step 4: Extracting work item IDs...")
        work_item_ids = []
        for item_data in all_test_items:
            # Unpack the tuple with 5 elements now
            item, item_suite_name, parent_lead, parent_module, test_type = item_data
            
            test_case = item.get('testCase', {})
            work_item_id = test_case.get('id') or item.get('testCaseReference', {}).get('id')
            if work_item_id:
                work_item_ids.append(work_item_id)
        
        # Fetch all work items in batch
        print(f"   Step 5: Fetching {len(set(work_item_ids))} unique work items in batch...")
        work_items_dict = self.get_work_items_batch(work_item_ids)
        print(f"   ✓ Retrieved {len(work_items_dict)} work items")
        
         # Process test data
        print(f"   Step 6: Processing test data...")
        all_test_data = []
        
        for item_data in all_test_items:
            # Unpack the tuple with 5 elements now
            item, item_suite_name, parent_lead, parent_module, suite_test_type = item_data
            
            test_case = item.get('testCase', {})
            work_item_id = test_case.get('id') or item.get('testCaseReference', {}).get('id')
            
            # Ensure work_item_id is int for dict lookup (API returns int keys)
            if work_item_id is not None:
                try:
                    work_item_id = int(work_item_id)
                except (ValueError, TypeError):
                    pass
            
            # Get work item details for automation status
            work_item = work_items_dict.get(work_item_id) if work_item_id else None
            automation_status = ''
            
            # Determine test type based on suite folder structure
            if suite_test_type:
                # Use test type from suite hierarchy (Automation or Manual folder)
                test_type = suite_test_type
            else:
                # Fallback: Check work item tags/automation status
                test_type = 'Manual'  # Default
                
                if work_item:
                    fields = work_item.get('fields', {})
                    tags = fields.get('System.Tags', '')
                    automation_status = fields.get('Microsoft.VSTS.TCM.AutomationStatus', '')
                    
                    if 'automation' in str(tags).lower() or 'automated' in str(automation_status).lower():
                        test_type = 'Automation'
            
            # Extract automation status if we have work item
            if work_item:
                fields = work_item.get('fields', {})
                automation_status = fields.get('Microsoft.VSTS.TCM.AutomationStatus', '')
            
            # Get outcome
            outcome = 'Not Run'
            if 'results' in item:
                outcome = item['results'].get('outcome', 'Not Run')
            elif 'lastResultOutcome' in item:
                outcome = item.get('lastResultOutcome', 'Not Run')
            elif 'outcome' in item:
                outcome = item.get('outcome', 'Not Run')
            
            # Get assigned to
            assigned_to = 'Unassigned'
            if 'assignedTo' in item and item['assignedTo']:
                assigned_to = item['assignedTo'].get('displayName', 'Unassigned')
            
            # Determine final lead and module names
            final_lead = parent_lead if parent_lead else 'Unassigned'
            final_module = parent_module if parent_module else item_suite_name
            
            # Override with assigned user's first name if available (only if no parent_lead)
            if assigned_to and assigned_to != 'Unassigned' and not parent_lead:
                try:
                    final_lead = assigned_to.split()[0]
                except:
                    final_lead = assigned_to
            
            all_test_data.append({
                'id': work_item_id or 'N/A',
                'name': test_case.get('name', 'N/A'),
                'suite': item_suite_name,
                'module': final_module,  # Use parent_module from hierarchy
                'lead': final_lead,      # Use parent_lead from hierarchy
                'type': test_type,       # Use test type from suite folder structure
                'outcome': outcome,
                'priority': item.get('priority', 2),
                'state': item.get('state', 'Active'),
                'automation_status': automation_status  # Include automation status
            })
        
        print(f"\n   ✅ Total test items collected: {len(all_test_data)}")
        return all_test_data

# ============================================================================
# HTML REPORT GENERATOR - CUSTOM FORMAT
# ============================================================================

class CustomHTMLReportGenerator:
    def __init__(self, test_data, plan_info=None, suite_name=None, insprint_data=None, 
                 planned_automation_ado=None, planned_automation_sp=None,
                 not_automated_ado=None, not_automated_sp=None,
                 us_count_all=None, us_count_planned=None, us_count_not_automated=None):
        self.test_data = test_data
        self.insprint_data = insprint_data or []
        self.plan_info = plan_info or {}
        self.suite_name = suite_name or 'Test Suite'
        self.timestamp = datetime.now().strftime("%B %d, %Y at %H:%M:%S")
        # Planned Automation data
        self.planned_automation_ado = planned_automation_ado or {}
        self.planned_automation_sp = planned_automation_sp or {}
        # Not Automated data
        self.not_automated_ado = not_automated_ado or {}
        self.not_automated_sp = not_automated_sp or {}
        # Suite counts from ADO
        self.us_count_all = us_count_all or {}
        self.us_count_planned = us_count_planned or {}
        self.us_count_not_automated = us_count_not_automated or {}
    
    def organize_data_by_lead_module(self):
        """Organize test data by Lead and Module"""
        organized = defaultdict(lambda: defaultdict(lambda: {
            'planned': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'not automated': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }))
        
        for test in self.test_data:
            lead = test['lead']
            module = test['module']
            test_type = test['type'].lower()
            outcome = test['outcome']
            
            # Map test type to match dictionary keys
            if 'planned' in test_type:
                test_type_key = 'planned'
            elif 'not automated' in test_type or 'manual' in test_type:
                test_type_key = 'not automated'
            else:
                test_type_key = 'not automated'  # Default to not automated
            
            # Increment total count
            organized[lead][module][test_type_key]['total'] += 1
            
            # Map outcome to correct category (case-insensitive matching)
            outcome_lower = outcome.lower()
            
            if outcome_lower in ['passed', 'pass']:
                organized[lead][module][test_type_key]['passed'] += 1
            elif outcome_lower in ['failed', 'fail']:
                organized[lead][module][test_type_key]['failed'] += 1
            elif outcome_lower in ['blocked', 'block']:
                organized[lead][module][test_type_key]['blocked'] += 1
            elif outcome_lower in ['not applicable', 'na', 'n/a', 'notapplicable']:
                organized[lead][module][test_type_key]['na'] += 1
            elif outcome_lower in ['not run', 'notrun', 'active', 'none', '']:
                organized[lead][module][test_type_key]['not_run'] += 1
            else:
                # Any unrecognized outcome goes to 'not_run'
                organized[lead][module][test_type_key]['not_run'] += 1
        
        return organized
    
    def calculate_percentages(self, data):
        """Calculate Pass% and Execution% for individual Lead/Module rows"""
        passed = data['passed']
        failed = data['failed']
        blocked = data['blocked']
        total = data['total']
        na = data['na']
        
        # Pass % = (Pass / (Pass + Fail + Blocked)) * 100
        denominator_pass = passed + failed + blocked
        pass_percentage = (passed / denominator_pass * 100) if denominator_pass > 0 else 0
        
        # Execution % = (Pass + Fail) / (Total - NA) * 100
        denominator_exec = total - na
        execution_percentage = ((passed + failed) / denominator_exec * 100) if denominator_exec > 0 else 0
        
        return pass_percentage, execution_percentage
    
    def calculate_grand_total_percentages(self, data):
        """Calculate Pass% and Execution% for Grand Total row (different formula)"""
        passed = data['passed']
        failed = data['failed']
        blocked = data['blocked']
        total = data['total']
        na = data['na']
        
        # Pass % = (Pass / (Pass + Fail + Blocked)) * 100
        denominator_pass = passed + failed + blocked
        pass_percentage = (passed / denominator_pass * 100) if denominator_pass > 0 else 0
        
        # Execution % = (Pass + Fail + Blocked) / (Total - NA) * 100
        numerator_exec = passed + failed + blocked
        denominator_exec = total - na
        execution_percentage = (numerator_exec / denominator_exec * 100) if denominator_exec > 0 else 0
        
        return pass_percentage, execution_percentage
    
    def calculate_grand_totals(self, organized_data):
        """Calculate grand totals"""
        grand_totals = {
            'planned': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'not automated': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }
        
        for lead_data in organized_data.values():
            for module_data in lead_data.values():
                for test_type in ['planned', 'not automated']:
                    for key in grand_totals[test_type]:
                        grand_totals[test_type][key] += module_data[test_type][key]
        
        return grand_totals
    
    def aggregate_data_by_lead(self, organized_data):
        """Aggregate data by lead (sum across all modules)"""
        lead_totals = defaultdict(lambda: {
            'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0, 'total': 0
        })
        
        for lead, modules in organized_data.items():
            for module, types in modules.items():
                # Sum planned and not automated counts
                for test_type in ['planned', 'not automated']:
                    data = types[test_type]
                    lead_totals[lead]['passed'] += data['passed']
                    lead_totals[lead]['failed'] += data['failed']
                    lead_totals[lead]['blocked'] += data['blocked']
                    lead_totals[lead]['na'] += data['na']
                    lead_totals[lead]['not_run'] += data['not_run']
                    lead_totals[lead]['total'] += data['total']
        
        return dict(lead_totals)
    
    def organize_data_by_lead_module_insprint(self):
        """Organize insprint test data by Lead and Module"""
        organized = defaultdict(lambda: defaultdict(lambda: {
            'planned': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'not automated': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }))
        
        for test in self.insprint_data:
            lead = test['lead']
            module = test['module']
            test_type = test['type'].lower()
            outcome = test['outcome']
            
            # Map test type to match dictionary keys
            if 'planned' in test_type:
                test_type_key = 'planned'
            elif 'not automated' in test_type or 'manual' in test_type:
                test_type_key = 'not automated'
            else:
                test_type_key = 'not automated'  # Default to not automated
            
            # Increment total count
            organized[lead][module][test_type_key]['total'] += 1
            
            # Map outcome to correct category (case-insensitive matching)
            outcome_lower = outcome.lower()
            
            if outcome_lower in ['passed', 'pass']:
                organized[lead][module][test_type_key]['passed'] += 1
            elif outcome_lower in ['failed', 'fail']:
                organized[lead][module][test_type_key]['failed'] += 1
            elif outcome_lower in ['blocked', 'block']:
                organized[lead][module][test_type_key]['blocked'] += 1
            elif outcome_lower in ['not applicable', 'na', 'n/a', 'notapplicable']:
                organized[lead][module][test_type_key]['na'] += 1
            elif outcome_lower in ['not run', 'notrun', 'active', 'none', '']:
                organized[lead][module][test_type_key]['not_run'] += 1
            else:
                # Any unrecognized outcome goes to 'not_run'
                organized[lead][module][test_type_key]['not_run'] += 1
        
        return organized
    
    def load_sharepoint_data_from_xlsx(self, xlsx_file, sheet_name='PT status'):
        """Load SharePoint test data from Excel file (with pre-aggregated counts)"""
        sp_data = []
        
        try:
            print(f"\n📥 Loading SharePoint data for Tab 2 from: {xlsx_file}")
            
            if not os.path.exists(xlsx_file):
                print(f"   ❌ ERROR: File not found: {xlsx_file}")
                print(f"   Current directory: {os.getcwd()}")
                return []
            
            wb = load_workbook(xlsx_file, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                print(f"   ⚠️ Sheet '{sheet_name}' not found in {xlsx_file}")
                print(f"   Available sheets: {wb.sheetnames}")
                wb.close()
                return []
            
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if not rows:
                print(f"   ⚠️ Sheet '{sheet_name}' is empty")
                return []
            
            # First row is header
            headers = [str(h).strip() if h else '' for h in rows[0]]
            print(f"   📋 Found {len(rows)-1} data rows")
            print(f"   📝 Headers: {', '.join(headers[:10])}..." if len(headers) > 10 else f"   📝 Headers: {', '.join(headers)}")
            
            # Check for required columns
            required_cols = ['PT Lead', 'Module (TextVerification)', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Total']
            missing_cols = [col for col in required_cols if col not in headers]
            if missing_cols:
                print(f"   ⚠️ WARNING: Missing expected columns: {', '.join(missing_cols)}")
                print(f"   This may cause data loading issues.")
            
            rows_with_data = 0
            rows_skipped = 0
            
            for row in rows[1:]:
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                
                lead = str(row_dict.get('PT Lead', '') or '').strip()
                module = str(row_dict.get('Module (TextVerification)', '') or '').strip()
                
                if not lead or not module:
                    rows_skipped += 1
                    continue
                
                try:
                    passed = int(row_dict.get('Passed', 0) or 0)
                    failed = int(row_dict.get('Failed', 0) or 0)
                    blocked = int(row_dict.get('Blocked', 0) or 0)
                    not_run = int(row_dict.get('Not Run', 0) or 0)
                    total = int(row_dict.get('Total', 0) or 0)
                except (ValueError, TypeError) as e:
                    rows_skipped += 1
                    continue
                
                rows_with_data += 1
                sp_data.append({
                    'lead': lead,
                    'module': module,
                    'type': 'Not Automated',
                    'passed': passed,
                    'failed': failed,
                    'blocked': blocked,
                    'na': 0,
                    'not_run': not_run,
                    'total': total
                })
            
            print(f"   ✅ Loaded {len(sp_data)} module records from SharePoint Excel (sheet: '{sheet_name}')")
            print(f"   📊 Processed: {rows_with_data} rows loaded, {rows_skipped} rows skipped")
            
            if sp_data:
                lead_summary = defaultdict(lambda: {'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0, 'total': 0})
                for item in sp_data:
                    lead = item['lead']
                    lead_summary[lead]['passed'] += item['passed']
                    lead_summary[lead]['failed'] += item['failed']
                    lead_summary[lead]['blocked'] += item['blocked']
                    lead_summary[lead]['not_run'] += item['not_run']
                    lead_summary[lead]['total'] += item['total']
                
                print("\n   📊 SharePoint Data Summary (Lead-wise):")
                print(f"   {'Lead':<20} {'Total':<8} {'Passed':<8} {'Failed':<8} {'Blocked':<8} {'Not Run':<8}")
                print("   " + "-" * 78)
                for lead in sorted(lead_summary.keys()):
                    counts = lead_summary[lead]
                    print(f"   {lead:<20} {counts['total']:<8} {counts['passed']:<8} {counts['failed']:<8} {counts['blocked']:<8} {counts['not_run']:<8}")
            else:
                print(f"   ⚠️ No valid data found in SharePoint Excel file")
            
            return sp_data
            
        except FileNotFoundError:
            print(f"   ⚠️ SharePoint Excel file not found: {xlsx_file}")
            print("   Report will show zeros for SharePoint data.")
            return []
        except Exception as e:
            print(f"   ❌ Error loading SharePoint data: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def organize_sp_data_by_lead_module(self, sp_data):
        """Organize SharePoint data by Lead and Module (data is already aggregated)"""
        organized = defaultdict(lambda: defaultdict(lambda: {
            'planned': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'not automated': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }))
        
        for item in sp_data:
            lead = item['lead']
            module = item['module']
            test_type = item['type'].lower()
            
            # Skip rows with empty lead or module
            if not lead or not module:
                continue
            
            # Ensure test_type is valid
            if 'planned' in test_type:
                test_type_key = 'planned'
            elif 'not automated' in test_type or 'manual' in test_type:
                test_type_key = 'not automated'
            else:
                test_type_key = 'not automated'  # Default
            
            # Add pre-aggregated counts to the organized structure
            organized[lead][module][test_type_key]['passed'] += item['passed']
            organized[lead][module][test_type_key]['failed'] += item['failed']
            organized[lead][module][test_type_key]['blocked'] += item['blocked']
            organized[lead][module][test_type_key]['na'] += item['na']
            organized[lead][module][test_type_key]['not_run'] += item['not_run']
            organized[lead][module][test_type_key]['total'] += item['total']
        
        return dict(organized)
    
    def load_sharepoint_insprint_yes_data(self, xlsx_file, sheet_name='PT status'):
        """Load SharePoint data where Insprint Yes/No = 'Yes'"""
        sp_insprint_data = []
        
        try:
            wb = load_workbook(xlsx_file, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                print(f"\n⚠️ Sheet '{sheet_name}' not found in {xlsx_file}")
                wb.close()
                return []
            
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if not rows:
                return []
            
            headers = [str(h).strip() if h else '' for h in rows[0]]
            
            for row in rows[1:]:
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                
                lead = str(row_dict.get('PT Lead', '') or '').strip()
                insprint_status = str(row_dict.get('Insprint Yes/No', '') or '').strip()
                
                if insprint_status.lower() != 'yes' or not lead:
                    continue
                
                try:
                    passed = int(row_dict.get('Passed', 0) or 0)
                    failed = int(row_dict.get('Failed', 0) or 0)
                    blocked = int(row_dict.get('Blocked', 0) or 0)
                    not_run = int(row_dict.get('Not Run', 0) or 0)
                    total = int(row_dict.get('Total', 0) or 0)
                except (ValueError, TypeError):
                    passed = 0
                    failed = 0
                    blocked = 0
                    not_run = 0
                    total = 1
                
                sp_insprint_data.append({
                    'lead': lead,
                    'passed': passed,
                    'failed': failed,
                    'blocked': blocked,
                    'not_run': not_run,
                    'total': total
                })
            
            print(f"\n✅ Loaded {len(sp_insprint_data)} Insprint='Yes' records from SharePoint Excel (sheet: '{sheet_name}')")
            return sp_insprint_data
            
        except (FileNotFoundError, Exception) as e:
            print(f"\n⚠️ Error loading SharePoint Insprint data: {e}")
            return []
    
    def generate_html(self):
        """Generate HTML report - Compact Design with Filters and Tabs"""
        organized_data = self.organize_data_by_lead_module()
        grand_totals = self.calculate_grand_totals(organized_data)
        manual_gt = grand_totals['not automated']
        auto_gt = grand_totals['planned']
        gt_pass_pct, gt_exec_pct = self.calculate_grand_total_percentages(manual_gt)
        auto_gt_pass_pct, auto_gt_exec_pct = self.calculate_grand_total_percentages(auto_gt)
        # Calculate insprint data for summary
        insprint_organized = self.organize_data_by_lead_module_insprint()
        insprint_grand_totals = self.calculate_grand_totals(insprint_organized)
        insprint_manual_gt = insprint_grand_totals['not automated']
        insprint_gt_pass_pct, insprint_gt_exec_pct = self.calculate_grand_total_percentages(insprint_manual_gt)
        plan_name = self.plan_info.get('name', f"Test Plan {ADO_CONFIG['plan_id']}")
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PT Execution Count Comparison Report - {self.suite_name}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: #f0f2f5;
            padding: 10px;
        }}
        .container {{
            max-width: 100%;
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px 15px;
            text-align: center;
        }}
        .header h1 {{ font-size: 18px; margin-bottom: 4px; color: white; }}
        .header p {{ font-size: 11px; opacity: 0.9; margin: 2px 0; color: white; }}
        .header a {{ color: white; text-decoration: underline; font-weight: 600; }}
        .header a:hover {{ color: #f0f0f0; text-decoration: underline; }}
        
        .tabs {{
            display: flex;
            background: #e9ecef;
            border-bottom: 2px solid #dee2e6;
        }}
        
        .tab {{
            padding: 10px 20px;
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            border: none;
            background: transparent;
            color: #495057;
            transition: all 0.3s;
            border-bottom: 3px solid transparent;
        }}
        
        .tab:hover {{
            background: #dee2e6;
        }}
        
        .tab.active {{
            background: white;
            color: #667eea;
            border-bottom: 3px solid #667eea;
        }}
        
        .tab-content {{
            display: none;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        .filter-section {{
            padding: 10px 15px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
            position: relative;
            overflow: visible;
        }}
        
        .filter-group {{
            display: flex;
            align-items: center;
            gap: 8px;
            position: relative;
        }}
        
        .filter-group label {{
            font-size: 11px;
            font-weight: 600;
            color: #495057;
        }}
        
        .filter-group select {{
            padding: 5px 10px;
            font-size: 11px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            background: white;
            cursor: pointer;
            min-width: 150px;
        }}
        
        .filter-group select:focus {{
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 2px rgba(102, 126, 234, 0.1);
        }}
        
        .reset-btn {{
            padding: 5px 15px;
            font-size: 11px;
            background: #667eea;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
        }}
        
        .reset-btn:hover {{
            background: #5568d3;
        }}
        
        .filter-info {{
            margin-left: auto;
            font-size: 11px;
            color: #6c757d;
            font-weight: 600;
        }}
        
        /* Custom Dropdown Styles */
        .custom-dropdown {{
            position: relative;
            display: inline-block;
            z-index: 100;
        }}
        
        .dropdown-toggle {{
            padding: 5px 10px;
            font-size: 11px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            background: white;
            cursor: pointer;
            min-width: 150px;
            text-align: left;
            font-weight: 600;
        }}
        
        .dropdown-toggle:hover {{
            background: #f8f9fa;
        }}
        
        .dropdown-menu {{
            display: none;
            position: absolute;
            top: 100%;
            left: 0;
            min-width: 200px;
            max-height: 350px;
            overflow-y: auto;
            overflow-x: hidden;
            background: white;
            border: 1px solid #ced4da;
            border-radius: 4px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
            z-index: 9999;
            margin-top: 2px;
        }}
        
        .dropdown-menu.show {{
            display: block;
        }}
        
        .dropdown-item {{
            padding: 6px 12px;
        }}
        
        .dropdown-item label {{
            display: flex;
            align-items: center;
            cursor: pointer;
            font-size: 11px;
            margin: 0;
            font-weight: normal;
        }}
        
        .dropdown-item label:hover {{
            background: #f8f9fa;
        }}
        
        .dropdown-item input[type="checkbox"] {{
            margin-right: 8px;
            cursor: pointer;
        }}
        
        .dropdown-divider {{
            height: 1px;
            background: #e9ecef;
            margin: 4px 0;
        }}
        
        .report-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 0;
            font-size: 10px;
        }}
        
        .report-table th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: black;
            padding: 6px 4px;
            text-align: center;
            font-size: 9px;
            font-weight: 600;
            border: 2px solid #000 !important;
            line-height: 1.2;
        }}
        
        .report-table th.main-header {{
            background: linear-gradient(135deg, #4299e1 0%, #667eea 100%);
            font-size: 10px;
            padding: 5px 3px;
        }}
        
        .report-table td {{
            padding: 4px 3px;
            text-align: center;
            border: 2px solid #000 !important;
            font-size: 10px;
            line-height: 1.3;
        }}
        
        .report-table tbody tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        
        .report-table tbody tr:hover {{
            background: #e9ecef;
        }}
        
        .report-table tbody tr.hidden {{
            display: none;
        }}
        
        .comparison-table tbody tr.hidden {{
            display: none !important;
        }}
        
        .sno-col {{ width: 30px; font-weight: 600; }}
        .lead-col {{ width: 70px; font-weight: 600; font-size: 9px; }}
        .module-col {{ width: 100px; font-size: 9px; }}
        .total-p1p2-col {{ width: 50px; font-weight: 700; background: #fd7e14 !important; color: white !important; }}
        .total-col {{ width: 40px; font-weight: 600; background: #e7f3ff !important; }}
        .pass-col {{ width: 35px; background: #d4edda !important; }}
        .fail-col {{ width: 35px; background: #f8d7da !important; }}
        .blocked-col {{ width: 35px; background: #fff3cd !important; }}
        .na-col {{ width: 35px; background: #e2e3e5 !important; }}
        .notrun-col {{ width: 40px; background: #cce5ff !important; }}
        .percentage-col {{ width: 45px; font-weight: 600; background: #fff9e6 !important; font-size: 9px; }}
        
        .grand-total-row {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            color: black !important;
            font-weight: 700 !important;
            font-size: 10px !important;
        }}

        /* ========================================================================
           LEADS WISE EXECUTION STATUS TABLE STYLES (TAB 3 & TAB 4)
           Compact Professional Design
           ======================================================================== */
        
        .leads-table {{
            width: auto;
            max-width: 700px;
            margin: 10px auto;
            border-collapse: collapse;
            font-size: 10px;
            background: white;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        }}
        
        /* Header Row - Compact Purple Gradient */
        .leads-table thead {{
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        .leads-table th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
            color: black !important;
            padding: 5px 6px;
            text-align: center;
            font-size: 9px;
            font-weight: 700;
            border: 2px solid #000 !important;
            line-height: 1.1;
            text-transform: uppercase;
            letter-spacing: 0.3px;
            white-space: nowrap;
        }}
        
        /* Data Cell Base Styles - Compact */
        .leads-table td {{
            padding: 4px 6px;
            text-align: center;
            border: 2px solid #000 !important;
            font-size: 10px;
            font-weight: 400;
            line-height: 1.2;
            color: #333 !important;
        }}
        
        /* Alternating Row Colors - Like Tab 1 */
        .leads-table tbody tr:nth-child(odd) td {{
            background: white !important;
        }}
        
        .leads-table tbody tr:nth-child(even) td {{
            background: #f8f9fa !important;
        }}
        
        /* Hover Effect - Matching Tab 1 */
        .leads-table tbody tr:not(.grand-total-row):hover td {{
            background: #e9ecef !important;
            cursor: pointer;
        }}
        
        /* Lead Name Column - Compact */
        .leads-table .lead-name-col {{
            width: 90px;
            font-weight: 700;
            text-align: left;
            padding-left: 8px;
            font-size: 10px;
            color: #1a1a1a !important;
            background: #e3f2fd !important;
        }}
        
        /* Pass Column - Compact */
        .leads-table .pass-col {{
            width: 40px;
            background: #d4edda !important;
            color: #155724 !important;
            font-weight: 600;
        }}
        
        /* Fail Column - Compact */
        .leads-table .fail-col {{
            width: 40px;
            background: #f8d7da !important;
            color: #721c24 !important;
            font-weight: 600;
        }}
        
        /* Blocked Column - Compact */
        .leads-table .blocked-col {{
            width: 40px;
            background: #fff3cd !important;
            color: #856404 !important;
            font-weight: 600;
        }}
        
        /* NA Column - Compact */
        .leads-table .na-col {{
            width: 40px;
            background: #e2e3e5 !important;
            color: #383d41 !important;
            font-weight: 600;
        }}
        
        /* Not Run Column - Compact */
        .leads-table .notrun-col {{
            width: 45px;
            background: #cce5ff !important;
            color: #004085 !important;
            font-weight: 600;
        }}
        
        /* Total Column - Compact */
        .leads-table .total-col {{
            width: 45px;
            font-weight: 700;
            background: #e7f3ff !important;
            color: #1a1a1a !important;
        }}
        
        /* Percentage Columns - Compact */
        .leads-table .percentage-col {{
            width: 50px;
            font-weight: 700;
            background: #fff9e6 !important;
            color: #1a1a1a !important;
            font-size: 10px;
        }}
        
        /* Grand Total Row - Compact Green Gradient */
        .leads-table .grand-total-row {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            font-weight: 700 !important;
            font-size: 10px !important;
        }}
        
        .leads-table .grand-total-row td {{
            border: 2px solid #000 !important;
            padding: 5px 6px;
            color: black !important;
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            font-weight: 700 !important;
            text-shadow: none;
        }}
        
        .leads-table .grand-total-row .lead-name-col {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            color: black !important;
            font-weight: 700 !important;
            text-align: left;
            padding-left: 8px;
            text-transform: uppercase;
        }}
        
        /* Override all status column colors for grand total */
        .leads-table .grand-total-row .pass-col,
        .leads-table .grand-total-row .fail-col,
        .leads-table .grand-total-row .blocked-col,
        .leads-table .grand-total-row .na-col,
        .leads-table .grand-total-row .notrun-col,
        .leads-table .grand-total-row .total-col,
        .leads-table .grand-total-row .percentage-col {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            color: black !important;
            font-weight: 700 !important;
        }}
        
        /* Ensure grand total ignores alternating colors */
        .leads-table tbody tr.grand-total-row:nth-child(even) td,
        .leads-table tbody tr.grand-total-row:nth-child(odd) td {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
        }}
        
        /* Grand total hover effect */
        .leads-table tbody tr.grand-total-row:hover td {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            cursor: default;
        }}
        
        /* ========================================================================
           SUMMARY TABLE STYLES (TAB 1)
           Compact Professional Design
           ======================================================================== */
        
        .summary-table {{
            width: 100%;
            max-width: 900px;
            margin: 15px auto;
            border-collapse: collapse;
            font-size: 10px;
            box-shadow: 0 3px 10px rgba(0,0,0,0.12);
            border-radius: 6px;
            overflow: hidden;
        }}
        
        /* Summary Table Header */
        .summary-table thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }}
        
        .summary-table th {{
            padding: 8px 10px;
            text-align: center;
            font-size: 9px;
            font-weight: 700;
            border: 2px solid #1e293b !important;
            color: black !important;
            text-transform: uppercase;
            letter-spacing: 0.6px;
            line-height: 1.2;
            text-shadow: none;
        }}
        
        /* Summary Table Data Cells */
        .summary-table td {{
            padding: 10px 8px;
            text-align: center;
            font-size: 11px;
            font-weight: 600;
            border: 2px solid #1e293b !important;
            color: white !important;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
            transition: all 0.3s ease;
        }}
        
        .summary-table tbody tr:hover td {{
            transform: scale(1.01);
            box-shadow: 0 3px 6px rgba(0,0,0,0.15);
            cursor: pointer;
        }}
        
        /* Pass Column - Enhanced Green */
        .summary-table .summary-pass {{
            background: linear-gradient(135deg, #10b981 0%, #34d399 100%) !important;
            box-shadow: inset 0 2px 4px rgba(16, 185, 129, 0.3);
        }}
        
        .summary-table .summary-pass:hover {{
            background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        }}
        
        /* Fail Column - Enhanced Red */
        .summary-table .summary-fail {{
            background: linear-gradient(135deg, #ef4444 0%, #f87171 100%) !important;
            box-shadow: inset 0 2px 4px rgba(239, 68, 68, 0.3);
        }}
        
        .summary-table .summary-fail:hover {{
            background: linear-gradient(135deg, #dc2626 0%, #ef4444 100%) !important;
        }}
        
        /* Blocked Column - Enhanced Yellow/Orange */
        .summary-table .summary-blocked {{
            background: linear-gradient(135deg, #f59e0b 0%, #fbbf24 100%) !important;
            box-shadow: inset 0 2px 4px rgba(245, 158, 11, 0.3);
        }}
        
        .summary-table .summary-blocked:hover {{
            background: linear-gradient(135deg, #d97706 0%, #f59e0b 100%) !important;
        }}
        
        /* NA Column - Enhanced Gray */
        .summary-table .summary-na {{
            background: linear-gradient(135deg, #6b7280 0%, #9ca3af 100%) !important;
            box-shadow: inset 0 2px 4px rgba(107, 114, 128, 0.3);
        }}
        
        .summary-table .summary-na:hover {{
            background: linear-gradient(135deg, #4b5563 0%, #6b7280 100%) !important;
        }}
        
        /* Not Run Column - Enhanced Blue */
        .summary-table .summary-notrun {{
            background: linear-gradient(135deg, #3b82f6 0%, #60a5fa 100%) !important;
            box-shadow: inset 0 2px 4px rgba(59, 130, 246, 0.3);
        }}
        
        .summary-table .summary-notrun:hover {{
            background: linear-gradient(135deg, #2563eb 0%, #3b82f6 100%) !important;
        }}
        
        /* Total Column - Enhanced Cyan */
        .summary-table .summary-total {{
            background: linear-gradient(135deg, #06b6d4 0%, #22d3ee 100%) !important;
            box-shadow: inset 0 2px 4px rgba(6, 182, 212, 0.3);
            font-size: 12px;
            font-weight: 700;
        }}
        
        .summary-table .summary-total:hover {{
            background: linear-gradient(135deg, #0891b2 0%, #06b6d4 100%) !important;
        }}
        
        /* Pass% Column - Enhanced Lime Green */
        .summary-table .summary-pass-pct {{
            background: linear-gradient(135deg, #84cc16 0%, #a3e635 100%) !important;
            box-shadow: inset 0 2px 4px rgba(132, 204, 22, 0.3);
            font-size: 12px;
            font-weight: 700;
        }}
        
        .summary-table .summary-pass-pct:hover {{
            background: linear-gradient(135deg, #65a30d 0%, #84cc16 100%) !important;
        }}
        
        /* Exec% Column - Enhanced Emerald Green */
        .summary-table .summary-exec-pct {{
            background: linear-gradient(135deg, #10b981 0%, #34d399 100%) !important;
            box-shadow: inset 0 2px 4px rgba(16, 185, 129, 0.3);
            font-size: 12px;
            font-weight: 700;
        }}
        
        .summary-table .summary-exec-pct:hover {{
            background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        }}
        
        /* Responsive Design for Summary Table */
        @media (max-width: 1400px) {{
            .summary-table {{
                font-size: 9px;
                margin: 12px auto;
                max-width: 800px;
            }}
            
            .summary-table th {{
                font-size: 8px;
                padding: 7px 8px;
            }}
            
            .summary-table td {{
                font-size: 10px;
                padding: 8px 6px;
            }}
            
            .summary-table .summary-total,
            .summary-table .summary-pass-pct,
            .summary-table .summary-exec-pct {{
                font-size: 11px;
            }}
        }}
        
        @media (max-width: 1024px) {{
            .summary-table {{
                font-size: 8px;
                margin: 10px auto;
                max-width: 700px;
            }}
            
            .summary-table th {{
                font-size: 7px;
                padding: 6px 7px;
            }}
            
            .summary-table td {{
                font-size: 9px;
                padding: 7px 5px;
            }}
            
            .summary-table .summary-total,
            .summary-table .summary-pass-pct,
            .summary-table .summary-exec-pct {{
                font-size: 10px;
            }}
        }}
        
        @media (max-width: 768px) {{
            .summary-table {{
                font-size: 7px;
                margin: 8px;
                max-width: 100%;
            }}
            
            .summary-table th {{
                font-size: 6px;
                padding: 5px 6px;
            }}
            
            .summary-table td {{
                font-size: 8px;
                padding: 6px 4px;
            }}
            
            .summary-table .summary-total,
            .summary-table .summary-pass-pct,
            .summary-table .summary-exec-pct {{
                font-size: 9px;
            }}
        }}
        
        /* Print Styles for Summary Table */
        @media print {{
            .summary-table {{
                font-size: 9px;
                box-shadow: none;
                border-radius: 0;
            }}
            
            .summary-table th {{
                font-size: 8px;
                padding: 6px 8px;
                background: #667eea !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .summary-table td {{
                padding: 7px 8px;
                font-size: 10px;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .summary-table tbody tr:hover td {{
                transform: none;
                box-shadow: none;
            }}
        }}
        
        /* Lead-wise Comparison Table Styles */
        .comparison-table {{
            width: 100%;
            max-width: 1200px;
            margin: 20px auto;
            border-collapse: collapse;
            font-size: 11px;
            background: white;
            box-shadow: 0 2px 6px rgba(0,0,0,0.1);
        }}
        
        .comparison-table thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: black;
        }}
        
        .comparison-table th {{
            padding: 10px 8px;
            text-align: center;
            font-size: 10px;
            font-weight: 700;
            border: 2px solid #000 !important;
            line-height: 1.2;
        }}
        
        .comparison-table td {{
            padding: 8px;
            text-align: center;
            border: 2px solid #000 !important;
            font-size: 11px;
        }}
        
        .comparison-table tbody tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        
        .comparison-table tbody tr:hover {{
            background: #e9ecef;
        }}
        
        .comparison-table .lead-name {{
            text-align: left;
            font-weight: 700;
            padding-left: 15px;
            background: #e3f2fd !important;
        }}
        
        .comparison-table .source-header {{
            background: #4a90e2 !important;
            font-size: 11px;
            font-weight: 800;
        }}
        
        .comparison-table .ado-col {{
            background: #d4edda !important;
        }}
        
        .comparison-table .sp-col {{
            background: #fff3cd !important;
        }}
        
        .comparison-table .mismatch {{
            background: #f8d7da !important;
            font-weight: 700;
            color: #721c24;
        }}
        
        .comparison-table .match {{
            background: #d4edda !important;
            font-weight: 700;
            color: #155724;
        }}

        /* Status Column */
        .status-matched {{
            background: #d4edda !important;
            color: #155724 !important;
            font-weight: 700;
            font-size: 10px;
            min-width: 95px;
            text-align: center;
        }}
        .status-not-matched {{
            background: #f8d7da !important;
            color: #721c24 !important;
            font-weight: 700;
            font-size: 10px;
            min-width: 95px;
            text-align: center;
        }}
        .comparison-filter-bar {{
            padding: 8px 15px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
            display: flex;
            gap: 12px;
            align-items: center;
            flex-wrap: wrap;
        }}
        .comparison-filter-bar label {{
            font-size: 11px;
            font-weight: 600;
            color: #495057;
        }}
        .comparison-filter-bar select {{
            padding: 5px 10px;
            font-size: 11px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            background: white;
            cursor: pointer;
            min-width: 155px;
        }}
        .filter-count-inline {{
            margin-left: auto;
            font-size: 11px;
            color: #6c757d;
            font-weight: 600;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 PT Execution Count Comparison Report</h1>
            <p><strong>{plan_name}</strong> | Generated: {self.timestamp}</p>
            <p style="margin-top: 8px; font-size: 12px;">
                📌 <a href="https://dev.azure.com/{ADO_CONFIG['organization']}/{ADO_CONFIG['project']}/_testPlans/execute?planId={ADO_CONFIG['plan_id']}&suiteId={ADO_CONFIG['suite_id']}" target="_blank">View in Azure DevOps</a> | 
                📊 <a href="https://ts.accenture.com/:x:/r/sites/mySPTesting/_layouts/15/Doc.aspx?sourcedoc=%7BF8AEAC81-9574-4452-A9F0-BE10A87EF602%7D&file=Mar%2028th_Release.xlsx&action=default&mobileredirect=true" target="_blank">View SharePoint Data</a>
            </p>
        </div>
        
        <!-- Tabs Navigation -->
        <div class="tabs">
            <button class="tab active" onclick="switchTab(event, 'summary-tab')">📊 Overall Summary</button>
            <button class="tab" onclick="switchTab(event, 'comparison-tab')">📋 Lead-wise Comparison</button>
            <button class="tab" onclick="switchTab(event, 'planned-automation-tab')">🎯 Insprint Automation Test Cases</button>
            <button class="tab" onclick="switchTab(event, 'not-automated-tab')">🚫 Manual Test Cases</button>
        </div>
        
        <!-- Tab 1: Overall Summary -->
        <div id="summary-tab" class="tab-content active">
            <div style="padding: 20px;">
                <h2 style="text-align: center; color: #495057; margin-bottom: 20px; font-size: 16px;">📊 Overall Test Count Summary (ADO vs SharePoint)</h2>
                
                <table class="comparison-table" style="max-width: 900px;">
                    <thead>
                        <tr>
                            <th style="width: 200px;">Data Source</th>
                            <th class="ado-col">Passed</th>
                            <th class="ado-col">Failed</th>
                            <th class="ado-col">Not Run</th>
                            <th class="ado-col">Blocked</th>
                            <th class="ado-col" style="font-weight: 800;">Total</th>
                        </tr>
                    </thead>
                    <tbody>
"""
        
        # Get lead-wise aggregated data from ADO
        organized_data = self.organize_data_by_lead_module()
        lead_totals_ado = self.aggregate_data_by_lead(organized_data)
        
        # Load and aggregate SharePoint data from Excel
        sp_data = self.load_sharepoint_data_from_xlsx(XLSX_FILE, 'PT status')
        if sp_data:
            organized_sp_data = self.organize_sp_data_by_lead_module(sp_data)
            lead_totals_sp = self.aggregate_data_by_lead(organized_sp_data)
        else:
            # Fallback to zeros if no SP data available
            lead_totals_sp = {}
        
        # Get all unique leads from both ADO and SP
        all_leads = set(lead_totals_ado.keys()) | set(lead_totals_sp.keys())
        
        # Calculate grand totals for summary
        grand_total_ado = {
            'passed': sum(lead['passed'] for lead in lead_totals_ado.values()),
            'failed': sum(lead['failed'] for lead in lead_totals_ado.values()),
            'not_run': sum(lead['not_run'] for lead in lead_totals_ado.values()),
            'na': sum(lead['na'] for lead in lead_totals_ado.values()),
            'blocked': sum(lead['blocked'] for lead in lead_totals_ado.values()),
            'total': sum(lead['total'] for lead in lead_totals_ado.values())
        }
        
        grand_total_sp = {
            'passed': sum(lead['passed'] for lead in lead_totals_sp.values()),
            'failed': sum(lead['failed'] for lead in lead_totals_sp.values()),
            'not_run': sum(lead['not_run'] for lead in lead_totals_sp.values()),
            'na': sum(lead['na'] for lead in lead_totals_sp.values()),
            'blocked': sum(lead['blocked'] for lead in lead_totals_sp.values()),
            'total': sum(lead['total'] for lead in lead_totals_sp.values())
        }
        
        # Add ADO row
        html += f"""
                        <tr>
                            <td class="lead-name" style="background: #d4edda !important;">Azure DevOps (ADO)</td>
                            <td class="ado-col" style="font-weight: 700; font-size: 12px;">{grand_total_ado['passed']}</td>
                            <td class="ado-col" style="font-weight: 700; font-size: 12px;">{grand_total_ado['failed']}</td>
                            <td class="ado-col" style="font-weight: 700; font-size: 12px;">{grand_total_ado['not_run']}</td>
                            <td class="ado-col" style="font-weight: 700; font-size: 12px;">{grand_total_ado['blocked']}</td>
                            <td class="ado-col" style="font-weight: 800; font-size: 13px;">{grand_total_ado['total']}</td>
                        </tr>
"""
        
        # Add SP row
        html += f"""
                        <tr>
                            <td class="lead-name" style="background: #fff3cd !important;">SharePoint (SP)</td>
                            <td class="sp-col" style="font-weight: 700; font-size: 12px;">{grand_total_sp['passed']}</td>
                            <td class="sp-col" style="font-weight: 700; font-size: 12px;">{grand_total_sp['failed']}</td>
                            <td class="sp-col" style="font-weight: 700; font-size: 12px;">{grand_total_sp['not_run']}</td>
                            <td class="sp-col" style="font-weight: 700; font-size: 12px;">{grand_total_sp['blocked']}</td>
                            <td class="sp-col" style="font-weight: 800; font-size: 13px;">{grand_total_sp['total']}</td>
                        </tr>
"""
        
        # Calculate difference
        diff_passed = grand_total_ado['passed'] - grand_total_sp['passed']
        diff_failed = grand_total_ado['failed'] - grand_total_sp['failed']
        diff_not_run = grand_total_ado['not_run'] - grand_total_sp['not_run']
        diff_blocked = grand_total_ado['blocked'] - grand_total_sp['blocked']
        diff_total = grand_total_ado['total'] - grand_total_sp['total']
        
        # Add Difference row
        html += f"""
                        <tr style="background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black; font-weight: 700;">
                            <td style="text-align: left; padding-left: 15px; background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black;">Difference (ADO - SP)</td>
                            <td style="background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black; font-size: 12px;">{diff_passed:+d}</td>
                            <td style="background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black; font-size: 12px;">{diff_failed:+d}</td>
                            <td style="background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black; font-size: 12px;">{diff_not_run:+d}</td>
                            <td style="background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black; font-size: 12px;">{diff_blocked:+d}</td>
                            <td style="background: linear-gradient(90deg, #6c757d 0%, #adb5bd 100%) !important; color: black; font-weight: 800; font-size: 13px;">{diff_total:+d}</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Tab 2: Lead-wise Count Comparison -->
        <div id="comparison-tab" class="tab-content">
            <div style="padding: 20px;">
                <h2 style="text-align: center; color: #495057; margin-bottom: 20px; font-size: 16px;">📊 Lead-wise Test Count Comparison (ADO vs SharePoint)</h2>
                
                <div class="comparison-filter-bar">
                    <label>Filter by Status:</label>
                    <select id="statusFilterTab2" onchange="applyStatusFilter('comparison-tab')">
                        <option value="all">-- All Status --</option>
                        <option value="Matched">Matched</option>
                        <option value="Not Matched">Not Matched</option>
                    </select>
                    <button class="reset-btn" onclick="resetStatusFilter('comparison-tab')">Reset</button>
                    <span id="filterCountTab2" class="filter-count-inline"></span>
                </div>
                <table class="comparison-table">
                    <thead>
                        <tr>
                            <th rowspan="2" style="width: 150px;">Lead Name</th>
                            <th rowspan="2" style="width: 80px;">Suite Count (ADO)</th>
                            <th colspan="5" class="source-header">Azure DevOps (ADO)</th>
                            <th colspan="5" class="source-header">SharePoint (SP)</th>
                            <th rowspan="2" style="min-width: 95px;">Status</th>
                        </tr>
                        <tr>
                            <th class="ado-col">Passed</th>
                            <th class="ado-col">Failed</th>
                            <th class="ado-col">Not Run</th>
                            <th class="ado-col">Blocked</th>
                            <th class="ado-col">Total</th>
                            <th class="sp-col">Total</th>
                            <th class="sp-col">Passed</th>
                            <th class="sp-col">Failed</th>
                            <th class="sp-col">Not Run</th>
                            <th class="sp-col">Blocked</th>
                        </tr>
                    </thead>
                    <tbody>
"""
        
        # Add rows for each lead (from both ADO and SP)
        for lead in sorted(all_leads):
            ado = lead_totals_ado.get(lead, {'passed': 0, 'failed': 0, 'not_run': 0, 'blocked': 0, 'total': 0})
            sp = lead_totals_sp.get(lead, {'passed': 0, 'failed': 0, 'not_run': 0, 'blocked': 0, 'total': 0})
            suite_count = self.us_count_all.get(lead, 0)
            
            # Highlight lead name if there's any count mismatch
            has_mismatch = (ado['total'] != sp['total'] or 
                           ado['passed'] != sp['passed'] or 
                           ado['failed'] != sp['failed'] or 
                           ado['not_run'] != sp['not_run'] or 
                           ado['blocked'] != sp['blocked'])
            lead_style = ' style="background-color: #800080; font-weight: 700;"' if has_mismatch else ''
            
            html += f"""
                    <tr>
                        <td class="lead-name"{lead_style}>{lead}</td>
                        <td style="text-align: center; font-weight: 600; background-color: #e3f2fd;">{suite_count}</td>
                        <td class="ado-col">{ado['passed']}</td>
                        <td class="ado-col">{ado['failed']}</td>
                        <td class="ado-col">{ado['not_run']}</td>
                        <td class="ado-col">{ado['blocked']}</td>
                        <td class="ado-col" style="font-weight: 700;">{ado['total']}</td>
                        <td class="sp-col" style="font-weight: 700;">{sp['total']}</td>
                        <td class="sp-col">{sp['passed']}</td>
                        <td class="sp-col">{sp['failed']}</td>
                        <td class="sp-col">{sp['not_run']}</td>
                        <td class="sp-col">{sp['blocked']}</td>
                        <td class="{'status-matched' if ado['passed'] == sp['passed'] and ado['failed'] == sp['failed'] and ado['not_run'] == sp['not_run'] and ado['blocked'] == sp['blocked'] and ado['total'] == sp['total'] else 'status-not-matched'}" style="border: 2px solid #000;">{'Matched' if ado['passed'] == sp['passed'] and ado['failed'] == sp['failed'] and ado['not_run'] == sp['not_run'] and ado['blocked'] == sp['blocked'] and ado['total'] == sp['total'] else 'Not Matched'}</td>
                    </tr>
"""
        
        # Grand Total Row
        grand_total_ado = {
            'passed': sum(lead['passed'] for lead in lead_totals_ado.values()),
            'failed': sum(lead['failed'] for lead in lead_totals_ado.values()),
            'not_run': sum(lead['not_run'] for lead in lead_totals_ado.values()),
            'blocked': sum(lead['blocked'] for lead in lead_totals_ado.values()),
            'total': sum(lead['total'] for lead in lead_totals_ado.values())
        }
        
        grand_total_sp = {
            'passed': sum(lead['passed'] for lead in lead_totals_sp.values()),
            'failed': sum(lead['failed'] for lead in lead_totals_sp.values()),
            'not_run': sum(lead['not_run'] for lead in lead_totals_sp.values()),
            'blocked': sum(lead['blocked'] for lead in lead_totals_sp.values()),
            'total': sum(lead['total'] for lead in lead_totals_sp.values())
        }
        
        # Calculate total suite count
        total_suite_count = sum(self.us_count_all.values())
        
        html += f"""
                    <tr style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700;">
                        <td style="text-align: left; padding-left: 15px; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">GRAND TOTAL</td>
                        <td style="text-align: center; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{total_suite_count}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_ado['passed']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_ado['failed']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_ado['not_run']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_ado['blocked']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; font-weight: 800; color: black;">{grand_total_ado['total']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; font-weight: 800; color: black;">{grand_total_sp['total']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_sp['passed']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_sp['failed']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_sp['not_run']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{grand_total_sp['blocked']}</td>
                        <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700; border: 2px solid #000;">-</td>
                    </tr>
                </tbody>
            </table>
        </div>
        </div>
        
        <!-- Tab 3: Insprint Automated Test Cases -->
        <div id="planned-automation-tab" class="tab-content">
            <div style="padding: 20px;">
                <h2 style="text-align: center; color: #495057; margin-bottom: 15px; font-size: 16px;">🎯 Planned Automation Lead-wise Comparison</h2>
                
                <div class="comparison-filter-bar">
                    <label>Filter by Status:</label>
                    <select id="statusFilterTab3" onchange="applyStatusFilter('planned-automation-tab')">
                        <option value="all">-- All Status --</option>
                        <option value="Matched">Matched</option>
                        <option value="Not Matched">Not Matched</option>
                    </select>
                    <button class="reset-btn" onclick="resetStatusFilter('planned-automation-tab')">Reset</button>
                    <span id="filterCountTab3" class="filter-count-inline"></span>
                </div>
                <table class="comparison-table">
                    <thead>
                        <tr>
                            <th rowspan="2" style="width: 150px;">Lead Name</th>
                            <th rowspan="2" style="width: 80px;">Suite Count (ADO)</th>
                            <th colspan="5" class="source-header">Azure DevOps</th>
                            <th colspan="5" class="source-header">SharePoint</th>
                            <th rowspan="2" style="min-width: 95px;">Status</th>
                        </tr>
                        <tr>
                            <th class="ado-col">Total</th>
                            <th class="ado-col">Passed</th>
                            <th class="ado-col">Failed</th>
                            <th class="ado-col">Blocked</th>
                            <th class="ado-col">Not Run</th>
                            <th class="sp-col">Total</th>
                            <th class="sp-col">Passed</th>
                            <th class="sp-col">Failed</th>
                            <th class="sp-col">Blocked</th>
                            <th class="sp-col">Not Run</th>
                        </tr>
                    </thead>
                    <tbody>
"""
        
        # Get all unique leads from both planned automation ADO and SP insprint data
        all_planned_leads = sorted(set(list(self.planned_automation_ado.keys()) + list(self.planned_automation_sp.keys())))
        
        # Add rows for each lead
        for lead in all_planned_leads:
            ado = self.planned_automation_ado.get(lead, {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0})
            sp = self.planned_automation_sp.get(lead, {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0})
            suite_count = self.us_count_planned.get(lead, 0)
            
            # Check if totals match
            total_class = 'match' if ado['total'] == sp['total'] else 'mismatch'
            
            html += f"""
                        <tr>
                            <td class="lead-name">{lead}</td>
                            <td style="text-align: center; font-weight: 600; background-color: #e3f2fd;">{suite_count}</td>
                            <td class="ado-col {total_class}" style="font-weight: 700;">{ado['total']}</td>
                            <td class="ado-col">{ado['passed']}</td>
                            <td class="ado-col">{ado['failed']}</td>
                            <td class="ado-col">{ado['blocked']}</td>
                            <td class="ado-col">{ado['not_run']}</td>
                            <td class="sp-col {total_class}" style="font-weight: 700;">{sp['total']}</td>
                            <td class="sp-col">{sp['passed']}</td>
                            <td class="sp-col">{sp['failed']}</td>
                            <td class="sp-col">{sp['blocked']}</td>
                            <td class="sp-col">{sp['not_run']}</td>
                            <td class="{'status-matched' if ado['passed'] == sp['passed'] and ado['failed'] == sp['failed'] and ado['not_run'] == sp['not_run'] and ado['blocked'] == sp['blocked'] and ado['total'] == sp['total'] else 'status-not-matched'}" style="border: 2px solid #000;">{'Matched' if ado['passed'] == sp['passed'] and ado['failed'] == sp['failed'] and ado['not_run'] == sp['not_run'] and ado['blocked'] == sp['blocked'] and ado['total'] == sp['total'] else 'Not Matched'}</td>
                        </tr>
"""
        
        # Calculate grand totals for planned automation
        planned_grand_total_ado = {
            'total': sum(lead['total'] for lead in self.planned_automation_ado.values()),
            'passed': sum(lead['passed'] for lead in self.planned_automation_ado.values()),
            'failed': sum(lead['failed'] for lead in self.planned_automation_ado.values()),
            'blocked': sum(lead['blocked'] for lead in self.planned_automation_ado.values()),
            'not_run': sum(lead['not_run'] for lead in self.planned_automation_ado.values())
        }
        
        planned_grand_total_sp = {
            'total': sum(lead['total'] for lead in self.planned_automation_sp.values()),
            'passed': sum(lead['passed'] for lead in self.planned_automation_sp.values()),
            'failed': sum(lead['failed'] for lead in self.planned_automation_sp.values()),
            'blocked': sum(lead['blocked'] for lead in self.planned_automation_sp.values()),
            'not_run': sum(lead['not_run'] for lead in self.planned_automation_sp.values())
        }
        
        grand_total_class = 'match' if planned_grand_total_ado['total'] == planned_grand_total_sp['total'] else 'mismatch'
        
        # Calculate total suite count for planned
        total_suite_count_planned = sum(self.us_count_planned.values())
        
        html += f"""
                        <tr style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700;">
                            <td style="text-align: left; padding-left: 15px; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">GRAND TOTAL</td>
                            <td style="text-align: center; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{total_suite_count_planned}</td>
                            <td class="{grand_total_class}" style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; font-weight: 800; color: black;">{planned_grand_total_ado['total']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_ado['passed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_ado['failed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_ado['blocked']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_ado['not_run']}</td>
                            <td class="{grand_total_class}" style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; font-weight: 800; color: black;">{planned_grand_total_sp['total']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_sp['passed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_sp['failed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_sp['blocked']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{planned_grand_total_sp['not_run']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700; border: 2px solid #000;">-</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Tab 4: Manual Test Cases -->
        <div id="not-automated-tab" class="tab-content">
            <div style="padding: 20px;">
                <h2 style="text-align: center; color: #495057; margin-bottom: 15px; font-size: 16px;">🚫 Not Automated Lead-wise Comparison</h2>
 
                <div class="comparison-filter-bar">
                    <label>Filter by Status:</label>
                    <select id="statusFilterTab4" onchange="applyStatusFilter('not-automated-tab')">
                        <option value="all">-- All Status --</option>
                        <option value="Matched">Matched</option>
                        <option value="Not Matched">Not Matched</option>
                    </select>
                    <button class="reset-btn" onclick="resetStatusFilter('not-automated-tab')">Reset</button>
                    <span id="filterCountTab4" class="filter-count-inline"></span>
                </div>
                <table class="comparison-table">
                    <thead>
                        <tr>
                            <th rowspan="2" style="width: 150px;">Lead Name</th>
                            <th rowspan="2" style="width: 80px;">Suite Count (ADO)</th>
                            <th colspan="5" class="source-header">Azure DevOps</th>
                            <th colspan="5" class="source-header">SharePoint</th>
                            <th rowspan="2" style="min-width: 95px;">Status</th>
                        </tr>
                        <tr>
                            <th class="ado-col">Total</th>
                            <th class="ado-col">Passed</th>
                            <th class="ado-col">Failed</th>
                            <th class="ado-col">Blocked</th>
                            <th class="ado-col">Not Run</th>
                            <th class="sp-col">Total</th>
                            <th class="sp-col">Passed</th>
                            <th class="sp-col">Failed</th>
                            <th class="sp-col">Blocked</th>
                            <th class="sp-col">Not Run</th>
                        </tr>
                    </thead>
                    <tbody>
"""
        
        # Get all unique leads from both not_automated ADO and SP insprint-no data
        all_not_automated_leads = sorted(set(list(self.not_automated_ado.keys()) + list(self.not_automated_sp.keys())))
        
        # Add rows for each lead
        for lead in all_not_automated_leads:
            ado = self.not_automated_ado.get(lead, {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0})
            sp = self.not_automated_sp.get(lead, {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0})
            suite_count = self.us_count_not_automated.get(lead, 0)
            
            # Check if totals match
            total_class = 'match' if ado['total'] == sp['total'] else 'mismatch'
            
            html += f"""
                        <tr>
                            <td class="lead-name">{lead}</td>
                            <td style="text-align: center; font-weight: 600; background-color: #e3f2fd;">{suite_count}</td>
                            <td class="ado-col {total_class}" style="font-weight: 700;">{ado['total']}</td>
                            <td class="ado-col">{ado['passed']}</td>
                            <td class="ado-col">{ado['failed']}</td>
                            <td class="ado-col">{ado['blocked']}</td>
                            <td class="ado-col">{ado['not_run']}</td>
                            <td class="sp-col {total_class}" style="font-weight: 700;">{sp['total']}</td>
                            <td class="sp-col">{sp['passed']}</td>
                            <td class="sp-col">{sp['failed']}</td>
                            <td class="sp-col">{sp['blocked']}</td>
                            <td class="sp-col">{sp['not_run']}</td>
                            <td class="{'status-matched' if ado['passed'] == sp['passed'] and ado['failed'] == sp['failed'] and ado['not_run'] == sp['not_run'] and ado['blocked'] == sp['blocked'] and ado['total'] == sp['total'] else 'status-not-matched'}" style="border: 2px solid #000;">{'Matched' if ado['passed'] == sp['passed'] and ado['failed'] == sp['failed'] and ado['not_run'] == sp['not_run'] and ado['blocked'] == sp['blocked'] and ado['total'] == sp['total'] else 'Not Matched'}</td>
                        </tr>
"""
        
        # Calculate grand totals for not automated
        not_automated_grand_total_ado = {
            'total': sum(lead['total'] for lead in self.not_automated_ado.values()),
            'passed': sum(lead['passed'] for lead in self.not_automated_ado.values()),
            'failed': sum(lead['failed'] for lead in self.not_automated_ado.values()),
            'blocked': sum(lead['blocked'] for lead in self.not_automated_ado.values()),
            'not_run': sum(lead['not_run'] for lead in self.not_automated_ado.values())
        }
        
        not_automated_grand_total_sp = {
            'total': sum(lead['total'] for lead in self.not_automated_sp.values()),
            'passed': sum(lead['passed'] for lead in self.not_automated_sp.values()),
            'failed': sum(lead['failed'] for lead in self.not_automated_sp.values()),
            'blocked': sum(lead['blocked'] for lead in self.not_automated_sp.values()),
            'not_run': sum(lead['not_run'] for lead in self.not_automated_sp.values())
        }
        
        grand_total_class_na = 'match' if not_automated_grand_total_ado['total'] == not_automated_grand_total_sp['total'] else 'mismatch'
        
        # Calculate total suite count for not automated
        total_suite_count_not_automated = sum(self.us_count_not_automated.values())
        
        html += f"""
                        <tr style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700;">
                            <td style="text-align: left; padding-left: 15px; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">GRAND TOTAL</td>
                            <td style="text-align: center; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{total_suite_count_not_automated}</td>
                            <td class="{grand_total_class_na}" style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; font-weight: 800; color: black;">{not_automated_grand_total_ado['total']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_ado['passed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_ado['failed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_ado['blocked']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_ado['not_run']}</td>
                            <td class="{grand_total_class_na}" style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; font-weight: 800; color: black;">{not_automated_grand_total_sp['total']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_sp['passed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_sp['failed']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_sp['blocked']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black;">{not_automated_grand_total_sp['not_run']}</td>
                            <td style="background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700; border: 2px solid #000;">-</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
        
        <div class="footer">
            <p><strong>© 2026 Test Count Report</strong> | """ + f"{ADO_CONFIG['organization']} / {ADO_CONFIG['project']}" + """</p>
        </div>
    </div>
    
    <script>
        // Tab Switching Function
        function switchTab(event, tabId) {
            // Hide all tab contents
            const tabContents = document.querySelectorAll('.tab-content');
            tabContents.forEach(content => {
                content.classList.remove('active');
            });
            
            // Remove active class from all tabs
            const tabs = document.querySelectorAll('.tab');
            tabs.forEach(tab => {
                tab.classList.remove('active');
            });
            
            // Show the selected tab content
            document.getElementById(tabId).classList.add('active');
            
            // Add active class to the clicked tab
            event.currentTarget.classList.add('active');
        }
        
        // Store original module options for each lead
        const leadModuleMap = {};
        
        // Initialize lead-module mapping from table data
        function initializeLeadModuleMap() {
            const rows = document.querySelectorAll('#reportTableBody tr:not(.grand-total-row)');
            rows.forEach(row => {
                const lead = row.getAttribute('data-lead');
                const module = row.getAttribute('data-module');
                
                if (!leadModuleMap[lead]) {
                    leadModuleMap[lead] = new Set();
                }
                leadModuleMap[lead].add(module);
            });
        }
        
        // Update module dropdown based on selected lead
        function updateModuleOptions() {
            const leadFilter = document.getElementById('leadFilter').value;
            const moduleFilter = document.getElementById('moduleFilter');
            const currentModule = moduleFilter.value;
            
            // Clear existing options except "All Modules"
            moduleFilter.innerHTML = '<option value="all">-- All Modules --</option>';
            
            if (leadFilter === 'all') {
                // Show all modules if "All Leads" is selected
                const allModules = new Set();
                Object.values(leadModuleMap).forEach(modules => {
                    modules.forEach(module => allModules.add(module));
                });
                
                Array.from(allModules).sort().forEach(module => {
                    const option = document.createElement('option');
                    option.value = module;
                    option.textContent = module;
                    moduleFilter.appendChild(option);
                });
            } else {
                // Show only modules for selected lead
                const modulesForLead = leadModuleMap[leadFilter] || new Set();
                Array.from(modulesForLead).sort().forEach(module => {
                    const option = document.createElement('option');
                    option.value = module;
                    option.textContent = module;
                    moduleFilter.appendChild(option);
                });
            }
            
            // Try to restore previous module selection if it's still available
            const availableOptions = Array.from(moduleFilter.options).map(opt => opt.value);
            if (availableOptions.includes(currentModule)) {
                moduleFilter.value = currentModule;
            } else {
                moduleFilter.value = 'all';
            }
            
            // Apply filters after updating options
            applyFilters();
        }
        
        function updateGrandTotal() {
            // Get all visible rows (not hidden and not grand total)
            const visibleRows = document.querySelectorAll('#reportTableBody tr:not(.grand-total-row):not(.hidden)');
            
            // Initialize totals
            let totalP1P2 = 0;
            let manualTotal = 0, manualPass = 0, manualFail = 0, manualBlock = 0, manualNA = 0, manualNotRun = 0;
            let autoTotal = 0, autoPass = 0, autoFail = 0, autoBlock = 0, autoNA = 0, autoNotRun = 0;
            
            // Sum up values from visible rows
            visibleRows.forEach(row => {
                const cells = row.querySelectorAll('td');
                // Column indices: 3=TotalP1P2, 4=ManTotal, 5=ManPass, 6=ManFail, 7=ManBlock, 8=ManNA, 9=ManNotRun
                // 12=AutoTotal, 13=AutoPass, 14=AutoFail, 15=AutoBlock, 16=AutoNA, 17=AutoNotRun
                totalP1P2 += parseInt(cells[3].textContent) || 0;
                manualTotal += parseInt(cells[4].textContent) || 0;
                manualPass += parseInt(cells[5].textContent) || 0;
                manualFail += parseInt(cells[6].textContent) || 0;
                manualBlock += parseInt(cells[7].textContent) || 0;
                manualNA += parseInt(cells[8].textContent) || 0;
                manualNotRun += parseInt(cells[9].textContent) || 0;
                autoTotal += parseInt(cells[12].textContent) || 0;
                autoPass += parseInt(cells[13].textContent) || 0;
                autoFail += parseInt(cells[14].textContent) || 0;
                autoBlock += parseInt(cells[15].textContent) || 0;
                autoNA += parseInt(cells[16].textContent) || 0;
                autoNotRun += parseInt(cells[17].textContent) || 0;
            });
            
            // Calculate percentages for Manual
            const manualDenomPass = manualPass + manualFail + manualBlock;
            const manualPassPct = manualDenomPass > 0 ? (manualPass / manualDenomPass * 100) : 0;
            const manualDenomExec = manualTotal - manualNA;
            const manualExecPct = manualDenomExec > 0 ? ((manualPass + manualFail + manualBlock) / manualDenomExec * 100) : 0;
            
            // Calculate percentages for Automation
            const autoDenomPass = autoPass + autoFail + autoBlock;
            const autoPassPct = autoDenomPass > 0 ? (autoPass / autoDenomPass * 100) : 0;
            const autoDenomExec = autoTotal - autoNA;
            const autoExecPct = autoDenomExec > 0 ? ((autoPass + autoFail + autoBlock) / autoDenomExec * 100) : 0;
            
            // Update grand total row
            const grandTotalRow = document.getElementById('grandTotalRow');
            if (grandTotalRow) {
                const cells = grandTotalRow.querySelectorAll('td');
                cells[1].textContent = totalP1P2;
                cells[2].textContent = manualTotal;
                cells[3].textContent = manualPass;
                cells[4].textContent = manualFail;
                cells[5].textContent = manualBlock;
                cells[6].textContent = manualNA;
                cells[7].textContent = manualNotRun;
                cells[8].textContent = manualExecPct.toFixed(2) + '%';
                cells[9].textContent = manualPassPct.toFixed(2) + '%';
                cells[10].textContent = autoTotal;
                cells[11].textContent = autoPass;
                cells[12].textContent = autoFail;
                cells[13].textContent = autoBlock;
                cells[14].textContent = autoNA;
                cells[15].textContent = autoNotRun;
                cells[16].textContent = autoExecPct.toFixed(2) + '%';
                cells[17].textContent = autoPassPct.toFixed(2) + '%';
            }
        }
        
        function applyFilters() {
            const leadFilter = document.getElementById('leadFilter').value;
            const moduleFilter = document.getElementById('moduleFilter').value;
            const rows = document.querySelectorAll('#reportTableBody tr:not(.grand-total-row)');
            
            let visibleCount = 0;
            
            rows.forEach(row => {
                const lead = row.getAttribute('data-lead');
                const module = row.getAttribute('data-module');
                
                const leadMatch = leadFilter === 'all' || lead === leadFilter;
                const moduleMatch = moduleFilter === 'all' || module === moduleFilter;
                
                if (leadMatch && moduleMatch) {
                    row.classList.remove('hidden');
                    visibleCount++;
                } else {
                    row.classList.add('hidden');
                }
            });
            
            // Update visible count
            const totalRows = rows.length;
            document.getElementById('visibleCount').textContent = 
                `Showing ${visibleCount} of ${totalRows} rows`;
            
            // Update grand total based on visible rows
            updateGrandTotal();
        }
        
        function resetFilters() {
            document.getElementById('leadFilter').value = 'all';
            document.getElementById('moduleFilter').value = 'all';
            updateModuleOptions();
        }
        
        // Insprint Tab Filter Functions
        const leadModuleMapInsprint = {};
        
        // Initialize leadModuleMapInsprint
        if (document.getElementById('reportTableBodyInsprint')) {
            const insprintRows = document.querySelectorAll('#reportTableBodyInsprint tr:not(.grand-total-row)');
            insprintRows.forEach(row => {
                const lead = row.getAttribute('data-lead');
                const module = row.getAttribute('data-module');
                
                if (!leadModuleMapInsprint[lead]) {
                    leadModuleMapInsprint[lead] = new Set();
                }
                leadModuleMapInsprint[lead].add(module);
            });
        }
        
        function updateModuleOptionsInsprint() {
            const leadFilter = document.getElementById('leadFilterInsprint').value;
            const moduleFilter = document.getElementById('moduleFilterInsprint');
            const currentModule = moduleFilter.value;
            
            // Clear existing options except "All Modules"
            moduleFilter.innerHTML = '<option value="all">-- All Modules --</option>';
            
            if (leadFilter === 'all') {
                // Show all modules if "All Leads" is selected
                const allModules = new Set();
                Object.values(leadModuleMapInsprint).forEach(modules => {
                    modules.forEach(module => allModules.add(module));
                });
                
                Array.from(allModules).sort().forEach(module => {
                    const option = document.createElement('option');
                    option.value = module;
                    option.textContent = module;
                    moduleFilter.appendChild(option);
                });
            } else {
                // Show only modules for selected lead
                const modulesForLead = leadModuleMapInsprint[leadFilter] || new Set();
                Array.from(modulesForLead).sort().forEach(module => {
                    const option = document.createElement('option');
                    option.value = module;
                    option.textContent = module;
                    moduleFilter.appendChild(option);
                });
            }
            
            // Try to restore previous module selection if it's still available
            const availableOptions = Array.from(moduleFilter.options).map(opt => opt.value);
            if (availableOptions.includes(currentModule)) {
                moduleFilter.value = currentModule;
            } else {
                moduleFilter.value = 'all';
            }
            
            // Apply filters after updating options
            applyFiltersInsprint();
        }
        
        function updateGrandTotalInsprint() {
            // Get all visible rows (not hidden and not grand total)
            const visibleRows = document.querySelectorAll('#reportTableBodyInsprint tr:not(.grand-total-row):not(.hidden)');
            
            // Initialize totals
            let totalSum = 0;
            let insprintTotal = 0, insprintPass = 0, insprintFail = 0, insprintBlock = 0, insprintNA = 0, insprintNotRun = 0;
            
            // Sum up values from visible rows
            visibleRows.forEach(row => {
                const cells = row.querySelectorAll('td');
                // Column indices: 3=Total, 4=InsprintTotal, 5=Pass, 6=Fail, 7=Block, 8=NA, 9=NotRun
                totalSum += parseInt(cells[3].textContent) || 0;
                insprintTotal += parseInt(cells[4].textContent) || 0;
                insprintPass += parseInt(cells[5].textContent) || 0;
                insprintFail += parseInt(cells[6].textContent) || 0;
                insprintBlock += parseInt(cells[7].textContent) || 0;
                insprintNA += parseInt(cells[8].textContent) || 0;
                insprintNotRun += parseInt(cells[9].textContent) || 0;
            });
            
            // Calculate percentages
            const denomPass = insprintPass + insprintFail + insprintBlock;
            const passPct = denomPass > 0 ? (insprintPass / denomPass * 100) : 0;
            const denomExec = insprintTotal - insprintNA;
            const execPct = denomExec > 0 ? ((insprintPass + insprintFail + insprintBlock) / denomExec * 100) : 0;
            
            // Update grand total row
            const grandTotalRow = document.getElementById('grandTotalRowInsprint');
            if (grandTotalRow) {
                const cells = grandTotalRow.querySelectorAll('td');
                cells[1].textContent = totalSum;
                cells[2].textContent = insprintTotal;
                cells[3].textContent = insprintPass;
                cells[4].textContent = insprintFail;
                cells[5].textContent = insprintBlock;
                cells[6].textContent = insprintNA;
                cells[7].textContent = insprintNotRun;
                cells[8].textContent = execPct.toFixed(2) + '%';
                cells[9].textContent = passPct.toFixed(2) + '%';
            }
        }
        
        function applyFiltersInsprint() {
            const leadFilter = document.getElementById('leadFilterInsprint').value;
            const moduleFilter = document.getElementById('moduleFilterInsprint').value;
            const rows = document.querySelectorAll('#reportTableBodyInsprint tr:not(.grand-total-row)');
            
            let visibleCount = 0;
            
            rows.forEach(row => {
                const lead = row.getAttribute('data-lead');
                const module = row.getAttribute('data-module');
                
                const leadMatch = leadFilter === 'all' || lead === leadFilter;
                const moduleMatch = moduleFilter === 'all' || module === moduleFilter;
                
                if (leadMatch && moduleMatch) {
                    row.classList.remove('hidden');
                    visibleCount++;
                } else {
                    row.classList.add('hidden');
                }
            });
            
            // Update visible count
            const totalRows = rows.length;
            document.getElementById('visibleCountInsprint').textContent = 
                `Showing ${visibleCount} of ${totalRows} rows`;
            
            // Update grand total based on visible rows
            updateGrandTotalInsprint();
        }
        
        function resetFiltersInsprint() {
            document.getElementById('leadFilterInsprint').value = 'all';
            document.getElementById('moduleFilterInsprint').value = 'all';
            updateModuleOptionsInsprint();
        }

        // ── Status column helpers ────────────────────────────────────────
        function addStatusCells(tabId, adoTotalIdx, spTotalIdx) {
            const tab = document.getElementById(tabId);
            if (!tab) return;
            const rows = tab.querySelectorAll('table.comparison-table tbody tr');
            let total = 0, visible = 0;
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (!cells.length) return;
                const isGrandTotal = cells[0].textContent.trim().toUpperCase() === 'GRAND TOTAL';
                const td = document.createElement('td');
                if (isGrandTotal) {
                    td.style.cssText = 'border: 2px solid #000 !important; background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important; color: black; font-weight: 700; font-size: 10px; text-align: center;';
                    td.textContent = '-';
                } else {
                    if (cells.length <= Math.max(adoTotalIdx, spTotalIdx)) return;
                    const adoTotal = parseInt(cells[adoTotalIdx].textContent.trim()) || 0;
                    const spTotal  = parseInt(cells[spTotalIdx].textContent.trim())  || 0;
                    const isMatched = adoTotal === spTotal;
                    td.className = isMatched ? 'status-matched' : 'status-not-matched';
                    td.style.border = '2px solid #000';
                    td.textContent = isMatched ? '\u2705 Matched' : '\u274c Not Matched';
                    row.setAttribute('data-status', isMatched ? 'Matched' : 'Not Matched');
                    total++;
                    visible++;
                }
                row.appendChild(td);
            });
            _updateFilterCount(tabId, visible, total);
        }

        function _updateComparisonGrandTotal(tabId) {
            const tab = document.getElementById(tabId);
            if (!tab) return;
            const rows = tab.querySelectorAll('table.comparison-table tbody tr');
            let grandTotalRow = null;
            const sums = {};
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (!cells.length) return;
                if (cells[0].textContent.trim().toUpperCase() === 'GRAND TOTAL') {
                    grandTotalRow = row;
                    return;
                }
                if (row.classList.contains('hidden')) return;
                cells.forEach((cell, i) => {
                    if (i === 0 || i === cells.length - 1) return;
                    sums[i] = (sums[i] || 0) + (parseInt(cell.textContent.trim()) || 0);
                });
            });
            if (!grandTotalRow) return;
            const gtCells = grandTotalRow.querySelectorAll('td');
            gtCells.forEach((cell, i) => {
                if (i === 0 || i === gtCells.length - 1) return;
                cell.textContent = sums[i] !== undefined ? sums[i] : 0;
            });
        }

        function applyStatusFilter(tabId) {
            const selectMap = {
                'comparison-tab':         'statusFilterTab2',
                'planned-automation-tab': 'statusFilterTab3',
                'not-automated-tab':      'statusFilterTab4'
            };
            const statusFilter = document.getElementById(selectMap[tabId]).value;
            const rows = document.querySelectorAll('#' + tabId + ' table.comparison-table tbody tr');
            let visible = 0, total = 0;
            rows.forEach(row => {
                const cells = row.querySelectorAll('td');
                if (!cells.length || cells[0].textContent.trim().toUpperCase() === 'GRAND TOTAL') return;
                total++;
                const status = cells[cells.length - 1].textContent.trim();
                if (statusFilter === 'all' || status === statusFilter) {
                    row.classList.remove('hidden');
                    visible++;
                } else {
                    row.classList.add('hidden');
                }
            });
            _updateComparisonGrandTotal(tabId);
            _updateFilterCount(tabId, visible, total);
        }

        function resetStatusFilter(tabId) {
            const selectMap = {
                'comparison-tab':         'statusFilterTab2',
                'planned-automation-tab': 'statusFilterTab3',
                'not-automated-tab':      'statusFilterTab4'
            };
            document.getElementById(selectMap[tabId]).value = 'all';
            applyStatusFilter(tabId);
        }

        function _updateFilterCount(tabId, visible, total) {
            const countMap = {
                'comparison-tab':         'filterCountTab2',
                'planned-automation-tab': 'filterCountTab3',
                'not-automated-tab':      'filterCountTab4'
            };
            const el = document.getElementById(countMap[tabId]);
            if (el) el.textContent = visible === total ? `${total} leads` : `Showing ${visible} of ${total} leads`;
        }

        // Initialize on page load
        document.addEventListener('DOMContentLoaded', function() {
            initializeLeadModuleMap();
            updateModuleOptions();
            // Initialize status filter counts for all comparison tabs
            ['comparison-tab', 'planned-automation-tab', 'not-automated-tab'].forEach(function(tabId) {
                applyStatusFilter(tabId);
            });
        });
    </script>
</body>
</html>
"""
        return html
    
    def generate_html_file(self, filename=None):
        """Generate and save HTML report to file"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(OUTPUT_DIR, f"pt_execution_count_comparison_report_{timestamp}.html")
        
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        html_content = self.generate_html()
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"\n✅ Report generated: {filename}")
        return filename


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution flow"""
    print("=" * 80)
    print("🚀 AZURE DEVOPS TEST EXECUTION REPORT GENERATOR")
    print("=" * 80)

    # ── Step 0: Download SharePoint file (skip if recent) ────────────────────
    print("\n📥 Step 0: Checking SharePoint Excel file...")
    print("-" * 80)
    _should_download = True
    if os.path.exists(XLSX_FILE):
        print(f"📋 Existing Excel found — re-downloading latest from SharePoint...")

    if _should_download:
        try:
            _sp_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "download_PT status_file.py")
            _spec = importlib.util.spec_from_file_location("download_PT_status_file", _sp_script)
            _sp_module = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_sp_module)
            _sp_module.main()
            print("✅ SharePoint download complete.")
        except Exception as _e:
            print(f"⚠️  SharePoint download failed: {_e}")
            print("   Continuing with existing local file if available...")
    print()

    # Initialize client
    client = AzureDevOpsClient(ADO_CONFIG)

    # ── Parallelise the 3 independent ADO setup calls ─────────────────────────
    print("\n🔗 Running parallel ADO setup calls (connection check, plan info, suites, suite verify)...")
    with ThreadPoolExecutor(max_workers=4) as _setup_ex:
        _conn_f   = _setup_ex.submit(client.test_connection)
        _plan_f   = _setup_ex.submit(client.get_test_plan)
        _suites_f = _setup_ex.submit(client.get_all_suites_in_plan)   # pre-warms cache
        _suite_f  = _setup_ex.submit(client.verify_suite_exists, ADO_CONFIG['suite_id'])

        ok         = _conn_f.result()
        plan_info  = _plan_f.result()
        _suites_f.result()          # result stored in client._cached_suites
        suite_info = _suite_f.result()

    if not ok:
        print("\n❌ Connection failed. Please check your configuration.")
        return

    if not plan_info:
        print("\n❌ Could not fetch test plan information.")
        return

    if not suite_info:
        print(f"\n⚠️  Warning: Suite {ADO_CONFIG['suite_id']} not directly accessible")
        print("   Attempting to find suite in plan hierarchy...")
        
        all_suites = client.get_all_suites_in_plan()
        target_suite = None
        
        for suite in all_suites:
            if str(suite.get('id')) == str(ADO_CONFIG['suite_id']):
                target_suite = suite
                break
            elif suite.get('name', '').lower() == ADO_CONFIG['target_suite_name'].lower():
                target_suite = suite
                ADO_CONFIG['suite_id'] = suite.get('id')
                break
        
        if target_suite:
            suite_info = target_suite
            print(f"   ✅ Found suite: {suite_info.get('name')}")
        else:
            print(f"\n❌ Could not find suite '{ADO_CONFIG['target_suite_name']}'")
            print("\n📋 Available suites in this plan:")
            for suite in all_suites[:20]:
                print(f"   - {suite.get('name')} (ID: {suite.get('id')})")
            return
    
    suite_name = suite_info.get('name', ADO_CONFIG['target_suite_name'])
    suite_id = suite_info.get('id', ADO_CONFIG['suite_id'])
    
    # Collect test data
    test_data = client.get_all_test_data_from_suite(suite_id, suite_name)
    
    if not test_data:
        print("\n⚠️  No test data found in the suite.")
        return
    
    # Collect Insprint test data
    insprint_suite_id = ADO_CONFIG.get('insprint_suite_id')
    insprint_suite_name = ADO_CONFIG.get('insprint_suite_name', 'Insprint Execution')
    insprint_data = []
    
    if insprint_suite_id:
        print(f"\n📊 Collecting Insprint Test Data from Suite {insprint_suite_id}...")
        insprint_suite_info = client.verify_suite_exists(insprint_suite_id)
        
        if insprint_suite_info:
            insprint_data = client.get_all_test_data_from_suite(insprint_suite_id, insprint_suite_name)
            print(f"   ✅ Found {len(insprint_data)} Insprint test items")
        else:
            print(f"   ⚠️  Could not access Insprint suite {insprint_suite_id}")
    
    # Process Planned Automation Comparison
    print(f"\n🎯 Processing Planned Automation Comparison...")
    print(f"   📋 Filtering test cases based on work item's 'Microsoft.VSTS.TCM.AutomationStatus' field...")
    
    # Filter test cases with Automation Status = "Planned" from ADO (case-insensitive)
    planned_automation_tests = []
    automation_status_summary = defaultdict(int)
    
    for test in test_data:
        automation_status = test.get('automation_status', '').strip()
        automation_status_summary[automation_status if automation_status else 'Not Set'] += 1
        
        # Case-insensitive comparison for "Planned"
        if automation_status.lower() == 'planned':
            planned_automation_tests.append(test)
    
    print(f"\n   📊 Work Item Automation Status Field Distribution:")
    for status, count in sorted(automation_status_summary.items()):
        print(f"      • {status}: {count}")
    
    print(f"\n   ✅ Found {len(planned_automation_tests)} test cases with Automation Status = 'Planned'")
    
    # Organize planned automation data by lead
    planned_automation_ado = defaultdict(lambda: {
        'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0
    })
    
    # Track unique suite names for planned automation from ADO (for Tab 2 combined count)
    suite_names_planned_ado = defaultdict(set)
    # Track unique work item IDs (US IDs) for planned automation from ADO (Tab 3)
    work_item_ids_planned_ado = defaultdict(set)
    
    for test in planned_automation_tests:
        lead = test['lead']
        outcome = test['outcome'].lower()
        suite_name = test.get('suite', '')
        work_item_id = str(test.get('id', ''))
        
        # Track unique suite names from ADO (for Tab 2)
        if suite_name and suite_name != 'N/A':
            suite_names_planned_ado[lead].add(suite_name)
        
        # Track unique work item IDs (US IDs) from ADO (for Tab 3)
        if work_item_id and work_item_id != 'N/A':
            work_item_ids_planned_ado[lead].add(work_item_id)
        
        planned_automation_ado[lead]['total'] += 1
        
        if outcome in ['passed', 'pass']:
            planned_automation_ado[lead]['passed'] += 1
        elif outcome in ['failed', 'fail']:
            planned_automation_ado[lead]['failed'] += 1
        elif outcome in ['blocked', 'block']:
            planned_automation_ado[lead]['blocked'] += 1
        else:
            planned_automation_ado[lead]['not_run'] += 1
    
    planned_automation_ado = dict(planned_automation_ado)
    print(f"   ✅ Tracked {sum(len(suites) for suites in suite_names_planned_ado.values())} unique suites for Planned Automation from ADO")
    print(f"   ✅ Tracked {sum(len(wids) for wids in work_item_ids_planned_ado.values())} unique US IDs (work items) for Planned Automation from ADO")
    
    # Load SharePoint Insprint=Yes/Partial data
    print(f"\n📥 Loading SharePoint Insprint='Yes' and 'Partial' data...")
    sp_insprint_data = []
    xlsx_file = XLSX_FILE
    sp_sheet_name = 'PT status'
    
    print(f"   📂 Looking for file: {xlsx_file}")
    print(f"   📍 Current directory: {os.getcwd()}")
    print(f"   📄 Full path: {os.path.abspath(xlsx_file)}")
    
    # Check if file exists
    if not os.path.exists(xlsx_file):
        print(f"   ❌ ERROR: SharePoint Excel file not found!")
        print(f"   Please ensure '{xlsx_file}' exists in: {os.getcwd()}")
        
        # List available Excel files to help troubleshooting
        excel_files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xls'))]
        if excel_files:
            print(f"   📁 Excel files found in current directory:")
            for ef in excel_files:
                print(f"      - {ef}")
        else:
            print(f"   📁 No Excel files found in current directory")
        
        all_sp_rows = []
    elif os.path.exists(xlsx_file):
        print(f"   ✅ File found!")
        try:
            wb = load_workbook(xlsx_file, read_only=True, data_only=True)
            
            if sp_sheet_name not in wb.sheetnames:
                print(f"   ⚠️  Sheet '{sp_sheet_name}' not found in {xlsx_file}")
                print(f"   Available sheets: {wb.sheetnames}")
                wb.close()
                all_sp_rows = []
            else:
                ws = wb[sp_sheet_name]
                all_sp_rows = list(ws.iter_rows(values_only=True))
                wb.close()
        except Exception as e:
            print(f"   ⚠️  Error reading {xlsx_file}: {e}")
            all_sp_rows = []
        
        if all_sp_rows:
            sp_headers = [str(h).strip() if h else '' for h in all_sp_rows[0]]
            print(f"   📋 Found {len(all_sp_rows)-1} data rows in SharePoint Excel")
            print(f"   📝 Total columns: {len(sp_headers)}")
            print(f"   📝 Column headers: {sp_headers}")
            
            # Check for required columns
            required_cols = ['PT Lead', 'Insprint Yes/No', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Total']
            missing_cols = [col for col in required_cols if col not in sp_headers]
            if missing_cols:
                print(f"   ⚠️  WARNING: Missing columns: {missing_cols}")
                print(f"   Available columns starting with 'PT': {[h for h in sp_headers if 'PT' in h.upper()]}")
                print(f"   Available columns with 'Lead': {[h for h in sp_headers if 'LEAD' in h.upper()]}")
                print(f"   Available columns with 'Insprint': {[h for h in sp_headers if 'INSPRINT' in h.upper()]}")
            
            rows_processed = 0
            rows_with_insprint_yes = 0
            rows_with_lead = 0
            sample_rows_shown = 0
            
            for row_idx, row in enumerate(all_sp_rows[1:], start=2):
                rows_processed += 1
                row_dict = {sp_headers[i]: row[i] for i in range(min(len(sp_headers), len(row)))}
                
                lead = str(row_dict.get('PT Lead', '') or '').strip()
                insprint_status = str(row_dict.get('Insprint Yes/No', '') or '').strip()
                
                if lead:
                    rows_with_lead += 1
                    
                # Show first few sample rows for debugging
                if sample_rows_shown < 3 and lead:
                    print(f"   🔍 Sample Row {row_idx}: Lead='{lead}', Insprint='{insprint_status}', Total={row_dict.get('Total', 'N/A')}")
                    sample_rows_shown += 1
                
                if insprint_status.lower() in ['yes', 'partial'] and lead:
                    rows_with_insprint_yes += 1
                
                if insprint_status.lower() not in ['yes', 'partial'] or not lead:
                    continue
                
                try:
                    passed = int(row_dict.get('Passed', 0) or 0)
                    failed = int(row_dict.get('Failed', 0) or 0)
                    blocked = int(row_dict.get('Blocked', 0) or 0)
                    not_run = int(row_dict.get('Not Run', 0) or 0)
                    total = int(row_dict.get('Total', 0) or 0)
                except (ValueError, TypeError):
                    passed = 0
                    failed = 0
                    blocked = 0
                    not_run = 0
                    total = 1
                
                sp_insprint_data.append({
                    'lead': lead,
                    'passed': passed,
                    'failed': failed,
                    'blocked': blocked,
                    'not_run': not_run,
                    'total': total
                })
            
            print(f"\n   📊 Processing Summary for Insprint='Yes'/'Partial':")
            print(f"   - Total rows processed: {rows_processed}")
            print(f"   - Rows with PT Lead: {rows_with_lead}")
            print(f"   - Rows with Insprint='Yes'/'Partial': {rows_with_insprint_yes}")
            print(f"   - Records loaded: {len(sp_insprint_data)}")
            
            if len(sp_insprint_data) > 0:
                print(f"   ✅ Successfully loaded {len(sp_insprint_data)} Insprint='Yes'/'Partial' records")
                # Show summary of loaded data
                leads_found = set(item['lead'] for item in sp_insprint_data)
                print(f"   👥 Unique leads found: {', '.join(sorted(leads_found))}")
                total_count = sum(item['total'] for item in sp_insprint_data)
                print(f"   📊 Total count across all records: {total_count}")
            else:
                print(f"   ❌ ERROR: No Insprint='Yes'/'Partial' data was loaded!")
                if rows_with_insprint_yes > 0:
                    print(f"   Found {rows_with_insprint_yes} rows with Insprint='Yes'/'Partial' but data parsing failed.")
                else:
                    print(f"   No rows found with Insprint='Yes' or 'Partial'. Check the Excel file content.")
                print(f"   🔍 Troubleshooting: Check if column names match exactly (case-sensitive)")
        else:
            print(f"   ⚠️  No data rows found in SharePoint Excel file")
    else:
        print(f"   ⚠️  (File check failed - this shouldn't happen)")
        all_sp_rows = []
    
    # Aggregate SharePoint Insprint data by lead
    planned_automation_sp = defaultdict(lambda: {
        'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0, 'total': 0
    })
    
    print(f"\n📊 Aggregating SharePoint Insprint='Yes'/'Partial' data by lead...")
    for item in sp_insprint_data:
        lead = item['lead']
        planned_automation_sp[lead]['passed'] += item['passed']
        planned_automation_sp[lead]['failed'] += item['failed']
        planned_automation_sp[lead]['blocked'] += item['blocked']
        planned_automation_sp[lead]['not_run'] += item['not_run']
        planned_automation_sp[lead]['total'] += item['total']
    
    planned_automation_sp = dict(planned_automation_sp)
    
    if planned_automation_sp:
        print(f"   ✅ Aggregated data for {len(planned_automation_sp)} leads from SharePoint")
        for lead, counts in list(planned_automation_sp.items())[:5]:  # Show first 5 leads
            print(f"   - {lead}: Total={counts['total']}, Passed={counts['passed']}, Failed={counts['failed']}")
    else:
        print(f"   ⚠️ No SharePoint data was aggregated (sp_insprint_data is empty)")
    
    # Print summary
    print(f"\n📊 Planned Automation Comparison Summary:")
    print(f"{'Lead':<20} {'ADO Total':<12} {'SP Total':<12} {'Match':<8}")
    print("-" * 52)
    all_planned_leads = sorted(set(list(planned_automation_ado.keys()) + list(planned_automation_sp.keys())))
    for lead in all_planned_leads:
        ado_total = planned_automation_ado.get(lead, {}).get('total', 0)
        sp_total = planned_automation_sp.get(lead, {}).get('total', 0)
        match = '✓' if ado_total == sp_total else '✗'
        print(f"{lead:<20} {ado_total:<12} {sp_total:<12} {match:<8}")
    
    # Process Not Automated Comparison
    print(f"\n🚫 Processing Not Automated Comparison...")
    print(f"   📋 Filtering test cases based on work item's 'Microsoft.VSTS.TCM.AutomationStatus' field...")
    
    # Filter test cases with Automation Status = "Not Automated" from ADO (case-insensitive)
    not_automated_tests = []
    for test in test_data:
        automation_status = test.get('automation_status', '').strip()
        
        # Case-insensitive comparison for "Not Automated"
        if automation_status.lower() == 'not automated':
            not_automated_tests.append(test)
    
    print(f"   ✅ Found {len(not_automated_tests)} test cases with Automation Status = 'Not Automated'")
    
    # Organize not automated data by lead
    not_automated_ado = defaultdict(lambda: {
        'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0
    })
    
    # Track unique suite names for not automated from ADO (Tab 4)
    suite_names_not_automated_ado = defaultdict(set)
    
    for test in not_automated_tests:
        lead = test['lead']
        outcome = test['outcome'].lower()
        suite_name = test.get('suite', '')
        
        # Track unique suite names from ADO
        if suite_name and suite_name != 'N/A':
            suite_names_not_automated_ado[lead].add(suite_name)
        
        not_automated_ado[lead]['total'] += 1
        
        if outcome in ['passed', 'pass']:
            not_automated_ado[lead]['passed'] += 1
        elif outcome in ['failed', 'fail']:
            not_automated_ado[lead]['failed'] += 1
        elif outcome in ['blocked', 'block']:
            not_automated_ado[lead]['blocked'] += 1
        else:
            not_automated_ado[lead]['not_run'] += 1
    
    not_automated_ado = dict(not_automated_ado)
    print(f"   ✅ Tracked {sum(len(suites) for suites in suite_names_not_automated_ado.values())} unique suites for Not Automated from ADO")
    
    # Load SharePoint Insprint=No data
    print(f"\n📥 Loading SharePoint Insprint='No' data...")
    sp_insprint_no_data = []
    
    if os.path.exists(xlsx_file) and all_sp_rows:
        rows_processed_no = 0
        rows_with_insprint_no = 0
        sample_rows_no_shown = 0
        
        for row_idx, row in enumerate(all_sp_rows[1:], start=2):
            rows_processed_no += 1
            row_dict = {sp_headers[i]: row[i] for i in range(min(len(sp_headers), len(row)))}
            
            lead = str(row_dict.get('PT Lead', '') or '').strip()
            insprint_status = str(row_dict.get('Insprint Yes/No', '') or '').strip()
            
            # Show first few sample rows for debugging
            if sample_rows_no_shown < 3 and lead and insprint_status.lower() == 'no':
                print(f"   🔍 Sample Row {row_idx}: Lead='{lead}', Insprint='{insprint_status}', Total={row_dict.get('Total', 'N/A')}")
                sample_rows_no_shown += 1
            
            if insprint_status.lower() == 'no' and lead:
                rows_with_insprint_no += 1
            
            if insprint_status.lower() != 'no' or not lead:
                continue
            
            try:
                passed = int(row_dict.get('Passed', 0) or 0)
                failed = int(row_dict.get('Failed', 0) or 0)
                blocked = int(row_dict.get('Blocked', 0) or 0)
                not_run = int(row_dict.get('Not Run', 0) or 0)
                total = int(row_dict.get('Total', 0) or 0)
            except (ValueError, TypeError):
                passed = 0
                failed = 0
                blocked = 0
                not_run = 0
                total = 1
            
            sp_insprint_no_data.append({
                'lead': lead,
                'passed': passed,
                'failed': failed,
                'blocked': blocked,
                'not_run': not_run,
                'total': total
            })
        
        print(f"\n   📊 Processing Summary for Insprint='No':")
        print(f"   - Total rows processed: {rows_processed_no}")
        print(f"   - Rows with Insprint='No': {rows_with_insprint_no}")
        print(f"   - Records loaded: {len(sp_insprint_no_data)}")
        
        if len(sp_insprint_no_data) > 0:
            print(f"   ✅ Successfully loaded {len(sp_insprint_no_data)} Insprint='No' records")
            leads_found = set(item['lead'] for item in sp_insprint_no_data)
            print(f"   👥 Unique leads found: {', '.join(sorted(leads_found))}")
            total_count = sum(item['total'] for item in sp_insprint_no_data)
            print(f"   📊 Total count across all records: {total_count}")
        else:
            print(f"   ❌ ERROR: No Insprint='No' data was loaded!")
            if rows_with_insprint_no > 0:
                print(f"   Found {rows_with_insprint_no} rows with Insprint='No' but data parsing failed.")
            else:
                print(f"   No rows found with Insprint='No'. Check the Excel file content.")
    elif not os.path.exists(xlsx_file):
        print(f"   ⚠️  SharePoint Excel file not found: {xlsx_file}")
    else:
        print(f"   ⚠️  No data rows available from Excel file (file may be empty or sheet not found)")
    
    # Aggregate SharePoint Insprint-No data by lead
    not_automated_sp = defaultdict(lambda: {
        'passed': 0, 'failed': 0, 'blocked': 0, 'not_run': 0, 'total': 0
    })
    
    print(f"\n📊 Aggregating SharePoint Insprint='No' data by lead...")
    for item in sp_insprint_no_data:
        lead = item['lead']
        not_automated_sp[lead]['passed'] += item['passed']
        not_automated_sp[lead]['failed'] += item['failed']
        not_automated_sp[lead]['blocked'] += item['blocked']
        not_automated_sp[lead]['not_run'] += item['not_run']
        not_automated_sp[lead]['total'] += item['total']
    
    not_automated_sp = dict(not_automated_sp)
    
    if not_automated_sp:
        print(f"   ✅ Aggregated data for {len(not_automated_sp)} leads from SharePoint")
        for lead, counts in list(not_automated_sp.items())[:5]:  # Show first 5 leads
            print(f"   - {lead}: Total={counts['total']}, Passed={counts['passed']}, Failed={counts['failed']}")
    else:
        print(f"   ⚠️ No SharePoint data was aggregated (sp_insprint_no_data is empty)")
    
    # Print summary
    print(f"\n📊 Not Automated Comparison Summary:")
    print(f"{'Lead':<20} {'ADO Total':<12} {'SP Total':<12} {'Match':<8}")
    print("-" * 52)
    all_not_automated_leads = sorted(set(list(not_automated_ado.keys()) + list(not_automated_sp.keys())))
    for lead in all_not_automated_leads:
        ado_total = not_automated_ado.get(lead, {}).get('total', 0)
        sp_total = not_automated_sp.get(lead, {}).get('total', 0)
        match = '✓' if ado_total == sp_total else '✗'
        print(f"{lead:<20} {ado_total:<12} {sp_total:<12} {match:<8}")
    
    # Calculate suite counts from ADO for each tab
    # Tab 2: Combined suite counts (Planned + Not Automated)
    all_leads = set(suite_names_planned_ado.keys()) | set(suite_names_not_automated_ado.keys())
    us_count_all = {}
    for lead in all_leads:
        planned_suites = suite_names_planned_ado.get(lead, set())
        not_automated_suites = suite_names_not_automated_ado.get(lead, set())
        # Combine both sets to get unique suites across both automation statuses
        combined_suites = planned_suites | not_automated_suites
        us_count_all[lead] = len(combined_suites)
    
    # Tab 3: US ID counts (work item IDs) for Planned Automation
    us_count_planned = {lead: len(suites) for lead, suites in suite_names_planned_ado.items()}
    
    # Tab 4: Suite counts for Not Automated
    us_count_not_automated = {lead: len(suites) for lead, suites in suite_names_not_automated_ado.items()}
    
    print(f"\n📊 Suite Count Summary:")
    print(f"   Tab 2 (Combined Suites): {sum(us_count_all.values())} total suites across all leads")
    print(f"   Tab 3 (Planned Suites): {sum(us_count_planned.values())} total suites across all leads")
    print(f"   Tab 4 (Not Automated Suites): {sum(us_count_not_automated.values())} total suites across all leads")
    
    # Generate HTML report
    print(f"\n📝 Generating HTML Report...")
    report_gen = CustomHTMLReportGenerator(
        test_data, 
        plan_info, 
        suite_name, 
        insprint_data,
        planned_automation_ado,
        planned_automation_sp,
        not_automated_ado,
        not_automated_sp,
        us_count_all,
        us_count_planned,
        us_count_not_automated
    )
    report_file = report_gen.generate_html_file()
    
    print("\n" + "=" * 80)
    print("✅ REPORT GENERATION COMPLETED")
    print("=" * 80)
    print(f"\n📄 Report Location: {report_file}")
    print(f"📊 Total Test Cases: {len(test_data)}")
    
    # Summary statistics
    not_automated_count = sum(1 for t in test_data if t['type'].lower() == 'not automated')
    planned_count = sum(1 for t in test_data if t['type'].lower() == 'planned')
    
    print(f"   - Not Automated: {not_automated_count}")
    print(f"   - Planned: {planned_count}")
    
    # Outcome summary
    outcomes = {}
    for test in test_data:
        outcome = test['outcome']
        outcomes[outcome] = outcomes.get(outcome, 0) + 1
    
    print(f"\n📈 Outcome Summary:")
    for outcome, count in sorted(outcomes.items()):
        print(f"   - {outcome}: {count}")
    
    print("\n" + "=" * 80)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
    except Exception as e:
        print(f"\n\n❌ An error occurred: {e}")
        import traceback
        traceback.print_exc()
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
from config import OUTPUT_DIR, UAT_DETAILED_REPORT_FILE

# Try to import win32com for Excel automation
try:
    import win32com.client
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: win32com not available. Install with: pip install pywin32")

# File path
base_dir = OUTPUT_DIR
input_file = UAT_DETAILED_REPORT_FILE

print("=" * 80)
print("UAT DETAILED REPORT PROCESSING")
print("=" * 80)
print()

# Step 0: Preprocess the file - Check sheet name and convert formulas to values
print("Step 0: Preprocessing - Checking sheet name and breaking external links...")
print("-" * 80)

# First, try to use Excel automation to break external links and convert formulas to values
if EXCEL_AVAILABLE:
    print("Using Excel automation to break external links and convert formulas to values...")
    try:
        # Get absolute path
        abs_input_file = os.path.abspath(input_file)
        
        # Start Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        print(f"Opening workbook: {abs_input_file}")
        wb_excel = excel.Workbooks.Open(abs_input_file)
        
        # Break all external links
        if wb_excel.LinkSources(1) is not None:  # 1 = xlExcelLinks
            print("Breaking external links...")
            links = wb_excel.LinkSources(1)
            for link in links:
                print(f"  Breaking link: {link}")
                wb_excel.BreakLink(Name=link, Type=1)  # 1 = xlExcelLinks
            print(f"✓ Broke {len(links)} external link(s)")
        else:
            print("No external links found")
        
        # Now check and rename sheet if needed
        print("Checking sheet names...")
        target_sheet_name = "UAT Detailed Report"
        sheet_found = False
        
        for sheet in wb_excel.Worksheets:
            if sheet.Name == target_sheet_name:
                sheet_found = True
                ws_excel = sheet
                break
        
        if not sheet_found:
            # Try to find similar sheet
            for sheet in wb_excel.Worksheets:
                name_lower = sheet.Name.lower()
                if 'uat' in name_lower and 'detailed' in name_lower and 'report' in name_lower:
                    print(f"Renaming sheet '{sheet.Name}' to '{target_sheet_name}'...")
                    sheet.Name = target_sheet_name
                    ws_excel = sheet
                    sheet_found = True
                    break
            
            # If still not found, use first sheet
            if not sheet_found:
                ws_excel = wb_excel.Worksheets(1)
                print(f"Renaming sheet '{ws_excel.Name}' to '{target_sheet_name}'...")
                ws_excel.Name = target_sheet_name
        else:
            ws_excel = wb_excel.Worksheets(target_sheet_name)
            print(f"✓ '{target_sheet_name}' sheet found")
        
        # Convert all formulas to values in the active sheet
        print("Converting formulas to values...")
        used_range = ws_excel.UsedRange
        
        # Copy the used range
        used_range.Copy()
        
        # Paste as values only
        # xlPasteValues = -4163
        used_range.PasteSpecial(Paste=-4163)
        
        # Clear clipboard
        excel.CutCopyMode = False
        
        print(f"✓ Converted all formulas to values in '{target_sheet_name}' sheet")
        
        # Save and close
        wb_excel.Save()
        wb_excel.Close(SaveChanges=True)
        excel.Quit()
        
        print("✓ File saved with values only and no external links")
        print()
        
    except Exception as e:
        print(f"Error using Excel automation: {e}")
        print("Falling back to openpyxl method...")
        try:
            excel.Quit()
        except:
            pass
        
        # Fallback to openpyxl method
        wb_prep = openpyxl.load_workbook(input_file)
        print(f"Available sheets: {wb_prep.sheetnames}")
        target_sheet_name = "UAT Detailed Report"
        if target_sheet_name not in wb_prep.sheetnames:
            found_sheet = wb_prep.sheetnames[0]
            for name in wb_prep.sheetnames:
                if 'uat' in name.lower() and 'detailed' in name.lower():
                    found_sheet = name
                    break
            print(f"Renaming sheet '{found_sheet}' to '{target_sheet_name}'...")
            wb_prep[found_sheet].title = target_sheet_name
        wb_prep.save(input_file)
        wb_prep.close()
        print("✓ Sheet name standardized (formulas may still exist)")
        print()
else:
    print("Excel automation not available. Please ensure the file has no formulas/external links.")
    print("You may need to manually:")
    print("  1. Open the file in Excel")
    print("  2. Go to Data > Edit Links > Break Links")
    print("  3. Select all cells and Copy > Paste Special > Values")
    print("  4. Save the file")
    print()
    
    # Still try to standardize sheet name
    wb_prep = openpyxl.load_workbook(input_file)
    print(f"Available sheets: {wb_prep.sheetnames}")
    target_sheet_name = "UAT Detailed Report"
    if target_sheet_name not in wb_prep.sheetnames:
        found_sheet = wb_prep.sheetnames[0]
        for name in wb_prep.sheetnames:
            if 'uat' in name.lower() and 'detailed' in name.lower():
                found_sheet = name
                break
        print(f"Renaming sheet '{found_sheet}' to '{target_sheet_name}'...")
        wb_prep[found_sheet].title = target_sheet_name
        wb_prep.save(input_file)
    wb_prep.close()
    print("✓ Sheet name standardized")
    print()

# Step 1: Read the Excel file and unmerge cells
print("Step 1: Reading UAT Release Detailed Report and unmerging cells...")
print("-" * 80)

# Load workbook with openpyxl to handle merged cells
wb = openpyxl.load_workbook(input_file)

# Check available sheet names
print(f"Available sheets: {wb.sheetnames}")

# Use the standardized sheet name from Step 0
sheet_name = "UAT Detailed Report"
print(f"Using sheet: {sheet_name}")
ws = wb[sheet_name]

# Find and unmerge all merged cells, filling them with the merged value
print("Unmerging cells and filling values...")
merged_ranges = list(ws.merged_cells.ranges)
for merged_range in merged_ranges:
    # Get the value from the top-left cell
    min_row, min_col = merged_range.min_row, merged_range.min_col
    cell_value = ws.cell(min_row, min_col).value
    
    # Unmerge the cells
    ws.unmerge_cells(str(merged_range))
    
    # Fill all cells in the previously merged range with the value
    for row in range(merged_range.min_row, merged_range.max_row + 1):
        for col in range(merged_range.min_col, merged_range.max_col + 1):
            ws.cell(row, col).value = cell_value

# Save the workbook temporarily to read with pandas
temp_file = os.path.join(base_dir, "temp_unmerged_detailed.xlsx")
wb.save(temp_file)
wb.close()

print(f"Unmerged cells saved to temporary file")
print()

# Step 2: Read with pandas and process the data
print("Step 2: Reading data with pandas...")
print("-" * 80)

# Read the unmerged data
df = pd.read_excel(temp_file, sheet_name=sheet_name)

# Check if the first row contains the actual headers
print(f"Checking for multi-row headers...")
first_row = df.iloc[0]

# If first row contains 'Feature ID', 'Pass %', etc., use it as header
if any('Feature ID' in str(val) or 'Parent ID' in str(val) for val in first_row.values):
    print("Found multi-row headers - using first data row as column names")
    # Set first row as column names
    df.columns = df.iloc[0]
    # Drop the first row since it's now the header
    df = df[1:].reset_index(drop=True)

# Display column names for debugging
print(f"Total rows in sheet: {len(df)}")
print(f"Columns found: {df.columns.tolist()}")
print()

# Find the actual column names (they might have spaces or variations)
feature_id_col = None
execution_pct_col = None
pass_pct_col = None
total_stories_col = None
stories_not_delivered_col = None

# Search for Feature ID column (might be Parent ID or Feature ID)
for col in df.columns:
    col_lower = str(col).lower()
    if ('feature' in col_lower and 'id' in col_lower) or ('parent' in col_lower and 'id' in col_lower):
        feature_id_col = col
        break

# Search for Pass% column
for col in df.columns:
    col_str = str(col).lower().strip()
    if ('pass' in col_str and '%' in col_str) or col_str == 'pass %':
        pass_pct_col = col
        break

# Search for Execution% column
for col in df.columns:
    col_str = str(col).lower().strip()
    if ('execution' in col_str and '%' in col_str) or col_str == 'execution %':
        execution_pct_col = col
        break

# Search for Total US Stories column
for col in df.columns:
    col_lower = str(col).lower().strip()
    if col_lower == 'total stories' or (('total' in col_lower or 'us' in col_lower or 'uat' in col_lower) and 'stories' in col_lower and 'not' not in col_lower):
        total_stories_col = col
        break

# Search for Stories Not Delivered column
for col in df.columns:
    col_lower = str(col).lower().strip()
    if col_lower == 'stories not delivered' or ('stories' in col_lower and 'not' in col_lower and 'delivered' in col_lower):
        stories_not_delivered_col = col
        break
    elif col_lower == 'not delivered' or ('not' in col_lower and 'delivered' in col_lower):
        stories_not_delivered_col = col
        break

if not feature_id_col:
    print("ERROR: Could not find 'Feature ID' or 'Parent ID' column")
    print("Available columns:", df.columns.tolist())
    print("\n⚠️  Script cannot continue without Feature/Parent ID column")
    # Clean up temp file before exit
    if os.path.exists(temp_file):
        os.remove(temp_file)
    exit(1)

if not pass_pct_col:
    print("WARNING: Could not find 'Pass%' column")

if not execution_pct_col:
    print("WARNING: Could not find 'Execution%' column")

if not total_stories_col:
    print("WARNING: Could not find 'Total Stories' column")
    
if not stories_not_delivered_col:
    print("WARNING: Could not find 'Stories Not Delivered' column")

print()
print(f"Feature/Parent ID Column: {feature_id_col}")
print(f"Pass% Column: {pass_pct_col}")
print(f"Execution% Column: {execution_pct_col}")
print(f"Total US Stories Column: {total_stories_col}")
print(f"Stories Not Delivered Column: {stories_not_delivered_col}")
print()

# Step 3: Process the data
print("Step 3: Processing data - grouping by Feature/Parent ID...")
print("-" * 80)

# Filter out rows where Feature ID is null/empty or is the header text
df_filtered = df[df[feature_id_col].notna()].copy()
df_filtered = df_filtered[df_filtered[feature_id_col] != '']
df_filtered = df_filtered[df_filtered[feature_id_col].astype(str).str.strip() != '']
# Remove header text if it appears in data
df_filtered = df_filtered[~df_filtered[feature_id_col].astype(str).str.lower().str.contains('feature|parent|id')]

print(f"Rows with valid Feature/Parent ID: {len(df_filtered)}")

# Convert percentage columns to numeric, handling % signs
def clean_percentage(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        # Remove % sign and convert to float
        value = value.replace('%', '').strip()
        try:
            num = float(value)
            # If it's a percentage value like 81 (from "81%"), convert to decimal 0.81
            # Values typically stored as 0-100 range need to be divided by 100
            if num > 1:
                return num / 100.0
            else:
                return num
        except:
            return None
    # If numeric value (could be 0.81 or 81)
    num = float(value)
    if num > 1:
        return num / 100.0
    else:
        return num

df_filtered['Pass%_numeric'] = df_filtered[pass_pct_col].apply(clean_percentage) if pass_pct_col else 0
df_filtered['Execution%_numeric'] = df_filtered[execution_pct_col].apply(clean_percentage) if execution_pct_col else 0

# Convert story counts to numeric
def clean_numeric(value):
    if pd.isna(value):
        return 0
    if isinstance(value, str):
        value = value.strip()
        try:
            return float(value)
        except:
            return 0
    return float(value)

if total_stories_col:
    df_filtered['Total_Stories_numeric'] = df_filtered[total_stories_col].apply(clean_numeric)
else:
    df_filtered['Total_Stories_numeric'] = 0
    
if stories_not_delivered_col:
    df_filtered['Not_Delivered_numeric'] = df_filtered[stories_not_delivered_col].apply(clean_numeric)
else:
    df_filtered['Not_Delivered_numeric'] = 0

# Group by Feature ID and calculate average for percentages, sum for story counts
agg_dict = {}

if pass_pct_col:
    agg_dict['Pass%_numeric'] = 'mean'
if execution_pct_col:
    agg_dict['Execution%_numeric'] = 'mean'
if total_stories_col:
    agg_dict['Total_Stories_numeric'] = 'sum'
if stories_not_delivered_col:
    agg_dict['Not_Delivered_numeric'] = 'sum'

# If no metrics found, use at least the story counts with default values
if not agg_dict:
    agg_dict = {
        'Total_Stories_numeric': 'sum',
        'Not_Delivered_numeric': 'sum'
    }

result = df_filtered.groupby(feature_id_col).agg(agg_dict).reset_index()

# Rename columns for output
new_columns = {'Feature ID': 'Feature ID'}
if 'Pass%_numeric' in result.columns:
    new_columns['Pass%_numeric'] = 'Pass%'
if 'Execution%_numeric' in result.columns:
    new_columns['Execution%_numeric'] = 'Execution%'
if 'Total_Stories_numeric' in result.columns:
    new_columns['Total_Stories_numeric'] = 'Total UAT Stories'
if 'Not_Delivered_numeric' in result.columns:
    new_columns['Not_Delivered_numeric'] = 'UAT Not Delivered'

result.rename(columns=new_columns, inplace=True)

# Ensure all required columns exist (set to 0 if missing)
if 'Pass%' not in result.columns:
    result['Pass%'] = 0.0
if 'Execution%' not in result.columns:
    result['Execution%'] = 0.0
if 'Total UAT Stories' not in result.columns:
    result['Total UAT Stories'] = 0
if 'UAT Not Delivered' not in result.columns:
    result['UAT Not Delivered'] = 0

# Calculate UAT Delivered
result['UAT Delivered'] = result['Total UAT Stories'] - result['UAT Not Delivered']

# Ensure percentage columns are numeric (safety check)
print("Converting percentage columns to numeric...")
result['Pass%'] = pd.to_numeric(result['Pass%'], errors='coerce').fillna(0)
result['Execution%'] = pd.to_numeric(result['Execution%'], errors='coerce').fillna(0)

# Round percentages to 2 decimal places
result['Pass%'] = result['Pass%'].round(2)
result['Execution%'] = result['Execution%'].round(2)

# Convert story counts to integers
result['Total UAT Stories'] = result['Total UAT Stories'].astype(int)
result['UAT Not Delivered'] = result['UAT Not Delivered'].astype(int)
result['UAT Delivered'] = result['UAT Delivered'].astype(int)

print(f"Unique Feature/Parent IDs found: {len(result)}")
print()
print("Sample output:")
print(result.head(10))
print()

# Step 4: Write output to new sheet
print("Step 4: Writing output to 'Consolidated UAT status' sheet...")
print("-" * 80)

# Load the original workbook to add new sheet
wb = openpyxl.load_workbook(input_file)

# Remove 'Consolidated UAT status' sheet if it already exists
if 'Consolidated UAT status' in wb.sheetnames:
    print("Removing existing 'Consolidated UAT status' sheet...")
    wb.remove(wb['Consolidated UAT status'])

# Create new sheet
ws_output = wb.create_sheet('Consolidated UAT status')

# Write headers
ws_output['A1'] = 'Parent ID'
ws_output['B1'] = 'Total UAT Stories'
ws_output['C1'] = 'UAT Not Delivered'
ws_output['D1'] = 'UAT Delivered'
ws_output['E1'] = 'Execution%'
ws_output['F1'] = 'Pass%'

# Style headers
from openpyxl.styles import Font, PatternFill, Alignment
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
header_alignment = Alignment(horizontal='center', vertical='center')

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws_output[f'{col}1']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Write data
for idx, row in result.iterrows():
    ws_output[f'A{idx+2}'] = row['Feature ID']
    ws_output[f'B{idx+2}'] = int(row['Total UAT Stories'])
    ws_output[f'C{idx+2}'] = int(row['UAT Not Delivered'])
    ws_output[f'D{idx+2}'] = int(row['UAT Delivered'])
    # Multiply by 100 to convert decimal to percentage (1.0 -> 100, 0.82 -> 82)
    ws_output[f'E{idx+2}'] = row['Execution%'] * 100
    ws_output[f'F{idx+2}'] = row['Pass%'] * 100

# Format Execution% and Pass% columns as number with % sign and 2 decimal places
from openpyxl.styles import numbers
for idx in range(2, len(result) + 2):
    ws_output[f'E{idx}'].number_format = '0.00"%"'
    ws_output[f'F{idx}'].number_format = '0.00"%"'

# Adjust column widths
ws_output.column_dimensions['A'].width = 30
ws_output.column_dimensions['B'].width = 18
ws_output.column_dimensions['C'].width = 18
ws_output.column_dimensions['D'].width = 15
ws_output.column_dimensions['E'].width = 15
ws_output.column_dimensions['F'].width = 15

print(f"✓ 'Consolidated UAT status' sheet created")

# Step 5: Create UAT Status sheet with test case data
print("\nStep 5: Creating 'UAT Status' sheet with overall totals...")
print("-" * 80)

# Initialize uat_status_df
uat_status_df = None

# Find test case related columns in the original dataframe
total_test_cases_col = None
completed_test_cases_col = None
pass_col = None
fail_col = None
not_tested_col = None

# Search for test case columns (columns J to N approximately)
for col in df.columns:
    col_lower = str(col).lower().strip()
    if 'total' in col_lower and 'test' in col_lower and 'cases' in col_lower:
        total_test_cases_col = col
    elif 'completed' in col_lower and 'test' in col_lower:
        completed_test_cases_col = col
    elif col_lower == 'pass' and 'test' not in col_lower:
        pass_col = col
    elif col_lower == 'fail' and 'test' not in col_lower:
        fail_col = col
    elif ('not tested' in col_lower or 'not run' in col_lower) and '%' not in col_lower:
        not_tested_col = col

print(f"Test case columns found:")
print(f"  Total Test Cases: {total_test_cases_col}")
print(f"  Pass: {pass_col}")
print(f"  Fail: {fail_col}")
print(f"  Not Tested: {not_tested_col}")

# Calculate overall totals
if total_test_cases_col and pass_col and fail_col:
    total_test_cases = df_filtered[total_test_cases_col].apply(clean_numeric).sum()
    passed = df_filtered[pass_col].apply(clean_numeric).sum()
    failed = df_filtered[fail_col].apply(clean_numeric).sum()
    not_run = df_filtered[not_tested_col].apply(clean_numeric).sum() if not_tested_col else 0
    
    # Read percentages from the Total row in the original sheet
    # Read the original Excel without headers to find the total row
    df_raw = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
    
    # First, find the header row to identify column positions
    header_row_idx = None
    pass_pct_col_idx = None
    execution_pct_col_idx = None
    
    # Search for header row (look for "Pass %" and "Execution %")
    for idx in range(min(10, len(df_raw))):  # Check first 10 rows for headers
        row = df_raw.iloc[idx]
        for col_idx in range(len(row)):
            cell_val = str(row.iloc[col_idx]).lower().strip()
            if 'pass' in cell_val and '%' in cell_val:
                header_row_idx = idx
                pass_pct_col_idx = col_idx
            if 'execution' in cell_val and '%' in cell_val:
                execution_pct_col_idx = col_idx
        if header_row_idx is not None and pass_pct_col_idx is not None and execution_pct_col_idx is not None:
            print(f"  Found header row at index {header_row_idx}")
            print(f"    Pass % column: {pass_pct_col_idx}")
            print(f"    Execution % column: {execution_pct_col_idx}")
            break
    
    # Look for the row that contains "Total" text or our total test cases value
    total_row = None
    total_row_idx = None
    
    for idx in range(len(df_raw) - 20, len(df_raw)):  # Check last 20 rows
        row = df_raw.iloc[idx]
        # Check if this row has "Total" text in first few columns
        has_total_text = False
        for col_idx in range(min(5, len(row))):
            cell_val = str(row.iloc[col_idx]).lower().strip()
            if cell_val == 'total' or cell_val.startswith('total'):
                has_total_text = True
                break
        
        # Or check if this row has the total test cases value
        has_total_value = False
        if not has_total_text:
            for col_idx in range(len(row)):
                cell_val = clean_numeric(row.iloc[col_idx])
                if cell_val == int(total_test_cases):
                    has_total_value = True
                    break
        
        if has_total_text or has_total_value:
            total_row = row
            total_row_idx = idx
            print(f"  Found Total row at index {idx}")
            break
    
    # Initialize default values
    uat_execution_pct = 0.0
    uat_pass_pct = 0.0
    
    # Extract Execution % and Pass % from the total row using the header column indices
    if total_row is not None and pass_pct_col_idx is not None and execution_pct_col_idx is not None:
        # Get values from the exact columns identified in the header
        pass_val = total_row.iloc[pass_pct_col_idx]
        execution_val = total_row.iloc[execution_pct_col_idx]
        
        # Clean the percentage values (might be stored as 0.81 or 81 or "81%")
        def clean_pct_value(val):
            if pd.isna(val):
                return 0.0
            if isinstance(val, str):
                val = val.replace('%', '').strip()
                try:
                    num = float(val)
                    # If value is > 1, it's stored as percentage (81 means 81%)
                    if num > 1:
                        return num / 100.0
                    else:
                        return num
                except:
                    return 0.0
            else:
                # If value is > 1, it's stored as percentage (81 means 81%)
                if val > 1:
                    return val / 100.0
                else:
                    return val
        
        uat_pass_pct = clean_pct_value(pass_val)
        uat_execution_pct = clean_pct_value(execution_val)
        
        print(f"  Using percentages from Total row: Pass% = {uat_pass_pct:.4f}, Execution% = {uat_execution_pct:.4f}")
    else:
        # Fallback to calculated average if columns not found
        # Note: df_filtered percentages are already in decimal format (0.81 not 81)
        if execution_pct_col:
            uat_execution_pct = df_filtered['Execution%_numeric'].mean()
        if pass_pct_col:
            uat_pass_pct = df_filtered['Pass%_numeric'].mean()
        print(f"  Could not find columns in Total row - using calculated averages")
    
    print(f"\n✓ Overall totals calculated:")
    print(f"  Total UAT Test Cases: {int(total_test_cases)}")
    print(f"  Passed: {int(passed)}")
    print(f"  Failed: {int(failed)}")
    print(f"  Not Run: {int(not_run)}")
    print(f"  UAT Execution %: {uat_execution_pct * 100:.2f}%")
    print(f"  UAT Pass %: {uat_pass_pct * 100:.2f}%")
    
    # Remove 'UAT Status' sheet if it already exists
    if 'UAT Status' in wb.sheetnames:
        print("\nRemoving existing 'UAT Status' sheet...")
        wb.remove(wb['UAT Status'])
    
    # Create UAT Status sheet with transposed format
    ws_uat = wb.create_sheet('UAT Status')
    
    # Write data in vertical format (transposed)
    # Column A: Metric names, Column B: Values
    ws_uat['A1'] = 'Total UAT Test Cases'
    ws_uat['B1'] = int(total_test_cases)
    
    ws_uat['A2'] = 'Passed'
    ws_uat['B2'] = int(passed)
    
    ws_uat['A3'] = 'Failed'
    ws_uat['B3'] = int(failed)
    
    ws_uat['A4'] = 'Blocked'
    ws_uat['B4'] = 0
    
    ws_uat['A5'] = 'Not Run'
    ws_uat['B5'] = int(not_run)
    
    ws_uat['A6'] = 'UAT Execution %'
    ws_uat['B6'] = uat_execution_pct * 100
    
    ws_uat['A7'] = 'UAT Pass %'
    ws_uat['B7'] = uat_pass_pct * 100
    
    # Style the metric names (Column A)
    from openpyxl.styles import Font, PatternFill, Alignment
    label_font = Font(bold=True, size=11)
    label_alignment = Alignment(horizontal='left', vertical='center')
    
    for row in range(1, 8):
        ws_uat[f'A{row}'].font = label_font
        ws_uat[f'A{row}'].alignment = label_alignment
    
    # Format percentage rows
    ws_uat['B6'].number_format = '0.00"%"'
    ws_uat['B7'].number_format = '0.00"%"'
    
    # Adjust column widths
    ws_uat.column_dimensions['A'].width = 25
    ws_uat.column_dimensions['B'].width = 15
    
    print(f"\n✓ 'UAT Status' sheet created with overall totals")
    uat_status_df = "created"  # Just a flag to indicate it was created
else:
    print("⚠️  Could not find test case columns - skipping UAT Status sheet")
    uat_status_df = None

# Save workbook to the same input file
try:
    wb.save(input_file)
    wb.close()
    
    print(f"\n✓ Output written to: {input_file}")
    print(f"✓ Sheets created:")
    print(f"  - Consolidated UAT status")
    if uat_status_df is not None:
        print(f"  - UAT Status")
    print()
except PermissionError:
    wb.close()
    print(f"\n❌ ERROR: Cannot save file - it is currently open in Excel!")
    print(f"\nPlease:")
    print(f"  1. Close the file: {input_file}")
    print(f"  2. Run this script again")
    print()
    # Clean up temp file before exit
    if os.path.exists(temp_file):
        os.remove(temp_file)
    exit(1)

# Clean up temporary file
if os.path.exists(temp_file):
    os.remove(temp_file)
    print("Temporary file cleaned up")

print()
print("=" * 80)
print("UAT DETAILED REPORT PROCESSING COMPLETE!")
print("=" * 80)
print()
print("Summary:")
print(f"  Input File: {input_file}")
print(f"  Input Sheet: {sheet_name}")
print(f"  Output Sheets: ")
print(f"    1. Consolidated UAT status - {len(result)} rows")
if uat_status_df is not None:
    print(f"    2. UAT Status - {len(uat_status_df)} rows")
print(f"  Unique Feature/Parent IDs: {len(result)}")
print(f"  Total rows processed: {len(df_filtered)}")
print()
print("'Consolidated UAT status' sheet columns:")
print("  1. Parent ID - Unique feature identifiers")
print("  2. Total UAT Stories - Sum of Total US Stories for each Parent ID")
print("  3. UAT Not Delivered - Sum of Stories Not Delivered for each Parent ID")
print("  4. UAT Delivered - Total UAT Stories minus UAT Not Delivered")
print("  5. Execution% - Average execution percentage across all occurrences")
print("  6. Pass% - Average pass percentage across all occurrences")
if uat_status_df is not None:
    print()
    print("'UAT Status' sheet columns:")
    print("  1. Parent ID - Unique feature identifiers")
    print("  2. Total UAT Test Cases - Sum of test cases for each Parent ID")
    print("  3. Passed - Sum of passed test cases")
    print("  4. Failed - Sum of failed test cases")
    print("  5. Blocked - Blocked test cases (currently 0)")
    print("  6. Not Run - Sum of not tested/not run test cases")
    print("  7. UAT Execution % - Average execution percentage")
    print("  8. UAT Pass % - Average pass percentage")

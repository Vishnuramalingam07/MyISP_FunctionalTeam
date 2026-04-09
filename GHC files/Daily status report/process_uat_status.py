import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

# File path
base_dir = r"C:\Users\vishnu.ramalingam\MyISP_Tools\GHC files\Daily status report"
input_file = os.path.join(base_dir, "UAT Status Excel.xlsx")

print("=" * 80)
print("UAT STATUS PROCESSING")
print("=" * 80)
print()

# Step 1: Read the Excel file and unmerge cells
print("Step 1: Reading UAT Status Excel and unmerging cells...")
print("-" * 80)

# Load workbook with openpyxl to handle merged cells
wb = openpyxl.load_workbook(input_file)

# Check available sheet names
print(f"Available sheets: {wb.sheetnames}")

# Try to find the UAT Status sheet
sheet_name = None
for name in wb.sheetnames:
    if 'uat' in name.lower() or 'status' in name.lower():
        sheet_name = name
        break

# If not found, use the first sheet
if not sheet_name:
    sheet_name = wb.sheetnames[0]

print(f"Using sheet: {sheet_name}")
ws = wb[sheet_name]

# Find and unmerge all merged cells, filling them with the merged value
print("Unmerging cells and filling values...")
merged_ranges = list(ws.merged_cells.ranges)
for merged_range in merged_ranges:
    # Get the value from the top-left cell
    min_row, min_col = merged_range.min_row, merged_range.min_col
    cell_value = ws.cell(min_row, min_col).value
    
    # Unmerge the cells
    ws.unmerge_cells(str(merged_range))
    
    # Fill all cells in the previously merged range with the value
    for row in range(merged_range.min_row, merged_range.max_row + 1):
        for col in range(merged_range.min_col, merged_range.max_col + 1):
            ws.cell(row, col).value = cell_value

# Save the workbook temporarily to read with pandas
temp_file = os.path.join(base_dir, "temp_unmerged.xlsx")
wb.save(temp_file)
wb.close()

print(f"Unmerged cells saved to temporary file")
print()

# Step 2: Read with pandas and process the data
print("Step 2: Reading data with pandas...")
print("-" * 80)

# Read the unmerged data
df = pd.read_excel(temp_file, sheet_name=sheet_name)

# Display column names for debugging
print(f"Total rows in sheet: {len(df)}")
print(f"Columns found: {df.columns.tolist()}")
print()

# Find the actual column names (they might have spaces or variations)
feature_id_col = None
execution_pct_col = None
pass_pct_col = None
total_stories_col = None
stories_not_delivered_col = None

# Search for Feature ID column
for col in df.columns:
    if 'feature' in str(col).lower() and 'id' in str(col).lower():
        feature_id_col = col
        break

# Search for Pass% column (column I)
for col in df.columns:
    if 'pass' in str(col).lower() and '%' in str(col):
        pass_pct_col = col
        break

# Search for Execution% column (column L)
for col in df.columns:
    if 'execution' in str(col).lower() and '%' in str(col):
        execution_pct_col = col
        break

# Search for Total US Stories column
for col in df.columns:
    if 'total' in str(col).lower() and 'us' in str(col).lower() and 'stories' in str(col).lower():
        total_stories_col = col
        break

# Search for Stories Not Delivered column
for col in df.columns:
    if 'stories' in str(col).lower() and 'not' in str(col).lower() and 'delivered' in str(col).lower():
        stories_not_delivered_col = col
        break

if not feature_id_col:
    print("ERROR: Could not find 'Feature ID' column")
    print("Available columns:", df.columns.tolist())
    # Try to use column index if name not found
    print("\nAttempting to use column by index...")
    # Typically Feature ID might be in early columns
    for i, col in enumerate(df.columns):
        print(f"Column {i}: {col} - Sample values: {df[col].head(3).tolist()}")

if not pass_pct_col:
    print("WARNING: Could not find 'Pass%' column by name, checking column I (index 8)...")
    if len(df.columns) > 8:
        pass_pct_col = df.columns[8]
        print(f"Using column: {pass_pct_col}")

if not execution_pct_col:
    print("WARNING: Could not find 'Execution%' column by name, checking column L (index 11)...")
    if len(df.columns) > 11:
        execution_pct_col = df.columns[11]
        print(f"Using column: {execution_pct_col}")

print()
print(f"Feature ID Column: {feature_id_col}")
print(f"Pass% Column: {pass_pct_col}")
print(f"Execution% Column: {execution_pct_col}")
print(f"Total US Stories Column: {total_stories_col}")
print(f"Stories Not Delivered Column: {stories_not_delivered_col}")
print()

# Step 3: Process the data
print("Step 3: Processing data - grouping by Feature ID...")
print("-" * 80)

# Filter out rows where Feature ID is null/empty or is the header text "Feature ID"
df_filtered = df[df[feature_id_col].notna()].copy()
df_filtered = df_filtered[df_filtered[feature_id_col] != '']
df_filtered = df_filtered[df_filtered[feature_id_col] != 'Feature ID']  # Remove header text if it appears in data

print(f"Rows with valid Feature ID: {len(df_filtered)}")

# Convert percentage columns to numeric, handling % signs
def clean_percentage(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        # Remove % sign and convert to float
        value = value.replace('%', '').strip()
        try:
            return float(value)
        except:
            return None
    return float(value)

df_filtered['Pass%_numeric'] = df_filtered[pass_pct_col].apply(clean_percentage)
df_filtered['Execution%_numeric'] = df_filtered[execution_pct_col].apply(clean_percentage)

# Convert story counts to numeric
def clean_numeric(value):
    if pd.isna(value):
        return 0
    if isinstance(value, str):
        value = value.strip()
        try:
            return float(value)
        except:
            return 0
    return float(value)

if total_stories_col:
    df_filtered['Total_Stories_numeric'] = df_filtered[total_stories_col].apply(clean_numeric)
else:
    df_filtered['Total_Stories_numeric'] = 0
    
if stories_not_delivered_col:
    df_filtered['Not_Delivered_numeric'] = df_filtered[stories_not_delivered_col].apply(clean_numeric)
else:
    df_filtered['Not_Delivered_numeric'] = 0

# Group by Feature ID and calculate average for percentages, sum for story counts
agg_dict = {
    'Pass%_numeric': 'mean',
    'Execution%_numeric': 'mean',
    'Total_Stories_numeric': 'sum',
    'Not_Delivered_numeric': 'sum'
}

result = df_filtered.groupby(feature_id_col).agg(agg_dict).reset_index()

# Rename columns for output
result.columns = ['Feature ID', 'Pass%', 'Execution%', 'Total UAT Stories', 'UAT Not Delivered']

# Calculate UAT Delivered
result['UAT Delivered'] = result['Total UAT Stories'] - result['UAT Not Delivered']

# Round percentages to 2 decimal places
result['Pass%'] = result['Pass%'].round(2)
result['Execution%'] = result['Execution%'].round(2)

# Convert story counts to integers
result['Total UAT Stories'] = result['Total UAT Stories'].astype(int)
result['UAT Not Delivered'] = result['UAT Not Delivered'].astype(int)
result['UAT Delivered'] = result['UAT Delivered'].astype(int)

print(f"Unique Feature IDs found: {len(result)}")
print()
print("Sample output:")
print(result.head(10))
print()

# Step 4: Write output to new sheet
print("Step 4: Writing output to 'Consolidated UAT status' sheet...")
print("-" * 80)

# Load the original workbook to add new sheet
wb = openpyxl.load_workbook(input_file)

# Remove 'Consolidated UAT status' sheet if it already exists
if 'Consolidated UAT status' in wb.sheetnames:
    print("Removing existing 'Consolidated UAT status' sheet...")
    wb.remove(wb['Consolidated UAT status'])

# Create new sheet
ws_output = wb.create_sheet('Consolidated UAT status')

# Write headers
ws_output['A1'] = 'Parent ID'
ws_output['B1'] = 'Total UAT Stories'
ws_output['C1'] = 'UAT Not Delivered'
ws_output['D1'] = 'UAT Delivered'
ws_output['E1'] = 'Execution%'
ws_output['F1'] = 'Pass%'

# Style headers
from openpyxl.styles import Font, PatternFill, Alignment
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
header_alignment = Alignment(horizontal='center', vertical='center')

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws_output[f'{col}1']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Write data
for idx, row in result.iterrows():
    ws_output[f'A{idx+2}'] = row['Feature ID']
    ws_output[f'B{idx+2}'] = int(row['Total UAT Stories'])
    ws_output[f'C{idx+2}'] = int(row['UAT Not Delivered'])
    ws_output[f'D{idx+2}'] = int(row['UAT Delivered'])
    # Multiply by 100 to convert decimal to percentage (1.0 -> 100, 0.82 -> 82)
    ws_output[f'E{idx+2}'] = row['Execution%'] * 100
    ws_output[f'F{idx+2}'] = row['Pass%'] * 100

# Format Execution% and Pass% columns as number with % sign and 2 decimal places
from openpyxl.styles import numbers
for idx in range(2, len(result) + 2):
    ws_output[f'E{idx}'].number_format = '0.00"%"'
    ws_output[f'F{idx}'].number_format = '0.00"%"'

# Adjust column widths
ws_output.column_dimensions['A'].width = 30
ws_output.column_dimensions['B'].width = 18
ws_output.column_dimensions['C'].width = 18
ws_output.column_dimensions['D'].width = 15
ws_output.column_dimensions['E'].width = 15
ws_output.column_dimensions['F'].width = 15

# Save workbook to the same input file
wb.save(input_file)
wb.close()

print(f"Output written to: {input_file}")
print(f"Sheet name: 'Consolidated UAT status'")
print()

# Clean up temporary file
if os.path.exists(temp_file):
    os.remove(temp_file)
    print("Temporary file cleaned up")

print()
print("=" * 80)
print("UAT STATUS PROCESSING COMPLETE!")
print("=" * 80)
print()
print("Summary:")
print(f"  Input File: {input_file}")
print(f"  Input Sheet: {sheet_name}")
print(f"  Output Sheet: Consolidated UAT status")
print(f"  Unique Feature IDs: {len(result)}")
print(f"  Total rows processed: {len(df_filtered)}")
print()
print("Columns in output:")
print("  1. Parent ID - Unique feature identifiers")
print("  2. Total UAT Stories - Sum of Total US Stories for each Parent ID")
print("  3. UAT Not Delivered - Sum of Stories Not Delivered for each Parent ID")
print("  4. UAT Delivered - Total UAT Stories minus UAT Not Delivered")
print("  5. Execution% - Average execution percentage across all occurrences")
print("  6. Pass% - Average pass percentage across all occurrences")

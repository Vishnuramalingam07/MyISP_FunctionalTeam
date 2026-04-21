import pandas as pd
import os
import warnings
warnings.filterwarnings("ignore", category=Warning, module="requests")
import requests
from datetime import datetime
from config import OUTPUT_DIR, POD_MAPPING_PATH, PT_STATUS_FILE, STORY_SUMMARY_FILE

# Azure DevOps configuration
org = "accenturecio08"
project = "AutomationProcess_29697"
query_id = "94403798-5713-49c0-b713-ce33fb41dafa"
api_url = f"https://dev.azure.com/{org}/{project}/_apis/wit/wiql/{query_id}?api-version=7.0"

# Get PAT from environment variable
pat = os.environ.get("AZURE_DEVOPS_PAT")
if not pat:
    raise RuntimeError("Azure DevOps PAT not found in environment variable 'AZURE_DEVOPS_PAT'. Please set it before running the script.")
headers = {"Content-Type": "application/json"}

print("Fetching work items from ADO query...")

# Fetch work item IDs from query
response = requests.get(api_url, auth=("", pat), headers=headers)
response.raise_for_status()
work_items = response.json()["workItems"]
ids = [str(item["id"]) for item in work_items]

if not ids:
    print("No work items found for this query.")
    exit()

print(f"Found {len(ids)} work items")

# Fetch work item details with specified fields
def fetch_details(ids):
    url = f"https://dev.azure.com/{org}/_apis/wit/workitemsbatch?api-version=7.0"
    payload = {
        "ids": ids,
        "fields": [
            "System.Id",
            "System.Title",
            "System.State",
            "System.AreaPath",
            "System.Tags",
            "System.Parent",
            "Microsoft.VSTS.Scheduling.StoryPoints",
            "Custom.mySPInitiative",
            "Custom.SubInitiative",
            "Custom.TextVerification",
            "Custom.TestingPOC",
            "Custom.PlannedforPTDate",
            "Custom.ActualPTDate",
            "Custom.PlannedUATDate",
            "Custom.ActualUATDate"
        ]
    }
    response = requests.post(url, auth=("", pat), headers=headers, json=payload)
    response.raise_for_status()
    return response.json()["value"]

# Fetch all work item details in batches
batch_size = 200
all_details = []
for i in range(0, len(ids), batch_size):
    batch_ids = ids[i:i + batch_size]
    all_details.extend(fetch_details(batch_ids))
    print(f"  Processed {min(i + batch_size, len(ids))}/{len(ids)} work items")

print("Processing work items...")

# Read mapping file
mapping_path = POD_MAPPING_PATH
print(f"Reading mapping file: {mapping_path}")
mapping_df = pd.read_csv(mapping_path)

# Create mapping dictionaries (case-insensitive)
node_to_m_poc = {str(k).strip().lower(): v for k, v in zip(mapping_df['Node Name'], mapping_df['M POC'])}
node_to_sm_poc = {str(k).strip().lower(): v for k, v in zip(mapping_df['Node Name'], mapping_df['SM Name'])}
node_to_ad_poc = {str(k).strip().lower(): v for k, v in zip(mapping_df['Node Name'], mapping_df['AD POC'])}

# Prepare data
data = []
parent_ids_set = set()
for item in all_details:
    fields = item["fields"]
    area_path = fields.get("System.AreaPath", "")
    node_name = area_path.split('\\')[-1] if area_path else ""
    
    # Extract Parent ID if it exists
    parent = fields.get("System.Parent")
    parent_id = ""
    if isinstance(parent, dict):
        parent_id = parent.get("id", "")
    elif parent:
        parent_id = str(parent)
    
    # Collect parent IDs for later fetching
    if parent_id:
        parent_ids_set.add(parent_id)
    
    # Extract Testing POC display name
    testing_poc = fields.get("Custom.TestingPOC", "")
    if isinstance(testing_poc, dict):
        testing_poc = testing_poc.get("displayName", "")
    elif testing_poc:
        testing_poc = str(testing_poc)
    
    # Get tags and check for exact "Testing NA" tag
    tags = fields.get("System.Tags", "")
    
    # Split tags by semicolon and check for exact match (case-insensitive)
    is_testing_na = False
    is_uat_not_testable = False
    
    if tags:
        tags_list = [tag.strip().lower() for tag in str(tags).split(';')]
        
        # Check for exact "testing na" tag (not "testing na - only for uat")
        if 'testing na' in tags_list:
            is_testing_na = True
        
        # Check for UAT not testable tags
        uat_na_tags = [
            'testing na - only for uat',
            'bpms bau',
            'ims bau',
            'uat testing not required'
        ]
        for uat_tag in uat_na_tags:
            if uat_tag in tags_list:
                is_uat_not_testable = True
                break
    
    # PT testable stories are those without "Testing NA" tag
    is_pt_testable = not is_testing_na
    
    # UAT classification only applies to PT testable stories
    uat_testable = 0
    uat_not_testable = 0
    if is_pt_testable:
        if is_uat_not_testable:
            uat_not_testable = 1
        else:
            uat_testable = 1
    
    # PT Delivery classification only applies to PT testable stories
    pt_delivered = 0
    pt_not_delivered = 0
    state = fields.get("System.State", "")
    state_lower = str(state).strip().lower() if state else ""
    
    if is_pt_testable:
        pt_delivered_states = [
            'ready for uat', 'ready to test', 'ready for test', 'in test', 
            'test complete', 'blocked', 'closed', 'ready for e2e test', 
            'blocked in pt', 'pt in test', 'blocked in uat', 'uat in test', 
            'awaiting uat deployment'
        ]
        if state_lower in pt_delivered_states:
            pt_delivered = 1
        else:
            pt_not_delivered = 1
    
    # UAT Delivery classification only applies to UAT testable stories
    uat_delivered = 0
    uat_not_delivered = 0
    if uat_testable:
        uat_delivered_states = [
            'ready for uat', 'test complete', 'closed', 
            'uat in test', 'blocked in uat'
        ]
        if state_lower in uat_delivered_states:
            uat_delivered = 1
        else:
            uat_not_delivered = 1
    
    # Get POC mapping based on Node Name
    node_name_lower = node_name.strip().lower()
    m_poc = node_to_m_poc.get(node_name_lower, "")
    sm_poc = node_to_sm_poc.get(node_name_lower, "")
    ad_poc = node_to_ad_poc.get(node_name_lower, "")
    
    data.append({
        "mySP Initiative": fields.get("Custom.mySPInitiative", ""),
        "Sub Initiative": fields.get("Custom.SubInitiative", ""),
        "TextVerification": fields.get("Custom.TextVerification", ""),
        "Parent": parent_id,
        "ID": fields.get("System.Id", ""),
        "Title": fields.get("System.Title", ""),
        "State": fields.get("System.State", ""),
        "Node Name": node_name,
        "M POC": m_poc,
        "SM POC": sm_poc,
        "AD POC": ad_poc,
        "Testing POC": testing_poc,
        "Planned for PT Date": fields.get("Custom.PlannedforPTDate", ""),
        "Actual PT Date": fields.get("Custom.ActualPTDate", ""),
        "Planned UAT Date": fields.get("Custom.PlannedUATDate", ""),
        "Actual UAT Date": fields.get("Custom.ActualUATDate", ""),
        "Story Points": fields.get("Microsoft.VSTS.Scheduling.StoryPoints", ""),
        "Tags": tags,
        "Testing NA stories": 1 if is_testing_na else 0,
        "PT testable stories": 1 if is_pt_testable else 0,
        "UAT Testable Stories": uat_testable,
        "UAT Testing NA stories": uat_not_testable,
        "PT delivered": pt_delivered,
        "PT NOT delivered": pt_not_delivered,
        "UAT delivered": uat_delivered,
        "UAT NOT delivered": uat_not_delivered
    })

# Fetch parent titles from ADO
print("\nFetching parent titles from ADO...")
parent_titles = {}
if parent_ids_set:
    parent_ids_list = list(parent_ids_set)
    # Fetch parent details in batches
    for i in range(0, len(parent_ids_list), batch_size):
        batch_ids = [int(pid) for pid in parent_ids_list[i:i + batch_size]]
        url = f"https://dev.azure.com/{org}/_apis/wit/workitemsbatch?api-version=7.0"
        payload = {
            "ids": batch_ids,
            "fields": ["System.Id", "System.Title"]
        }
        response = requests.post(url, auth=("", pat), headers=headers, json=payload)
        response.raise_for_status()
        parent_details = response.json()["value"]
        for parent in parent_details:
            parent_id = parent["id"]
            parent_title = parent["fields"].get("System.Title", "")
            parent_titles[parent_id] = parent_title
        print(f"  Fetched {min(i + batch_size, len(parent_ids_list))}/{len(parent_ids_list)} parent titles")

# Add Parent Title to data
for record in data:
    parent_id = record.get("Parent", "")
    if parent_id:
        # Convert parent_id to int for matching
        try:
            parent_id_int = int(parent_id)
            record["Parent Title"] = parent_titles.get(parent_id_int, "")
        except (ValueError, TypeError):
            record["Parent Title"] = ""
    else:
        record["Parent Title"] = ""

df = pd.DataFrame(data)

# Reorder columns to place Parent Title after Parent
columns = list(df.columns)
parent_idx = columns.index("Parent")
columns.insert(parent_idx + 1, columns.pop(columns.index("Parent Title")))
df = df[columns]

# Define output directory
output_dir = OUTPUT_DIR

# Read PT Status excel for lookup (replacing Status excel.xlsx dependency)
pt_status_excel_path = PT_STATUS_FILE
print(f"\nReading status excel: {pt_status_excel_path}")
# Read PT status sheet (case insensitive sheet name matching)
try:
    status_df = pd.read_excel(pt_status_excel_path, sheet_name='PT status')
except ValueError:
    # Try case-insensitive match if exact name doesn't work
    xl = pd.ExcelFile(pt_status_excel_path)
    sheet_names_lower = {name.lower(): name for name in xl.sheet_names}
    actual_sheet_name = sheet_names_lower.get('pt status', 'PT status')
    status_df = pd.read_excel(pt_status_excel_path, sheet_name=actual_sheet_name)

# Create lookup dictionary from Status excel by ID
status_lookup = {}
for _, row in status_df.iterrows():
    work_id = row.get('ID', '')
    if work_id:
        # Convert to int first to remove decimal, then to string for consistent matching
        try:
            work_id_clean = str(int(float(work_id)))
        except (ValueError, TypeError):
            work_id_clean = str(work_id)
        
        status_lookup[work_id_clean] = {
            'Passed': row.get('Passed', 0),
            'Failed': row.get('Failed', 0),
            'Blocked': row.get('Blocked', 0),
            'Not Run': row.get('Not Run', 0)
        }

# Add status columns to dataframe - convert to numeric, default to 0
# Convert IDs to int then to string to match the lookup keys
df['ID_str'] = df['ID'].astype(str).str.replace('.0', '', regex=False)
df['Passed'] = df['ID_str'].map(lambda x: status_lookup.get(x, {}).get('Passed', 0))
df['Failed'] = df['ID_str'].map(lambda x: status_lookup.get(x, {}).get('Failed', 0))
df['Blocked'] = df['ID_str'].map(lambda x: status_lookup.get(x, {}).get('Blocked', 0))
df['Not Run'] = df['ID_str'].map(lambda x: status_lookup.get(x, {}).get('Not Run', 0))

# Drop the temporary ID_str column
df.drop('ID_str', axis=1, inplace=True)

# Convert to numeric to handle any non-numeric values
df['Passed'] = pd.to_numeric(df['Passed'], errors='coerce').fillna(0).astype(int)
df['Failed'] = pd.to_numeric(df['Failed'], errors='coerce').fillna(0).astype(int)
df['Blocked'] = pd.to_numeric(df['Blocked'], errors='coerce').fillna(0).astype(int)
df['Not Run'] = pd.to_numeric(df['Not Run'], errors='coerce').fillna(0).astype(int)

# Save to Excel with multiple sheets
output_path = os.path.join(output_dir, "Story_summary.xlsx")

# Filter PT testable stories
testable_stories_df = df[df['PT testable stories'] == 1].copy()

# Write to Excel with multiple sheets
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Story Summary')
    testable_stories_df.to_excel(writer, index=False, sheet_name='Testable stories')
    
    # Get the Testable stories sheet to add formulas
    workbook = writer.book
    testable_sheet = writer.sheets['Testable stories']
    
    # Find column indices for Passed, Failed, Blocked, Not Run
    headers = list(testable_stories_df.columns)
    passed_col = headers.index('Passed') + 1  # +1 because Excel columns are 1-indexed
    failed_col = headers.index('Failed') + 1
    blocked_col = headers.index('Blocked') + 1
    not_run_col = headers.index('Not Run') + 1
    
    # Add new column headers
    total_col = len(headers) + 1
    exec_pct_col = len(headers) + 2
    pass_pct_col = len(headers) + 3
    
    testable_sheet.cell(row=1, column=total_col, value='Total')
    testable_sheet.cell(row=1, column=exec_pct_col, value='Execution %')
    testable_sheet.cell(row=1, column=pass_pct_col, value='Pass %')
    
    # Add formulas for each data row
    from openpyxl.styles import numbers
    from openpyxl.utils import get_column_letter
    
    for row_num in range(2, len(testable_stories_df) + 2):  # Start from row 2 (after header)
        # Total formula
        total_formula = f"={get_column_letter(passed_col)}{row_num}+{get_column_letter(failed_col)}{row_num}+{get_column_letter(blocked_col)}{row_num}+{get_column_letter(not_run_col)}{row_num}"
        testable_sheet.cell(row=row_num, column=total_col, value=total_formula)
        
        # Execution % formula
        exec_formula = f"=IF({get_column_letter(passed_col)}{row_num}+{get_column_letter(failed_col)}{row_num}=0,0,({get_column_letter(passed_col)}{row_num}+{get_column_letter(failed_col)}{row_num})/{get_column_letter(total_col)}{row_num})"
        exec_cell = testable_sheet.cell(row=row_num, column=exec_pct_col, value=exec_formula)
        exec_cell.number_format = '0.00%'
        
        # Pass % formula
        pass_formula = f"=IF({get_column_letter(passed_col)}{row_num}=0,0,{get_column_letter(passed_col)}{row_num}/({get_column_letter(passed_col)}{row_num}+{get_column_letter(failed_col)}{row_num}+{get_column_letter(blocked_col)}{row_num}))"
        pass_cell = testable_sheet.cell(row=row_num, column=pass_pct_col, value=pass_formula)
        pass_cell.number_format = '0.00%'
    
    # Create Consolidated PT_UAT Status sheet
    print("\nCreating Consolidated PT_UAT Status sheet...")
    
    # Filter testable stories with non-empty Parent IDs
    testable_with_parent = testable_stories_df[testable_stories_df['Parent'].notna() & (testable_stories_df['Parent'] != '')].copy()
    
    # Convert Parent to string and remove any decimal points
    testable_with_parent['Parent'] = testable_with_parent['Parent'].astype(str).str.replace('.0', '', regex=False)
    
    # Convert test status columns to numeric (handle empty strings and non-numeric values)
    testable_with_parent['Passed'] = pd.to_numeric(testable_with_parent['Passed'], errors='coerce').fillna(0)
    testable_with_parent['Failed'] = pd.to_numeric(testable_with_parent['Failed'], errors='coerce').fillna(0)
    testable_with_parent['Blocked'] = pd.to_numeric(testable_with_parent['Blocked'], errors='coerce').fillna(0)
    testable_with_parent['Not Run'] = pd.to_numeric(testable_with_parent['Not Run'], errors='coerce').fillna(0)
    
    # Convert Execution % and Pass % to numeric (they are currently formulas in the Excel)
    # We need to calculate them from the data
    testable_with_parent['Execution%_calc'] = testable_with_parent.apply(
        lambda row: (row['Passed'] + row['Failed']) / (row['Passed'] + row['Failed'] + row['Blocked'] + row['Not Run']) 
        if (row['Passed'] + row['Failed'] + row['Blocked'] + row['Not Run']) > 0 else 0, axis=1
    )
    testable_with_parent['Pass%_calc'] = testable_with_parent.apply(
        lambda row: row['Passed'] / (row['Passed'] + row['Failed'] + row['Blocked']) 
        if (row['Passed'] + row['Failed'] + row['Blocked']) > 0 else 0, axis=1
    )
    
    # Group by Parent and calculate averages and sums
    consolidated_pt = testable_with_parent.groupby('Parent').agg({
        'Parent Title': 'first',  # Take the first title (they should all be the same)
        'PT testable stories': 'sum',
        'PT delivered': 'sum',
        'PT NOT delivered': 'sum',
        'Execution%_calc': 'mean',
        'Pass%_calc': 'mean'
    }).reset_index()
    
    # Rename columns
    consolidated_pt.columns = ['Parent ID', 'Parent Title', 'Total PT stories', 'PT Delivered', 'PT Not Delivered', 'PT Execution%', 'PT Pass%']
    
    # Read PO Details Excel for Product Owner lookup
    po_details_path = os.path.join(output_dir, "PO Details.xlsx")
    try:
        print(f"Reading PO Details Excel: {po_details_path}")
        po_df = pd.read_excel(po_details_path)
        
        # Convert Parent ID to string and remove decimals for matching
        if 'Parent ID' in po_df.columns:
            po_df['Parent ID'] = po_df['Parent ID'].astype(str).str.replace('.0', '', regex=False)
        elif 'Parent' in po_df.columns:
            po_df['Parent'] = po_df['Parent'].astype(str).str.replace('.0', '', regex=False)
            po_df.rename(columns={'Parent': 'Parent ID'}, inplace=True)
        
        # Create lookup dictionary for Product Owner
        po_lookup = {}
        for _, row in po_df.iterrows():
            parent_id = str(row['Parent ID']).strip()
            # Check various possible column names for Product Owner
            product_owner = row.get('Product Owner', row.get('PO', row.get('ProductOwner', 'N/A')))
            po_lookup[parent_id] = product_owner
        
        # Add Product Owner column after Parent Title
        consolidated_pt.insert(2, 'Product Owner', consolidated_pt['Parent ID'].apply(
            lambda x: po_lookup.get(str(x).strip(), 'N/A')
        ))
        
        print(f"Successfully loaded Product Owner data for {len(po_lookup)} Parent IDs")
    except FileNotFoundError:
        print(f"Warning: PO Details Excel not found at {po_details_path}")
        print("Setting Product Owner to 'N/A' for all Parent IDs")
        consolidated_pt.insert(2, 'Product Owner', 'N/A')
    except Exception as e:
        print(f"Warning: Error reading PO Details Excel: {e}")
        print("Setting Product Owner to 'N/A' for all Parent IDs")
        consolidated_pt.insert(2, 'Product Owner', 'N/A')
    
    # Read UAT Status Excel for lookup
    uat_status_path = os.path.join(output_dir, "UAT Release Detailed Report.xlsx")
    try:
        print(f"Reading UAT Status Excel: {uat_status_path}")
        uat_df = pd.read_excel(uat_status_path, sheet_name='Consolidated UAT status')
        
        # Convert Parent ID to string and remove decimals for matching
        uat_df['Parent ID'] = uat_df['Parent ID'].astype(str).str.replace('.0', '', regex=False)
        
        # Create lookup dictionary for UAT data
        uat_lookup = {}
        for _, row in uat_df.iterrows():
            parent_id = str(row['Parent ID']).strip()
            uat_lookup[parent_id] = {
                'Total UAT Stories': row.get('Total UAT Stories', 0),
                'UAT Delivered': row.get('UAT Delivered', 0),
                'UAT Not Delivered': row.get('UAT Not Delivered', 0),
                'Execution%': row.get('Execution%', 0),
                'Pass%': row.get('Pass%', 0)
            }
        
        # Add UAT columns using lookup
        consolidated_pt['Total UAT stories'] = consolidated_pt['Parent ID'].apply(
            lambda x: uat_lookup.get(str(x).strip(), {}).get('Total UAT Stories', 'UAT NOT Applicable')
        )
        consolidated_pt['UAT Delivered'] = consolidated_pt['Parent ID'].apply(
            lambda x: uat_lookup.get(str(x).strip(), {}).get('UAT Delivered', 'UAT NOT Applicable')
        )
        consolidated_pt['UAT Not Delivered'] = consolidated_pt['Parent ID'].apply(
            lambda x: uat_lookup.get(str(x).strip(), {}).get('UAT Not Delivered', 'UAT NOT Applicable')
        )
        consolidated_pt['UAT Execution%'] = consolidated_pt['Parent ID'].apply(
            lambda x: uat_lookup.get(str(x).strip(), {}).get('Execution%', 'UAT NOT Applicable')
        )
        consolidated_pt['UAT Pass%'] = consolidated_pt['Parent ID'].apply(
            lambda x: uat_lookup.get(str(x).strip(), {}).get('Pass%', 'UAT NOT Applicable')
        )
        
        print(f"Successfully loaded UAT data for {len(uat_lookup)} Parent IDs")
    except FileNotFoundError:
        print(f"Warning: UAT Status Excel not found at {uat_status_path}")
        print("Setting all UAT columns to 'UAT NOT Applicable'")
        consolidated_pt['Total UAT stories'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Delivered'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Not Delivered'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Execution%'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Pass%'] = 'UAT NOT Applicable'
    except Exception as e:
        print(f"Warning: Error reading UAT Status Excel: {e}")
        print("Setting all UAT columns to 'UAT NOT Applicable'")
        consolidated_pt['Total UAT stories'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Delivered'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Not Delivered'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Execution%'] = 'UAT NOT Applicable'
        consolidated_pt['UAT Pass%'] = 'UAT NOT Applicable'
    
    # Read Bug Summary for defect counts by Parent ID
    bug_summary_path = os.path.join(output_dir, "Open_Bug_summary.xlsx")
    try:
        print(f"Reading Bug Summary Excel: {bug_summary_path}")
        bug_df = pd.read_excel(bug_summary_path)
        
        # Convert Defect Record to string and remove decimals for matching
        bug_df['Defect Record'] = bug_df['Defect Record'].astype(str).str.replace('.0', '', regex=False)
        
        # Create defect count dictionary by Parent ID (Defect Record) and Severity
        defect_counts = {}
        for _, row in bug_df.iterrows():
            defect_record = str(row['Defect Record']).strip()
            severity = str(row.get('Severity', '')).strip()
            
            if defect_record not in defect_counts:
                defect_counts[defect_record] = {
                    'Total': 0,
                    'Critical': 0,
                    'High': 0,
                    'Medium': 0,
                    'Low': 0
                }
            
            defect_counts[defect_record]['Total'] += 1
            
            # Count by severity
            if '1 - Critical' in severity or severity == '1':
                defect_counts[defect_record]['Critical'] += 1
            elif '2 - High' in severity or severity == '2':
                defect_counts[defect_record]['High'] += 1
            elif '3 - Medium' in severity or severity == '3':
                defect_counts[defect_record]['Medium'] += 1
            elif '4 - Low' in severity or severity == '4':
                defect_counts[defect_record]['Low'] += 1
        
        # Add defect count columns
        consolidated_pt['Total Open Bugs for Parent ID'] = consolidated_pt['Parent ID'].apply(
            lambda x: defect_counts.get(str(x).strip(), {}).get('Total', 0)
        )
        consolidated_pt['Critical'] = consolidated_pt['Parent ID'].apply(
            lambda x: defect_counts.get(str(x).strip(), {}).get('Critical', 0)
        )
        consolidated_pt['High'] = consolidated_pt['Parent ID'].apply(
            lambda x: defect_counts.get(str(x).strip(), {}).get('High', 0)
        )
        consolidated_pt['Medium'] = consolidated_pt['Parent ID'].apply(
            lambda x: defect_counts.get(str(x).strip(), {}).get('Medium', 0)
        )
        consolidated_pt['Low'] = consolidated_pt['Parent ID'].apply(
            lambda x: defect_counts.get(str(x).strip(), {}).get('Low', 0)
        )
        
        print(f"Successfully loaded defect counts for {len(defect_counts)} Parent IDs")
    except FileNotFoundError:
        print(f"Warning: Bug Summary Excel not found at {bug_summary_path}")
        print("Setting all defect count columns to 0")
        consolidated_pt['Total Open Bugs for Parent ID'] = 0
        consolidated_pt['Critical'] = 0
        consolidated_pt['High'] = 0
        consolidated_pt['Medium'] = 0
        consolidated_pt['Low'] = 0
    except Exception as e:
        print(f"Warning: Error reading Bug Summary Excel: {e}")
        print("Setting all defect count columns to 0")
        consolidated_pt['Total Open Bugs for Parent ID'] = 0
        consolidated_pt['Critical'] = 0
        consolidated_pt['High'] = 0
        consolidated_pt['Medium'] = 0
        consolidated_pt['Low'] = 0
    
    # Write to Excel
    consolidated_pt.to_excel(writer, index=False, sheet_name='Consolidated PT_UAT Status')
    
    # Format columns
    # Columns: Parent ID, Parent Title, Product Owner, Total PT stories, PT Delivered, PT Not Delivered, PT Execution%, PT Pass%, ...
    consolidated_sheet = writer.sheets['Consolidated PT_UAT Status']
    pt_delivered_col_idx = 5      # Column E (PT Delivered) - number format
    pt_not_delivered_col_idx = 6  # Column F (PT Not Delivered) - number format
    pt_exec_col_idx = 7           # Column G (PT Execution%) - percentage format
    pt_pass_col_idx = 8           # Column H (PT Pass%) - percentage format
    uat_exec_col_idx = 12         # Column L (UAT Execution%) - percentage format
    uat_pass_col_idx = 13         # Column M (UAT Pass%) - percentage format
    
    for row_num in range(2, len(consolidated_pt) + 2):
        # Format PT Delivered and PT Not Delivered as numbers
        pt_delivered_cell = consolidated_sheet.cell(row=row_num, column=pt_delivered_col_idx)
        pt_delivered_cell.number_format = '0'
        
        pt_not_delivered_cell = consolidated_sheet.cell(row=row_num, column=pt_not_delivered_col_idx)
        pt_not_delivered_cell.number_format = '0'
        
        # Format PT Execution% and PT Pass% as percentages (values are already decimals 0-1)
        pt_exec_cell = consolidated_sheet.cell(row=row_num, column=pt_exec_col_idx)
        pt_exec_cell.number_format = '0.00%'
        
        pt_pass_cell = consolidated_sheet.cell(row=row_num, column=pt_pass_col_idx)
        pt_pass_cell.number_format = '0.00%'
        
        # Format UAT percentages (only if numeric)
        uat_exec_val = consolidated_sheet.cell(row=row_num, column=uat_exec_col_idx).value
        if isinstance(uat_exec_val, (int, float)):
            uat_exec_cell = consolidated_sheet.cell(row=row_num, column=uat_exec_col_idx)
            uat_exec_cell.number_format = '0.00%'
            uat_exec_cell.value = uat_exec_val / 100  # Convert back to decimal for percentage format
        
        uat_pass_val = consolidated_sheet.cell(row=row_num, column=uat_pass_col_idx).value
        if isinstance(uat_pass_val, (int, float)):
            uat_pass_cell = consolidated_sheet.cell(row=row_num, column=uat_pass_col_idx)
            uat_pass_cell.number_format = '0.00%'
            uat_pass_cell.value = uat_pass_val / 100  # Convert back to decimal for percentage format
    
    print(f"Consolidated PT_UAT Status sheet created with {len(consolidated_pt)} unique Parent IDs")

print(f"\nData exported successfully!")
print(f"Output file: {output_path}")
print(f"\nSummary:")
print(f"Total Work Items: {len(df)}")
print(f"Testing NA stories: {df['Testing NA stories'].sum()}")
print(f"PT testable stories: {df['PT testable stories'].sum()}")
print(f"PT delivered: {df['PT delivered'].sum()}")
print(f"PT NOT delivered: {df['PT NOT delivered'].sum()}")
print(f"UAT Testable Stories: {df['UAT Testable Stories'].sum()}")
print(f"UAT Testing NA stories: {df['UAT Testing NA stories'].sum()}")
print(f"UAT delivered: {df['UAT delivered'].sum()}")
print(f"UAT NOT delivered: {df['UAT NOT delivered'].sum()}")
if 'State' in df.columns:
    print("\nBy State:")
    for state, count in df['State'].value_counts().items():
        print(f"  {state}: {count}")

import os
import base64
import re
from html import escape
from datetime import datetime
from collections import defaultdict
import requests
import certifi
from typing import Dict, List
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from openpyxl import load_workbook


def generate_enhanced_html_report(
    output_path: str,
    detail_rows: List[dict],
    org: str,
    project: str,
) -> str:
    """Generate HTML report matching the Missing_Data_Report format exactly."""
    os.makedirs(output_path if os.path.isdir(output_path) else os.path.dirname(output_path), exist_ok=True)
    file_path = os.path.join(output_path, "Missing_Data_Report.html") if os.path.isdir(output_path) else output_path

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ba_rows = [r for r in detail_rows if str(r.get("Supervisor", "")).strip().lower() != "not available"]
    po_rows = [r for r in detail_rows if str(r.get("Supervisor", "")).strip().lower() == "not available"]
    total_items = len(detail_rows)
    total_ba = len(ba_rows)
    total_po = len(po_rows)
    total_categories = len(dict.fromkeys(
        str(r.get("Missing") or r.get("Missing Field") or "Unknown").strip() for r in detail_rows
    ))

    def _dropdown(rows: List[dict]) -> str:
        names = sorted({str(r.get("Created By", "")).strip() for r in rows if str(r.get("Created By", "")).strip()})
        opts = '<option value="All">All</option>'
        for n in names:
            opts += f'<option value="{escape(n)}">{escape(n)}</option>'
        return opts

    def _summary_rows(rows: List[dict]) -> str:
        counts: Dict[tuple, int] = {}
        for r in rows:
            key = (str(r.get("Created By", "")).strip(), str(r.get("Missing") or r.get("Missing Field") or "Unknown").strip())
            counts[key] = counts.get(key, 0) + 1
        parts = []
        for (cb, cat), cnt in sorted(counts.items(), key=lambda x: (x[0][0], x[0][1])):
            parts.append(
                f"<tr class='summary-row' data-created-by='{escape(cb)}' data-category='{escape(cat)}'>"
                f"<td title='{escape(cb)}'>{escape(cb)}</td>"
                f"<td title='{escape(cat)}'>{escape(cat)}</td>"
                f"<td class='count-cell'>{cnt}</td></tr>"
            )
        return "".join(parts)

    def _details_rows(rows: List[dict]) -> str:
        cats = list(dict.fromkeys(str(r.get("Missing") or r.get("Missing Field") or "Unknown").strip() for r in rows))
        grouped: Dict[str, list] = {}
        for r in rows:
            cat = str(r.get("Missing") or r.get("Missing Field") or "Unknown").strip()
            grouped.setdefault(cat, []).append(r)
        parts = []
        for cat in cats:
            cat_rows = sorted(grouped[cat], key=lambda r: str(r.get("US ID", "")))
            parts.append(
                f"<tr class='category-header' data-category='{escape(cat)}'>"
                f"<td colspan='5'><strong class='details-dark-title'>{escape(cat)}</strong> "
                f"<span class='category-count'>({len(cat_rows)} items)</span></td></tr>"
            )
            for item in cat_rows:
                cb = str(item.get("Created By", ""))
                sup = str(item.get("Supervisor", ""))
                us_id = str(item.get("US ID", ""))
                title = str(item.get("Title", ""))
                link = (f'<a href="https://dev.azure.com/{org}/{project}/_workitems/edit/{us_id}" '
                        f'target="_blank" class="open-link-btn">OPEN</a>') if us_id.strip() else '<span class="na-link">N/A</span>'
                parts.append(
                    f"<tr class='data-row' data-created-by='{escape(cb)}' data-category='{escape(cat)}'>"
                    f"<td>{escape(us_id)}</td><td>{escape(title)}</td>"
                    f"<td>{escape(cb)}</td><td>{escape(sup)}</td>"
                    f"<td class='link-cell'>{link}</td></tr>"
                )
        return "".join(parts)

    ba_sum_html   = _summary_rows(ba_rows)
    po_sum_html   = _summary_rows(po_rows)
    ba_sum_count  = sum(1 for r in ba_sum_html.split("<tr class='summary-row'") if r.strip()) if ba_sum_html else 0
    po_sum_count  = sum(1 for r in po_sum_html.split("<tr class='summary-row'") if r.strip()) if po_sum_html else 0
    ba_det_html   = _details_rows(ba_rows)
    po_det_html   = _details_rows(po_rows)
    ba_drop       = _dropdown(ba_rows)
    po_drop       = _dropdown(po_rows)

    css = """* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: Arial, sans-serif; background-color: #f5f5f5; }
.header { background-color: #2c3e50; color: white; padding: 16px 24px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); display: flex; justify-content: space-between; align-items: center; }
.header h1 { margin: 0; font-size: 28px; }
.header .header-info { text-align: right; font-size: 13px; color: #ecf0f1; line-height: 1.5; }
.header .header-info div { margin-bottom: 4px; }
.header .header-info div:last-child { margin-bottom: 0; }
.main-tab-nav { background-color: #34495e; display: flex; box-shadow: 0 2px 5px rgba(0,0,0,0.2); }
.main-tab-button { background-color: #34495e; color: white; border: none; cursor: pointer; padding: 16px 32px; transition: all 0.3s; font-size: 16px; font-weight: 600; flex: 1; text-align: center; border-right: 1px solid #2c3e50; }
.main-tab-button:hover { background-color: #4a6278; }
.main-tab-button.active { background-color: #3498db; border-bottom: 4px solid #2980b9; }
.main-tab-content { display: none; background-color: white; }
.main-tab-content.active { display: block; }
.toggle-container { background-color: #ecf0f1; padding: 12px 16px; margin: 0; border-radius: 0; display: flex; justify-content: center; align-items: center; gap: 16px; }
.toggle-label { font-size: 15px; font-weight: 600; color: #2c3e50; }
.toggle-switch { position: relative; display: inline-flex; background-color: #2F3E4E; border-radius: 30px; padding: 4px; }
.toggle-option { position: relative; padding: 8px 20px; border-radius: 25px; font-weight: 500; font-size: 14px; transition: all 0.3s ease; cursor: pointer; white-space: nowrap; color: #FFFFFF; background-color: transparent; border: none; }
.toggle-option.active { background-color: #3FA2DB; color: #FFFFFF; border-radius: 25px; }
.toggle-option:hover:not(.active) { background-color: rgba(255,255,255,0.1); }
.content-section { display: none; padding: 10px 10px; animation: fadeIn 0.3s; }
.content-section.active { display: block; }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }
.summary-header { color: #2c3e50; margin: 0 0 12px 0; padding-bottom: 8px; border-bottom: 2px solid #3498db; font-size: 20px; font-weight: 600; }
h2 { color: #2c3e50; margin: 0 0 16px 0; padding-bottom: 8px; border-bottom: 3px solid #3498db; font-size: 22px; }
.filter-container { margin: 6px 0 8px 0; padding: 8px 10px; background-color: #f8f9fa; border-radius: 4px; display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
.filter-label { font-weight: 600; font-size: 14px; color: #2c3e50; }
.filter-dropdown { height: 36px; padding: 0 10px; font-size: 14px; border: 1px solid #3FA2DB; border-radius: 6px; background-color: white; cursor: pointer; min-width: 200px; transition: border 0.2s ease; }
.filter-dropdown:hover { border-color: #2980b9; }
.filter-dropdown:focus { outline: none; border-color: #3FA2DB; box-shadow: 0 0 0 2px rgba(63,162,219,0.2); }
.clear-filter-btn { padding: 8px 16px; background-color: #7F8C8D; color: #FFFFFF; border: none; border-radius: 6px; font-size: 14px; font-weight: 600; cursor: pointer; transition: background-color 0.2s ease; }
.clear-filter-btn:hover { background-color: #6c7a7d; }
.filter-count { font-size: 14px; color: #34495e; margin-left: auto; font-weight: 500; }
table { border-collapse: collapse; width: 100%; margin-top: 6px; background-color: white; border: 1px solid #ddd; }
thead { background-color: #3FA2DB; color: #FFFFFF; }
th { padding: 8px 6px; text-align: left; font-weight: 600; font-size: 13px; color: #FFFFFF; position: sticky; top: 0; z-index: 10; white-space: nowrap; }
th.count-header { text-align: right; }
td { padding: 7px 6px; border-bottom: 1px solid #E0E0E0; font-size: 13px; font-weight: 400; color: #2C3E50; word-wrap: break-word; overflow-wrap: break-word; line-height: 1.35; }
tbody tr { transition: background-color 0.2s; }
tbody tr:nth-child(even) { background-color: #f8f9fa; }
tbody tr:hover { background-color: #e3f2fd; }
.na { background-color: #ffe6e6 !important; color: #c0392b; font-weight: bold; }
.category-header { background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%) !important; color: white !important; font-size: 15px; cursor: default; }
.category-header td { padding: 8px 6px !important; font-weight: bold; color: #FFFFFF !important; border: none !important; }
.category-header:hover { transform: none !important; box-shadow: none !important; }
.category-count { font-weight: normal; opacity: 0.9; }
.details-dark-title { color: #FFFFFF !important; font-weight: bold; }
.link-cell { text-align: center; }
.open-link-btn { display: inline-block; padding: 6px 14px; background-color: #3498db; color: white; text-decoration: none; border-radius: 4px; font-size: 13px; font-weight: 600; transition: background-color 0.2s ease; border: none; cursor: pointer; }
.open-link-btn:hover { background-color: #2980b9; }
.na-link { color: #95a5a6; font-style: italic; font-size: 13px; }
.summary-row { transition: background-color 0.2s ease; }
.count-cell { text-align: right; font-weight: 700; color: #2C3E50; font-size: 13px; padding-right: 10px !important; }
.table-wrapper { overflow-x: auto; width: 100%; }
table.summary-table { margin-top: 4px; table-layout: auto; width: 100%; border-spacing: 0; }
table.summary-table th, table.summary-table td { padding: 6px 10px; }
table.summary-table th:nth-child(1), table.summary-table td:nth-child(1) { word-wrap: break-word; overflow-wrap: break-word; white-space: normal; color: #2c3e50; }
table.summary-table th:nth-child(2), table.summary-table td:nth-child(2) { white-space: nowrap; color: #5f6368; }
table.summary-table th:nth-child(3), table.summary-table td:nth-child(3) { width: 80px; min-width: 60px; text-align: right; white-space: nowrap; }
.no-records { text-align: center; padding: 32px; font-size: 16px; color: #7f8c8d; background-color: #f8f9fa; border-radius: 4px; margin: 16px 0; border: 2px dashed #bdc3c7; }
@media (max-width: 900px) { table.summary-table th:nth-child(3), table.summary-table td:nth-child(3) { width: 70px; } }
@media (max-width: 768px) { .main-tab-nav { flex-direction: column; } .main-tab-button { border-right: none; border-bottom: 1px solid #2c3e50; } .toggle-container { flex-direction: column; } .toggle-switch { width: 100%; max-width: 280px; } .filter-container { flex-direction: column; align-items: stretch; } .filter-dropdown { width: 100%; min-width: 100%; } .content-section { padding: 8px 6px; } table { font-size: 12px; } th, td { padding: 5px 4px; } table.summary-table th, table.summary-table td { padding: 4px 3px; font-size: 12px; } }
@media (max-width: 480px) { table.summary-table th:nth-child(2), table.summary-table td:nth-child(2) { white-space: normal; } table.summary-table th:nth-child(3), table.summary-table td:nth-child(3) { width: 60px; } }"""

    js = """
function switchMainTab(event, tabId) {
  document.querySelectorAll('.main-tab-content').forEach(function(c) { c.classList.remove('active'); });
  document.querySelectorAll('.main-tab-button').forEach(function(b) { b.classList.remove('active'); });
  document.getElementById(tabId).classList.add('active');
  event.currentTarget.classList.add('active');
}
function toggleSummaryView(view) {
  document.querySelectorAll('#summaryToggle .toggle-option').forEach(function(b) { b.classList.remove('active'); });
  document.querySelector('#summaryToggle [data-view="' + view + '"]').classList.add('active');
  document.getElementById('baSummaryContent').classList.toggle('active', view === 'ba');
  document.getElementById('poDevSummaryContent').classList.toggle('active', view === 'podev');
}
function toggleDetailsView(view) {
  document.querySelectorAll('#detailsToggle .toggle-option').forEach(function(b) { b.classList.remove('active'); });
  document.querySelector('#detailsToggle [data-view="' + view + '"]').classList.add('active');
  document.getElementById('baDetailsContent').classList.toggle('active', view === 'ba');
  document.getElementById('poDevDetailsContent').classList.toggle('active', view === 'podev');
}
function _filterSummary(tableId, filterId, countId, noRecId) {
  var val = document.getElementById(filterId).value;
  var rows = document.querySelectorAll('#' + tableId + ' .summary-row');
  var count = 0;
  rows.forEach(function(r) {
    var show = val === 'All' || r.getAttribute('data-created-by') === val;
    r.style.display = show ? '' : 'none';
    if (show) count++;
  });
  document.getElementById(countId).textContent = count;
  document.getElementById(noRecId).style.display = count === 0 ? 'block' : 'none';
}
function filterBASummary() { _filterSummary('baSummaryTable','baSummaryFilter','baSummaryCount','noRecordsBASummary'); }
function clearBASummaryFilter() { document.getElementById('baSummaryFilter').value = 'All'; filterBASummary(); }
function filterPODevSummary() { _filterSummary('poDevSummaryTable','poDevSummaryFilter','poDevSummaryCount','noRecordsPODevSummary'); }
function clearPODevSummaryFilter() { document.getElementById('poDevSummaryFilter').value = 'All'; filterPODevSummary(); }
function _filterDetails(tableId, filterId, countId, noRecId) {
  var val = document.getElementById(filterId).value;
  var rows = document.querySelectorAll('#' + tableId + ' .data-row');
  var count = 0;
  rows.forEach(function(r) {
    var show = val === 'All' || r.getAttribute('data-created-by') === val;
    r.style.display = show ? '' : 'none';
    if (show) count++;
  });
  document.getElementById(countId).textContent = count;
  document.getElementById(noRecId).style.display = count === 0 ? 'block' : 'none';
  document.querySelectorAll('#' + tableId + ' .category-header').forEach(function(hdr) {
    var cat = hdr.getAttribute('data-category');
    var visibleRows = Array.from(document.querySelectorAll('#' + tableId + ' .data-row[data-category="' + CSS.escape(cat) + '"]')).filter(function(r) { return r.style.display !== 'none'; });
    hdr.style.display = visibleRows.length === 0 ? 'none' : '';
    var countSpan = hdr.querySelector('.category-count');
    if (countSpan) { countSpan.textContent = '(' + visibleRows.length + ' items)'; }
  });
}
function filterBADetails() { _filterDetails('baDetailsTable','baDetailsFilter','baDetailsCount','noRecordsBA'); }
function clearBADetailsFilter() { document.getElementById('baDetailsFilter').value = 'All'; filterBADetails(); }
function filterPODevDetails() { _filterDetails('poDevDetailsTable','poDevDetailsFilter','poDevDetailsCount','noRecordsPODev'); }
function clearPODevDetailsFilter() { document.getElementById('poDevDetailsFilter').value = 'All'; filterPODevDetails(); }"""

    html = (
        "<!DOCTYPE html>\n<html>\n<head>\n<meta charset='utf-8'>\n"
        "<title>Missing Data Report for Scopes</title>\n"
        f"<style>\n{css}\n</style>\n</head>\n<body>\n"

        "<div class='header'>\n"
        "<h1>Missing Data Report for Scopes</h1>\n"
        "<div class='header-info'>\n"
        f"<div>Generated on: {generated_at}</div>\n"
        f"<div>Total Categories: {total_categories} | BA Created: {total_ba} | PO/Dev Created: {total_po} | Total Items: {total_items}</div>\n"
        "</div>\n</div>\n\n"

        "<div class='main-tab-nav'>\n"
        "<button class=\"main-tab-button active\" onclick=\"switchMainTab(event, 'summary')\">Summary</button>\n"
        "<button class=\"main-tab-button\" onclick=\"switchMainTab(event, 'details')\">Details</button>\n"
        "</div>\n\n"

        # ---- SUMMARY TAB ----
        "<div id='summary' class='main-tab-content active'>\n"
        "<div class='toggle-container'>\n"
        "<span class='toggle-label'>View:</span>\n"
        "<div class='toggle-switch' id='summaryToggle'>\n"
        "<button class='toggle-option active' data-view='ba' onclick=\"toggleSummaryView('ba')\">BA Created</button>\n"
        "<button class='toggle-option' data-view='podev' onclick=\"toggleSummaryView('podev')\">PO/Dev Team Created</button>\n"
        "</div>\n</div>\n\n"

        # BA Summary
        "<div id='baSummaryContent' class='content-section active'>\n"
        "<div class='summary-header'>BA Summary</div>\n"
        "<div class='filter-container'>\n"
        "<label for='baSummaryFilter' class='filter-label'>Created By:</label>\n"
        f"<select id='baSummaryFilter' class='filter-dropdown' onchange='filterBASummary()'>{ba_drop}</select>\n"
        "<button class='clear-filter-btn' onclick='clearBASummaryFilter()'>Clear Filter</button>\n"
        f"<span class='filter-count'>Showing: <span id='baSummaryCount'>{ba_sum_count}</span> of {ba_sum_count} items</span>\n"
        "</div>\n"
        "<div id='noRecordsBASummary' class='no-records' style='display:none;'>No records found for the selected filter.</div>\n"
        "<div class='table-wrapper'>\n"
        "<table class='summary-table' id='baSummaryTable'>\n"
        "<thead><tr><th>Created By</th><th>Missing Category</th><th class='count-header'>Count</th></tr></thead>\n"
        f"<tbody>{ba_sum_html}</tbody>\n"
        "</table>\n</div>\n</div>\n\n"

        # PO/Dev Summary
        "<div id='poDevSummaryContent' class='content-section'>\n"
        "<div class='summary-header'>PO/Dev Summary</div>\n"
        "<div class='filter-container'>\n"
        "<label for='poDevSummaryFilter' class='filter-label'>Created By:</label>\n"
        f"<select id='poDevSummaryFilter' class='filter-dropdown' onchange='filterPODevSummary()'>{po_drop}</select>\n"
        "<button class='clear-filter-btn' onclick='clearPODevSummaryFilter()'>Clear Filter</button>\n"
        f"<span class='filter-count'>Showing: <span id='poDevSummaryCount'>{po_sum_count}</span> of {po_sum_count} items</span>\n"
        "</div>\n"
        "<div id='noRecordsPODevSummary' class='no-records' style='display:none;'>No records found for the selected filter.</div>\n"
        "<div class='table-wrapper'>\n"
        "<table class='summary-table' id='poDevSummaryTable'>\n"
        "<thead><tr><th>Created By</th><th>Missing Category</th><th class='count-header'>Count</th></tr></thead>\n"
        f"<tbody>{po_sum_html}</tbody>\n"
        "</table>\n</div>\n</div>\n"
        "</div>\n\n"

        # ---- DETAILS TAB ----
        "<div id='details' class='main-tab-content'>\n"
        "<div class='toggle-container'>\n"
        "<span class='toggle-label'>View:</span>\n"
        "<div class='toggle-switch' id='detailsToggle'>\n"
        "<button class='toggle-option active' data-view='ba' onclick=\"toggleDetailsView('ba')\">BA Created</button>\n"
        "<button class='toggle-option' data-view='podev' onclick=\"toggleDetailsView('podev')\">PO/Dev Team Created</button>\n"
        "</div>\n</div>\n\n"

        # BA Details
        "<div id='baDetailsContent' class='content-section active'>\n"
        "<h2>Details - BA Created</h2>\n"
        "<div class='filter-container'>\n"
        "<label for='baDetailsFilter' class='filter-label'>Created By:</label>\n"
        f"<select id='baDetailsFilter' class='filter-dropdown' onchange='filterBADetails()'>{ba_drop}</select>\n"
        "<button class='clear-filter-btn' onclick='clearBADetailsFilter()'>Clear Filter</button>\n"
        f"<span class='filter-count'>Showing: <span id='baDetailsCount'>{total_ba}</span> of {total_ba} items</span>\n"
        "</div>\n"
        "<div id='noRecordsBA' class='no-records' style='display:none;'>No records found for the selected filter.</div>\n"
        "<table id='baDetailsTable'>\n"
        "<thead><tr><th>US ID</th><th>Title</th><th>Created By</th><th>Supervisor</th><th>Link</th></tr></thead>\n"
        f"<tbody>{ba_det_html}</tbody>\n"
        "</table>\n</div>\n\n"

        # PO/Dev Details
        "<div id='poDevDetailsContent' class='content-section'>\n"
        "<h2>Details - PO/Dev Team Created</h2>\n"
        "<div class='filter-container'>\n"
        "<label for='poDevDetailsFilter' class='filter-label'>Created By:</label>\n"
        f"<select id='poDevDetailsFilter' class='filter-dropdown' onchange='filterPODevDetails()'>{po_drop}</select>\n"
        "<button class='clear-filter-btn' onclick='clearPODevDetailsFilter()'>Clear Filter</button>\n"
        f"<span class='filter-count'>Showing: <span id='poDevDetailsCount'>{total_po}</span> of {total_po} items</span>\n"
        "</div>\n"
        "<div id='noRecordsPODev' class='no-records' style='display:none;'>No records found for the selected filter.</div>\n"
        "<table id='poDevDetailsTable'>\n"
        "<thead><tr><th>US ID</th><th>Title</th><th>Created By</th><th>Supervisor</th><th>Link</th></tr></thead>\n"
        f"<tbody>{po_det_html}</tbody>\n"
        "</table>\n</div>\n"
        "</div>\n\n"

        f"<script>{js}\n</script>\n"
        "</body>\n</html>"
    )

    with open(file_path, "w", encoding="utf-8") as f:
        f.write(html)
    return file_path

ORG = "accenturecio08"
PROJECT = "AutomationProcess_29697"
CREATED_QUERY_ID = "730dc08e-2b34-4d8f-a74e-6b7c74a05071"
CLOSED_QUERY_ID = os.environ.get("ADO_CLOSED_QUERY_ID", "").strip()

MISSING_FIELD_NAMES = [
    ##"Defect Record",
    "mySP Initiative",
    "Sub Initiative",
    "TextVerification",
    ##"TextVerification1",
    "Parent",
    "Requirement Requestor",
    ##"Broken",
    ##"Identified in mySP Release Name",
]

CLOSED_MISSING_FIELD_NAMES = MISSING_FIELD_NAMES.copy()



OUTPUT_PATH = r"C:\Users\vishnu.ramalingam\MyISP_Tools\Missing_Data_Report_For_Scope"
PT_LEAD_MAPPING_PATH = r"C:\Users\vishnu.ramalingam\MyISP_Tools\Missing_Data_Report_For_Scope\BA_Team_Names 1.xlsx"
HTML_OUTPUT_DIR = os.path.join(os.path.dirname(OUTPUT_PATH), "category_reports")


def get_ssl_verify_setting() -> bool | str:
    verify_ssl = os.environ.get("ADO_VERIFY_SSL", "true").strip().lower()
    if verify_ssl in {"0", "false", "no", "off"}:
        return False

    explicit_bundle = (
        os.environ.get("ADO_CA_BUNDLE")
        or os.environ.get("REQUESTS_CA_BUNDLE")
        or os.environ.get("CURL_CA_BUNDLE")
    )
    if explicit_bundle:
        explicit_bundle = explicit_bundle.strip()
        if os.path.isfile(explicit_bundle):
            return explicit_bundle
        if os.path.isdir(explicit_bundle):
            print(
                f"Ignoring CA directory '{explicit_bundle}' and using certifi bundle to avoid SSL context load issues."
            )
            return certifi.where()
        print(
            f"CA bundle path '{explicit_bundle}' is invalid. Falling back to certifi bundle."
        )

    return certifi.where()


SSL_VERIFY = get_ssl_verify_setting()

def create_session_with_retries() -> requests.Session:
    """Create a session with retry logic and connection pooling."""
    session = requests.Session()
    
    # Configure retry strategy
    retry_strategy = Retry(
        total=3,  # Total number of retries
        backoff_factor=1,  # Wait 1, 2, 4 seconds between retries
        status_forcelist=[429, 500, 502, 503, 504],  # HTTP status codes to retry on
        allowed_methods=["GET", "POST"]  # HTTP methods to retry
    )
    
    adapter = HTTPAdapter(
        max_retries=retry_strategy,
        pool_connections=10,  # Number of connection pools
        pool_maxsize=10  # Max number of connections in pool
    )
    
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    return session

def build_headers(pat: str) -> Dict[str, str]:
    token = base64.b64encode(f":{pat}".encode("ascii")).decode("ascii")
    return {"Authorization": f"Basic {token}"}


def get_query_work_item_ids(base_uri: str, headers: Dict[str, str], query_id: str, session: requests.Session) -> List[int]:
    url = f"{base_uri}/_apis/wit/wiql/{query_id}?api-version=7.1"
    print(f"Fetching work item IDs from query {query_id}...")
    try:
        resp = session.get(url, headers=headers, timeout=30, verify=SSL_VERIFY)
        resp.raise_for_status()
        data = resp.json()
        work_items = [item["id"] for item in data.get("workItems", [])]
        print(f"Found {len(work_items)} work items.")
        return work_items
    except requests.exceptions.Timeout:
        print(f"ERROR: Request timed out while fetching query {query_id}")
        raise
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Failed to fetch work items from query {query_id}: {e}")
        raise


def get_field_reference_map(base_uri: str, headers: Dict[str, str], session: requests.Session) -> Dict[str, str]:
    url = f"{base_uri}/_apis/wit/fields?api-version=7.1"
    print("Fetching field reference map...")
    try:
        resp = session.get(url, headers=headers, timeout=30, verify=SSL_VERIFY)
        resp.raise_for_status()
        fields = resp.json().get("value", [])
        name_to_ref = {}
        for field in fields:
            name = field.get("name")
            ref = field.get("referenceName")
            if name and ref:
                name_to_ref[name.lower()] = ref
        print(f"Loaded {len(name_to_ref)} field mappings.")
        return name_to_ref
    except requests.exceptions.Timeout:
        print("ERROR: Request timed out while fetching field reference map")
        raise
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Failed to fetch field reference map: {e}")
        raise


def chunk_list(items: List[int], size: int) -> List[List[int]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def is_missing_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def normalize_text(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"\s+", " ", value)
    return value


def normalize_name(value: str) -> str:
    value = value.lower()
    value = re.sub(r"\[[^\]]*\]", " ", value)
    value = re.sub(r"[^a-z,\s]", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ,")


def extract_email(value: str) -> str:
    match = re.search(r"([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})", value or "")
    return match.group(1).lower() if match else ""


def build_supervisor_mapping(mapping_path: str) -> tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    if not os.path.exists(mapping_path):
        return {}, {}, {}

    workbook = load_workbook(mapping_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in worksheet[1]]

    created_by_idx = None
    supervisor_idx = None
    for idx, header in enumerate(headers, start=1):
        if header in {"created by", "ba team"}:
            created_by_idx = idx
        if header == "supervisor":
            supervisor_idx = idx

    if not created_by_idx or not supervisor_idx:
        return {}, {}, {}

    email_map: Dict[str, str] = {}
    full_map: Dict[str, str] = {}
    name_buckets: Dict[str, set[str]] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        created_by_raw = row[created_by_idx - 1] if created_by_idx - 1 < len(row) else None
        supervisor_raw = row[supervisor_idx - 1] if supervisor_idx - 1 < len(row) else None
        if created_by_raw is None:
            continue

        created_by_value = str(created_by_raw).strip()
        supervisor_value = str(supervisor_raw).strip() if supervisor_raw is not None else ""
        if not created_by_value:
            continue

        email = extract_email(created_by_value)
        if email and supervisor_value:
            email_map[email] = supervisor_value

        full_key = normalize_text(created_by_value)
        if full_key and supervisor_value:
            full_map[full_key] = supervisor_value

        name_part = created_by_value.split("@", 1)[0]
        name_key = normalize_name(name_part)
        if name_key and supervisor_value:
            name_buckets.setdefault(name_key, set()).add(supervisor_value)

    unique_name_map: Dict[str, str] = {}
    for name_key, supervisors in name_buckets.items():
        if len(supervisors) == 1:
            unique_name_map[name_key] = next(iter(supervisors))

    return email_map, full_map, unique_name_map


def resolve_supervisor(
    created_by_value: str,
    email_map: Dict[str, str],
    full_map: Dict[str, str],
    unique_name_map: Dict[str, str],
) -> str:
    value = created_by_value or ""

    # Try 1: Email extraction
    email = extract_email(value)
    if email and email in email_map:
        return email_map[email]

    # Try 2: Full text match
    full_key = normalize_text(value)
    if full_key and full_key in full_map:
        return full_map[full_key]

    # Try 3: Name-only match
    no_prefix = re.sub(r"\[[^\]]*\]", " ", value)
    no_prefix = re.sub(r"\b[A-Z]{2,}\s*-\s*", " ", no_prefix)
    name_key = normalize_name(no_prefix.split("@", 1)[0])
    if name_key and name_key in unique_name_map:
        return unique_name_map[name_key]

    # Try 4: Handle "LastName, FirstName" format by searching for partial surname match
    # Example: "Infancia Sermugam, Angel" should match "Infancia Sermugam, a.infancia.sermugam"
    if "," in value and "@" not in value:
        # Extract surname (part before comma)
        surname_part = value.split(",")[0].strip().lower()
        
        # Search in full_map for entries that start with this surname
        for full_key_candidate, supervisor in full_map.items():
            if full_key_candidate.startswith(surname_part + ","):
                return supervisor
        
        # Search in unique_name_map for entries that start with this surname
        for name_key_candidate, supervisor in unique_name_map.items():
            if name_key_candidate.startswith(surname_part):
                return supervisor

    return "Not Available"


def build_summary_and_details(
    all_items: List[dict],
    missing_ref_names: Dict[str, str],
    missing_field_names: List[str],
    pt_lead_ref: str,
    email_map: Dict[str, str],
    full_map: Dict[str, str],
    unique_name_map: Dict[str, str],
) -> tuple[List[dict], List[dict], dict]:
    rows = []
    for item in all_items:
        fields_data = item.get("fields", {})
        created_by = fields_data.get("System.CreatedBy", {})
        created_by_unique = created_by.get("uniqueName", "") if isinstance(created_by, dict) else ""
        created_by_display = created_by.get("displayName", "") if isinstance(created_by, dict) else ""

        # Get PT Lead directly from Azure DevOps field
        pt_lead = fields_data.get(pt_lead_ref, "")
        if not pt_lead or (isinstance(pt_lead, str) and not pt_lead.strip()):
            pt_lead = "Unmapped"
        elif isinstance(pt_lead, str):
            pt_lead = pt_lead.strip()

        row = {
            "Id": item.get("id"),
            "Title": fields_data.get("System.Title", ""),
            "PT lead": pt_lead,
            "Created By": created_by_display or created_by_unique,
            "Supervisor": resolve_supervisor(
                created_by_display or created_by_unique,
                email_map,
                full_map,
                unique_name_map,
            ),
        }
        for name, ref in missing_ref_names.items():
            row[name] = fields_data.get(ref)
        rows.append(row)

    summary_rows = []
    detail_rows = []
    group_ranges = {}
    detail_start_row = 2

    for name in missing_field_names:
        missing_rows = [row for row in rows if is_missing_value(row.get(name))]
        count = len(missing_rows)
        summary_rows.append(
            {
                "Missing Field": name,
                "Count of defect Title": count,
            }
        )

        if count > 0:
            start_row = detail_start_row
            for r in missing_rows:
                detail_rows.append(
                    {
                        "Missing": name,
                        "US ID": r.get("Id"),
                        "Title": r.get("Title"),
                        "Created By": r.get("Created By"),
                        "Supervisor": r.get("Supervisor"),
                    }
                )
                detail_start_row += 1
            end_row = detail_start_row - 1
            group_ranges[name] = (start_row, end_row)

    return summary_rows, detail_rows, group_ranges


def resolve_missing_ref_names(
    missing_field_names: List[str],
    name_to_ref: Dict[str, str],
) -> Dict[str, str]:
    missing_ref_names = {}
    for name in missing_field_names:
        ref = name_to_ref.get(name.lower())
        if ref:
            missing_ref_names[name] = ref
        else:
            missing_ref_names[name] = None
    return missing_ref_names


def fetch_items_for_query(
    base_uri: str,
    headers: Dict[str, str],
    query_id: str,
    fields: List[str],
    session: requests.Session,
) -> List[dict]:
    work_item_ids = get_query_work_item_ids(base_uri, headers, query_id, session)
    if not work_item_ids:
        return []

    all_items = []
    batches = chunk_list(work_item_ids, 200)
    print(f"Fetching work item details in {len(batches)} batches...")
    
    for i, batch in enumerate(batches, 1):
        try:
            ids_str = ",".join(str(item_id) for item_id in batch)
            fields_str = ",".join(fields)
            url = f"{base_uri}/_apis/wit/workitems?ids={ids_str}&fields={fields_str}&api-version=7.1"
            print(f"  Batch {i}/{len(batches)}: Fetching {len(batch)} items...")
            resp = session.get(url, headers=headers, timeout=30, verify=SSL_VERIFY)
            resp.raise_for_status()
            batch_items = resp.json().get("value", [])
            all_items.extend(batch_items)
            print(f"  Batch {i}/{len(batches)}: Fetched {len(batch_items)} items successfully.")
        except requests.exceptions.Timeout:
            print(f"  ERROR: Batch {i}/{len(batches)} timed out. Skipping batch.")
            continue
        except requests.exceptions.RequestException as e:
            print(f"  ERROR: Batch {i}/{len(batches)} failed: {e}. Skipping batch.")
            continue
    
    print(f"Total items fetched: {len(all_items)}")
    return all_items


def write_html_report_grouped_by_missing(
    output_path: str,
    detail_rows: List[dict],
) -> str:
    """Generate enhanced HTML report with two main tabs (Summary & Details) and toggle for BA/PO-Dev."""
    return generate_enhanced_html_report(output_path, detail_rows, ORG, PROJECT)


def sanitize_category_filename(category: str) -> str:
    value = (category or "Unknown").strip()
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^A-Za-z0-9_-]", "", value)
    return value or "Unknown"


def write_category_html_reports(detail_rows: List[dict], output_dir: str) -> None:
    os.makedirs(output_dir, exist_ok=True)

    grouped_rows: Dict[str, List[dict]] = {}
    po_dev_team_records: Dict[str, List[dict]] = {}  # Records with Supervisor = "Not Available"
    
    for row in detail_rows:
        category = str(row.get("Missing") or row.get("Missing Field") or "Unknown").strip() or "Unknown"
        supervisor = str(row.get("Supervisor", "")).strip().lower()
        
        if supervisor == "not available":
            # Add to PO/Dev Team Created report
            po_dev_team_records.setdefault(category, []).append(row)
        else:
            # Add to regular category report
            grouped_rows.setdefault(category, []).append(row)

    # Generate regular category reports (excluding Supervisor = "Not Available")
    for category, rows in grouped_rows.items():
        sorted_rows = sorted(rows, key=lambda item: str(item.get("US ID", "")))
        safe_category = escape(category)
        body_rows = []

        for item in sorted_rows:
            supervisor = str(item.get("Supervisor", ""))
            supervisor_class = " class=\"na\"" if supervisor.strip().lower() == "not available" else ""
            us_id = str(item.get('US ID', ''))
            
            # Generate Azure DevOps work item link
            if us_id and us_id.strip():
                ado_link = f"https://dev.azure.com/{ORG}/{PROJECT}/_workitems/edit/{us_id}"
                link_button = f'<a href="{ado_link}" target="_blank" class="open-link-btn">OPEN</a>'
            else:
                link_button = '<span class="na-link">N/A</span>'
            
            body_rows.append(
                "<tr>"
                f"<td>{escape(str(item.get('Missing') or item.get('Missing Field') or ''))}</td>"
                f"<td>{escape(us_id)}</td>"
                f"<td>{escape(str(item.get('Title', '')))}</td>"
                f"<td>{escape(str(item.get('Created By', '')))}</td>"
                f"<td{supervisor_class}>{escape(supervisor)}</td>"
                f"<td class='link-cell'>{link_button}</td>"
                "</tr>"
            )

        html_content = (
            "<html>"
            "<head>"
            f"<title>Category Report - {safe_category}</title>"
            "<style>"
            "table { border-collapse: collapse; width: 100%; }"
            "th, td { border: 1px solid black; padding: 8px; text-align: left; }"
            "th { background-color: #f2f2f2; }"
            ".na { background-color: #fff8dc; }"
            ".link-cell { text-align: center; }"
            ".open-link-btn { display: inline-block; padding: 6px 12px; background-color: #3498db; color: white; text-decoration: none; "
            "border-radius: 4px; font-size: 12px; font-weight: 600; transition: all 0.3s ease; }"
            ".open-link-btn:hover { background-color: #2980b9; transform: translateY(-1px); box-shadow: 0 3px 6px rgba(0,0,0,0.2); }"
            ".na-link { color: #95a5a6; font-style: italic; font-size: 12px; }"
            "</style>"
            "</head>"
            "<body>"
            f"<h2>Category: {safe_category}</h2>"
            f"<p>Total rows: {len(sorted_rows)}</p>"
            "<table>"
            "<tr>"
            "<th>Missing</th>"
            "<th>US ID</th>"
            "<th>Title</th>"
            "<th>Created By</th>"
            "<th>Supervisor</th>"
            "<th>Link</th>"
            "</tr>"
            + "".join(body_rows)
            + "</table>"
            "</body>"
            "</html>"
        )

        file_name = f"{sanitize_category_filename(category)}_Report.html"
        file_path = os.path.join(output_dir, file_name)
        with open(file_path, "w", encoding="utf-8") as html_file:
            html_file.write(html_content)
    
    # Generate PO/Dev Team Created report
    if po_dev_team_records:
        po_dev_total = sum(len(rows) for rows in po_dev_team_records.values())
        
        # Extract unique "Created By" values for dropdown
        created_by_set = set()
        for rows in po_dev_team_records.values():
            for item in rows:
                created_by = str(item.get('Created By', '')).strip()
                if created_by:
                    created_by_set.add(created_by)
        
        # Sort Created By values alphabetically
        sorted_created_by = sorted(created_by_set)
        
        # Sort categories alphabetically and prepare tab data
        sorted_categories = sorted(po_dev_team_records.keys())
        
        # Generate tab buttons
        tab_buttons = []
        for idx, category in enumerate(sorted_categories):
            rows = po_dev_team_records[category]
            active_class = " active" if idx == 0 else ""
            safe_category_id = sanitize_category_filename(category)
            tab_buttons.append(
                f'<button class="tab-button{active_class}" '
                f'onclick="openTab(event, \'{safe_category_id}\')" '
                f'data-category="{escape(category)}">'
                f'{escape(category)} <span class="tab-count" id="count-{safe_category_id}">({len(rows)})</span>'
                f'</button>'
            )
        
        # Generate tab content sections
        tab_contents = []
        for idx, category in enumerate(sorted_categories):
            rows = po_dev_team_records[category]
            sorted_rows = sorted(rows, key=lambda item: str(item.get("US ID", "")))
            safe_category_id = sanitize_category_filename(category)
            display_style = "block" if idx == 0 else "none"
            
            # Build rows for this category
            category_rows = []
            for item in sorted_rows:
                supervisor = str(item.get("Supervisor", ""))
                supervisor_class = " class=\"na\"" if supervisor.strip().lower() == "not available" else ""
                us_id = str(item.get('US ID', ''))
                created_by = str(item.get('Created By', ''))
                
                # Generate Azure DevOps work item link
                if us_id and us_id.strip():
                    ado_link = f"https://dev.azure.com/{ORG}/{PROJECT}/_workitems/edit/{us_id}"
                    link_button = f'<a href="{ado_link}" target="_blank" class="open-link-btn">OPEN</a>'
                else:
                    link_button = '<span class="na-link">N/A</span>'
                
                category_rows.append(
                    f"<tr class='data-row' data-created-by='{escape(created_by)}' data-category='{escape(category)}'>"
                    f"<td>{escape(str(item.get('Missing') or item.get('Missing Field') or ''))}</td>"
                    f"<td>{escape(us_id)}</td>"
                    f"<td>{escape(str(item.get('Title', '')))}</td>"
                    f"<td>{escape(created_by)}</td>"
                    f"<td{supervisor_class}>{escape(supervisor)}</td>"
                    f"<td class='link-cell'>{link_button}</td>"
                    "</tr>"
                )
            
            # Create tab content
            tab_contents.append(
                f'<div id="{safe_category_id}" class="tab-content" style="display:{display_style};" '
                f'data-category="{escape(category)}">'
                f'<h3>{escape(category)}</h3>'
                f'<div id="noRecords-{safe_category_id}" class="no-records" style="display:none;">'
                f'No records found for the selected filter.</div>'
                f'<table class="category-table">'
                f'<tr>'
                f'<th>Original Category</th>'
                f'<th>US ID</th>'
                f'<th>Title</th>'
                f'<th>Created By</th>'
                f'<th>Supervisor</th>'
                f'<th>Link</th>'
                f'</tr>'
                + "".join(category_rows)
                + '</table>'
                + '</div>'
            )
        
        # Generate dropdown options
        dropdown_options = '<option value="All">All</option>'
        for created_by in sorted_created_by:
            dropdown_options += f'<option value="{escape(created_by)}">{escape(created_by)}</option>'
        
        html_content = (
            "<html>"
            "<head>"
            "<title>PO/Dev Team Created Report</title>"
            "<style>"
            "body { font-family: Arial, sans-serif; padding: 20px; background-color: #f5f7fa; }"
            "h2 { color: #2c3e50; margin-bottom: 10px; }"
            "h3 { color: #34495e; margin-bottom: 15px; font-size: 18px; }"
            "table { border-collapse: collapse; width: 100%; margin-top: 20px; background-color: white; "
            "box-shadow: 0 1px 3px rgba(0,0,0,0.1); }"
            "th, td { border: 1px solid #dce1e8; padding: 12px 8px; text-align: left; }"
            "th { background-color: #3498db; color: white; font-weight: 600; position: sticky; top: 0; }"
            ".na { background-color: #ffe6e6; color: #c0392b; font-weight: bold; }"
            ".link-cell { text-align: center; }"
            ".open-link-btn { display: inline-block; padding: 6px 12px; background-color: #3498db; color: white; "
            "text-decoration: none; border-radius: 4px; font-size: 12px; font-weight: 600; transition: all 0.3s ease; }"
            ".open-link-btn:hover { background-color: #2980b9; transform: translateY(-1px); box-shadow: 0 3px 6px rgba(0,0,0,0.2); }"
            ".na-link { color: #95a5a6; font-style: italic; font-size: 12px; }"
            ".filter-container { margin: 20px 0; padding: 15px; background-color: white; border-radius: 8px; "
            "box-shadow: 0 2px 4px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 15px; }"
            ".filter-label { font-weight: 600; font-size: 14px; color: #2c3e50; }"
            ".filter-dropdown { padding: 8px 12px; font-size: 14px; border: 2px solid #3498db; border-radius: 6px; "
            "background-color: white; cursor: pointer; min-width: 250px; transition: all 0.3s ease; }"
            ".filter-dropdown:hover { border-color: #2980b9; }"
            ".filter-dropdown:focus { outline: none; border-color: #2980b9; box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1); }"
            ".clear-filter-btn { padding: 8px 16px; background-color: #95a5a6; color: white; border: none; "
            "border-radius: 6px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.3s ease; }"
            ".clear-filter-btn:hover { background-color: #7f8c8d; }"
            ".total-count { font-size: 16px; color: #34495e; margin: 10px 0; padding: 10px 15px; "
            "background-color: white; border-radius: 6px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }"
            ".no-records { text-align: center; padding: 30px; font-size: 16px; color: #7f8c8d; "
            "background-color: #f8f9fa; border-radius: 8px; margin: 20px 0; }"
            ".data-row { transition: background-color 0.2s ease; }"
            ".data-row:hover { background-color: #f0f8ff; }"
            ""
            "/* Tab Styles */"
            ".tab-container { margin: 20px 0; background-color: white; border-radius: 8px; "
            "box-shadow: 0 2px 4px rgba(0,0,0,0.1); overflow: hidden; }"
            ".tab-buttons { display: flex; flex-wrap: wrap; background-color: #ecf0f1; border-bottom: 3px solid #3498db; "
            "padding: 0; margin: 0; gap: 0; overflow-x: auto; }"
            ".tab-button { background-color: #ecf0f1; color: #34495e; border: none; padding: 14px 20px; "
            "cursor: pointer; font-size: 14px; font-weight: 600; transition: all 0.3s ease; "
            "border-right: 1px solid #bdc3c7; white-space: nowrap; flex-shrink: 0; }"
            ".tab-button:hover { background-color: #d5dbdb; }"
            ".tab-button.active { background-color: #3498db; color: white; border-bottom: 3px solid #2980b9; "
            "margin-bottom: -3px; }"
            ".tab-count { font-weight: normal; opacity: 0.9; }"
            ".tab-content { display: none; padding: 20px; animation: fadeIn 0.3s ease; }"
            ".tab-content.active { display: block; }"
            "@keyframes fadeIn { from { opacity: 0; transform: translateY(-10px); } to { opacity: 1; transform: translateY(0); } }"
            ""
            "@media (max-width: 768px) {"
            "  body { padding: 10px; }"
            "  .filter-container { flex-direction: column; align-items: flex-start; }"
            "  .filter-dropdown { width: 100%; min-width: 100%; }"
            "  table { font-size: 12px; }"
            "  th, td { padding: 6px 4px; }"
            "  .tab-buttons { flex-direction: column; }"
            "  .tab-button { border-right: none; border-bottom: 1px solid #bdc3c7; }"
            "}"
            "</style>"
            "</head>"
            "<body>"
            "<h2>PO/Dev Team Created</h2>"
            "<div class='filter-container'>"
            "<label for='createdByFilter' class='filter-label'>Created By:</label>"
            f"<select id='createdByFilter' class='filter-dropdown' onchange='filterByCreatedBy()'>{dropdown_options}</select>"
            "<button class='clear-filter-btn' onclick='clearFilter()'>Clear Filter</button>"
            "</div>"
            f"<p class='total-count'>Total: <span id='totalCount'>{po_dev_total}</span> items</p>"
            ""
            "<div class='tab-container'>"
            "<div class='tab-buttons'>"
            + "".join(tab_buttons)
            + "</div>"
            + "".join(tab_contents)
            + "</div>"
            ""
            "<script>"
            "// Tab switching functionality"
            "function openTab(evt, tabId) {"
            "  const tabContents = document.getElementsByClassName('tab-content');"
            "  for (let i = 0; i < tabContents.length; i++) {"
            "    tabContents[i].style.display = 'none';"
            "  }"
            "  "
            "  const tabButtons = document.getElementsByClassName('tab-button');"
            "  for (let i = 0; i < tabButtons.length; i++) {"
            "    tabButtons[i].classList.remove('active');"
            "  }"
            "  "
            "  document.getElementById(tabId).style.display = 'block';"
            "  evt.currentTarget.classList.add('active');"
            "}"
            ""
            "// Filter by Created By functionality"
            "function filterByCreatedBy() {"
            "  const filterValue = document.getElementById('createdByFilter').value;"
            "  const tabContents = document.getElementsByClassName('tab-content');"
            "  let totalVisibleCount = 0;"
            "  "
            "  // Iterate through each tab"
            "  for (let tabContent of tabContents) {"
            "    const category = tabContent.getAttribute('data-category');"
            "    const table = tabContent.querySelector('.category-table');"
            "    const rows = table.getElementsByTagName('tr');"
            "    const noRecordsMsg = tabContent.querySelector('.no-records');"
            "    const tabId = tabContent.id;"
            "    let visibleCount = 0;"
            "    "
            "    // Filter rows in this tab"
            "    for (let i = 1; i < rows.length; i++) {"
            "      const row = rows[i];"
            "      if (row.classList.contains('data-row')) {"
            "        const createdBy = row.getAttribute('data-created-by');"
            "        "
            "        if (filterValue === 'All' || createdBy === filterValue) {"
            "          row.style.display = '';"
            "          visibleCount++;"
            "        } else {"
            "          row.style.display = 'none';"
            "        }"
            "      }"
            "    }"
            "    "
            "    totalVisibleCount += visibleCount;"
            "    "
            "    // Update tab count"
            "    const countSpan = document.getElementById('count-' + tabId);"
            "    if (countSpan) {"
            "      countSpan.textContent = '(' + visibleCount + ')';"
            "    }"
            "    "
            "    // Show/hide no records message"
            "    if (visibleCount === 0) {"
            "      noRecordsMsg.style.display = 'block';"
            "      table.style.display = 'none';"
            "    } else {"
            "      noRecordsMsg.style.display = 'none';"
            "      table.style.display = 'table';"
            "    }"
            "  }"
            "  "
            "  // Update total count"
            "  document.getElementById('totalCount').textContent = totalVisibleCount;"
            "}"
            ""
            "function clearFilter() {"
            "  document.getElementById('createdByFilter').value = 'All';"
            "  filterByCreatedBy();"
            "}"
            "</script>"
            "</body>"
            "</html>"
        )
        
        file_path = os.path.join(output_dir, "PODevTeamCreated_Report.html")
        with open(file_path, "w", encoding="utf-8") as html_file:
            html_file.write(html_content)


def write_category_index_report(output_dir: str) -> None:
    report_files = [
        file_name
        for file_name in os.listdir(output_dir)
        if file_name.endswith("_Report.html")
    ]

    report_entries = []
    for file_name in report_files:
        display_name = file_name.replace("_Report.html", "").replace("_", " ")
        report_path = os.path.join(output_dir, file_name)
        row_count = 0
        with open(report_path, "r", encoding="utf-8") as report_file:
            report_content = report_file.read()
        row_count_match = re.search(r"<p>Total rows:\s*(\d+)</p>", report_content)
        if row_count_match:
            row_count = int(row_count_match.group(1))
        report_entries.append((file_name, display_name, row_count))

    report_entries.sort(key=lambda entry: entry[2], reverse=True)

    links = [
        f'<li><a href="{escape(file_name)}">{escape(display_name)}</a> ({row_count})</li>'
        for file_name, display_name, row_count in report_entries
    ]

    html_content = (
        "<html>"
        "<head>"
        "<title>Category Reports Index</title>"
        "<style>"
        "body { font-family: Arial, sans-serif; }"
        "li { margin: 6px 0; }"
        "</style>"
        "</head>"
        "<body>"
        "<h2>Category Reports Index</h2>"
        f"<p>Total reports: {len(report_entries)}</p>"
        "<ul>"
        + "".join(links)
        + "</ul>"
        "</body>"
        "</html>"
    )

    index_path = os.path.join(output_dir, "index.html")
    with open(index_path, "w", encoding="utf-8") as index_file:
        index_file.write(html_content)


def main() -> None:
    from dotenv import load_dotenv
    load_dotenv('ADO_SECRETS.env')
    pat = os.getenv("ADO_PAT_MAIN", "")
    if not pat:
        raise SystemExit("ADO_PAT_MAIN not found in ADO_SECRETS.env file. Please set your Azure DevOps PAT token.")

    base_uri = f"https://dev.azure.com/{ORG}/{PROJECT}"
    headers = build_headers(pat)
    
    # Create session with retry logic
    print("Creating session with retry logic...")
    session = create_session_with_retries()
    
    print("Loading supervisor mapping...")
    email_map, full_map, unique_name_map = build_supervisor_mapping(PT_LEAD_MAPPING_PATH)

    name_to_ref = get_field_reference_map(base_uri, headers, session)
    missing_ref_names = resolve_missing_ref_names(MISSING_FIELD_NAMES, name_to_ref)
    closed_missing_ref_names = (
        resolve_missing_ref_names(CLOSED_MISSING_FIELD_NAMES, name_to_ref)
        if CLOSED_QUERY_ID
        else {}
    )

    # Get PT Lead field reference - try multiple variations
    pt_lead_ref = name_to_ref.get("owner") or name_to_ref.get("pt lead") or name_to_ref.get("ptlead") or name_to_ref.get("pt-lead") or name_to_ref.get("pt lead name")
    if not pt_lead_ref:
        # List all available fields to help debug
        print("All available custom fields:")
        for field_name, field_ref in sorted(name_to_ref.items()):
            if field_ref.startswith("Custom.") or "PT" in field_ref or "Lead" in field_ref or "owner" in field_name.lower():
                print(f"  {field_name} -> {field_ref}")
        raise SystemExit("\nPT Lead field was not found in Azure DevOps fields list. Please specify the exact field name.")

    missing_not_found = [k for k, v in missing_ref_names.items() if v is None]
    closed_missing_not_found = [
        k for k, v in closed_missing_ref_names.items() if v is None
    ]
    if missing_not_found:
        raise SystemExit(
            "These fields were not found in Azure DevOps fields list: "
            + ", ".join(missing_not_found)
        )
    if closed_missing_not_found:
        raise SystemExit(
            "These fields were not found in Azure DevOps fields list: "
            + ", ".join(closed_missing_not_found)
        )

    all_missing_refs = set(missing_ref_names.values()) | set(closed_missing_ref_names.values())
    fields = ["System.Id", "System.Title", "System.CreatedBy", pt_lead_ref] + sorted(
        ref for ref in all_missing_refs if ref
    )

    # Allow runtime override of the query ID (e.g. passed from the web UI)
    runtime_query_id = os.environ.get("ADO_QUERY_ID", "").strip()
    active_query_id = runtime_query_id if runtime_query_id else CREATED_QUERY_ID
    print(f"\nUsing Query ID: {active_query_id}")

    print("\nFetching CREATED query items...")
    created_items = fetch_items_for_query(base_uri, headers, active_query_id, fields, session)
    
    closed_items = []
    if CLOSED_QUERY_ID:
        print("\nFetching CLOSED query items...")
        closed_items = fetch_items_for_query(base_uri, headers, CLOSED_QUERY_ID, fields, session)

    # Close the session
    session.close()

    if not created_items and not closed_items:
        raise SystemExit("No work items found for the provided queries.")

    combined_summary_rows = []
    combined_detail_rows = []
    combined_group_ranges = {}

    if created_items:
        print("\nProcessing CREATED items...")
        summary_rows, detail_rows, group_ranges = build_summary_and_details(
            created_items,
            missing_ref_names,
            MISSING_FIELD_NAMES,
            pt_lead_ref,
            email_map,
            full_map,
            unique_name_map,
        )
        combined_summary_rows.extend(summary_rows)
        combined_detail_rows.extend(detail_rows)
        combined_group_ranges.update(group_ranges)
        
        # Log unmapped users for debugging
        unmapped_users = set()
        for row in detail_rows:
            if row.get("Supervisor") == "Not Available":
                unmapped_users.add(row.get("Created By", "Unknown"))
        
        if unmapped_users:
            print(f"\n⚠ Warning: {len(unmapped_users)} users could not be mapped to BA team:")
            for user in sorted(unmapped_users):
                print(f"  - {user}")
            print("  These items will be classified as PO/Dev Team Created.")
            print("  To fix: Add these names to BA_Team_Names.xlsx with their supervisors.")

    if closed_items:
        print("Processing CLOSED items...")
        closed_summary_rows, closed_detail_rows, closed_group_ranges = build_summary_and_details(
            closed_items,
            closed_missing_ref_names,
            CLOSED_MISSING_FIELD_NAMES,
            pt_lead_ref,
            email_map,
            full_map,
            unique_name_map,
        )
        combined_summary_rows.extend(closed_summary_rows)

        # Adjust row numbers for closed items detail rows
        offset = len(combined_detail_rows)
        adjusted_closed_group_ranges = {}
        for key, (start, end) in closed_group_ranges.items():
            adjusted_closed_group_ranges[key] = (start + offset, end + offset)

        combined_detail_rows.extend(closed_detail_rows)
        combined_group_ranges.update(adjusted_closed_group_ranges)

    if combined_detail_rows:
        print("\nGenerating HTML report...")
        saved_output_path = write_html_report_grouped_by_missing(
            OUTPUT_PATH,
            combined_detail_rows,
        )
        print("Writing additional HTML category reports...")
        write_category_html_reports(combined_detail_rows, HTML_OUTPUT_DIR)
        write_category_index_report(HTML_OUTPUT_DIR)
        print(f"\n✓ Reports generated successfully!")
        print(f"\nMain Report: {saved_output_path}")
        print(f"Category Reports: {os.path.join(HTML_OUTPUT_DIR, 'index.html')}")
    else:
        print("No data to generate reports.")
        saved_output_path = OUTPUT_PATH


if __name__ == "__main__":
    main()

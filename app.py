"""
MyISP Internal Tools - Flask Server
Handles report generation and serves static files
"""
from flask import Flask, request, jsonify, send_from_directory, redirect, session, abort, send_file
import subprocess
import sys
import os
import glob
import calendar
import re
import webbrowser
from pathlib import Path
from datetime import datetime
import threading
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from functools import wraps
import secrets
from werkzeug.utils import secure_filename

# Database integration – Local PostgreSQL
try:
    from postgres_client import postgres as _supabase
    SUPABASE_ENABLED = True
    print("✓ PostgreSQL database connected")
except Exception as _sb_err:
    _supabase = None
    SUPABASE_ENABLED = False
    print(f"⚠️  Database disabled (PostgreSQL connection failed): {_sb_err}")

# Cell background colors matching the HTML tracker
STATUS_COLORS = {
    'P':  'C8F7C5',  # soft mint green
    'VG': 'B3DDF2',  # soft sky blue
    'VR': 'D4EBF2',  # soft powder blue
    'C':  '9EE5DB',  # soft turquoise
    'SL': 'FFB3B3',  # soft red
    'OP': 'FFE69C',  # soft gold
    'MH': 'FFF9C4',  # soft yellow
    'UP': 'FFB3D9',  # soft pink
    'CL': 'D4C5F9',  # soft lavender
    'CG': 'FFB3D1',  # soft rose
    'T':  'E8D5F5',  # soft orchid
    'AT': 'C5B3E6',  # soft purple
    'ML': 'F0D5F0',  # soft plum
}

def get_fill(status):
    """Return a PatternFill for the given status, or None for blank/unknown."""
    code = (status or '').strip().upper()
    if code in STATUS_COLORS:
        return PatternFill(fill_type='solid', fgColor=STATUS_COLORS[code])
    # Clear fill for empty / unknown status
    return PatternFill(fill_type=None)

app = Flask(__name__, static_folder='.')
# Secret key for session management (generate new key on first run)
app.config['SECRET_KEY'] = secrets.token_hex(32)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB upload limit

@app.errorhandler(413)
def handle_413(e):
    """Return JSON when uploaded file exceeds MAX_CONTENT_LENGTH."""
    return jsonify({'success': False, 'error': 'File too large. Maximum upload size is 50 MB.'}), 413

MASTER_ATTENDANCE_FILE = os.path.join(os.getcwd(), 'Attendance', 'Master_Attendance.xlsx')
MASTER_ATTENDANCE_LOG_FILE = os.path.join(os.getcwd(), 'Attendance', 'Master_Attendance_Logs.xlsx')


# ══════════════════════════════════════════════════════════════════════════
# AUTHENTICATION HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════

def get_authorized_users():
    """Load authorized users from Supabase (falls back to Access.csv when Supabase is unavailable)."""
    # ── Supabase path ────────────────────────────────────────────────────
    if SUPABASE_ENABLED:
        try:
            result = _supabase.table('authorized_users').select('username').execute()
            return [row['username'].lower().strip() for row in (result.data or []) if row.get('username')]
        except Exception as sb_err:
            print(f"⚠️  Supabase authorized_users query failed, falling back to CSV: {sb_err}")

    # ── CSV fallback ─────────────────────────────────────────────────────
    try:
        import csv, io
        access_file = os.path.join(os.getcwd(), 'Attendance', 'Access.csv')

        if not os.path.exists(access_file):
            print(f"⚠️ Access.csv not found at {access_file}")
            return []

        authorized_users = []
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(access_file, 'r', encoding=encoding, newline='') as f:
                    csv_content = f.read()
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                for row in csv_reader:
                    username_col = next((k for k in row if 'username' in k.lower()), None)
                    if username_col and row[username_col]:
                        user = row[username_col].lower().strip()
                        if user:
                            authorized_users.append(user)
                break
            except UnicodeDecodeError:
                continue
        return authorized_users
    except Exception as e:
        print(f"❌ Error loading authorized users: {e}")
        return []

def is_user_authorized(username):
    """Check if username is in authorized list"""
    if not username:
        return False
    username = username.lower().strip()
    authorized_users = get_authorized_users()
    return username in authorized_users

def require_auth(f):
    """Decorator to require authentication for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if user is authenticated in session
        username = session.get('authenticated_user')
        
        if not username:
            print(f"❌ Unauthenticated access attempt to {request.path}")
            return jsonify({
                "error": "Authentication required",
                "message": "Please authenticate to access this resource"
            }), 401
        
        # Verify user is still authorized
        if not is_user_authorized(username):
            print(f"❌ Unauthorized user trying to access: {username}")
            session.clear()
            return jsonify({
                "error": "Access denied",
                "message": "Your account is not authorized to access this resource"
            }), 403
        
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def enforce_attendance_authentication():
    """Enforce authentication on ALL Attendance tracker requests"""
    path = request.path
    
    # Skip auth check for these paths
    skip_paths = [
        '/api/auth/check-access',
        '/api/auth/login',
        '/api/auth/logout',
        '/Attendance/Start-Auth-Service.bat',
        '/Attendance/Launch-Attendance-Tracker.bat',
        '/Attendance/auto-login.html',
        '/',
        '/favicon.ico'
    ]
    
    # Check if this is an Attendance-related request
    if '/Attendance/' in path or path.startswith('/api/attendance/'):
        # Skip if it's an allowed path
        if any(path.startswith(skip) or path == skip for skip in skip_paths):
            return None
        
        # Check session authentication
        username = session.get('authenticated_user')
        
        if not username:
            print(f"🚫 BLOCKED unauthenticated access to: {path} from IP: {request.remote_addr}")
            if request.path.endswith('.html'):
                # Redirect HTML requests to login
                return redirect('/Attendance/auto-login.html?error=auth_required')
            else:
                # Return 401 for API requests
                return jsonify({
                    "error": "Authentication required",
                    "message": "Please authenticate first"
                }), 401
        
        # Verify user is still authorized
        if not is_user_authorized(username):
            print(f"🚫 BLOCKED unauthorized user: {username} trying to access: {path}")
            session.clear()
            if request.path.endswith('.html'):
                return redirect('/Attendance/auto-login.html?error=access_denied')
            else:
                return jsonify({
                    "error": "Access denied",
                    "message": "Your account is not authorized"
                }), 403
        
        print(f"✅ Authenticated access: {username} → {path}")

@app.before_request
def handle_options_preflight():
    """Respond to CORS preflight OPTIONS requests immediately."""
    if request.method == 'OPTIONS':
        resp = app.make_default_options_response()
        origin = request.headers.get('Origin')
        if origin:
            resp.headers['Access-Control-Allow-Origin'] = origin
        else:
            resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Credentials'] = 'true'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Windows-Username'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        return resp

@app.after_request
def add_cors_headers(response):
    """Allow browser clients from other local origins (port/IP) to call API endpoints."""
    # For session cookies to work with CORS, we need specific origin instead of *
    origin = request.headers.get('Origin')
    if origin:
        response.headers['Access-Control-Allow-Origin'] = origin
    else:
        response.headers['Access-Control-Allow-Origin'] = '*'
    
    response.headers['Access-Control-Allow-Credentials'] = 'true'  # Allow cookies/sessions
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Windows-Username'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    
    # Prevent caching for HTML files to ensure updates are immediately visible
    if request.path.endswith('.html'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    
    return response

# ══════════════════════════════════════════════════════════════════════════
# ACCESS CONTROL API
# ══════════════════════════════════════════════════════════════════════════

def extract_username_from_request():
    """
    Automatically extract Windows username from HTTP request headers.
    Tries multiple methods in order of reliability.
    """
    username = None
    detection_method = None
    
    # Method 1: Check for custom X-Windows-Username header (from our launcher/proxy)
    if request.headers.get('X-Windows-Username'):
        username = request.headers.get('X-Windows-Username')
        detection_method = "X-Windows-Username header"
    
    # Method 2: Check for REMOTE_USER (Windows Authentication via IIS/Apache)
    elif request.environ.get('REMOTE_USER'):
        username = request.environ.get('REMOTE_USER')
        # Remove domain prefix if present (DOMAIN\user -> user)
        if '\\' in username:
            username = username.split('\\')[-1]
        elif '@' in username:
            username = username.split('@')[0]
        detection_method = "Windows Authentication (REMOTE_USER)"
    
    # Method 3: Check for X-Forwarded-User header (from reverse proxy)
    elif request.headers.get('X-Forwarded-User'):
        username = request.headers.get('X-Forwarded-User')
        if '\\' in username:
            username = username.split('\\')[-1]
        detection_method = "X-Forwarded-User header"
    
    # Method 4: Check Authorization header for Basic auth or custom token
    elif request.headers.get('Authorization'):
        auth_header = request.headers.get('Authorization')
        if auth_header.startswith('Bearer '):
            # Custom token format: Bearer base64(username)
            try:
                import base64
                token = auth_header.replace('Bearer ', '')
                username = base64.b64decode(token).decode('utf-8')
                detection_method = "Bearer token"
            except:
                pass
    
    # Method 5: Check for custom cookie with username
    elif request.cookies.get('windows_user'):
        try:
            import base64
            encoded_user = request.cookies.get('windows_user')
            username = base64.b64decode(encoded_user).decode('utf-8')
            detection_method = "Secure cookie"
        except:
            pass
    
    # Method 6: POST request with username in body (manual entry)
    if not username and request.method == 'POST':
        data = request.get_json()
        if data and data.get('username'):
            username = data.get('username')
            detection_method = "POST body (manual)"
    
    # Method 7: Server environment (local access only)
    if not username:
        server_user = (
            os.environ.get('USERNAME') or
            os.environ.get('USER') or
            os.environ.get('LOGNAME')
        )
        if server_user and request.remote_addr in ['127.0.0.1', '::1', 'localhost']:
            username = server_user
            detection_method = "Server environment (local)"
    
    if username:
        # Clean up username
        username = username.lower().strip()
        # Remove domain/email suffix if present
        if '\\' in username:
            username = username.split('\\')[-1]
        if '@' in username:
            username = username.split('@')[0]
    
    return username, detection_method

@app.route('/api/auth/check-access', methods=['GET', 'POST'])
def check_attendance_access():
    """Check if current user has access to attendance tracker"""
    try:
        import csv
        
        # Automatically detect username from request
        username, detection_method = extract_username_from_request()
        
        if username:
            print(f"\n🔐 Username detected: {username} (via {detection_method})")
        else:
            print(f"\n❌ No username detected - all automatic methods failed")
        
        # Validate username format
        if not username or username == 'unknown' or len(username) < 3:
            print(f"❌ Invalid username: {username}")
            return jsonify({
                'success': False,
                'has_access': False,
                'username': username,
                'message': 'Invalid username provided'
            }), 400
        
        # Path to Access.csv
        access_file = os.path.join(os.getcwd(), 'Attendance', 'Access.csv')
        
        if not os.path.exists(access_file):
            print(f"❌ Access.csv not found at {access_file}")
            print(f"❌ Access denied for user: {username} (Access.csv missing)")
            
            return jsonify({
                'success': True,
                'has_access': False,
                'username': username,
                'message': 'Access denied: Access.csv file not found. Please contact administrator.'
            }), 403
        
        # Read Access.csv and check if user exists
        # Try multiple encodings to handle different file formats
        authorized_users = []
        # Try utf-8-sig first to handle BOM (Byte Order Mark)
        encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        csv_content = None
        encoding_used = None
        for encoding in encodings_to_try:
            try:
                with open(access_file, 'r', encoding=encoding, newline='') as f:
                    csv_content = f.read()
                encoding_used = encoding
                print(f"   ✓ Successfully read Access.csv with {encoding} encoding")
                break
            except UnicodeDecodeError:
                continue
        
        if csv_content is None:
            print(f"❌ Could not read Access.csv with any encoding")
            print(f"❌ Access denied for user: {username} (Access.csv unreadable)")
            
            return jsonify({
                'success': True,
                'has_access': False,
                'username': username,
                'message': 'Access denied: Access.csv file is corrupted. Please contact administrator.'
            }), 403
        
        # Parse the CSV content
        import io
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        for row in csv_reader:
            print(f"   📋 CSV row: {row}")
            # Handle different possible header names (with/without BOM, with/without spaces)
            username_col = None
            for key in row.keys():
                if 'username' in key.lower():
                    username_col = key
                    break
            
            if username_col and row[username_col]:
                user_from_csv = row[username_col].lower().strip()
                authorized_users.append(user_from_csv)
                print(f"      ✓ Added user: {user_from_csv}")
        
        print(f"   📊 Total authorized users: {len(authorized_users)}")
        print(f"   👥 Authorized users: {', '.join(authorized_users) if authorized_users else '(none)'}")
        
        has_access = username in authorized_users
        
        # Log access attempt
        try:
            log_file = os.path.join(os.path.dirname(access_file), 'access_log.csv')
            log_exists = os.path.exists(log_file)
            
            with open(log_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if not log_exists:
                    writer.writerow(['Timestamp', 'Username', 'IP Address', 'User Agent', 'Access Granted', 'Method'])
                
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ip_address = request.remote_addr or 'unknown'
                user_agent = request.headers.get('User-Agent', 'unknown')[:100]
                access_method = 'Remote' if request.method == 'POST' else 'Local'
                
                writer.writerow([timestamp, username, ip_address, user_agent, 'Yes' if has_access else 'No', access_method])
        except Exception as log_error:
            print(f"⚠️ Failed to log access attempt: {log_error}")
        
        if has_access:
            print(f"✅ Access granted for user: {username} (IP: {request.remote_addr})")
            
            # Store authenticated user in session
            session['authenticated_user'] = username
            session['auth_timestamp'] = datetime.now().isoformat()
            session['auth_method'] = detection_method
            session.permanent = True  # Make session persistent
            
            return jsonify({
                'success': True,
                'has_access': True,
                'username': username,
                'message': 'Access granted'
            })
        else:
            print(f"❌ Access denied for user: {username} (IP: {request.remote_addr})")
            print(f"   Authorized users: {', '.join(authorized_users)}")
            return jsonify({
                'success': True,
                'has_access': False,
                'username': username,
                'message': 'Access denied: User not found in authorized list'
            }), 403
    
    except Exception as e:
        print(f"❌ Error checking access: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Logout user and clear session"""
    username = session.get('authenticated_user', 'unknown')
    session.clear()
    print(f"👋 User logged out: {username}")
    return jsonify({
        'success': True,
        'message': 'Logged out successfully'
    })

@app.route('/api/auth/session', methods=['GET'])
def check_session():
    """Check if user has valid session"""
    username = session.get('authenticated_user')
    if username and is_user_authorized(username):
        return jsonify({
            'authenticated': True,
            'username': username,
            'auth_method': session.get('auth_method'),
            'auth_timestamp': session.get('auth_timestamp')
        })
    else:
        session.clear()
        return jsonify({
            'authenticated': False
        }), 401

# ══════════════════════════════════════════════════════════════════════════
# PROTECTED ATTENDANCE ROUTES  
# ══════════════════════════════════════════════════════════════════════════

@app.route('/Attendance/team-attendance-tracker-sharepoint.html')
def serve_attendance_tracker():
    """Serve attendance tracker HTML with authentication check (handled by middleware)"""
    # Authentication is enforced by before_request middleware
    # If we reach here, user is authenticated
    username = session.get('authenticated_user')
    print(f"📄 Serving attendance tracker to authenticated user: {username}")
    return send_file(os.path.join(os.getcwd(), 'Attendance', 'team-attendance-tracker-sharepoint.html'))

@app.route('/ado-testcase-upload.html')
def serve_ado_testcase_upload_page():
    """Serve the ADO Testcase Upload page directly to avoid any static-serving ambiguity."""
    return send_file(os.path.join(os.getcwd(), 'ado-testcase-upload.html'))

@app.route('/daily-report.html')
def serve_daily_report_page():
    """Serve the Main Release Daily Status Report page."""
    return send_file(os.path.join(os.getcwd(), 'daily-report.html'))

@app.route('/hotfix-daily-report.html')
def serve_hotfix_daily_report_page():
    """Serve the Hot Fix Daily Status Report page."""
    return send_file(os.path.join(os.getcwd(), 'hotfix-daily-report.html'))

@app.route('/api/download-daily-status-report', methods=['GET'])
def download_daily_status_report():
    """Download the Main Release Daily Status Dashboard HTML as a file."""
    report_file = _find_main_release_dashboard_file()
    if not os.path.exists(report_file):
        return jsonify({'success': False, 'error': 'Report not yet generated'}), 404
    return send_file(report_file, as_attachment=True, download_name='Daily_Status_Dashboard.html')

@app.route('/api/download-hotfix-daily-status-report', methods=['GET'])
def download_hotfix_daily_status_report():
    """Download the Hot Fix Daily Status Dashboard HTML as a file."""
    report_file = os.path.join(os.getcwd(), 'Hot_Fix_Daily_Status_Report', 'Daily_Status_Dashboard.html')
    if not os.path.exists(report_file):
        return jsonify({'success': False, 'error': 'Report not yet generated'}), 404
    return send_file(report_file, as_attachment=True, download_name='HotFix_Daily_Status_Dashboard.html')

# ══════════════════════════════════════════════════════════════════════════
# REPORT GENERATION APIs
# ══════════════════════════════════════════════════════════════════════════

# API endpoint to generate regression report
@app.route('/api/generate-regression-report', methods=['POST'])
def generate_regression_report():
    """Triggers the regression report Python script"""
    try:
        print("\n" + "="*80)
        print("🚀 Starting Regression Report Generation...")
        print("="*80)
        
        # Get the script path
        script_path = os.path.join(os.getcwd(), 'Regression_Report', 'RegReport 1_Updated 1.py')
        
        if not os.path.exists(script_path):
            return jsonify({
                'success': False,
                'error': f'Script not found: {script_path}'
            }), 404
        
        # Run the Python script with UTF-8 encoding
        print(f"📝 Executing: {script_path}")
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            ['python', script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=300,  # 5 minute timeout
            env=env
        )
        
        if result.returncode != 0:
            print(f"❌ Script failed with error:")
            print(result.stderr)
            return jsonify({
                'success': False,
                'error': result.stderr
            }), 500
        
        # Find the most recently created report file
        report_pattern = 'regression_execution_report_*.html'
        report_files = glob.glob(report_pattern)
        
        if not report_files:
            return jsonify({
                'success': False,
                'error': 'Report file not found after generation'
            }), 500
        
        # Get the most recent file
        latest_report = max(report_files, key=os.path.getctime)
        
        print(f"✅ Report generated successfully: {latest_report}")
        print("="*80 + "\n")
        
        return jsonify({
            'success': True,
            'report_file': latest_report,
            'message': 'Report generated successfully'
        })
        
    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'Report generation timed out (>5 minutes)'
        }), 500
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ── Concurrency locks ────────────────────────────────────────────────────
# Each runner gets its own lock so the in_progress check+set is atomic,
# preventing two simultaneous requests from both passing the guard.
_daily_report_lock   = threading.Lock()
_hotfix_report_lock  = threading.Lock()
_tc_compare_lock     = threading.Lock()
_mpoc_extref_lock    = threading.Lock()
_missing_fields_lock = threading.Lock()
_missing_scope_lock  = threading.Lock()
_ado_upload_lock     = threading.Lock()
# Single shared lock protecting all settings file reads + writes
_settings_file_lock  = threading.Lock()
# ─────────────────────────────────────────────────────────────────────────

# Global variable to track report generation status
report_generation_status = {'in_progress': False, 'last_generated': None, 'error': None}

# Global variable to track missing-data-scope report generation status
missing_scope_report_status = {'in_progress': False, 'last_generated': None, 'error': None}

# ══════════════════════════════════════════════════════════════════════════
# DAILY STATUS REPORT APIs
# ══════════════════════════════════════════════════════════════════════════

daily_status_report_status = {'in_progress': False, 'last_generated': None, 'error': None, 'log': ''}

def _find_main_release_dashboard_file():
    """Resolve the generated dashboard HTML path using the same logic as config.py."""
    project_root = Path(__file__).resolve().parent
    candidates = [
        # Priority 1: Script directory (works in Docker) - MATCHES config.py
        project_root / "Main_Release_Daily_Status_Report" / "Daily_Status_Dashboard.html",
        # Priority 2: GHC files in project root
        project_root / "GHC files" / "Daily status report - Integrated" / "Daily_Status_Dashboard.html",
        # Priority 3: User home directory
        Path.home() / "GHC files" / "Daily status report - Integrated" / "Daily_Status_Dashboard.html",
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    # Return the project-local path as default (script will create it)
    return str(candidates[0])

def run_daily_status_script():
    """Background thread to run the daily status dashboard generator"""
    global daily_status_report_status
    try:
        print("\n" + "="*80)
        print("Starting Daily Status Report Generation...")
        print("="*80)

        script_path = os.path.join(os.getcwd(), 'Main_Release_Daily_Status_Report',
                                   'generate_daily_status_dashboard Integrated.py')

        if not os.path.exists(script_path):
            daily_status_report_status['error'] = f'Script not found: {script_path}'
            daily_status_report_status['in_progress'] = False
            return

        print(f"Executing: {script_path}")
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            ['python', script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=600,   # 10 minute timeout
            env=env
        )

        log_output = (result.stdout or '') + (result.stderr or '')
        daily_status_report_status['log'] = log_output

        if result.returncode != 0:
            print(f"Script failed:\n{result.stderr}")
            daily_status_report_status['error'] = result.stderr or 'Script exited with error'
            daily_status_report_status['in_progress'] = False
            return

        report_file = _find_main_release_dashboard_file()
        if os.path.exists(report_file):
            print(f"Report generated successfully: {report_file}")
            daily_status_report_status['last_generated'] = report_file
            daily_status_report_status['error'] = None
            # Auto-open the dashboard in the default browser
            try:
                webbrowser.open(f'file:///{report_file.replace(os.sep, "/")}')
                print(f"Opened dashboard in browser: {report_file}")
            except Exception as e:
                print(f"Warning: Could not open browser: {e}")
        else:
            daily_status_report_status['error'] = 'Dashboard HTML not found after generation'

        daily_status_report_status['in_progress'] = False
        print("="*80 + "\n")

    except subprocess.TimeoutExpired:
        daily_status_report_status['error'] = 'Report generation timed out (>10 minutes)'
        daily_status_report_status['in_progress'] = False
    except Exception as e:
        print(f"Error: {str(e)}")
        daily_status_report_status['error'] = str(e)
        daily_status_report_status['in_progress'] = False


@app.route('/api/generate-daily-status-report', methods=['POST'])
def generate_daily_status_report():
    """Start daily status report generation in background"""
    global daily_status_report_status
    with _daily_report_lock:
        if daily_status_report_status['in_progress']:
            return jsonify({'success': False, 'error': 'Report generation already in progress'}), 409
        daily_status_report_status['in_progress'] = True
        daily_status_report_status['error'] = None
        daily_status_report_status['log'] = ''

    thread = threading.Thread(target=run_daily_status_script, daemon=True)
    thread.start()

    return jsonify({'success': True, 'message': 'Daily status report generation started'})


@app.route('/api/check-daily-status-report', methods=['GET'])
def check_daily_status_report():
    """Poll status of daily status report generation"""
    return jsonify({
        'in_progress': daily_status_report_status['in_progress'],
        'last_generated': daily_status_report_status['last_generated'],
        'error': daily_status_report_status['error'],
        'log': daily_status_report_status['log'][-3000:] if daily_status_report_status['log'] else ''
    })

# Map of field keys → (script relative path, variable name, regex pattern)
DAILY_REPORT_SCRIPT_FIELDS = {
    'pt_folder_path': (
        os.path.join('Main_Release_Daily_Status_Report', 'download_PT status_file.py'),
        'FILE_FOLDER_RELATIVE_URL',
        r'FILE_FOLDER_RELATIVE_URL\s*=\s*"[^"]*"'
    ),
    'pt_file_name': (
        os.path.join('Main_Release_Daily_Status_Report', 'download_PT status_file.py'),
        'FILE_NAME_BASE',
        r'FILE_NAME_BASE\s*=\s*"[^"]*"'
    ),
    'uat_folder_path': (
        os.path.join('Main_Release_Daily_Status_Report', 'download_UAT status_file.py'),
        'FILE_FOLDER_RELATIVE_URL',
        r'FILE_FOLDER_RELATIVE_URL\s*=\s*"[^"]*"'
    ),
    'uat_file_name': (
        os.path.join('Main_Release_Daily_Status_Report', 'download_UAT status_file.py'),
        'FILE_NAME_BASE',
        r'FILE_NAME_BASE\s*=\s*"[^"]*"'
    ),
    'po_query_id': (
        os.path.join('Main_Release_Daily_Status_Report', 'generate_product_owner_details Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
    'bugs_query_id': (
        os.path.join('Main_Release_Daily_Status_Report', 'generate_open bug_summary Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
    'stories_query_id': (
        os.path.join('Main_Release_Daily_Status_Report', 'generate_story_summary_detailed Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
    'defects_query_id': (
        os.path.join('Main_Release_Daily_Status_Report', 'generate_overall_defect_summary Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
}

@app.route('/api/get-daily-report-settings', methods=['GET'])
def get_daily_report_settings():
    """Read current configurable values from all 6 Daily Report Python scripts"""
    try:
        values = {}
        for key, (rel_path, var_name, pattern) in DAILY_REPORT_SCRIPT_FIELDS.items():
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            match = re.search(pattern, content)
            if match:
                # Extract the quoted value from the matched line
                val_match = re.search(r'"([^"]*)"', match.group())
                values[key] = val_match.group(1) if val_match else ''
            else:
                values[key] = ''
        return jsonify({'success': True, 'settings': values})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-daily-report-settings', methods=['POST'])
def update_daily_report_settings():
    """Update configurable values in the Daily Report Python scripts"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400

        updated = []
        for key, new_value in data.items():
            if key not in DAILY_REPORT_SCRIPT_FIELDS:
                continue
            new_value = new_value.strip()
            if not new_value:
                return jsonify({'success': False, 'error': f'Value for {key} cannot be empty'}), 400

            rel_path, var_name, pattern = DAILY_REPORT_SCRIPT_FIELDS[key]
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404

            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()

            replacement = f'{var_name} = "{new_value}"'
            new_content = re.sub(pattern, replacement, content)

            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(new_content)

            updated.append(key)
            print(f"✅ Updated {key} in {rel_path}")

        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════
# HOT FIX DAILY STATUS REPORT APIs
# ══════════════════════════════════════════════════════════════════════════

hotfix_daily_report_status = {'in_progress': False, 'last_generated': None, 'error': None, 'log': ''}

def run_hotfix_daily_status_script():
    """Background thread to run the Hot Fix daily status dashboard generator"""
    global hotfix_daily_report_status
    try:
        print("\n" + "="*80)
        print("Starting Hot Fix Daily Status Report Generation...")
        print("="*80)

        script_path = os.path.join(os.getcwd(), 'Hot_Fix_Daily_Status_Report',
                                   'generate_daily_status_dashboard Integrated.py')

        if not os.path.exists(script_path):
            hotfix_daily_report_status['error'] = f'Script not found: {script_path}'
            hotfix_daily_report_status['in_progress'] = False
            return

        hotfix_dir = os.path.join(os.getcwd(), 'Hot_Fix_Daily_Status_Report')
        print(f"Executing: {script_path}")
        print(f"Hot Fix output dir: {hotfix_dir}")
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        env['MYISP_HOTFIX_OUTPUT_DIR'] = hotfix_dir
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=600,
            env=env,
            cwd=hotfix_dir
        )

        log_output = (result.stdout or '') + (result.stderr or '')
        hotfix_daily_report_status['log'] = log_output

        if result.returncode != 0:
            print(f"Script failed:\n{result.stderr}")
            hotfix_daily_report_status['error'] = result.stderr or 'Script exited with error'
            hotfix_daily_report_status['in_progress'] = False
            return

        report_file = os.path.join(os.getcwd(), 'Hot_Fix_Daily_Status_Report',
                                   'Daily_Status_Dashboard.html')
        if os.path.exists(report_file):
            print(f"Report generated successfully: {report_file}")
            hotfix_daily_report_status['last_generated'] = 'Hot_Fix_Daily_Status_Report/Daily_Status_Dashboard.html'
            hotfix_daily_report_status['error'] = None
        else:
            hotfix_daily_report_status['error'] = 'Dashboard HTML not found after generation'

        hotfix_daily_report_status['in_progress'] = False
        print("="*80 + "\n")

    except subprocess.TimeoutExpired:
        hotfix_daily_report_status['error'] = 'Report generation timed out (>10 minutes)'
        hotfix_daily_report_status['in_progress'] = False
    except Exception as e:
        print(f"Error: {str(e)}")
        hotfix_daily_report_status['error'] = str(e)
        hotfix_daily_report_status['in_progress'] = False


@app.route('/api/generate-hotfix-daily-report', methods=['POST'])
def generate_hotfix_daily_report():
    """Start Hot Fix daily status report generation in background"""
    global hotfix_daily_report_status
    with _hotfix_report_lock:
        if hotfix_daily_report_status['in_progress']:
            return jsonify({'success': False, 'error': 'Report generation already in progress'}), 409
        hotfix_daily_report_status['in_progress'] = True
        hotfix_daily_report_status['error'] = None
        hotfix_daily_report_status['log'] = ''

    thread = threading.Thread(target=run_hotfix_daily_status_script, daemon=True)
    thread.start()

    return jsonify({'success': True, 'message': 'Hot Fix daily status report generation started'})


@app.route('/api/check-hotfix-daily-report', methods=['GET'])
def check_hotfix_daily_report():
    """Poll status of Hot Fix daily status report generation"""
    return jsonify({
        'in_progress': hotfix_daily_report_status['in_progress'],
        'last_generated': hotfix_daily_report_status['last_generated'],
        'error': hotfix_daily_report_status['error'],
        'log': hotfix_daily_report_status['log'][-3000:] if hotfix_daily_report_status['log'] else ''
    })


# Map of field keys → (script relative path, variable name, regex pattern) for Hot Fix
HOTFIX_REPORT_SCRIPT_FIELDS = {
    'pt_folder_path': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'download_PT status_file.py'),
        'FILE_FOLDER_RELATIVE_URL',
        r'FILE_FOLDER_RELATIVE_URL\s*=\s*"[^"]*"'
    ),
    'pt_file_name': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'download_PT status_file.py'),
        'FILE_NAME_BASE',
        r'FILE_NAME_BASE\s*=\s*"[^"]*"'
    ),
    'uat_folder_path': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'download_UAT status_file.py'),
        'FILE_FOLDER_RELATIVE_URL',
        r'FILE_FOLDER_RELATIVE_URL\s*=\s*"[^"]*"'
    ),
    'uat_file_name': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'download_UAT status_file.py'),
        'FILE_NAME_BASE',
        r'FILE_NAME_BASE\s*=\s*"[^"]*"'
    ),
    'po_query_id': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'generate_product_owner_details Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
    'bugs_query_id': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'generate_open bug_summary Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
    'stories_query_id': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'generate_story_summary_detailed Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
    'defects_query_id': (
        os.path.join('Hot_Fix_Daily_Status_Report', 'generate_overall_defect_summary Integrated.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
}


@app.route('/api/get-hotfix-daily-report-settings', methods=['GET'])
def get_hotfix_daily_report_settings():
    """Read current configurable values from all Hot Fix Daily Report Python scripts"""
    try:
        values = {}
        for key, (rel_path, var_name, pattern) in HOTFIX_REPORT_SCRIPT_FIELDS.items():
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            match = re.search(pattern, content)
            if match:
                val_match = re.search(r'"([^"]*)"', match.group())
                values[key] = val_match.group(1) if val_match else ''
            else:
                values[key] = ''
        return jsonify({'success': True, 'settings': values})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-hotfix-daily-report-settings', methods=['POST'])
def update_hotfix_daily_report_settings():
    """Update configurable values in the Hot Fix Daily Report Python scripts"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400

        updated = []
        for key, new_value in data.items():
            if key not in HOTFIX_REPORT_SCRIPT_FIELDS:
                continue
            new_value = new_value.strip()
            if not new_value:
                return jsonify({'success': False, 'error': f'Value for {key} cannot be empty'}), 400

            rel_path, var_name, pattern = HOTFIX_REPORT_SCRIPT_FIELDS[key]
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404

            with _settings_file_lock:
                with open(full_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                replacement = f'{var_name} = "{new_value}"'
                new_content = re.sub(pattern, replacement, content)

                with open(full_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)

            updated.append(key)
            print(f"✅ Updated {key} in {rel_path}")

        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════
# TC COMPARE (MATCH TFS & SP) SETTINGS APIs
# ══════════════════════════════════════════════════════════════════════════

# Map of field keys → (script relative path, variable name, regex pattern, style)
# style='dict'   → value is a dict entry:  'key': 'value'
# style='assign' → value is an assignment: VAR = "value"
TC_COMPARE_SCRIPT_FIELDS = {
    'plan_id': (
        os.path.join('TC_Compare', 'Count Mismatch Report_V5_Final.py'),
        'plan_id',
        r"'plan_id'\s*:\s*'[^']*'",
        'dict'
    ),
    'suite_id': (
        os.path.join('TC_Compare', 'Count Mismatch Report_V5_Final.py'),
        'suite_id',
        r"'suite_id'\s*:\s*'[^']*'",
        'dict'
    ),
    'sp_folder_path': (
        os.path.join('TC_Compare', 'download_sharepoint_file.py'),
        'FILE_FOLDER_RELATIVE_URL',
        r'FILE_FOLDER_RELATIVE_URL\s*=\s*(?:"[^"]*")+',
        'assign'
    ),
    'file_name': (
        os.path.join('TC_Compare', 'download_sharepoint_file.py'),
        'FILE_NAME_BASE',
        r'FILE_NAME_BASE\s*=\s*"[^"]*"',
        'assign'
    ),
}


@app.route('/api/get-tc-compare-settings', methods=['GET'])
def get_tc_compare_settings():
    """Read current configurable values from TC Compare Python scripts"""
    try:
        values = {}
        for key, (rel_path, var_name, pattern, style) in TC_COMPARE_SCRIPT_FIELDS.items():
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            match = re.search(pattern, content)
            if match:
                if style == 'dict':
                    # Extract value from  'key': 'value'  — last single-quoted group
                    val_matches = re.findall(r"'([^']*)'", match.group())
                    values[key] = val_matches[-1] if val_matches else ''
                else:
                    # Join all double-quoted strings (handles implicit string concatenation)
                    val_matches = re.findall(r'"([^"]*)"', match.group())
                    values[key] = ''.join(val_matches) if val_matches else ''
            else:
                values[key] = ''
        return jsonify({'success': True, 'settings': values})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-tc-compare-settings', methods=['POST'])
def update_tc_compare_settings():
    """Update configurable values in the TC Compare Python scripts"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400

        updated = []
        for key, new_value in data.items():
            if key not in TC_COMPARE_SCRIPT_FIELDS:
                continue
            new_value = new_value.strip()
            if not new_value:
                return jsonify({'success': False, 'error': f'Value for {key} cannot be empty'}), 400

            rel_path, var_name, pattern, style = TC_COMPARE_SCRIPT_FIELDS[key]
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404

            with _settings_file_lock:
                with open(full_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                if style == 'dict':
                    replacement = f"'{var_name}': '{new_value}'"
                else:
                    replacement = f'{var_name} = "{new_value}"'

                new_content = re.sub(pattern, replacement, content)

                with open(full_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)

            updated.append(key)
            print(f"✅ Updated {key} in {rel_path}")

        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════
# TC COMPARE (MATCH TFS & SP) REPORT APIs
# ══════════════════════════════════════════════════════════════════════════

tc_compare_report_status = {'in_progress': False, 'last_generated': None, 'error': None, 'log': ''}

def run_tc_compare_script():
    """Background thread to run the TC Compare (Match TFS & SP) report"""
    global tc_compare_report_status
    try:
        print("\n" + "="*80)
        print("Starting TC Compare (Match TFS & SP) Report Generation...")
        print("="*80)

        script_path = os.path.join(os.getcwd(), 'TC_Compare', 'Count Mismatch Report_V5_Final.py')

        if not os.path.exists(script_path):
            tc_compare_report_status['error'] = f'Script not found: {script_path}'
            tc_compare_report_status['in_progress'] = False
            return

        print(f"Executing: {script_path}")
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            ['python', script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=900,   # 15-minute timeout (SharePoint download + ADO API calls)
            env=env
        )

        log_output = (result.stdout or '') + (result.stderr or '')
        tc_compare_report_status['log'] = log_output

        if result.returncode != 0:
            print(f"Script failed:\n{result.stderr}")
            tc_compare_report_status['error'] = result.stderr or 'Script exited with error'
            tc_compare_report_status['in_progress'] = False
            return

        # Find the most recently generated report file
        tc_compare_dir = os.path.join(os.getcwd(), 'TC_Compare')
        report_files = glob.glob(os.path.join(tc_compare_dir, 'pt_execution_count_comparison_report_*.html'))
        if report_files:
            latest = max(report_files, key=os.path.getctime)
            rel_path = 'TC_Compare/' + os.path.basename(latest)
            print(f"Report generated successfully: {latest}")
            tc_compare_report_status['last_generated'] = rel_path
            tc_compare_report_status['error'] = None
        else:
            tc_compare_report_status['error'] = 'Report HTML not found after generation'

        tc_compare_report_status['in_progress'] = False
        print("="*80 + "\n")

    except subprocess.TimeoutExpired:
        tc_compare_report_status['error'] = 'Report generation timed out (>15 minutes)'
        tc_compare_report_status['in_progress'] = False
    except Exception as e:
        print(f"Error: {str(e)}")
        tc_compare_report_status['error'] = str(e)
        tc_compare_report_status['in_progress'] = False


@app.route('/api/generate-tc-compare-report', methods=['POST'])
def generate_tc_compare_report():
    """Start TC Compare report generation in background"""
    global tc_compare_report_status
    with _tc_compare_lock:
        if tc_compare_report_status['in_progress']:
            return jsonify({'success': False, 'error': 'Report generation already in progress'}), 409
        tc_compare_report_status['in_progress'] = True
        tc_compare_report_status['error'] = None
        tc_compare_report_status['log'] = ''

    thread = threading.Thread(target=run_tc_compare_script, daemon=True)
    thread.start()

    return jsonify({'success': True, 'message': 'TC Compare report generation started'})


@app.route('/api/check-tc-compare-report', methods=['GET'])
def check_tc_compare_report():
    """Poll status of TC Compare report generation"""
    return jsonify({
        'in_progress': tc_compare_report_status['in_progress'],
        'last_generated': tc_compare_report_status['last_generated'],
        'error': tc_compare_report_status['error'],
        'log': tc_compare_report_status['log'][-3000:] if tc_compare_report_status['log'] else ''
    })

# ══════════════════════════════════════════════════════════════════════════
# US EXTERNAL REFERENCE ID UPDATE (M_POC) APIs
# ══════════════════════════════════════════════════════════════════════════

MPOC_EXTREF_SCRIPT_FIELDS = {
    'query_id': (
        os.path.join('M_POC', 'Auto_Update_ExternalRef ID.py'),
        'query_id',
        r'query_id\s*=\s*"[^"]*"'
    ),
}

mpoc_extref_status = {'in_progress': False, 'last_run': None, 'error': None, 'log': '', 'last_csv': None}


def run_mpoc_extref_script(query_id=None):
    """Background thread to run the US ExternalRef ID update script"""
    global mpoc_extref_status
    try:
        print("\n" + "="*80)
        print("Starting US External Reference ID Update...")
        print("="*80)

        script_path = os.path.join(os.getcwd(), 'M_POC', 'Auto_Update_ExternalRef ID.py')

        if not os.path.exists(script_path):
            mpoc_extref_status['error'] = f'Script not found: {script_path}'
            mpoc_extref_status['in_progress'] = False
            return

        # Use provided query_id; fall back to reading from the script file
        if not query_id:
            with open(script_path, 'r', encoding='utf-8') as f:
                content = f.read()
            match = re.search(r'query_id\s*=\s*"([^"]*)"', content)
            query_id = match.group(1) if match else ''

        if not query_id:
            mpoc_extref_status['error'] = 'No Query ID provided. Please enter it in the ADO Query ID field.'
            mpoc_extref_status['in_progress'] = False
            return

        print(f"Executing: {script_path}")
        print(f"Query ID: {query_id}")
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            [sys.executable, script_path, query_id],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=600,
            env=env,
            cwd=os.path.join(os.getcwd(), 'M_POC')
        )

        log_output = (result.stdout or '') + (result.stderr or '')
        mpoc_extref_status['log'] = log_output

        if result.returncode != 0:
            print(f"Script failed:\n{result.stderr}")
            mpoc_extref_status['error'] = result.stderr or 'Script exited with error'
            mpoc_extref_status['in_progress'] = False
            return

        mpoc_extref_status['last_run'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        mpoc_extref_status['error'] = None
        mpoc_extref_status['in_progress'] = False

        # Find the most recently generated CSV results file
        mpoc_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'M_POC')
        csv_files = glob.glob(os.path.join(mpoc_dir, 'ADO_Update_Results_*.csv'))
        if csv_files:
            latest_csv = max(csv_files, key=os.path.getctime)
            mpoc_extref_status['last_csv'] = latest_csv  # Store full absolute path
            print(f"Results CSV: {latest_csv}")

        print("US ExternalRef ID update completed successfully.")
        print("="*80 + "\n")

    except subprocess.TimeoutExpired:
        mpoc_extref_status['error'] = 'Update timed out (>10 minutes)'
        mpoc_extref_status['in_progress'] = False
    except Exception as e:
        print(f"Error: {str(e)}")
        mpoc_extref_status['error'] = str(e)
        mpoc_extref_status['in_progress'] = False


@app.route('/mpoc-extref.html')
def serve_mpoc_extref_page():
    """Serve the US External Reference ID Update page."""
    return send_file(os.path.join(os.getcwd(), 'm-poc-extref.html'))


@app.route('/api/generate-mpoc-extref', methods=['POST'])
def generate_mpoc_extref():
    """Start the US ExternalRef ID update in background"""
    global mpoc_extref_status
    data = request.get_json(silent=True) or {}
    query_id = (data.get('query_id') or '').strip()

    with _mpoc_extref_lock:
        if mpoc_extref_status['in_progress']:
            return jsonify({'success': False, 'error': 'Update already in progress'}), 409
        mpoc_extref_status['in_progress'] = True
        mpoc_extref_status['error'] = None
        mpoc_extref_status['log'] = ''

    thread = threading.Thread(target=run_mpoc_extref_script, kwargs={'query_id': query_id}, daemon=True)
    thread.start()

    return jsonify({'success': True, 'message': 'US ExternalRef ID update started'})


@app.route('/api/check-mpoc-extref', methods=['GET'])
def check_mpoc_extref():
    """Poll status of US ExternalRef ID update"""
    return jsonify({
        'in_progress': mpoc_extref_status['in_progress'],
        'last_run': mpoc_extref_status['last_run'],
        'error': mpoc_extref_status['error'],
        'log': mpoc_extref_status['log'][-3000:] if mpoc_extref_status['log'] else '',
        'last_csv': mpoc_extref_status.get('last_csv')
    })


@app.route('/api/download-mpoc-extref-results', methods=['GET'])
def download_mpoc_extref_results():
    """Download the latest ADO_Update_Results CSV file"""
    csv_path = mpoc_extref_status.get('last_csv')
    if not csv_path or not os.path.exists(csv_path):
        # Fall back to scanning the M_POC directory using app.py location as base
        mpoc_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'M_POC')
        csv_files = glob.glob(os.path.join(mpoc_dir, 'ADO_Update_Results_*.csv'))
        if not csv_files:
            return jsonify({'success': False, 'error': 'No results file found yet'}), 404
        csv_path = max(csv_files, key=os.path.getctime)
    return send_file(csv_path, as_attachment=True, download_name=os.path.basename(csv_path), mimetype='text/csv')


@app.route('/api/get-mpoc-extref-settings', methods=['GET'])
def get_mpoc_extref_settings():
    """Read current query_id from M_POC script"""
    try:
        values = {}
        for key, (rel_path, var_name, pattern) in MPOC_EXTREF_SCRIPT_FIELDS.items():
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            match = re.search(pattern, content)
            if match:
                val_match = re.search(r'"([^"]*)"', match.group())
                values[key] = val_match.group(1) if val_match else ''
            else:
                values[key] = ''
        return jsonify({'success': True, 'settings': values})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-mpoc-extref-settings', methods=['POST'])
def update_mpoc_extref_settings():
    """Update query_id in M_POC script"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400

        updated = []
        for key, new_value in data.items():
            if key not in MPOC_EXTREF_SCRIPT_FIELDS:
                continue
            new_value = new_value.strip()
            if not new_value:
                return jsonify({'success': False, 'error': f'Value for {key} cannot be empty'}), 400

            rel_path, var_name, pattern = MPOC_EXTREF_SCRIPT_FIELDS[key]
            full_path = os.path.join(os.getcwd(), rel_path)
            if not os.path.exists(full_path):
                return jsonify({'success': False, 'error': f'Script not found: {rel_path}'}), 404

            with _settings_file_lock:
                with open(full_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                replacement = f'{var_name} = "{new_value}"'
                new_content = re.sub(pattern, replacement, content)

                with open(full_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)

            updated.append(key)
            print(f"✅ Updated {key} in {rel_path}")

        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════

def run_missing_fields_script():
    """Background thread to run the missing fields report script"""
    global report_generation_status
    try:
        print("\n" + "="*80)
        print("🚀 Starting Missing Fields Report Generation...")
        print("="*80)
        
        script_path = os.path.join(os.getcwd(), 'Missing_Filed_Report', 'Generate_missing_fields_report .py')
        
        print(f"📝 Executing: {script_path}")
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            ['python', script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=300,
            env=env
        )
        
        if result.returncode != 0:
            print(f"❌ Script failed with error:")
            print(result.stderr)
            report_generation_status['error'] = result.stderr
            report_generation_status['in_progress'] = False
            return
        
        report_file = os.path.join('Missing_Filed_Report', 'Missing_Fields_Report.html')
        
        if os.path.exists(report_file):
            print(f"✅ Report generated successfully: {report_file}")
            print("="*80 + "\n")
            report_generation_status['last_generated'] = report_file
            report_generation_status['error'] = None
        else:
            report_generation_status['error'] = 'Report file not found after generation'
        
        report_generation_status['in_progress'] = False
        
    except subprocess.TimeoutExpired:
        report_generation_status['error'] = 'Report generation timed out (>5 minutes)'
        report_generation_status['in_progress'] = False
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        report_generation_status['error'] = str(e)
        report_generation_status['in_progress'] = False

# API endpoint to run attendance tracker update script
@app.route('/api/run-attendance-update', methods=['POST'])
def run_attendance_update():
    """Runs update_attendance_tracker.py to refresh team data and holidays in the HTML tracker"""
    try:
        print("\n" + "="*80)
        print("🚀 Starting Attendance Tracker Update...")
        print("="*80)

        script_path = os.path.join(os.getcwd(), 'Attendance', 'update_attendance_tracker.py')

        if not os.path.exists(script_path):
            return jsonify({'success': False, 'error': f'Script not found: {script_path}'}), 404

        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'

        result = subprocess.run(
            ['python', script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=120,
            env=env,
            input='\n'   # automatically answer the "Press Enter to exit" prompt
        )

        output = (result.stdout or '') + (result.stderr or '')

        if result.returncode != 0:
            print(f"❌ Script failed:\n{result.stderr}")
            return jsonify({'success': False, 'error': result.stderr or 'Script exited with error'}), 500

        print(f"✅ Attendance tracker updated successfully")
        print("="*80 + "\n")

        return jsonify({
            'success': True,
            'message': 'Attendance tracker updated successfully',
            'output': output
        })

    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'Script timed out (120 s)'}), 500
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# API endpoint to start missing fields report generation
@app.route('/api/generate-missing-fields-report', methods=['POST'])
def generate_missing_fields_report():
    """Starts the missing fields report generation in background"""
    global report_generation_status
    
    # Get the script path
    script_path = os.path.join(os.getcwd(), 'Missing_Filed_Report', 'Generate_missing_fields_report .py')
    
    if not os.path.exists(script_path):
        return jsonify({
            'success': False,
            'error': f'Script not found: {script_path}'
        }), 404
    
    # Check if already in progress (atomic check+set under lock)
    with _missing_fields_lock:
        if report_generation_status['in_progress']:
            return jsonify({
                'success': False,
                'error': 'Report generation already in progress. Please wait.'
            }), 429
        report_generation_status['in_progress'] = True
        report_generation_status['error'] = None

    thread = threading.Thread(target=run_missing_fields_script)
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'message': 'Report generation started',
        'status': 'in_progress'
    })

# API endpoint to check report generation status
@app.route('/api/check-missing-fields-report-status', methods=['GET'])
def check_missing_fields_report_status():
    """Check the status of missing fields report generation"""
    global report_generation_status
    
    if report_generation_status['in_progress']:
        return jsonify({
            'status': 'in_progress',
            'message': 'Report is being generated. Please wait...'
        })
    elif report_generation_status['error']:
        return jsonify({
            'status': 'error',
            'error': report_generation_status['error']
        })
    elif report_generation_status['last_generated']:
        return jsonify({
            'status': 'completed',
            'report_file': report_generation_status['last_generated'],
            'message': 'Report generated successfully'
        })
    else:
        return jsonify({
            'status': 'not_started',
            'message': 'No report has been generated yet'
        })

# API endpoint to download the missing fields HTML report to the browser client
@app.route('/api/download-missing-fields-report', methods=['GET'])
def download_missing_fields_report():
    """Serve Missing_Fields_Report.html as a file download to the requesting browser client.

    Using send_file with as_attachment=True ensures the browser's Save-As dialog
    opens on the CLIENT machine (the user's desktop), not on the server host.
    """
    try:
        from flask import send_file
        report_path = os.path.join(
            os.getcwd(), 'Missing_Filed_Report', 'Missing_Fields_Report.html'
        )

        if not os.path.exists(report_path):
            return jsonify({
                'success': False,
                'error': 'Report file not found. Please generate the report first.'
            }), 404

        # as_attachment=True forces a download on the browser/client side
        return send_file(
            report_path,
            as_attachment=True,
            download_name='Missing_Fields_Report.html',
            mimetype='text/html'
        )

    except Exception as e:
        print(f'❌ Error downloading missing fields report: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════
# MISSING DATA FOR US SCOPE REPORT APIs
# ══════════════════════════════════════════════════════════════════════════

def run_missing_scope_script(query_id=None):
    """Background thread to run the missing data scope report script."""
    global missing_scope_report_status
    try:
        print("\n" + "=" * 80)
        print("🚀 Starting Missing Data for US Scope Report Generation...")
        print("=" * 80)

        script_path = os.path.join(
            os.getcwd(), 'Missing_Data_Report_For_Scope', 'generate_missing_field_tables.py'
        )
        print(f"📝 Executing: {script_path}")
        if query_id:
            print(f"🔑 Using Query ID: {query_id}")

        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        if query_id:
            env['ADO_QUERY_ID'] = query_id

        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=300,
            env=env,
        )

        if result.returncode != 0:
            print(f"❌ Script failed:\n{result.stderr}")
            missing_scope_report_status['error'] = result.stderr or 'Script exited with non-zero status'
            missing_scope_report_status['in_progress'] = False
            return

        report_file = os.path.join(
            'Missing_Data_Report_For_Scope', 'Missing_Data_Report.html'
        )

        if os.path.exists(os.path.join(os.getcwd(), report_file)):
            print(f"✅ Report generated successfully: {report_file}")
            print("=" * 80 + "\n")
            missing_scope_report_status['last_generated'] = report_file
            missing_scope_report_status['error'] = None
        else:
            missing_scope_report_status['error'] = 'Report file not found after generation'

        missing_scope_report_status['in_progress'] = False

    except subprocess.TimeoutExpired:
        missing_scope_report_status['error'] = 'Report generation timed out (>5 minutes)'
        missing_scope_report_status['in_progress'] = False
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        missing_scope_report_status['error'] = str(e)
        missing_scope_report_status['in_progress'] = False


@app.route('/api/generate-missing-scope-report', methods=['POST'])
def generate_missing_scope_report():
    """Start the missing data scope report generation in a background thread."""
    global missing_scope_report_status

    script_path = os.path.join(
        os.getcwd(), 'Missing_Data_Report_For_Scope', 'generate_missing_field_tables.py'
    )
    if not os.path.exists(script_path):
        return jsonify({'success': False, 'error': f'Script not found: {script_path}'}), 404

    data = request.get_json(silent=True) or {}
    query_id = (data.get('query_id') or '').strip() or None

    with _missing_scope_lock:
        if missing_scope_report_status['in_progress']:
            return jsonify({
                'success': False,
                'error': 'Report generation already in progress. Please wait.'
            }), 429
        missing_scope_report_status['in_progress'] = True
        missing_scope_report_status['error'] = None

    thread = threading.Thread(target=run_missing_scope_script, args=(query_id,))
    thread.daemon = True
    thread.start()

    return jsonify({'success': True, 'message': 'Report generation started', 'status': 'in_progress'})


@app.route('/api/check-missing-scope-report-status', methods=['GET'])
def check_missing_scope_report_status():
    """Check the status of missing data scope report generation."""
    global missing_scope_report_status

    if missing_scope_report_status['in_progress']:
        return jsonify({'status': 'in_progress', 'message': 'Report is being generated. Please wait...'})
    elif missing_scope_report_status['error']:
        return jsonify({'status': 'error', 'error': missing_scope_report_status['error']})
    elif missing_scope_report_status['last_generated']:
        return jsonify({
            'status': 'completed',
            'report_file': missing_scope_report_status['last_generated'],
            'message': 'Report generated successfully',
        })
    else:
        return jsonify({'status': 'not_started', 'message': 'No report has been generated yet'})


@app.route('/api/download-missing-scope-report', methods=['GET'])
def download_missing_scope_report():
    """Serve Missing_Data_Report.html as a file download to the browser client."""
    try:
        report_path = os.path.join(
            os.getcwd(), 'Missing_Data_Report_For_Scope', 'Missing_Data_Report.html'
        )
        if not os.path.exists(report_path):
            return jsonify({
                'success': False,
                'error': 'Report file not found. Please generate the report first.'
            }), 404

        return send_file(
            report_path,
            as_attachment=True,
            download_name='Missing_Data_Report.html',
            mimetype='text/html',
        )
    except Exception as e:
        print(f'❌ Error downloading missing scope report: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500


# Check if report generation is available
@app.route('/api/check-script', methods=['GET'])
def check_script():
    """Check if the regression script exists"""
    script_path = os.path.join(os.getcwd(), 'Regression_Report', 'RegReport 1_Updated 1.py')
    exists = os.path.exists(script_path)
    return jsonify({
        'available': exists,
        'path': script_path
    })

# API endpoint to update Query ID in Missing Field Report script
@app.route('/api/update-query-id', methods=['POST'])
def update_query_id():
    """Updates the QUERY_ID in the Missing Field Report Python script"""
    try:
        print("\n" + "="*60)
        print("🔄 Updating Query ID...")
        print("="*60)
        
        data = request.get_json()
        if not data:
            print("❌ No JSON data received")
            return jsonify({'success': False, 'error': 'No data received'}), 400
        
        new_query_id = data.get('query_id', '').strip()
        print(f"📝 New Query ID: {new_query_id}")
        
        if not new_query_id:
            print("❌ Query ID is empty")
            return jsonify({'success': False, 'error': 'Query ID cannot be empty'}), 400
        
        script_path = os.path.join(os.getcwd(), 'Missing_Filed_Report', 'Generate_missing_fields_report .py')
        print(f"📁 Script path: {script_path}")
        
        if not os.path.exists(script_path):
            print(f"❌ Script not found at: {script_path}")
            return jsonify({'success': False, 'error': f'Script not found at: {script_path}'}), 404
        
        # Read the file
        print("\U0001f4d6 Reading file...")
        with _settings_file_lock:
            with open(script_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Replace the QUERY_ID line (line 12)
            lines = content.split('\n')
            print(f"\U0001f4c4 File has {len(lines)} lines")

            if len(lines) >= 12:
                print(f"\U0001f50d Line 12 before: {lines[11]}")
                if 'QUERY_ID' in lines[11]:
                    lines[11] = f'QUERY_ID = "{new_query_id}"'
                    print(f"\u270f\ufe0f Line 12 after: {lines[11]}")

                    # Write back to file
                    print("\U0001f4be Writing file...")
                    with open(script_path, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(lines))

                    print(f"\u2705 Query ID updated successfully!")
                    print("="*60 + "\n")
                    return jsonify({'success': True, 'message': 'Query ID updated successfully'})
                else:
                    print(f"\u274c QUERY_ID not found in line 12")
                    return jsonify({'success': False, 'error': f'Line 12 does not contain QUERY_ID. Found: {lines[11][:50]}'}), 500
            else:
                print(f"\u274c File has fewer than 12 lines")
                return jsonify({'success': False, 'error': f'File has only {len(lines)} lines, expected at least 12'}), 500
            
    except Exception as e:
        print(f'❌ Error updating Query ID: {str(e)}')
        print(f"Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# API endpoint to download AI Data Transfer Excel file
@app.route('/api/download-macro', methods=['GET'])
def download_macro():
    """Downloads the AI Data Transfer Excel macro file"""
    try:
        excel_file = os.path.join(os.getcwd(), 'AI_Data_Transfer', 'Regression_Data.xlsm')
        
        if not os.path.exists(excel_file):
            return jsonify({
                'success': False,
                'error': f'Excel file not found: {excel_file}'
            }), 404
        
        print(f"\n📥 Serving Excel file for download: {excel_file}")
        
        # Serve the file for download with proper headers
        from flask import send_file
        return send_file(
            excel_file,
            as_attachment=True,
            download_name='Regression_Data.xlsm',
            mimetype='application/vnd.ms-excel.sheet.macroEnabled.12'
        )
        
    except Exception as e:
        print(f"❌ Error downloading macro file: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API endpoint to run AI Data Transfer Excel macro
@app.route('/api/run-macro', methods=['GET'])
def run_macro():
    """Opens the AI Data Transfer Excel macro file from Downloads folder"""
    try:
        # Get user's Downloads folder using environment variables
        if os.name == 'nt':  # Windows
            user_profile = os.environ.get('USERPROFILE')
            downloads_folder = os.path.join(user_profile, 'Downloads')
        else:  # Mac/Linux
            downloads_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
        
        excel_file = os.path.join(downloads_folder, 'Regression_Data.xlsm')
        
        print(f"\n▶️ Looking for Excel file in: {excel_file}")
        
        if not os.path.exists(excel_file):
            error_msg = f'Excel file not found in Downloads folder. Please download the macro file first.'
            print(f"❌ {error_msg}")
            print(f"   Expected location: {excel_file}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 404
        
        print(f"✅ Found Excel file, unblocking and opening...")
        
        # Unblock the file (remove "Mark of the Web" - Zone.Identifier)
        try:
            zone_identifier_file = excel_file + ':Zone.Identifier'
            if os.path.exists(zone_identifier_file):
                os.remove(zone_identifier_file)
                print(f"   ✓ File unblocked (removed Zone.Identifier)")
            
            # Also try using PowerShell Unblock-File command
            unblock_cmd = f'powershell.exe -Command "Unblock-File -Path \'{excel_file}\'"'
            subprocess.run(unblock_cmd, shell=True, capture_output=True, timeout=5)
            print(f"   ✓ PowerShell Unblock-File executed")
        except Exception as unblock_error:
            print(f"   ⚠️ Could not unblock file: {unblock_error}")
            # Continue anyway - file might not need unblocking
        
        # Open the file with the default application (Excel)
        if os.name == 'nt':  # Windows
            os.startfile(excel_file)
        else:  # Mac/Linux
            subprocess.run(['xdg-open', excel_file])
        
        print(f"✅ Excel file opened successfully from Downloads")
        
        return jsonify({
            'success': True,
            'message': 'Excel file opened successfully (file has been unblocked)',
            'file_path': excel_file
        })
        
    except Exception as e:
        print(f"❌ Error opening macro file: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API endpoint to get regression report configuration
@app.route('/api/get-regression-config', methods=['GET'])
def get_regression_config():
    """Read current configuration from RegReport Python script"""
    try:
        script_path = os.path.join(os.getcwd(), 'Regression_Report', 'RegReport 1_Updated 1.py')
        
        if not os.path.exists(script_path):
            return jsonify({
                'success': False,
                'error': 'Configuration file not found'
            }), 404
        
        # Read the Python file
        with open(script_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract configuration values using simple string parsing
        config = {}
        
        # Find the ADO_CONFIG dictionary section
        import re
        
        # Extract each field value
        patterns = {
            'organization': r"'organization':\s*'([^']*)'",
            'project': r"'project':\s*'([^']*)'",
            'plan_id': r"'plan_id':\s*'([^']*)'",
            'suite_id': r"'suite_id':\s*'([^']*)'",
            'insprint_suite_id': r"'insprint_suite_id':\s*'([^']*)'",
            'target_suite_name': r"'target_suite_name':\s*'([^']*)'",
            'insprint_suite_name': r"'insprint_suite_name':\s*'([^']*)'",
            'pat_token': r"'pat_token':\s*'([^']*)'",
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, content)
            if match:
                config[key] = match.group(1)
            else:
                config[key] = ''
        
        print(f"📖 Retrieved configuration from RegReport script")
        
        return jsonify({
            'success': True,
            'config': config
        })
        
    except Exception as e:
        print(f"❌ Error reading config: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API endpoint to update regression report configuration
@app.route('/api/update-regression-config', methods=['POST'])
def update_regression_config():
    """Update configuration in RegReport Python script"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400
        
        script_path = os.path.join(os.getcwd(), 'Regression_Report', 'RegReport 1_Updated 1.py')
        
        if not os.path.exists(script_path):
            return jsonify({
                'success': False,
                'error': 'Configuration file not found'
            }), 404
        
        with _settings_file_lock:
            # Read the current file
            with open(script_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Update each field in ADO_CONFIG
            import re

            # Define fields to update
            fields = {
                'plan_id': data.get('plan_id'),
                'suite_id': data.get('suite_id'),
                'insprint_suite_id': data.get('insprint_suite_id'),
                'target_suite_name': data.get('target_suite_name'),
                'insprint_suite_name': data.get('insprint_suite_name'),
                'pat_token': data.get('pat_token'),
            }

            # Update each field value
            for key, value in fields.items():
                if value is not None:  # Only update if value is provided
                    # Pattern to match the field in ADO_CONFIG
                    pattern = rf"('{key}':\s*')([^']*)(')"
                    replacement = rf"\g<1>{value}\g<3>"
                    content = re.sub(pattern, replacement, content)

            # Create backup before writing
            backup_path = script_path + '.backup'
            with open(backup_path, 'w', encoding='utf-8') as f:
                with open(script_path, 'r', encoding='utf-8') as original:
                    f.write(original.read())

            # Write the updated content
            with open(script_path, 'w', encoding='utf-8') as f:
                f.write(content)
        
        print(f"✅ Configuration updated successfully")
        print(f"   Backup saved to: {backup_path}")
        
        return jsonify({
            'success': True,
            'message': 'Configuration updated successfully',
            'backup_created': True
        })
        
    except Exception as e:
        print(f"❌ Error updating config: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ── Supabase helpers for attendance ──────────────────────────────────────────

def _supabase_save_attendance(entries, lead_name, user_id, year, month, month_name, client_ip):
    """Upsert attendance records and write a log row to Supabase.

    Called after the Excel save succeeds so Supabase mirrors the authoritative data.
    Errors here are logged but do not fail the HTTP response.
    """
    # Generated by GitHub Copilot
    if not SUPABASE_ENABLED:
        return
    try:
        days_in_month = calendar.monthrange(year, month)[1]
        upsert_rows = []
        log_rows = []
        saved_at = datetime.now().isoformat()

        for entry in entries:
            member_name = (entry.get('member_name') or '').strip()
            location    = (entry.get('location') or '').strip()
            day         = entry.get('day')
            status      = (entry.get('status') or '').strip()

            if not member_name or not isinstance(day, int) or day < 1 or day > days_in_month:
                continue

            upsert_rows.append({
                'member_name': member_name,
                'lead_name':   lead_name,
                'location':    location,
                'year':        year,
                'month':       month,
                'day':         day,
                'status':      status,
                'updated_at':  saved_at,
            })
            log_rows.append({
                'saved_at':   saved_at,
                'user_id':    user_id,
                'lead_name':  lead_name,
                'member_name': member_name,
                'location':   location,
                'month_name': month_name,
                'year':       year,
                'day':        day,
                'new_value':  status,
                'changed':    'YES',
                'client_ip':  client_ip,
            })

        if upsert_rows:
            # Upsert in batches of 100 to avoid large request payloads
            batch_size = 100
            for i in range(0, len(upsert_rows), batch_size):
                _supabase.table('attendance_records').upsert(
                    upsert_rows[i:i + batch_size],
                    on_conflict='member_name,year,month,day'
                ).execute()

        if log_rows:
            for i in range(0, len(log_rows), batch_size):
                _supabase.table('attendance_logs').insert(log_rows[i:i + batch_size]).execute()

        print(f"✅ Supabase: upserted {len(upsert_rows)} attendance records")
    except Exception as e:
        print(f"⚠️  Supabase attendance save error (non-fatal): {e}")


def _supabase_load_attendance(lead_name, year, month):
    """Return attendance dict from Supabase or None if unavailable.

    Format matches the Excel loader: { 'MemberName|Location': { '1': 'P', ... } }
    """
    # Generated by GitHub Copilot
    if not SUPABASE_ENABLED:
        return None
    try:
        result = (
            _supabase.table('attendance_records')
            .select('member_name,location,day,status')
            .eq('year', year)
            .eq('month', month)
            .execute()
        )
        rows = result.data or []
        if not rows:
            return None

        attendance = {}
        for row in rows:
            key = f"{row['member_name']}|{row.get('location', '')}"
            if key not in attendance:
                attendance[key] = {}
            attendance[key][str(row['day'])] = row.get('status', '')

        return attendance
    except Exception as e:
        print(f"⚠️  Supabase attendance load error (non-fatal): {e}")
        return None


@app.route('/api/attendance/save-master-excel', methods=['POST'])
def save_attendance_to_master_excel():
    """In-place update of Master_Attendance.xlsx — finds existing member rows/day columns and writes only the changed cells."""
    print("\n" + "="*80)
    print("🔵 SAVE ENDPOINT CALLED (NEW VALIDATION CODE - VERSION 2.0)")
    print("="*80)
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        lead_name  = (payload.get('lead_name') or '').strip()
        # Capture Windows system username of whoever is running the server
        user_id = (payload.get('user_id') or '').strip()
        if not user_id:
            user_id = (
                os.environ.get('USERNAME') or
                os.environ.get('USER') or
                os.environ.get('LOGNAME') or
                request.remote_addr or
                'unknown'
            )
        year       = payload.get('year')
        month      = payload.get('month')
        entries    = payload.get('entries', [])
        
        print(f"📥 Request data: Lead={lead_name}, Year={year}, Month={month}, Entries={len(entries)}")

        if not lead_name:
            return jsonify({'success': False, 'error': 'lead_name is required'}), 400
        if not isinstance(year, int) or not isinstance(month, int):
            return jsonify({'success': False, 'error': 'year and month must be integers'}), 400
        if month < 1 or month > 12:
            return jsonify({'success': False, 'error': 'month must be between 1 and 12'}), 400
        if not isinstance(entries, list):
            return jsonify({'success': False, 'error': 'entries must be an array'}), 400

        os.makedirs(os.path.dirname(MASTER_ATTENDANCE_FILE), exist_ok=True)

        month_name    = calendar.month_name[month]
        report_label  = f'Attendance_{month_name}_{year}'
        # Sheet name uses spaces to match format: "Attendance January 2026"
        sheet_name    = f'Attendance {month_name} {year}'[:31]
        # Legacy underscore variant (from previous runs)
        sheet_name_legacy = re.sub(r'[\\/*?:\[\]]', '_', f'Attendance_{month_name}_{year}').strip()[:31]

        days_in_month = calendar.monthrange(year, month)[1]
        saved_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ══════════════════════════════════════════════════════════════════════
        # ── VALIDATION: Check file and sheet exist BEFORE any processing ─────
        # ══════════════════════════════════════════════════════════════════════
        
        # Step 1: Check if Master_Attendance.xlsx exists
        if not os.path.exists(MASTER_ATTENDANCE_FILE):
            error_msg = f'Master_Attendance.xlsx file does not exist. Please use the "Create Master Sheet" button to create a new sheet for {month_name} {year}.'
            print(f"\n❌ SAVE BLOCKED: File does not exist")
            print(f"   📁 Required file: {MASTER_ATTENDANCE_FILE}")
            print(f"   📋 Action: Click 'Create Master Sheet' button")
            
            return jsonify({
                'success': False,
                'error': error_msg,
                'sheet_name': sheet_name,
                'month': month_name,
                'year': year,
                'require_creation': True,
                'file_exists': False,
                'sheet_exists': False
            }), 400
        
        # Step 2: Load workbook and check if sheet exists
        try:
            workbook = load_workbook(MASTER_ATTENDANCE_FILE)
        except Exception as e:
            error_msg = f'Failed to load Master_Attendance.xlsx: {str(e)}'
            print(f"\n❌ SAVE BLOCKED: Cannot load file")
            print(f"   Error: {str(e)}")
            
            return jsonify({
                'success': False,
                'error': error_msg,
                'sheet_name': sheet_name,
                'month': month_name,
                'year': year,
                'require_creation': False,
                'file_exists': True,
                'sheet_exists': False
            }), 500
        
        # Step 3: Check if the required month sheet exists (check both formats)
        actual_sheet_name = None
        if sheet_name in workbook.sheetnames:
            actual_sheet_name = sheet_name
        elif sheet_name_legacy in workbook.sheetnames:
            actual_sheet_name = sheet_name_legacy
        
        if actual_sheet_name is None:
            # Sheet does NOT exist - block save
            error_msg = f'Sheet for {month_name} {year} does not exist in Master_Attendance.xlsx. Please use the "Create Master Sheet" button first.'
            print(f"\n❌ SAVE BLOCKED: Sheet not found")
            print(f"   📋 Required sheet: {sheet_name}")
            print(f"   📎 Available sheets: {', '.join(workbook.sheetnames)}")
            print(f"   📋 Action: Click 'Create Master Sheet' button")
            
            return jsonify({
                'success': False,
                'error': error_msg,
                'sheet_name': sheet_name,
                'month': month_name,
                'year': year,
                'require_creation': True,
                'file_exists': True,
                'sheet_exists': False,
                'available_sheets': workbook.sheetnames
            }), 400
        
        # ══════════════════════════════════════════════════════════════════════
        # ── VALIDATION PASSED: Sheet exists, proceed with update ─────────────
        # ══════════════════════════════════════════════════════════════════════
        
        print(f"\n✅ Validation passed - proceeding with save")
        print(f"   📋 Sheet: {actual_sheet_name}")
        print(f"   📅 Month: {month_name} {year}")
        
        # Get the sheet (use whichever format exists)
        sheet = workbook[actual_sheet_name]
        
        # If using legacy format, rename it to canonical format for consistency
        if actual_sheet_name == sheet_name_legacy and actual_sheet_name != sheet_name:
            try:
                sheet.title = sheet_name
                actual_sheet_name = sheet_name
                print(f"   ℹ️  Renamed legacy sheet to: {sheet_name}")
            except Exception as e:
                print(f"   ⚠️  Could not rename sheet: {e}")
        
        # Remove any duplicate legacy sheet
        if sheet_name in workbook.sheetnames and sheet_name_legacy in workbook.sheetnames and sheet_name != sheet_name_legacy:
            try:
                workbook.remove(workbook[sheet_name_legacy])
                print(f"   ℹ️  Removed duplicate legacy sheet")
            except Exception as e:
                print(f"   ⚠️  Could not remove duplicate: {e}")
        # Unmerge all merged cells so we can write freely
        merged_ranges = list(sheet.merged_cells.ranges)
        for mr in merged_ranges:
            sheet.unmerge_cells(str(mr))
        # Update metadata rows
        sheet.cell(row=2, column=1, value=f'Latest Update Source: {lead_name}')
        sheet.cell(row=3, column=1, value=f'Last Saved At: {saved_timestamp}')
        
        # Find header row - look for "Team Member Names" in column 1
        # In the new format, Row 1 has the main headers, and data starts at row 4
        header_row_idx = None
        for row_idx in range(1, min(sheet.max_row, 15) + 1):
            c1 = sheet.cell(row=row_idx, column=1).value
            if c1 and str(c1).strip().lower() in ['team member', 'team member names']:
                # This is row 1 with main headers, so data starts 3 rows later
                header_row_idx = row_idx + 2  # If row 1 has headers, data starts at row 4, so header_row_idx = 3
                break
        if header_row_idx is None:
            # Fallback: treat row 3 as the last metadata row
            header_row_idx = 3

        # ── Build member-name → row-index map from existing data rows ────────
        # Key is lowercase name only — location formats may differ between Excel and HTML
        member_row_map = {}   # key: name.lower().strip() → row_idx
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            name = sheet.cell(row=row_idx, column=1).value
            if name and str(name).strip():
                # Skip any row that looks like a metadata/header row
                name_str = str(name).strip()
                if name_str.lower().startswith('team member') or name_str.lower().startswith('latest update') or name_str.lower().startswith('last saved') or name_str.lower().startswith('team attendance'):
                    continue
                # Keep only the first occurrence; skip duplicates
                if name_str.lower() not in member_row_map:
                    member_row_map[name_str.lower()] = row_idx

        # ── Group incoming entries by member ─────────────────────────────────
        incoming_members = {}
        for entry in entries:
            member_name = (entry.get('member_name') or '').strip()
            location    = (entry.get('location')    or '').strip()
            day         = entry.get('day')
            status      = (entry.get('status')      or '').strip()

            if not member_name or not isinstance(day, int) or day < 1 or day > days_in_month:
                continue

            key = member_name.lower()   # match by name only
            if key not in incoming_members:
                incoming_members[key] = {'member_name': member_name, 'location': location, 'days': {}}
            incoming_members[key]['days'][day] = status

        # ── In-place cell updates ─────────────────────────────────────────────
        members_updated   = 0
        non_empty_count   = 0
        all_cell_changes  = []   # cell-level audit trail

        for key, data in incoming_members.items():
            member_name = data['member_name']
            location    = data['location']

            if key not in member_row_map:
                # Truly new member not in sheet yet — append a new row
                new_row_idx = sheet.max_row + 1
                sheet.cell(row=new_row_idx, column=1, value=member_name)
                sheet.cell(row=new_row_idx, column=2, value=lead_name)
                sheet.cell(row=new_row_idx, column=3, value=location)
                member_row_map[key] = new_row_idx

            row_idx = member_row_map[key]
            members_updated += 1
            
            # Update lead name and location for existing members
            sheet.cell(row=row_idx, column=2, value=lead_name)
            sheet.cell(row=row_idx, column=3, value=location)

            # Write only the specific day cells that were sent
            cell_changes = []   # collect (member, day, old, new) for the log
            for day, status in data['days'].items():
                col_idx = day + 3
                old_val = (sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                new_val = status
                cell = sheet.cell(row=row_idx, column=col_idx, value=new_val)
                cell.fill = get_fill(new_val)
                if new_val:
                    non_empty_count += 1
                cell_changes.append((member_name, location, day, old_val, new_val))

            # Recalculate all summary columns for this row
            # MH (Mandatory Holiday) and OP (Optional Holiday) are NOT counted as leaves
            HOLIDAY_CODES = {'MH', 'OP'}
            present_count  = 0
            total_leaves   = 0
            unplanned      = 0
            sick_leave     = 0
            for d in range(1, days_in_month + 1):
                val = (sheet.cell(row=row_idx, column=d + 3).value or '').strip().upper()
                if val in ('P', 'C'):
                    present_count += 1
                elif val and val not in HOLIDAY_CODES:  # non-empty, not present/comp-off, not a holiday
                    total_leaves += 1
                if val == 'UP':
                    unplanned += 1
                if val == 'SL':
                    sick_leave += 1
            sheet.cell(row=row_idx, column=days_in_month + 4, value=present_count)
            sheet.cell(row=row_idx, column=days_in_month + 5, value=total_leaves)
            sheet.cell(row=row_idx, column=days_in_month + 6, value=unplanned)
            sheet.cell(row=row_idx, column=days_in_month + 7, value=sick_leave)

            all_cell_changes.extend(cell_changes)

        rows_written = len(member_row_map)

        LOG_HEADERS = [
            'Timestamp', 'Date', 'Time', 'User ID', 'Lead Name',
            'Member Name', 'Location', 'Month', 'Year', 'Day',
            'Old Value', 'New Value', 'Changed', 'Sheet Name', 'Client IP'
        ]
        save_date  = datetime.now().strftime('%Y-%m-%d')
        save_time  = datetime.now().strftime('%H:%M:%S')
        client_ip  = request.remote_addr

        # ── Append a log row ──────────────────────────────────────────────────
        if os.path.exists(MASTER_ATTENDANCE_LOG_FILE):
            try:
                logs_workbook = load_workbook(MASTER_ATTENDANCE_LOG_FILE)
                logs_sheet = logs_workbook['Logs'] if 'Logs' in logs_workbook.sheetnames else logs_workbook.active
                # Ensure headers match new format
                existing_header = [logs_sheet.cell(row=1, column=c).value for c in range(1, len(LOG_HEADERS)+1)]
                if existing_header != LOG_HEADERS:
                    # Rename old sheet and create fresh one
                    logs_sheet.title = 'Logs_Old'
                    logs_sheet = logs_workbook.create_sheet('Logs')
                    for ci, h in enumerate(LOG_HEADERS, 1):
                        logs_sheet.cell(row=1, column=ci, value=h)
            except Exception:
                os.remove(MASTER_ATTENDANCE_LOG_FILE)
                logs_workbook = Workbook()
                logs_sheet = logs_workbook.active
                logs_sheet.title = 'Logs'
                for ci, h in enumerate(LOG_HEADERS, 1):
                    logs_sheet.cell(row=1, column=ci, value=h)
        else:
            logs_workbook = Workbook()
            logs_sheet = logs_workbook.active
            logs_sheet.title = 'Logs'
            for ci, h in enumerate(LOG_HEADERS, 1):
                logs_sheet.cell(row=1, column=ci, value=h)

        # One row per changed cell
        for (m_name, m_loc, day, old_val, new_val) in all_cell_changes:
            changed_flag = 'YES' if old_val != new_val else 'NO'
            logs_sheet.append([
                saved_timestamp, save_date, save_time, user_id, lead_name,
                m_name, m_loc, month_name, year, day,
                old_val, new_val, changed_flag, sheet_name, client_ip
            ])

        # If nothing changed at cell level, still write one summary row
        if not all_cell_changes:
            logs_sheet.append([
                saved_timestamp, save_date, save_time, user_id, lead_name,
                '(full sync)', '', month_name, year, '',
                '', '', 'SYNC', sheet_name, client_ip
            ])

        workbook.save(MASTER_ATTENDANCE_FILE)
        logs_workbook.save(MASTER_ATTENDANCE_LOG_FILE)

        # Mirror to Supabase (non-blocking, errors are logged but not fatal)
        _supabase_save_attendance(entries, lead_name, user_id, year, month, month_name, client_ip)

        print(f"\n✅ Saved to Master Excel:")
        print(f"   Sheet: {actual_sheet_name}")
        print(f"   Members updated: {members_updated}")
        print(f"   Total rows: {rows_written}")
        print(f"   Non-empty cells: {non_empty_count}")

        return jsonify({
            'success': True,
            'message': 'Master Excel updated in-place successfully',
            'file_path': MASTER_ATTENDANCE_FILE,
            'sheet_name': actual_sheet_name,
            'report_label': report_label,
            'members_updated': members_updated,
            'rows_written': rows_written,
            'logs_file_path': MASTER_ATTENDANCE_LOG_FILE
        })

    except Exception as e:
        print(f"❌ Error saving attendance to master Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/attendance/create-master-sheet', methods=['POST'])
def create_master_sheet():
    """Create a new master sheet with all team members if it doesn't exist"""
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        year = payload.get('year')
        month = payload.get('month')

        if not isinstance(year, int) or not isinstance(month, int):
            return jsonify({'success': False, 'error': 'year and month must be integers'}), 400
        if month < 1 or month > 12:
            return jsonify({'success': False, 'error': 'month must be between 1 and 12'}), 400

        # File paths
        attendance_dir = os.path.join(os.path.dirname(__file__), 'Attendance')
        master_file = os.path.join(attendance_dir, 'Master_Attendance.xlsx')
        team_details_file = os.path.join(attendance_dir, 'Team Details.xlsx')

        # Check if Team Details file exists
        if not os.path.exists(team_details_file):
            return jsonify({
                'success': False,
                'error': f'Team Details.xlsx not found at {team_details_file}'
            }), 500

        # Generate sheet name
        month_name = calendar.month_name[month]
        sheet_name = f'Attendance {month_name} {year}'[:31]
        days_in_month = calendar.monthrange(year, month)[1]

        # Load or create workbook
        os.makedirs(attendance_dir, exist_ok=True)
        if os.path.exists(master_file):
            try:
                workbook = load_workbook(master_file)
            except Exception:
                os.remove(master_file)
                workbook = Workbook()
        else:
            workbook = Workbook()

        # Remove blank default sheet only when workbook is brand-new
        if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) == 1:
            default_sheet = workbook['Sheet']
            if default_sheet.max_row <= 1 and default_sheet['A1'].value is None:
                workbook.remove(default_sheet)

        # Check if sheet already exists
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            member_count = sheet.max_row - 3  # Exclude 3 header rows
            return jsonify({
                'success': True,
                'already_exists': True,
                'sheet_name': sheet_name,
                'member_count': member_count
            })

        # Load team details
        import pandas as pd
        df_team = pd.read_excel(team_details_file)
        
        # Create member-to-lead mapping
        team_members = []
        lead_names_set = set()
        for _, row in df_team.iterrows():
            if pd.notna(row['Team members']):
                member_name = str(row['Team members']).strip()
                lead_name = str(row['Lead']).strip() if pd.notna(row['Lead']) else ''
                location = str(row['Location']).strip() if pd.notna(row.get('Location', '')) else ''
                team_members.append({
                    'name': member_name,
                    'lead': lead_name,
                    'location': location
                })
                if lead_name:
                    lead_names_set.add(lead_name)

        # Sort team members by lead name, then by member name
        team_members.sort(key=lambda x: (x['lead'], x['name']))

        # Define colorful styles for attractive formatting
        # Header row 1 style - Vibrant purple gradient
        header_fill = PatternFill(start_color='7c3aed', end_color='a78bfa', fill_type='solid')
        header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Row 2 style - Bright indigo
        row2_fill = PatternFill(start_color='4f46e5', end_color='6366f1', fill_type='solid')
        row2_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        row2_alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 3 style - Cyan/turquoise
        row3_fill = PatternFill(start_color='06b6d4', end_color='22d3ee', fill_type='solid')
        row3_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        row3_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data row styles - Alternating colors
        even_row_fill = PatternFill(start_color='f0f9ff', end_color='e0f2fe', fill_type='solid')  # Light blue
        odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')   # White
        data_font = Font(name='Segoe UI', size=10, color='1f2937')
        data_alignment = Alignment(horizontal='left', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Border style - Thin gray borders
        thin_border = Border(
            left=Side(style='thin', color='cbd5e1'),
            right=Side(style='thin', color='cbd5e1'),
            top=Side(style='thin', color='cbd5e1'),
            bottom=Side(style='thin', color='cbd5e1')
        )
        
        # Create new sheet with standardized format
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Calculate total columns
        total_cols = days_in_month + 7  # Name, Lead, Location, Days, 4 summary columns
        
        # Row 1: Main column headers
        sheet.cell(row=1, column=1, value='Team Member Names')
        sheet.cell(row=1, column=2, value='Lead Name')
        
        # Apply styling to Row 1
        for col in range(1, total_cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Row 2: Metadata + Location header + Date numbers
        sheet.cell(row=2, column=1, value='Created via UI')
        sheet.cell(row=2, column=3, value='📍 Location')
        for d in range(1, days_in_month + 1):
            cell = sheet.cell(row=2, column=d + 3, value=d)
            cell.fill = row2_fill
            cell.font = row2_font
            cell.alignment = row2_alignment
            cell.border = thin_border
        
        # Summary column headers in Row 2
        summary_headers = ['Total Present', 'Total Leaves', 'UnPlanned', 'SickLeave']
        for idx, header in enumerate(summary_headers):
            cell = sheet.cell(row=2, column=days_in_month + 4 + idx, value=header)
            cell.fill = row2_fill
            cell.font = row2_font
            cell.alignment = row2_alignment
            cell.border = thin_border
        
        # Apply styling to Row 2 columns 1-2
        for col in [1, 2, 3]:
            cell = sheet.cell(row=2, column=col)
            cell.fill = row2_fill
            cell.font = row2_font
            cell.alignment = row2_alignment
            cell.border = thin_border
        
        # Row 3: Timestamp + Day names
        saved_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=3, column=1, value=f'Created At: {saved_timestamp}')
        import calendar as cal_module
        for d in range(1, days_in_month + 1):
            date_obj = datetime(year, month, d)
            day_name = cal_module.day_abbr[date_obj.weekday()]
            cell = sheet.cell(row=3, column=d + 3, value=day_name)
            cell.fill = row3_fill
            cell.font = row3_font
            cell.alignment = row3_alignment
            cell.border = thin_border
        
        # Apply styling to Row 3 columns 1-3 and summary columns
        for col in range(1, total_cols + 1):
            cell = sheet.cell(row=3, column=col)
            if not cell.value:  # If no value set yet
                cell.value = ''
            cell.fill = row3_fill
            cell.font = row3_font
            cell.alignment = row3_alignment
            cell.border = thin_border
        
        # Set column widths
        sheet.column_dimensions['A'].width = 30.0
        sheet.column_dimensions['B'].width = 25.0
        sheet.column_dimensions['C'].width = 22.0
        for d in range(1, days_in_month + 1):
            col_letter = sheet.cell(row=1, column=d + 3).column_letter
            sheet.column_dimensions[col_letter].width = 8.0
        for offset in range(4, 8):
            col_letter = sheet.cell(row=1, column=days_in_month + offset).column_letter
            sheet.column_dimensions[col_letter].width = 12.0
        
        # Freeze panes at D4 (row 4, column 4) - keeps headers and name columns visible
        sheet.freeze_panes = 'D4'

        # Add all team members starting from row 4 with alternating colors
        for idx, member in enumerate(team_members, start=4):
            # Determine row color (alternating)
            row_fill = even_row_fill if (idx % 2 == 0) else odd_row_fill
            
            # Add member data
            cell_name = sheet.cell(row=idx, column=1, value=member['name'])
            cell_lead = sheet.cell(row=idx, column=2, value=member['lead'])
            cell_location = sheet.cell(row=idx, column=3, value=member['location'])
            
            # Style the first 3 columns
            for cell in [cell_name, cell_lead, cell_location]:
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
            
            # Add empty cells for date columns with styling
            for d in range(1, days_in_month + 1):
                cell = sheet.cell(row=idx, column=d + 3, value='')
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Add empty cells for summary columns with styling
            for offset in range(4, 8):
                cell = sheet.cell(row=idx, column=days_in_month + offset, value='')
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border

        # Save the workbook
        workbook.save(master_file)

        print(f"\n✅ Created Master Sheet:")
        print(f"   Sheet: {sheet_name}")
        print(f"   Team members: {len(team_members)}")
        print(f"   Unique leads: {len(lead_names_set)}")

        return jsonify({
            'success': True,
            'already_exists': False,
            'sheet_name': sheet_name,
            'member_count': len(team_members),
            'lead_count': len(lead_names_set)
        })

    except Exception as e:
        print(f"❌ Error creating master sheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/attendance/load-master-excel', methods=['POST'])
def load_attendance_from_master_excel():
    """Load attendance data for a specific lead/month/year.

    Tries Supabase first; falls back to Master_Attendance.xlsx on failure.
    """
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        lead_name = (payload.get('lead_name') or '').strip()
        year  = payload.get('year')
        month = payload.get('month')

        if not isinstance(year, int) or not isinstance(month, int):
            return jsonify({'success': False, 'error': 'year and month must be integers'}), 400
        if month < 1 or month > 12:
            return jsonify({'success': False, 'error': 'month must be between 1 and 12'}), 400

        month_name = calendar.month_name[month]
        sheet_name = f'Attendance {month_name} {year}'[:31]

        # ── Try Supabase first ────────────────────────────────────────────
        sb_attendance = _supabase_load_attendance(lead_name, year, month)
        if sb_attendance is not None:
            print(f"✅ Supabase: loaded {len(sb_attendance)} member records for {month_name} {year}")
            return jsonify({
                'success': True,
                'has_data': bool(sb_attendance),
                'attendance': sb_attendance,
                'last_update_time': None,
                'sheet_name': sheet_name,
                'month': month,
                'year': year,
                'source': 'supabase',
            })

        # ── Fall back to Excel (with retry logic) ─────────────────────────
    except Exception as _pre_err:
        print(f"⚠️  Pre-load error: {_pre_err}")
        return jsonify({'success': False, 'error': str(_pre_err)}), 500

    max_retries = 5
    retry_delay = 1.0  # seconds
    workbook = None

    for attempt in range(max_retries):
        try:
            payload = request.get_json()
            lead_name = (payload.get('lead_name') or '').strip()
            year  = payload.get('year')
            month = payload.get('month')
            month_name = calendar.month_name[month]
            sheet_name = f'Attendance {month_name} {year}'[:31]

            # Check if Master Excel file exists
            if not os.path.exists(MASTER_ATTENDANCE_FILE):
                return jsonify({
                    'success': True,
                    'has_data': False,
                    'message': 'No master attendance file found',
                    'attendance': {}
                })

            sheet_name_legacy = re.sub(r'[\\/*?:\[\]]', '_', f'Attendance_{month_name}_{year}').strip()[:31]

            print(f"\n🔍 Loading attendance from Excel for {month_name} {year} (attempt {attempt + 1}/{max_retries})")
            print(f"   Looking for sheet: '{sheet_name}'")

            # Add delay on retry to let file operations complete
            if attempt > 0:
                delay = retry_delay * attempt
                print(f"   ⏳ Waiting {delay}s before retry...")
                time.sleep(delay)
            
            try:
                workbook = load_workbook(MASTER_ATTENDANCE_FILE, read_only=True, data_only=True)
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"   ⚠️  File busy, retrying... ({str(e)[:50]})")
                    workbook = None
                    continue
                else:
                    print(f"❌ Failed to open file after {max_retries} attempts: {str(e)}")
                    return jsonify({'success': False, 'error': f'Failed to open master file: {str(e)}'}), 500

            print(f"   Available sheets: {workbook.sheetnames}")

            # Try to find the sheet (try both naming conventions)
            sheet = None
            actual_sheet_name = None
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                actual_sheet_name = sheet_name
                print(f"   ✓ Found sheet: '{sheet_name}'")
            elif sheet_name_legacy in workbook.sheetnames:
                sheet = workbook[sheet_name_legacy]
                actual_sheet_name = sheet_name_legacy
                print(f"   ✓ Found legacy sheet: '{sheet_name_legacy}'")

            if not sheet:
                print(f"   ✗ Sheet not found")
                workbook.close()
                return jsonify({
                    'success': True,
                    'has_data': False,
                    'message': f'No sheet found for {month_name} {year}',
                    'attendance': {}
                })

            print(f"   Sheet has {sheet.max_row} rows")

            # Find header row - support multiple formats
            header_row_idx = None
            data_starts_at_col_3 = False  # Flag for new format where location is in column 3
            
            for row_idx in range(1, min(sheet.max_row, 15) + 1):
                c1 = sheet.cell(row=row_idx, column=1).value
                c2 = sheet.cell(row=row_idx, column=2).value
                c3 = sheet.cell(row=row_idx, column=3).value if row_idx >= 2 else None
                
                print(f"   Row {row_idx}: '{c1}' | '{c2}' | '{c3}'")
                
                if c1 and c2:
                    c1_str = str(c1).strip().lower()
                    c2_str = str(c2).strip().lower()
                    
                    # New format: "Team Member Names" in row 1, then data in col 1,2,3
                    # Row 1 has: Team Member Names | Lead Name
                    # Row 2 has: Created via UI | ... | 📍 Location | dates...
                    # Row 3 has: Created At: timestamp | day names...
                    # Row 4+: data
                    if 'team member' in c1_str and ('lead' in c2_str or c2_str == ''):
                        # Check if row 2 has location in column 3
                        if row_idx + 1 <= sheet.max_row:
                            row2_c3 = sheet.cell(row=row_idx + 1, column=3).value
                            if row2_c3 and 'location' in str(row2_c3).lower():
                                header_row_idx = row_idx + 2  # Data starts after row 3 (Created At)
                                data_starts_at_col_3 = True
                                print(f"   ✓ Found new format header at row {row_idx}, data starts at row {header_row_idx + 1}")
                                break
                    
                    # Old format: "Team Member" and "Location" in same row
                    if c1_str == 'team member' and c2_str == 'location':
                        header_row_idx = row_idx
                        print(f"   ✓ Found old format header at row {row_idx}")
                        break
                    
                    # Legacy format detection (no header row, data starts directly)
                    if row_idx >= 3 and c1_str not in ['team attendance', 'latest update', 'last saved', 'created at']:
                        if c2 and (',' in c2_str or 'bengaluru' in c2_str or 'chennai' in c2_str or 
                                   'hyderabad' in c2_str or 'mumbai' in c2_str):
                            header_row_idx = row_idx - 1
                            print(f"   ⚠️  Legacy format detected, data starts at row {row_idx}")
                            break

            # If no header row found, the sheet might be empty (no data saved yet)
            if header_row_idx is None:
                print(f"   ✗ No header row found - sheet may be newly created with no data")
                workbook.close()
                return jsonify({
                    'success': True,
                    'has_data': False,
                    'message': f'No attendance data saved yet for {month_name} {year}. Save some data first.',
                    'attendance': {}
                })

            # Read days_in_month from header row to determine columns
            days_in_month = calendar.monthrange(year, month)[1]
            
            # Build attendance data: { "member_name|location" : { "1": "P", "2": "VG", ... } }
            attendance = {}
            last_update_time = None
            
            # Try to read last update time from row 3
            try:
                last_saved_cell = sheet.cell(row=3, column=1).value
                if last_saved_cell and 'Last Saved At:' in str(last_saved_cell):
                    last_update_time = str(last_saved_cell).replace('Last Saved At:', '').strip()
            except:
                pass

            # Read all data into memory at once to avoid lazy loading issues
            print(f"   📥 Reading all data into memory...")
            print(f"   Format: {'New (Name|Lead|Location|Days)' if data_starts_at_col_3 else 'Old (Name|Location|Days)'}")
            
            all_rows = []
            try:
                # Adjust column count based on format
                max_col = days_in_month + 3 if data_starts_at_col_3 else days_in_month + 2
                for row in sheet.iter_rows(min_row=header_row_idx + 1, max_row=sheet.max_row,
                                          min_col=1, max_col=max_col, values_only=True):
                    all_rows.append(row)
            except Exception as e:
                raise Exception(f"Failed to read sheet data: {str(e)}")
            
            workbook.close()
            print(f"   ✓ Read {len(all_rows)} rows into memory, workbook closed")

            data_rows = 0
            for row_data in all_rows:
                if not row_data or not row_data[0]:
                    continue
                
                # Extract member info based on format
                if data_starts_at_col_3:
                    # New format: Name | Lead | Location | Days...
                    member_name = str(row_data[0]).strip()
                    lead_name = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ''
                    location = str(row_data[2]).strip() if len(row_data) > 2 and row_data[2] else ''
                    day_start_col = 3  # Days start at column index 3
                else:
                    # Old format: Name | Location | Days...
                    member_name = str(row_data[0]).strip()
                    location = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ''
                    day_start_col = 2  # Days start at column index 2
                
                # Skip empty member names or header-like rows
                if not member_name or member_name.lower() in ['team member', 'team member names', 'created at', 'created via ui']:
                    continue
                
                member_key = f"{member_name}|{location}"
                attendance[member_key] = {}
                
                for day in range(1, days_in_month + 1):
                    col_idx = day_start_col + day - 1  # Calculate correct column index
                    status = row_data[col_idx] if len(row_data) > col_idx else None
                    attendance[member_key][str(day)] = str(status).strip() if status else ''
                
                data_rows += 1

            print(f"   ✓ Processed {data_rows} member records with attendance data")
            
            return jsonify({
                'success': True,
                'has_data': True,
                'attendance': attendance,
                'last_update_time': last_update_time,
                'sheet_name': sheet_name,
                'month': month,
                'year': year
            })

        except Exception as e:
            if attempt < max_retries - 1:
                print(f"   ⚠️  Error on attempt {attempt + 1}, retrying... ({str(e)[:50]})")
                try:
                    if workbook:
                        workbook.close()
                except:
                    pass
                workbook = None
                continue
            else:
                print(f"❌ Error loading attendance from master Excel: {str(e)}")
                import traceback
                traceback.print_exc()
                try:
                    if workbook:
                        workbook.close()
                except:
                    pass
                return jsonify({'success': False, 'error': str(e)}), 500
    
    # If we get here, all retries failed
    return jsonify({'success': False, 'error': 'All retry attempts failed'}), 500


# Serve static files (must be last to not interfere with API routes)
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    # Don't serve API paths as static files
    if path.startswith('api/'):
        return jsonify({'error': 'API route not found'}), 404
    return send_from_directory('.', path)


# ══════════════════════════════════════════════════════════════════════════
# ADO TEST CASE UPLOAD API
# ══════════════════════════════════════════════════════════════════════════

ADO_UPLOAD_DIR = os.path.join(os.getcwd(), 'ADO_Testcase_Upload', 'uploads')
ADO_SUMMARY_DIR = os.path.join(os.getcwd(), 'ADO_Testcase_Upload', 'summaries')
ALLOWED_ADO_EXTENSIONS = {'.xlsx', '.xlsm', '.xls'}

ado_upload_status = {'in_progress': False, 'result': None, 'error': None}


def _generate_ado_summary_html(timestamp: str, output: str, success_count: int,
                               failed_titles: list, total: int, excel_name: str) -> str:
    """Build an HTML upload-summary page from script output."""
    rows = []
    current_tc = None
    tc_id = None
    suite_id = None
    us_id = None
    status_class = 'success'

    for line in output.splitlines():
        line = line.strip()
        if line.startswith('[') and '/' in line and 'Processing:' in line:
            # Flush previous
            if current_tc:
                row_class = 'fail' if current_tc in failed_titles else 'success'
                rows.append((current_tc, tc_id or '-', suite_id or '-', us_id or '-', row_class))
            # Parse: [1/N] Processing: <title>
            try:
                current_tc = line.split('Processing:')[1].strip()
            except Exception:
                current_tc = line
            tc_id = suite_id = us_id = None
        elif 'Created  ->' in line and 'Work Item ID' in line:
            try:
                tc_id = line.split(':')[-1].strip()
            except Exception:
                pass
        elif 'Added    ->' in line and 'Suite ID' in line:
            try:
                suite_id = line.split(':')[-1].strip()
            except Exception:
                pass
        elif 'Linked   ->' in line and 'User Story' in line:
            try:
                us_id = line.split(':')[-1].strip()
            except Exception:
                pass

    # Flush last
    if current_tc:
        row_class = 'fail' if current_tc in failed_titles else 'success'
        rows.append((current_tc, tc_id or '-', suite_id or '-', us_id or '-', row_class))

    table_rows_html = ''
    for i, (title, wid, sid, usid, rc) in enumerate(rows, start=1):
        icon = '✅' if rc == 'success' else '❌'
        table_rows_html += f'''
        <tr class="{rc}">
            <td>{i}</td>
            <td>{icon}</td>
            <td class="tc-title">{title}</td>
            <td>{wid}</td>
            <td>{sid}</td>
            <td>{usid}</td>
        </tr>'''

    failed_section = ''
    if failed_titles:
        failed_items = ''.join(f'<li>{t}</li>' for t in failed_titles)
        failed_section = f'''
        <div class="failed-section">
            <h3>❌ Failed Test Cases ({len(failed_titles)})</h3>
            <ul>{failed_items}</ul>
        </div>'''

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ADO Upload Summary – {timestamp}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f0f4ff; margin: 0; padding: 20px; color: #333; }}
  .container {{ max-width: 1100px; margin: 0 auto; background: #fff; border-radius: 12px;
                box-shadow: 0 4px 24px rgba(80,80,200,0.12); overflow: hidden; }}
  .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #fff;
             padding: 28px 32px; }}
  .header h1 {{ margin: 0 0 6px; font-size: 1.6rem; }}
  .header p  {{ margin: 0; opacity: 0.9; font-size: 0.95rem; }}
  .summary-bar {{ display: flex; gap: 20px; padding: 20px 32px; background: #f8f9ff;
                  border-bottom: 1px solid #e0e4f0; flex-wrap: wrap; }}
  .stat {{ background: #fff; border-radius: 10px; padding: 14px 22px; text-align: center;
           box-shadow: 0 2px 8px rgba(0,0,0,0.07); min-width: 120px; }}
  .stat .num {{ font-size: 2rem; font-weight: 800; }}
  .stat .lbl {{ font-size: 0.78rem; color: #666; margin-top: 2px; }}
  .stat.green .num {{ color: #10b981; }}
  .stat.red   .num {{ color: #ef4444; }}
  .stat.blue  .num {{ color: #3b82f6; }}
  .content {{ padding: 24px 32px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.88rem; }}
  th {{ background: #667eea; color: #fff; padding: 10px 12px; text-align: left; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #eee; }}
  tr.success {{ background: #f0fdf4; }}
  tr.fail    {{ background: #fff1f2; }}
  .tc-title  {{ font-weight: 600; }}
  .failed-section {{ margin-top: 24px; background: #fff1f2; border-radius: 8px;
                     padding: 16px 20px; border-left: 4px solid #ef4444; }}
  .failed-section h3 {{ margin: 0 0 10px; color: #dc2626; }}
  .failed-section ul {{ margin: 0; padding-left: 20px; line-height: 1.8; }}
  .back-btn {{ display: inline-block; margin-top: 20px; padding: 10px 22px;
               background: linear-gradient(135deg, #667eea, #764ba2); color: #fff;
               text-decoration: none; border-radius: 8px; font-weight: 600; font-size: 0.9rem; }}
  .log-section {{ margin-top: 24px; }}
  .log-section h3 {{ color: #555; margin-bottom: 8px; }}
  pre {{ background: #1e1e2e; color: #cdd6f4; padding: 16px; border-radius: 8px;
         font-size: 0.78rem; overflow-x: auto; white-space: pre-wrap; max-height: 300px;
         overflow-y: auto; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>📋 ADO Test Case Upload Summary</h1>
    <p>File: <strong>{excel_name}</strong> &nbsp;|&nbsp; Uploaded: {timestamp}</p>
  </div>
  <div class="summary-bar">
    <div class="stat blue"><div class="num">{total}</div><div class="lbl">Total</div></div>
    <div class="stat green"><div class="num">{success_count}</div><div class="lbl">Succeeded</div></div>
    <div class="stat red"><div class="num">{len(failed_titles)}</div><div class="lbl">Failed</div></div>
  </div>
  <div class="content">
    <table>
      <thead>
        <tr><th>#</th><th>Status</th><th>Test Case Title</th>
            <th>Work Item ID</th><th>Suite ID</th><th>User Story ID</th></tr>
      </thead>
      <tbody>{table_rows_html}</tbody>
    </table>
    {failed_section}
    <div class="log-section">
      <h3>📄 Full Script Output</h3>
      <pre>{output}</pre>
    </div>
    <a href="/ado-testcase-upload.html" class="back-btn">← Back to Upload</a>
  </div>
</div>
</body>
</html>'''


def _run_ado_upload_thread(save_path, fname, timestamp, plan_id, suite_id):
    """Background thread: run the ADO upload script and store results in ado_upload_status."""
    global ado_upload_status
    try:
        script_path = os.path.join(os.getcwd(), 'ADO_Testcase_Upload', 'upload_testcases_to_ADO.py')
        if not os.path.exists(script_path):
            ado_upload_status = {'in_progress': False, 'result': None,
                                 'error': f'Upload script not found: {script_path}',
                                 'output': '', 'log': ''}
            return

        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'

        cmd = ['python', script_path, '--excel-file', save_path]
        if plan_id and plan_id.isdigit():
            cmd += ['--plan-id', plan_id]
        if suite_id and suite_id.isdigit():
            cmd += ['--suite-id', suite_id]

        print(f'▶️  Running: {" ".join(cmd)}')
        ado_upload_status['log'] = f'Running: {" ".join(cmd)}\n'

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=1800,  # 30 minutes
            env=env
        )
        output = (result.stdout or '') + (result.stderr or '')
        print(f'🖥️  ADO script output:\n{output}')

        # Parse summary line: "X/Y test case(s) succeeded"
        success_count = 0
        total_count = 0
        failed_titles = []
        summary_match = re.search(r'(\d+)/(\d+) test case\(s\) succeeded', output)
        if summary_match:
            success_count = int(summary_match.group(1))
            total_count = int(summary_match.group(2))

        # Extract failed titles block
        in_failed = False
        for line in output.splitlines():
            if line.strip().startswith('Failed ('):
                in_failed = True
                continue
            if in_failed:
                stripped = line.strip()
                if stripped.startswith('- '):
                    failed_titles.append(stripped[2:])
                elif stripped.startswith('='):
                    in_failed = False

        # Build and save summary HTML
        html_content = _generate_ado_summary_html(
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            output=output,
            success_count=success_count,
            failed_titles=failed_titles,
            total=total_count,
            excel_name=fname
        )
        summary_file = f'ado_upload_summary_{timestamp}.html'
        summary_path = os.path.join(ADO_SUMMARY_DIR, summary_file)
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        script_success = result.returncode == 0
        ado_upload_status = {
            'in_progress': False,
            'result': summary_file,
            'error': None if script_success else f'Script exited with code {result.returncode}',
            'success': script_success,
            'message': f'{success_count}/{total_count} test case(s) uploaded successfully.',
            'summary_url': f'/api/ado-summary/{summary_file}',
            'success_count': success_count,
            'failed_count': len(failed_titles),
            'total': total_count,
            'output': output,
            'log': output,
        }

    except subprocess.TimeoutExpired:
        ado_upload_status = {'in_progress': False, 'result': None,
                             'error': 'Upload script timed out (>10 min)',
                             'output': '', 'log': '', 'success': False}
    except Exception as exc:
        print(f'❌ ADO upload thread error: {exc}')
        import traceback
        traceback.print_exc()
        ado_upload_status = {'in_progress': False, 'result': None,
                             'error': str(exc), 'output': '', 'log': '', 'success': False}


@app.route('/api/ado-upload', methods=['POST'])
def ado_upload():
    """Accept an Excel file, start background upload, return immediately so browser doesn't timeout."""
    global ado_upload_status
    try:
        with _ado_upload_lock:
            if ado_upload_status.get('in_progress'):
                return jsonify({'success': False, 'error': 'An upload is already in progress. Please wait.'}), 429
            # Claim the slot immediately inside the lock to prevent races
            ado_upload_status = {'in_progress': True, 'result': None, 'error': None,
                                 'output': '', 'log': 'Upload initializing...\n', 'success': False}

        if 'file' not in request.files:
            ado_upload_status['in_progress'] = False
            return jsonify({'success': False, 'error': 'No file part in the request'}), 400

        uploaded = request.files['file']
        if not uploaded or uploaded.filename == '':
            ado_upload_status['in_progress'] = False
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        fname = secure_filename(uploaded.filename)
        if not fname:
            ado_upload_status['in_progress'] = False
            return jsonify({'success': False, 'error': 'Invalid filename'}), 400

        ext = os.path.splitext(fname)[1].lower()
        if ext not in ALLOWED_ADO_EXTENSIONS:
            ado_upload_status['in_progress'] = False
            return jsonify({'success': False,
                            'error': f'Invalid file type "{ext}". Only .xlsx / .xlsm / .xls are accepted.'}), 400

        os.makedirs(ADO_UPLOAD_DIR, exist_ok=True)
        os.makedirs(ADO_SUMMARY_DIR, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_path = os.path.join(ADO_UPLOAD_DIR, f'{timestamp}_{fname}')
        uploaded.save(save_path)
        print(f'\n📁 ADO upload saved: {save_path}')

        plan_id  = request.form.get('plan_id',  '').strip()
        suite_id = request.form.get('suite_id', '').strip()

        ado_upload_status['log'] = 'Upload started...\n'

        thread = threading.Thread(
            target=_run_ado_upload_thread,
            args=(save_path, fname, timestamp, plan_id, suite_id),
            daemon=True
        )
        thread.start()

        return jsonify({'success': True, 'started': True,
                        'message': 'Upload started. Poll /api/ado-upload-status for progress.'})

    except Exception as exc:
        ado_upload_status = {'in_progress': False, 'result': None, 'error': str(exc),
                             'output': '', 'log': '', 'success': False}
        print(f'❌ ADO upload error: {exc}')
        return jsonify({'success': False, 'error': str(exc)}), 500


@app.route('/api/ado-upload-status', methods=['GET'])
def ado_upload_status_check():
    """Return current ADO upload status."""
    return jsonify(ado_upload_status)


@app.errorhandler(404)
def handle_404(e):
    """Return JSON 404 for API routes, HTML 404 for everything else."""
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': f'API endpoint not found: {request.path}'}), 404
    return str(e), 404


@app.errorhandler(500)
def handle_500(e):
    """Return JSON 500 for API routes, HTML 500 for everything else."""
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': f'Internal server error: {str(e)}'}), 500
    return str(e), 500


@app.errorhandler(405)
def handle_405(e):
    """Return JSON 405 for API routes."""
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': f'Method not allowed: {request.method} {request.path}'}), 405
    return str(e), 405


@app.route('/api/ado-summary/<filename>', methods=['GET'])
def serve_ado_summary(filename):
    """Serve a previously generated ADO upload summary HTML file."""
    safe = secure_filename(filename)
    if safe != filename or not safe.startswith('ado_upload_summary_') or not safe.endswith('.html'):
        return jsonify({'error': 'Invalid summary filename'}), 400
    summary_dir = os.path.join(os.getcwd(), 'ADO_Testcase_Upload', 'summaries')
    full_path = os.path.join(summary_dir, safe)
    if not os.path.exists(full_path):
        return jsonify({'error': 'Summary file not found'}), 404
    return send_file(full_path, mimetype='text/html')


if __name__ == '__main__':
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')
    print("\n" + "="*80)
    print("🌐 MyISP Internal Tools Server")
    print("="*80)
    print("\n✓ Server is starting...")
    print("✓ Access at: http://localhost:8000")
    print("✓ Team access at: http://192.168.1.2:8000")
    print("\n⚠️  Press Ctrl+C to stop the server\n")
    print("="*80 + "\n")
    
    app.run(host='0.0.0.0', port=8000, debug=False)
